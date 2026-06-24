import calendar
import base64
import binascii
import hashlib
import json
import logging
import math
import re
import secrets
import string
from urllib.parse import quote
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional
import requests

from django.http import HttpResponse, JsonResponse, FileResponse
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.signing import BadSignature, SignatureExpired, TimestampSigner
from django.views.decorators.http import require_http_methods
from django.db import DatabaseError, IntegrityError, OperationalError, transaction
from django.db.models import Q, Min, Max
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_time
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from apps.backend.common_auth.models import User
from apps.backend.brand.models import SiteBrandSettings
from apps.backend.core_platform.assistant_registry import build_assistant_training_context, get_assistant_product_profile
from apps.backend.products.models import Product
from core.models import BillingProfile, InvoiceSellerProfile, Organization, OrganizationSettings, UserProductAccess, UserProfile, Subscription as OrgSubscription, BusinessAutopilotChatHistory, log_admin_activity
from core.email_utils import send_templated_email
from core.notification_emails import mark_email_verified
from apps.backend.hrm.services.face_recognition_service import (
    FaceRecognitionUnavailable,
    FaceRecognitionValidationError,
    compress_uploaded_photo,
    encrypt_embeddings,
    generate_embedding,
    verify_employee_face,
)
from .site_admin_ai import (
    build_site_admin_instruction_context,
    get_site_admin_enabled_modules,
    get_site_admin_module_hints,
)

from .models import (
    Module,
    OrganizationModule,
    OrganizationUser,
    OrganizationEmployeeRole,
    OrganizationDepartment,
    CrmContact,
    CrmDeal,
    CrmLead,
    CrmLeadModification,
    CrmLeadProposalDocument,
    CrmMeeting,
    CrmSalesOrder,
    AccountsWorkspace,
    EmployeeSalaryHistory,
    AttendanceEntry,
    AttendanceGeoSetting,
    FaceRecognitionSetting,
    EmployeeFaceProfile,
    AttendancePhotoProof,
    PayrollEntry,
    PayrollSettings,
    Payslip,
    SalaryStructure,
    Subscription,
    SubscriptionCategory,
    SubscriptionSubCategory,
    BusinessAutopilotUserCrmReassignmentSnapshot,
    QuickEstimate,
    QuickEstimateHistory,
    QuickEstimateItem,
    QuickEstimateSequence,
    SiteAdminChatState,
)


MODULE_PATHS = {
    "crm": "/crm",
    "hrm": "/hrm",
    "projects": "/projects",
    "accounts": "/accounts",
    "subscriptions": "/subscriptions",
    "ticketing": "/ticketing",
    "stocks": "/stocks",
}

DEFAULT_ERP_MODULES = [
    {"name": "CRM", "slug": "crm", "sort_order": 1},
    {"name": "HR Management", "slug": "hrm", "sort_order": 2},
    {"name": "Project Management", "slug": "projects", "sort_order": 3},
    {"name": "Accounts", "slug": "accounts", "sort_order": 4},
    {"name": "Subscriptions", "slug": "subscriptions", "sort_order": 5},
    {"name": "Ticketing System", "slug": "ticketing", "sort_order": 6},
    {"name": "Inventory", "slug": "stocks", "sort_order": 7},
]

BUSINESS_AUTOPILOT_USER_TYPE_DEFAULTS = [
    {
        "key": "crm_user",
        "label": "CRM User",
        "description": "CRM-focused sales and follow-up user",
        "monthly_price_inr": 550,
        "monthly_price_usd": 7,
        "allowed_modules": ["dashboard", "inbox", "crm", "users", "profile"],
    },
    {
        "key": "hrm_user",
        "label": "HRM User",
        "description": "HR operations user for employees, attendance and payroll",
        "monthly_price_inr": 50,
        "monthly_price_usd": 1,
        "allowed_modules": ["dashboard", "hr", "users", "profile"],
    },
    {
        "key": "full_access_user",
        "label": "Full Access User",
        "description": "Full Business Autopilot access across enabled modules",
        "monthly_price_inr": 650,
        "monthly_price_usd": 8,
        "allowed_modules": ["dashboard", "inbox", "crm", "hr", "projects", "accounts", "subscriptions", "ticketing", "stocks", "users", "billing", "plans", "profile"],
    },
]
BUSINESS_AUTOPILOT_USER_TYPE_MAP = {
    row["key"]: row for row in BUSINESS_AUTOPILOT_USER_TYPE_DEFAULTS
}

ERP_EMPLOYEE_ROLES = {"company_admin", "org_user", "hr_view"}
DELETE_PROTECTED_PROFILE_ROLES = {"org_admin", "owner", "superadmin", "super_admin"}
ACCOUNTS_ALLOWED_ROOT_KEYS = {"customers", "vendors", "itemMasters", "gstTemplates", "billingTemplates", "estimates", "invoices", "quickEstimateContacts"}
QUICK_ESTIMATE_HEADER_TEXT_MAX_LENGTH = 200
ERP_MODULE_SLUG_SET = set(MODULE_PATHS.keys())
BUSINESS_AUTOPILOT_PRODUCT_SLUG = "business-autopilot-erp"
BUSINESS_AUTOPILOT_PRODUCT_SLUG_ALIASES = {"business-autopilot-erp", "business-autopilot"}
OPENAI_CHAT_COMPLETIONS_URL = "https://api.openai.com/v1/chat/completions"
OPENAI_AUDIO_TRANSCRIPTIONS_URL = "https://api.openai.com/v1/audio/transcriptions"
OPENAI_AUDIO_SPEECH_URL = "https://api.openai.com/v1/audio/speech"
DEFAULT_BA_OPENAI_AGENT_NAME = "Work Zilla AI Assistant"
PUBLIC_LOGIN_URL = "https://getworkzilla.com/auth/login/"
SUBSCRIPTION_STATUS_OPTIONS = ("Active", "Expired", "Cancelled")
ROLE_ACCESS_SECTION_KEYS = {
    "dashboard",
    "inbox",
    "crm",
    "hr",
    "projects",
    "accounts",
    "subscriptions",
    "ticketing",
    "stocks",
    "users",
    "billing",
    "plans",
    "profile",
}
ROLE_ACCESS_LEVELS = {
    "No Access",
    "View",
    "View and Edit",
    "Create, View and Edit Own",
    "Create, View and Edit All",
    "Full Access",
}
ROLE_ACCESS_LEVEL_ALIASES = {
    "No Access": "No Access",
    "View": "View",
    "Create/Edit": "Create, View and Edit Own",
    "View and Edit": "View and Edit",
    "Create, View and Edit": "Create, View and Edit Own",
    "Create, View and Edit Own": "Create, View and Edit Own",
    "Create, View and Edit All": "Create, View and Edit All",
    "Full Access": "Full Access",
}
logger = logging.getLogger(__name__)
MIN_GEO_RADIUS_METERS = 20
MAX_GPS_ACCURACY_METERS = 200

TEMP_PASSWORD_ALPHABET = string.ascii_letters + string.digits
TEMP_PASSWORD_LENGTH = 10
_BA_UNICODE_PDF_FONT = None
SITE_ADMIN_RESET_COMMANDS = {"cancel", "reset", "stop"}
SITE_ADMIN_QUICK_ESTIMATE_HINTS = frozenset(get_site_admin_module_hints("quick_estimate"))


def _ba_hex_to_rgb(value, default="#22c55e"):
    raw = str(value or default).strip()
    if not raw.startswith("#"):
        raw = default
    raw = raw.lstrip("#")
    if len(raw) == 3:
        raw = "".join(ch * 2 for ch in raw)
    if len(raw) != 6:
        raw = default.lstrip("#")
    try:
        return tuple(int(raw[index:index + 2], 16) / 255 for index in (0, 2, 4))
    except ValueError:
        fallback = default.lstrip("#")
        return tuple(int(fallback[index:index + 2], 16) / 255 for index in (0, 2, 4))


def _ba_pdf_image_from_data_url(data_url):
    raw = str(data_url or "").strip()
    if not raw.startswith("data:image/") or "," not in raw:
        return None
    _, encoded = raw.split(",", 1)
    try:
        return ImageReader(BytesIO(base64.b64decode(encoded)))
    except (ValueError, TypeError, binascii.Error):
        return None


def _ba_pdf_image_from_file(file_field):
    if not file_field:
        return None
    try:
        if getattr(file_field, "path", ""):
            return ImageReader(file_field.path)
    except Exception:
        pass
    try:
        file_field.open("rb")
        try:
            return ImageReader(BytesIO(file_field.read()))
        finally:
            file_field.close()
    except Exception:
        return None


BA_HEX_COLOR_RE = re.compile(r"^#(?:[0-9a-fA-F]{6})$")


def _ba_normalize_hex_color(value, fallback=""):
    color = str(value or "").strip()
    if not color:
        return fallback
    if not BA_HEX_COLOR_RE.match(color):
        return fallback
    return color.lower()


def _ba_effective_theme_primary_hex(org: Organization, fallback="#1f6f8b"):
    theme_primary = ""
    try:
        settings_obj = OrganizationSettings.objects.filter(organization=org).only("theme_primary_color").first()
        theme_primary = _ba_normalize_hex_color(getattr(settings_obj, "theme_primary_color", ""), "")
    except Exception:
        theme_primary = ""

    if not theme_primary:
        try:
            global_theme = SiteBrandSettings.get_active()
            theme_primary = _ba_normalize_hex_color(getattr(global_theme, "primary_color", ""), "")
        except Exception:
            theme_primary = ""

    return theme_primary or _ba_normalize_hex_color(fallback, "#1f6f8b") or "#1f6f8b"


def _ba_hex_to_rgb01(hex_color: str):
    raw = _ba_normalize_hex_color(hex_color, "")
    if not raw:
        return (0.12, 0.44, 0.55)
    value = raw.lstrip("#")
    r = int(value[0:2], 16) / 255.0
    g = int(value[2:4], 16) / 255.0
    b = int(value[4:6], 16) / 255.0
    return (r, g, b)


def _ba_blend_rgb(rgb, other_rgb, other_weight=0.0):
    w = float(other_weight or 0.0)
    w = 0.0 if w < 0 else 1.0 if w > 1 else w
    r, g, b = rgb
    or_, og, ob = other_rgb
    return (
        r * (1 - w) + or_ * w,
        g * (1 - w) + og * w,
        b * (1 - w) + ob * w,
    )


def _ba_resolve_org_logo_image(org, current_user=None):
    if not org:
        return None
    seen_profile_ids = set()
    profile_candidates = []

    if current_user and getattr(current_user, "is_authenticated", False):
        current_profile = (
            UserProfile.objects
            .filter(user=current_user, organization=org)
            .only("id", "profile_photo")
            .first()
        )
        if current_profile and current_profile.pk:
            profile_candidates.append(current_profile)
            seen_profile_ids.add(int(current_profile.pk))

    admin_profiles = (
        UserProfile.objects
        .filter(organization=org, profile_photo__isnull=False)
        .exclude(profile_photo="")
        .only("id", "profile_photo")
        .order_by("id")[:10]
    )
    for profile in admin_profiles:
        if not profile or not profile.pk:
            continue
        profile_id = int(profile.pk)
        if profile_id in seen_profile_ids:
            continue
        profile_candidates.append(profile)
        seen_profile_ids.add(profile_id)

    for profile in profile_candidates:
        logo_image = _ba_pdf_image_from_file(getattr(profile, "profile_photo", None))
        if logo_image:
            return logo_image

    try:
        global_brand = SiteBrandSettings.get_active()
        logo_image = _ba_pdf_image_from_file(getattr(global_brand, "logo", None))
        if logo_image:
            return logo_image
    except Exception:
        pass

    return None


def _ba_draw_pdf_logo(pdf, image, x, top_y, max_width, max_height):
    if not image:
        return 0
    try:
        image_width, image_height = image.getSize()
    except Exception:
        return 0
    if not image_width or not image_height:
        return 0
    scale = min(max_width / image_width, max_height / image_height)
    draw_width = image_width * scale
    draw_height = image_height * scale
    pdf.drawImage(
        image,
        x,
        top_y - draw_height,
        width=draw_width,
        height=draw_height,
        mask="auto",
        preserveAspectRatio=True,
        anchor="nw",
    )
    return draw_height


def _ba_wrap_pdf_text(pdf, text, max_width, font_name="Helvetica", font_size=9):
    words = str(text or "").split()
    if not words:
        return [""]
    lines = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if pdf.stringWidth(candidate, font_name, font_size) <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def _ba_get_unicode_pdf_font():
    global _BA_UNICODE_PDF_FONT
    if _BA_UNICODE_PDF_FONT is not None:
        return _BA_UNICODE_PDF_FONT
    font_candidates = [
        ("ArialUnicodeWZ", "/Library/Fonts/Arial Unicode.ttf"),
        ("ArialUnicodeWZ", "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        ("TimesNewRomanWZ", "/System/Library/Fonts/Supplemental/Times New Roman.ttf"),
        ("ArialWZ", "/System/Library/Fonts/Supplemental/Arial.ttf"),
    ]
    for font_name, path in font_candidates:
        try:
            pdfmetrics.getFont(font_name)
            _BA_UNICODE_PDF_FONT = font_name
            return _BA_UNICODE_PDF_FONT
        except KeyError:
            pass
        try:
            pdfmetrics.registerFont(TTFont(font_name, path))
            _BA_UNICODE_PDF_FONT = font_name
            return _BA_UNICODE_PDF_FONT
        except Exception:
            continue
    _BA_UNICODE_PDF_FONT = ""
    return _BA_UNICODE_PDF_FONT



def _get_business_autopilot_product():
    return Product.objects.filter(slug=BUSINESS_AUTOPILOT_PRODUCT_SLUG).first()


@require_http_methods(["POST"])
def crm_activity_log(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)
    action = str(payload.get("action") or "").strip()
    details = str(payload.get("details") or "").strip()
    if not action:
        return JsonResponse({"detail": "action_required"}, status=400)
    log_admin_activity(
        request.user,
        action[:120],
        details[:500],
        product_slug=BUSINESS_AUTOPILOT_PRODUCT_SLUG,
        request=request,
    )
    return JsonResponse({"logged": True})


def _generate_temp_login_password(length: int = TEMP_PASSWORD_LENGTH):
    safe_length = max(8, int(length or TEMP_PASSWORD_LENGTH))
    password = "".join(secrets.choice(TEMP_PASSWORD_ALPHABET) for _ in range(safe_length - 2))
    # Ensure at least one digit and one uppercase character.
    password += secrets.choice(string.digits)
    password += secrets.choice(string.ascii_uppercase)
    return "".join(secrets.SystemRandom().sample(password, len(password)))


def _get_business_autopilot_permission(role: str):
    normalized_role = str(role or "").strip().lower()
    if normalized_role == "company_admin":
        return UserProductAccess.PERMISSION_FULL
    if normalized_role == "org_user":
        return UserProductAccess.PERMISSION_EDIT
    return UserProductAccess.PERMISSION_VIEW


def _grant_business_autopilot_access(user: User, granted_by: User, role: str):
    product = _get_business_autopilot_product()
    if not product:
        return
    UserProductAccess.objects.update_or_create(
        user=user,
        product=product,
        defaults={
            "permission": _get_business_autopilot_permission(role),
            "granted_by": granted_by,
        },
    )


def _revoke_business_autopilot_access(user: User):
    product = _get_business_autopilot_product()
    if not product:
        return
    UserProductAccess.objects.filter(user=user, product=product).delete()


def _get_user_granted_products(user: User):
    rows = (
        UserProductAccess.objects
        .filter(user=user)
        .select_related("product")
        .order_by("product__name", "product__slug")
    )
    return [
        {
            "slug": row.product.slug,
            "name": row.product.name or row.product.slug,
            "permission": row.permission,
        }
        for row in rows
        if row.product_id
    ]


def _sync_business_autopilot_membership_access(org: Organization, granted_by: Optional[User] = None):
    product = _get_business_autopilot_product()
    if not org or not product:
        return
    _ensure_org_admin_memberships(org)
    memberships = list(
        OrganizationUser.objects
        .filter(organization=org, role__in=ERP_EMPLOYEE_ROLES, is_deleted=False)
        .select_related("user")
    )
    active_user_ids = set()
    for membership in memberships:
        if _normalize_membership_status(membership) == OrganizationUser.STATUS_ACTIVE and membership.user_id:
            active_user_ids.add(membership.user_id)
            UserProductAccess.objects.update_or_create(
                user=membership.user,
                product=product,
                defaults={
                    "permission": _get_business_autopilot_permission(membership.role),
                    "granted_by": granted_by,
                },
            )
    if memberships:
        UserProductAccess.objects.filter(product=product, user_id__in=[row.user_id for row in memberships if row.user_id]).exclude(
            user_id__in=active_user_ids
        ).delete()


def _resolve_org(user: User, request=None):
    if not user or not getattr(user, "is_authenticated", False):
        return None

    profile = UserProfile.objects.filter(user=user).select_related("organization").first()
    if request is not None:
        session_org_id = str(request.session.get("active_org_id") or "").strip()
        if session_org_id.isdigit():
            session_org_int = int(session_org_id)
            session_org = Organization.objects.filter(id=session_org_int).first()
            if session_org:
                if user.is_superuser or user.is_staff:
                    return session_org
                if profile and profile.organization_id == session_org_int:
                    return profile.organization or session_org
                if Organization.objects.filter(id=session_org_int, owner=user).exists():
                    return session_org
                if OrganizationUser.objects.filter(
                    organization_id=session_org_int,
                    user=user,
                    is_deleted=False,
                ).exists():
                    return session_org

    active_membership = (
        OrganizationUser.objects
        .filter(user=user, is_active=True, is_deleted=False)
        .select_related("organization")
        .order_by("-updated_at", "-id")
        .first()
    )
    if active_membership and active_membership.organization:
        return active_membership.organization
    if profile and profile.organization:
        return profile.organization
    any_membership = (
        OrganizationUser.objects
        .filter(user=user, is_deleted=False)
        .select_related("organization")
        .order_by("-is_active", "-updated_at", "-id")
        .first()
    )
    if any_membership and any_membership.organization:
        return any_membership.organization
    legacy_membership = (
        OrganizationUser.objects
        .filter(user=user)
        .select_related("organization")
        .order_by("-is_active", "is_deleted", "-updated_at", "-id")
        .first()
    )
    if legacy_membership and legacy_membership.organization:
        return legacy_membership.organization
    return Organization.objects.filter(owner=user).first()


def _normalize_admin_role(value):
    raw_value = str(value or "").strip().lower()
    compact_value = re.sub(r"[^a-z0-9]+", "", raw_value)
    alias_map = {
        "companyadmin": "company_admin",
        "orgadmin": "org_admin",
        "organizationadmin": "org_admin",
        "companyowner": "owner",
        "orgowner": "owner",
        "superadmin": "superadmin",
        "superuser": "superadmin",
        "admin": "company_admin",
        "owner": "owner",
    }
    normalized = alias_map.get(compact_value)
    if normalized:
        return normalized
    return raw_value.replace("-", "_").replace(" ", "_")


def _ensure_org_admin_memberships(org: Organization):
    if not org:
        return
    candidate_user_ids = set()
    if getattr(org, "owner_id", None):
        candidate_user_ids.add(org.owner_id)

    admin_profiles = (
        UserProfile.objects
        .filter(organization=org)
        .only("user_id", "role")
    )
    for profile in admin_profiles:
        normalized_role = _normalize_admin_role(getattr(profile, "role", ""))
        if normalized_role in {"company_admin", "org_admin", "owner"} and profile.user_id:
            candidate_user_ids.add(profile.user_id)

    for user_id in candidate_user_ids:
        membership, created = OrganizationUser.objects.get_or_create(
            organization=org,
            user_id=user_id,
            defaults={
                "role": "company_admin",
                "is_active": True,
                "is_deleted": False,
                "deleted_at": None,
            },
        )
        if created:
            continue
        update_fields = []
        if _normalize_admin_role(membership.role) != "company_admin":
            membership.role = "company_admin"
            update_fields.append("role")
        if membership.status != OrganizationUser.STATUS_ACTIVE:
            membership.status = OrganizationUser.STATUS_ACTIVE
            membership.status_changed_at = timezone.now()
            membership.resigned_at = None
            membership.resigned_by = None
            update_fields.extend(["status", "status_changed_at", "resigned_at", "resigned_by"])
        if not membership.is_active:
            membership.is_active = True
            update_fields.append("is_active")
        if membership.is_deleted:
            membership.is_deleted = False
            membership.deleted_at = None
            update_fields.extend(["is_deleted", "deleted_at"])
        if not membership.user.is_active:
            membership.user.is_active = True
            membership.user.save(update_fields=["is_active"])
        if update_fields:
            membership.save(update_fields=[*update_fields, "updated_at"])


def _can_manage_modules(user: User):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    profile = UserProfile.objects.filter(user=user).only("role").first()
    if not profile:
        return False
    normalized_role = _normalize_admin_role(profile.role)
    return normalized_role in {"company_admin", "org_admin", "owner", "org_user", "superadmin", "super_admin"}


def _can_manage_users(user: User, org: Organization = None):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    profile = UserProfile.objects.filter(user=user).only("role").first()
    profile_role = _normalize_admin_role(getattr(profile, "role", ""))
    if profile_role in {"company_admin", "org_admin", "owner", "superadmin", "super_admin"}:
        return True
    if org:
        membership = _get_org_membership(user, org)
        if membership:
            return _normalize_admin_role(membership.role) in {"company_admin", "org_admin", "owner"}
    return False


def _can_manage_openai(user: User, org: Organization = None):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    profile = UserProfile.objects.filter(user=user).only("role").first()
    profile_role = _normalize_admin_role(getattr(profile, "role", ""))
    if profile_role in {"company_admin", "org_admin", "owner", "superadmin", "super_admin"}:
        return True
    membership = _get_org_membership(user, org)
    if not membership:
        return False
    membership_role = _normalize_admin_role(getattr(membership, "role", ""))
    return membership_role in {"company_admin", "org_admin", "owner", "superadmin", "super_admin"}


def _mask_secret(value: str):
    secret = str(value or "").strip()
    if len(secret) <= 8:
        return "*" * len(secret)
    return f"{secret[:4]}{'*' * max(0, len(secret) - 8)}{secret[-4:]}"


def _normalize_role_access_map(payload):
    raw_map = payload if isinstance(payload, dict) else {}
    normalized = {}
    user_sub_section_keys = ("employee", "clients", "vendors")
    section_feature_keys = {
        "dashboard": ("overview", "widgets", "analytics"),
        "inbox": ("tickets", "messages", "followups"),
        "crm": ("leads", "contacts", "deals", "sales_orders", "meetings", "reports"),
        "hr": ("employees", "attendance", "payroll"),
        "projects": ("projects", "tasks", "timeline"),
        "accounts": ("customers", "vendors", "invoices"),
        "subscriptions": ("plans", "renewals", "alerts"),
        "ticketing": ("tickets", "categories", "assignments"),
        "stocks": ("assets", "stock_entries", "reports"),
        "users": ("directory", "roles", "permissions"),
        "billing": ("summary", "transactions", "addons"),
        "plans": ("plan_list", "upgrades", "history"),
        "profile": ("account", "security", "activity"),
    }
    for raw_key, raw_record in raw_map.items():
        key = str(raw_key or "").strip()
        if not key or len(key) > 200:
            continue
        record = raw_record if isinstance(raw_record, dict) else {}
        sections = record.get("sections") if isinstance(record.get("sections"), dict) else {}
        raw_section_features = (
            record.get("section_features")
            if isinstance(record.get("section_features"), dict)
            else {}
        )
        raw_user_sub_sections = (
            record.get("user_sub_sections")
            if isinstance(record.get("user_sub_sections"), dict)
            else {}
        )
        normalized_sections = {}
        for section_key in ROLE_ACCESS_SECTION_KEYS:
            raw_level = str(sections.get(section_key) or "No Access").strip()
            normalized_sections[section_key] = ROLE_ACCESS_LEVEL_ALIASES.get(raw_level, "No Access")
        normalized_section_features = {}
        for section_key, feature_keys in section_feature_keys.items():
            raw_feature_record = (
                raw_section_features.get(section_key)
                if isinstance(raw_section_features.get(section_key), dict)
                else {}
            )
            normalized_section_features[section_key] = {
                feature_key: bool(raw_feature_record.get(feature_key, True))
                for feature_key in feature_keys
            }
        normalized_user_sub_sections = {}
        for sub_key in user_sub_section_keys:
            raw_sub_record = (
                raw_user_sub_sections.get(sub_key)
                if isinstance(raw_user_sub_sections.get(sub_key), dict)
                else {}
            )
            raw_sub_access_level = str(raw_sub_record.get("access_level") or "No Access").strip()
            access_level = ROLE_ACCESS_LEVEL_ALIASES.get(raw_sub_access_level, "No Access")
            enabled = bool(raw_sub_record.get("enabled")) and access_level != "No Access"
            normalized_user_sub_sections[sub_key] = {
                "enabled": enabled,
                "access_level": access_level if enabled else "No Access",
            }
        normalized[key] = {
            "sections": normalized_sections,
            "section_features": normalized_section_features,
            "user_sub_sections": normalized_user_sub_sections,
            "can_export": bool(record.get("can_export")),
            "can_delete": bool(record.get("can_delete")),
            "attendance_self_service": bool(record.get("attendance_self_service")),
            "remarks": str(record.get("remarks") or "").strip()[:500],
        }
    return normalized


def _derive_yearly_from_monthly(value):
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return 0.0
    if amount < 0:
        return 0.0
    return round(amount * 10, 2)


def _normalize_business_autopilot_user_types(config):
    rows = config if isinstance(config, list) else []
    normalized = []
    seen = set()
    for default_row in BUSINESS_AUTOPILOT_USER_TYPE_DEFAULTS:
        key = default_row["key"]
        candidate = next(
            (
                row for row in rows
                if isinstance(row, dict) and str(row.get("key") or "").strip().lower() == key
            ),
            {},
        )
        allowed_modules = [
            section
            for section in (
                candidate.get("allowed_modules")
                if isinstance(candidate.get("allowed_modules"), list)
                else default_row.get("allowed_modules", [])
            )
            if str(section or "").strip().lower() in ROLE_ACCESS_SECTION_KEYS
        ]
        if not allowed_modules:
            allowed_modules = list(default_row.get("allowed_modules", []))
        monthly_inr = candidate.get("monthly_price_inr", default_row.get("monthly_price_inr", 0))
        monthly_usd = candidate.get("monthly_price_usd", default_row.get("monthly_price_usd", 0))
        normalized_row = {
            "key": key,
            "label": str(candidate.get("label") or default_row.get("label") or key).strip(),
            "description": str(candidate.get("description") or default_row.get("description") or "").strip(),
            "monthly_price_inr": round(float(monthly_inr or 0), 2),
            "yearly_price_inr": round(float(candidate.get("yearly_price_inr") or _derive_yearly_from_monthly(monthly_inr)), 2),
            "monthly_price_usd": round(float(monthly_usd or 0), 2),
            "yearly_price_usd": round(float(candidate.get("yearly_price_usd") or _derive_yearly_from_monthly(monthly_usd)), 2),
            "allowed_modules": list(dict.fromkeys(allowed_modules)),
        }
        normalized.append(normalized_row)
        seen.add(key)
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = str(row.get("key") or "").strip().lower()
        if not key or key in seen:
            continue
        allowed_modules = [
            section for section in (row.get("allowed_modules") if isinstance(row.get("allowed_modules"), list) else [])
            if str(section or "").strip().lower() in ROLE_ACCESS_SECTION_KEYS
        ]
        normalized.append(
            {
                "key": key,
                "label": str(row.get("label") or key).strip(),
                "description": str(row.get("description") or "").strip(),
                "monthly_price_inr": round(float(row.get("monthly_price_inr") or 0), 2),
                "yearly_price_inr": round(float(row.get("yearly_price_inr") or _derive_yearly_from_monthly(row.get("monthly_price_inr"))), 2),
                "monthly_price_usd": round(float(row.get("monthly_price_usd") or 0), 2),
                "yearly_price_usd": round(float(row.get("yearly_price_usd") or _derive_yearly_from_monthly(row.get("monthly_price_usd"))), 2),
                "allowed_modules": list(dict.fromkeys(allowed_modules)),
            }
        )
    return normalized


def _get_plan_user_type_config(plan):
    features = dict(getattr(plan, "features", {}) or {})
    return _normalize_business_autopilot_user_types(features.get("business_autopilot_user_types"))


def _get_subscription_user_types(subscription):
    plan = getattr(subscription, "plan", None)
    if not subscription or not plan:
        return _normalize_business_autopilot_user_types(None)
    return _get_plan_user_type_config(plan)


def _user_type_allowed_sections(user_type, *, org=None, plan=None):
    config_rows = _get_subscription_user_types(_get_active_erp_subscription(org)) if org is not None else _get_plan_user_type_config(plan)
    normalized_key = _normalize_user_type_key(user_type)
    row = next((item for item in config_rows if item["key"] == normalized_key), None)
    if not row:
        return set()
    return {str(section or "").strip().lower() for section in row.get("allowed_modules", [])}


def _normalize_user_type_key(value):
    raw = str(value or "").strip().lower()
    if raw in BUSINESS_AUTOPILOT_USER_TYPE_MAP:
        return raw
    if raw in {"crm", "crmuser"}:
        return "crm_user"
    if raw in {"hrm", "hrmuser", "hr"}:
        return "hrm_user"
    if raw in {"full", "fullaccess", "full_access"}:
        return "full_access_user"
    return "full_access_user"


def _normalize_user_type_counts(value, *, config_rows=None):
    config_rows = config_rows or _normalize_business_autopilot_user_types(None)
    allowed_keys = {row["key"] for row in config_rows}
    source = value if isinstance(value, dict) else {}
    normalized = {}
    for key in allowed_keys:
        try:
            count = int(source.get(key) or 0)
        except (TypeError, ValueError):
            count = 0
        normalized[key] = max(0, count)
    return normalized


def _empty_user_type_counts(config_rows=None):
    config_rows = config_rows or _normalize_business_autopilot_user_types(None)
    return {row["key"]: 0 for row in config_rows}


def _get_org_subscription_user_type_seats(org):
    active_sub = _get_active_erp_subscription(org)
    config_rows = _get_subscription_user_types(active_sub)
    if not active_sub or not getattr(active_sub, "plan", None):
        return {
            "config": config_rows,
            "counts": _empty_user_type_counts(config_rows),
            "total": 0,
        }
    counts = _normalize_user_type_counts(getattr(active_sub, "user_type_counts", {}), config_rows=config_rows)
    return {
        "config": config_rows,
        "counts": counts,
        "total": sum(counts.values()),
    }


def _get_org_user_type_usage(org, *, memberships=None):
    config_rows = _get_org_subscription_user_type_seats(org)["config"]
    counts = _empty_user_type_counts(config_rows)
    rows = memberships if isinstance(memberships, list) else _list_org_user_memberships(org)
    for membership in rows:
        if _normalize_membership_status(membership) != OrganizationUser.STATUS_ACTIVE:
            continue
        user_type = _normalize_user_type_key(getattr(membership, "user_type", ""))
        counts[user_type] = counts.get(user_type, 0) + 1
    return counts


def _serialize_user_type_pricing_rows(config_rows, *, seat_counts=None, active_counts=None):
    seat_counts = seat_counts or _empty_user_type_counts(config_rows)
    active_counts = active_counts or _empty_user_type_counts(config_rows)
    rows = []
    for row in config_rows:
        key = row["key"]
        seats = max(0, int(seat_counts.get(key) or 0))
        active = max(0, int(active_counts.get(key) or 0))
        rows.append(
            {
                **row,
                "seat_count": seats,
                "active_count": active,
                "remaining_count": max(0, seats - active),
            }
        )
    return rows


def _crm_resolve_role_access_record(role_access_map, profile_role, employee_role, user_type=""):
    safe_map = role_access_map if isinstance(role_access_map, dict) else {}
    normalized_profile_role = _normalize_admin_role(profile_role)
    normalized_employee_role = _normalize_admin_role(employee_role)
    normalized_user_type = _normalize_user_type_key(user_type)
    entries = [(key, value) for key, value in safe_map.items() if isinstance(value, dict)]

    if normalized_employee_role:
        for raw_key, value in entries:
            composite_key = str(raw_key or "").strip()
            key_without_user_type, _, raw_user_type = composite_key.partition("__")
            scope, raw_role = (key_without_user_type.split(":", 1) + [""])[:2]
            if (
                scope == "employee_role"
                and _normalize_admin_role(raw_role) == normalized_employee_role
                and (
                    (raw_user_type and _normalize_user_type_key(raw_user_type) == normalized_user_type)
                    or not raw_user_type
                )
            ):
                return value

    if normalized_profile_role:
        for raw_key, value in entries:
            composite_key = str(raw_key or "").strip()
            key_without_user_type, _, raw_user_type = composite_key.partition("__")
            scope, raw_role = (key_without_user_type.split(":", 1) + [""])[:2]
            if (
                scope == "system"
                and _normalize_admin_role(raw_role) == normalized_profile_role
                and (
                    (raw_user_type and _normalize_user_type_key(raw_user_type) == normalized_user_type)
                    or not raw_user_type
                )
            ):
                return value

    return None


def _crm_section_access_level(user: User, org: Organization, section_key: str = "crm"):
    if _crm_is_admin(user, org):
        return "Full Access"
    if not user or not user.is_authenticated or not org:
        return "No Access"
    settings_obj = OrganizationSettings.objects.filter(organization=org).only("business_autopilot_role_access_map").first()
    role_access_map = _normalize_role_access_map(getattr(settings_obj, "business_autopilot_role_access_map", {}) or {})
    profile = UserProfile.objects.filter(user=user).only("role").first()
    membership = _get_org_membership(user, org)
    role_access_record = _crm_resolve_role_access_record(
        role_access_map,
        getattr(profile, "role", ""),
        getattr(membership, "employee_role", ""),
        getattr(membership, "user_type", ""),
    )
    if not role_access_record:
        return "No Access"
    allowed_sections = _user_type_allowed_sections(getattr(membership, "user_type", ""), org=org)
    if section_key not in allowed_sections:
        return "No Access"
    sections = role_access_record.get("sections") if isinstance(role_access_record.get("sections"), dict) else {}
    return ROLE_ACCESS_LEVEL_ALIASES.get(str(sections.get(section_key) or "No Access").strip(), "No Access")


def _crm_has_view_access(user: User, org: Organization):
    access_level = _crm_section_access_level(user, org, "crm")
    return access_level in {"View", "View and Edit", "Create, View and Edit Own", "Create, View and Edit All", "Full Access"}


def _crm_has_edit_access(user: User, org: Organization):
    access_level = _crm_section_access_level(user, org, "crm")
    return access_level in {"View and Edit", "Create, View and Edit Own", "Create, View and Edit All", "Full Access"}


def _crm_has_create_access(user: User, org: Organization):
    access_level = _crm_section_access_level(user, org, "crm")
    return access_level in {"Create, View and Edit Own", "Create, View and Edit All", "Full Access"}


def _crm_has_unrestricted_row_access(user: User, org: Organization):
    # "All" scope must be explicit. Default to own-only access for "View and Edit"
    # and "Create, View and Edit Own". Only these grant full org-wide visibility:
    # - Create, View and Edit All
    # - Full Access
    return _crm_section_access_level(user, org, "crm") in {"Create, View and Edit All", "Full Access"}


def _crm_has_full_access(user: User, org: Organization):
    return _crm_section_access_level(user, org, "crm") == "Full Access"


def _crm_business_autopilot_permission(user: User):
    if not user or not user.is_authenticated:
        return ""
    product = _get_business_autopilot_product()
    if not product:
        return ""
    access_row = (
        UserProductAccess.objects
        .filter(user=user, product=product)
        .only("permission")
        .first()
    )
    return str(getattr(access_row, "permission", "") or "").strip().lower()


def _crm_has_product_view_access(user: User):
    return _crm_business_autopilot_permission(user) in {
        UserProductAccess.PERMISSION_VIEW,
        UserProductAccess.PERMISSION_EDIT,
        UserProductAccess.PERMISSION_FULL,
    }


def _crm_has_product_edit_access(user: User):
    return _crm_business_autopilot_permission(user) in {
        UserProductAccess.PERMISSION_EDIT,
        UserProductAccess.PERMISSION_FULL,
    }


def _crm_row_matches_user(user: User, row):
    user_ids = set(_crm_clean_user_id_list(getattr(row, "assigned_user_ids", [])))
    user_ids.update(_crm_clean_user_id_list(getattr(row, "owner_user_ids", [])))
    assigned_user_id = getattr(row, "assigned_user_id", None)
    if assigned_user_id:
        user_ids.add(int(assigned_user_id))
    created_by_id = getattr(row, "created_by_id", None)
    if created_by_id:
        user_ids.add(int(created_by_id))
    return bool(user and user.is_authenticated and int(user.id) in user_ids)


def _crm_can_view_row(user: User, org: Organization, row):
    if _crm_is_admin(user, org):
        return True
    if _crm_has_unrestricted_row_access(user, org):
        return True
    if _crm_has_view_access(user, org):
        return _crm_row_matches_user(user, row)
    # Legacy fallback: allow users with product access to view only rows they own/are assigned to.
    if _crm_has_product_view_access(user):
        return _crm_row_matches_user(user, row)
    return False


def _crm_can_edit_row(user: User, org: Organization, row):
    if _crm_is_admin(user, org):
        return True
    if _crm_has_unrestricted_row_access(user, org):
        return True
    if _crm_has_edit_access(user, org):
        return _crm_row_matches_user(user, row)
    # Legacy fallback: users with product edit access can edit only rows they own/are assigned to.
    if _crm_has_product_edit_access(user):
        return _crm_row_matches_user(user, row)
    return False


def _hr_section_access_level(user: User, org: Organization, section_key: str = "hr"):
    if _crm_is_admin(user, org):
        return "Full Access"
    if not user or not user.is_authenticated or not org:
        return "No Access"
    settings_obj = OrganizationSettings.objects.filter(organization=org).only("business_autopilot_role_access_map").first()
    role_access_map = _normalize_role_access_map(getattr(settings_obj, "business_autopilot_role_access_map", {}) or {})
    profile = UserProfile.objects.filter(user=user).only("role").first()
    membership = _get_org_membership(user, org)
    role_access_record = _crm_resolve_role_access_record(
        role_access_map,
        getattr(profile, "role", ""),
        getattr(membership, "employee_role", ""),
        getattr(membership, "user_type", ""),
    )
    if not role_access_record:
        return "No Access"
    allowed_sections = _user_type_allowed_sections(getattr(membership, "user_type", ""), org=org)
    if section_key not in allowed_sections:
        return "No Access"
    sections = role_access_record.get("sections") if isinstance(role_access_record.get("sections"), dict) else {}
    return ROLE_ACCESS_LEVEL_ALIASES.get(str(sections.get(section_key) or "No Access").strip(), "No Access")


def _serialize_openai_settings(settings_obj: OrganizationSettings):
    voice_gender = str(getattr(settings_obj, "business_autopilot_ai_voice_gender", "") or "female").strip().lower()
    if voice_gender not in {"male", "female"}:
        voice_gender = "female"
    return {
        "enabled": bool(settings_obj.business_autopilot_openai_enabled),
        "agent_name": str(settings_obj.business_autopilot_ai_agent_name or DEFAULT_BA_OPENAI_AGENT_NAME).strip() or DEFAULT_BA_OPENAI_AGENT_NAME,
        "account_email": str(settings_obj.business_autopilot_openai_account_email or "").strip(),
        "model": str(settings_obj.business_autopilot_openai_model or "gpt-4o-mini").strip() or "gpt-4o-mini",
        "has_api_key": bool(str(settings_obj.business_autopilot_openai_api_key or "").strip()),
        "masked_api_key": _mask_secret(settings_obj.business_autopilot_openai_api_key),
        "voice_gender": voice_gender,
        "wake_word_enabled": bool(getattr(settings_obj, "business_autopilot_ai_wake_word_enabled", False)),
        "wake_phrase": str(getattr(settings_obj, "business_autopilot_ai_wake_phrase", "") or "").strip(),
        "silence_gap_seconds": 5,
    }


def _business_autopilot_ai_scope(user: User, org: Organization):
    membership = _get_org_membership(user, org)
    user_type = str(getattr(membership, "user_type", "") or "").strip() or OrganizationUser.USER_TYPE_FULL
    allowed_sections = set(_user_type_allowed_sections(user_type, org=org))
    is_admin = _crm_is_admin(user, org)
    if is_admin:
        user_type = OrganizationUser.USER_TYPE_FULL
        allowed_sections.update({"crm", "hr", "accounts", "billing", "projects", "dashboard", "profile", "users"})
    scope = {
        "membership": membership,
        "user_type": user_type,
        "label": (BUSINESS_AUTOPILOT_USER_TYPE_MAP.get(user_type) or {}).get("label", "Business User"),
        "is_admin": is_admin,
        "allowed_sections": sorted(allowed_sections),
        "can_access_crm": is_admin or ("crm" in allowed_sections and _crm_has_view_access(user, org)),
        "can_access_hr": is_admin or ("hr" in allowed_sections and _hr_section_access_level(user, org, "hr") != "No Access"),
        "can_access_accounts": is_admin or "accounts" in allowed_sections,
        "can_access_billing": is_admin or "billing" in allowed_sections,
    }
    scope["can_chat"] = any((
        scope["can_access_crm"],
        scope["can_access_hr"],
        scope["can_access_accounts"],
        scope["can_access_billing"],
    ))
    return scope


def _format_decimal_text(value):
    try:
        amount = Decimal(str(value or 0))
    except (InvalidOperation, TypeError, ValueError):
        amount = Decimal("0")
    return f"{amount.quantize(Decimal('0.01'))}"


def _ba_to_decimal_amount(value, default="0.00"):
    try:
        return Decimal(str(value if value is not None else default))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(str(default or "0"))


def _ba_currency_symbol(currency_code):
    code = str(currency_code or "").strip().upper()
    mapping = {
        "INR": "Rs.",
        "USD": "$",
        "EUR": "EUR",
        "GBP": "GBP",
        "AED": "AED",
        "SAR": "SAR",
    }
    return mapping.get(code, code or "INR")


def _ba_number_words_under_1000(number: int):
    units = [
        "zero", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine",
        "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen",
        "seventeen", "eighteen", "nineteen",
    ]
    tens = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]
    value = int(number or 0)
    if value < 20:
        return units[value]
    if value < 100:
        return tens[value // 10] + (f"-{units[value % 10]}" if value % 10 else "")
    return (
        f"{units[value // 100]} hundred"
        + (f" { _ba_number_words_under_1000(value % 100) }" if value % 100 else "")
    )


def _ba_integer_to_indian_words(number: int):
    value = int(number or 0)
    if value == 0:
        return "zero"
    parts = []
    units = [
        (10_000_000, "crore"),
        (100_000, "lakh"),
        (1_000, "thousand"),
        (100, "hundred"),
    ]
    remainder = value
    for divisor, label in units:
        chunk = remainder // divisor
        if chunk:
            if divisor == 100:
                parts.append(f"{_ba_number_words_under_1000(chunk)} {label}")
            else:
                parts.append(f"{_ba_number_words_under_1000(chunk)} {label}")
            remainder %= divisor
    if remainder:
        parts.append(_ba_number_words_under_1000(remainder))
    return " ".join(parts).strip()


def _ba_integer_to_international_words(number: int):
    value = int(number or 0)
    if value == 0:
        return "zero"
    parts = []
    units = [
        (1_000_000_000, "billion"),
        (1_000_000, "million"),
        (1_000, "thousand"),
    ]
    remainder = value
    for divisor, label in units:
        chunk = remainder // divisor
        if chunk:
            parts.append(f"{_ba_number_words_under_1000(chunk)} {label}")
            remainder %= divisor
    if remainder:
        parts.append(_ba_number_words_under_1000(remainder))
    return " ".join(parts).strip()


def _ba_format_currency_amount_display(currency_code, amount, country=""):
    code = str(currency_code or "INR").strip().upper() or "INR"
    country_name = str(country or "").strip().lower()
    if code == "INR" or "india" in country_name:
        amount_text = _format_indian_currency_value(amount, "0.00")
    else:
        amount_text = _decimal_to_string(amount)
    return f"{_ba_currency_symbol(code)} {amount_text}"


def _ba_currency_amount_spoken(currency_code, amount, country=""):
    code = str(currency_code or "INR").strip().upper() or "INR"
    country_name = str(country or "").strip().lower()
    amount_value = _ba_to_decimal_amount(amount)
    sign = "minus " if amount_value < 0 else ""
    absolute_value = abs(amount_value)
    integer_part = int(absolute_value)
    decimals = int((absolute_value - Decimal(integer_part)).quantize(Decimal("0.01")) * 100)
    if code == "INR" or "india" in country_name:
        integer_words = _ba_integer_to_indian_words(integer_part)
    else:
        integer_words = _ba_integer_to_international_words(integer_part)
    currency_word = {
        "INR": "rupees",
        "USD": "dollars",
        "EUR": "euros",
        "GBP": "pounds",
        "AED": "dirhams",
        "SAR": "riyals",
    }.get(code, code.lower())
    if decimals:
        paise_word = "paise" if code == "INR" else "cents"
        return f"{sign}{integer_words} {currency_word} and {_ba_number_words_under_1000(decimals)} {paise_word}"
    return f"{sign}{integer_words} {currency_word}"


def _ba_direct_response(reply, tts_text=""):
    response_text = str(reply or "").strip()
    return {
        "reply": response_text,
        "tts_text": str(tts_text or response_text).strip() or response_text,
    }


def _attendance_work_minutes(entry):
    if not getattr(entry, "checkin_time", None) or not getattr(entry, "checkout_time", None):
        return None
    minutes = int((entry.checkout_time - entry.checkin_time).total_seconds() // 60)
    return minutes if minutes >= 0 else None


def _build_ba_assistant_scope_context(user: User, org: Organization, scope):
    today = timezone.localdate()
    assistant_profile = get_assistant_product_profile("business_autopilot")
    active_subscription = _get_active_erp_subscription(org)
    active_plan = getattr(active_subscription, "plan", None)
    modules, enabled_modules = _serialize_modules(org)
    active_user_type_rows = _get_subscription_user_types(active_subscription)
    meeting_enabled_user_types = []
    for row in active_user_type_rows:
        allowed_modules = {
            str(item or "").strip().lower()
            for item in (row.get("allowed_modules") or [])
            if str(item or "").strip()
        }
        if "crm" in allowed_modules:
            meeting_enabled_user_types.append({
                "key": str(row.get("key") or "").strip(),
                "label": str(row.get("label") or "").strip(),
            })
    context = {
        "today": today.isoformat(),
        "scope": {
            "user_type": scope["user_type"],
            "label": scope["label"],
            "allowed_sections": scope["allowed_sections"],
        },
        "product_capabilities": {
            "product_key": "business_autopilot",
            "product_label": getattr(assistant_profile, "label", "Business Autopilot"),
            "operating_mode": getattr(assistant_profile, "operating_mode", "internal business copilot"),
            "active_plan_name": str(getattr(active_plan, "name", "") or "").strip(),
            "active_subscription_status": str(getattr(active_subscription, "status", "") or "").strip().lower(),
            "enabled_modules": [
                {
                    "slug": str(row.get("slug") or "").strip(),
                    "name": str(row.get("name") or "").strip(),
                }
                for row in enabled_modules
            ],
            "eligible_modules": [
                {
                    "slug": str(row.get("slug") or "").strip(),
                    "name": str(row.get("name") or "").strip(),
                }
                for row in modules
                if row.get("eligible")
            ],
            "crm_meetings_available": any(str(row.get("slug") or "").strip().lower() == "crm" for row in enabled_modules),
            "meeting_enabled_user_types": meeting_enabled_user_types,
        },
    }
    active_memberships = list(
        OrganizationUser.objects.filter(
            organization=org,
            is_deleted=False,
            is_active=True,
        ).select_related("user")
    )
    context["org_summary"] = {
        "organization_name": str(getattr(org, "name", "") or "").strip(),
        "country": str(getattr(org, "country", "") or "India").strip(),
        "currency": str(getattr(org, "currency", "") or "INR").strip().upper() or "INR",
        "total_users": len(active_memberships),
        "admin_users": len([row for row in active_memberships if _is_org_admin_account_member(org, row)]),
        "employee_users": len([row for row in active_memberships if not _is_org_admin_account_member(org, row)]),
        "sample_user_names": [
            " ".join(
                part for part in [
                    str(getattr(getattr(row, "user", None), "first_name", "") or "").strip(),
                    str(getattr(getattr(row, "user", None), "last_name", "") or "").strip(),
                ]
                if part
            ).strip()
            or str(getattr(getattr(row, "user", None), "username", "") or "").strip()
            for row in active_memberships[:5]
            if (
                " ".join(
                    part for part in [
                        str(getattr(getattr(row, "user", None), "first_name", "") or "").strip(),
                        str(getattr(getattr(row, "user", None), "last_name", "") or "").strip(),
                    ]
                    if part
                ).strip()
                or str(getattr(getattr(row, "user", None), "username", "") or "").strip()
            )
        ],
    }

    if scope["can_access_hr"]:
        attendance_rows = list(
            AttendanceEntry.objects.filter(
                organization=org,
                attendance_date__month=today.month,
                attendance_date__year=today.year,
            ).select_related("employee_membership")
        )
        today_rows = [row for row in attendance_rows if row.attendance_date == today]
        overtime_rows = []
        for row in attendance_rows:
            work_minutes = _attendance_work_minutes(row)
            if work_minutes and work_minutes > 480:
                overtime_rows.append({
                    "employee_name": row.employee_name,
                    "attendance_date": row.attendance_date.isoformat(),
                    "overtime_minutes": work_minutes - 480,
                })
        increments = EmployeeSalaryHistory.objects.filter(
            organization=org,
            effective_from__gte=today - timedelta(days=365),
        ).order_by("-effective_from", "-id")[:10]
        context["hr"] = {
            "today_attendance_count": len(today_rows),
            "employees_checked_in_today": [row.employee_name for row in today_rows[:20]],
            "recent_increments": [
                {
                    "employee_name": row.employee_name,
                    "effective_from": row.effective_from.isoformat() if row.effective_from else "",
                    "increment_type": row.increment_type,
                    "increment_value": _format_decimal_text(row.increment_value),
                    "increment_amount": _format_decimal_text(row.increment_amount),
                    "new_salary": _format_decimal_text(row.new_salary),
                }
                for row in increments
            ],
            "overtime_this_month": overtime_rows[:20],
        }

    if scope["can_access_crm"]:
        leads_rows = list(CrmLead.objects.filter(organization=org, is_deleted=False).order_by("-created_at")[:50])
        deals_rows = list(CrmDeal.objects.filter(organization=org, is_deleted=False).order_by("-created_at")[:50])
        meetings_rows = list(CrmMeeting.objects.filter(organization=org, is_deleted=False).order_by("meeting_date", "meeting_time", "id")[:50])
        sales_rows = list(CrmSalesOrder.objects.filter(organization=org, is_deleted=False).order_by("-created_at")[:50])
        if not scope["is_admin"] and not _crm_has_unrestricted_row_access(user, org):
            leads_rows = [row for row in leads_rows if _crm_can_view_row(user, org, row)]
            deals_rows = [row for row in deals_rows if _crm_can_view_row(user, org, row)]
            meetings_rows = [row for row in meetings_rows if _crm_can_view_row(user, org, row)]
            sales_rows = [row for row in sales_rows if _crm_can_view_row(user, org, row)]
        lead_assignments = {}
        lead_amounts = {}
        status_counts = {}
        stage_counts = {}
        for row in leads_rows:
            owner_name = str(
                getattr(getattr(row, "assigned_user", None), "first_name", "") or ""
            ).strip()
            owner_last_name = str(
                getattr(getattr(row, "assigned_user", None), "last_name", "") or ""
            ).strip()
            full_name = " ".join(part for part in [owner_name, owner_last_name] if part).strip()
            if not full_name:
                full_name = str(getattr(getattr(row, "assigned_user", None), "username", "") or "").strip()
            if not full_name:
                full_name = str(row.assigned_team or "").strip()
            if not full_name:
                full_name = "Unassigned"
            lead_assignments[full_name] = lead_assignments.get(full_name, 0) + 1
            lead_amounts[full_name] = lead_amounts.get(full_name, Decimal("0")) + _crm_to_decimal(getattr(row, "lead_amount", 0))
            status_name = str(getattr(row, "status", "") or "").strip() or "Unknown"
            stage_name = str(getattr(row, "stage", "") or "").strip() or "Unknown"
            status_counts[status_name] = status_counts.get(status_name, 0) + 1
            stage_counts[stage_name] = stage_counts.get(stage_name, 0) + 1
        lead_assignment_summary = [
            {
                "name": name,
                "lead_count": count,
                "total_amount": _format_decimal_text(lead_amounts.get(name, Decimal("0"))),
            }
            for name, count in sorted(lead_assignments.items(), key=lambda item: (-item[1], item[0].lower()))
        ]
        lead_status_summary = [
            {"status": name, "count": count}
            for name, count in sorted(status_counts.items(), key=lambda item: (-item[1], item[0].lower()))
        ]
        lead_stage_summary = [
            {"stage": name, "count": count}
            for name, count in sorted(stage_counts.items(), key=lambda item: (-item[1], item[0].lower()))
        ]
        context["crm"] = {
            "currency": str(getattr(org, "currency", "") or "INR").strip().upper() or "INR",
            "country": str(getattr(org, "country", "") or "India").strip(),
            "visible_leads_count": len(leads_rows),
            "open_leads_count": sum(1 for row in leads_rows if str(row.status or "").strip().lower() == "open"),
            "today_meetings": [
                {
                    "title": row.title,
                    "related_to": row.related_to,
                    "meeting_date": row.meeting_date.isoformat() if row.meeting_date else "",
                    "meeting_time": row.meeting_time.isoformat() if row.meeting_time else "",
                    "status": row.status,
                    "owners": row.owner_names,
                }
                for row in meetings_rows
                if row.meeting_date == today
            ][:12],
            "upcoming_meetings": [
                {
                    "title": row.title,
                    "related_to": row.related_to,
                    "meeting_date": row.meeting_date.isoformat() if row.meeting_date else "",
                    "meeting_time": row.meeting_time.isoformat() if row.meeting_time else "",
                    "status": row.status,
                }
                for row in meetings_rows
                if row.meeting_date and row.meeting_date >= today
            ][:12],
            "recent_leads": [
                {
                    "lead_name": row.lead_name,
                    "company": row.company,
                    "stage": row.stage,
                    "status": row.status,
                    "amount": _format_decimal_text(row.lead_amount),
                }
                for row in leads_rows[:10]
            ],
            "visible_leads": [
                {
                    "lead_name": row.lead_name,
                    "company": row.company,
                    "stage": row.stage,
                    "status": row.status,
                    "assigned_to": (
                        " ".join(
                            part
                            for part in [
                                str(getattr(getattr(row, "assigned_user", None), "first_name", "") or "").strip(),
                                str(getattr(getattr(row, "assigned_user", None), "last_name", "") or "").strip(),
                            ]
                            if part
                        ).strip()
                        or str(getattr(getattr(row, "assigned_user", None), "username", "") or "").strip()
                        or str(row.assigned_team or "").strip()
                        or "Unassigned"
                    ),
                    "amount": _format_decimal_text(row.lead_amount),
                }
                for row in leads_rows[:30]
            ],
            "lead_assignment_summary": lead_assignment_summary[:20],
            "lead_status_summary": lead_status_summary[:20],
            "lead_stage_summary": lead_stage_summary[:20],
            "top_lead_owner": lead_assignment_summary[0] if lead_assignment_summary else None,
            "won_deals_count": sum(1 for row in deals_rows if str(row.status or "").strip().lower() == "won"),
            "sales_orders_count": len(sales_rows),
        }

    if scope["can_access_accounts"] or scope["can_access_billing"]:
        accounts_workspace = _get_accounts_workspace(org)
        workspace_data = accounts_workspace.data if isinstance(getattr(accounts_workspace, "data", None), dict) else {}
        invoice_rows = workspace_data.get("invoices") if isinstance(workspace_data.get("invoices"), list) else []
        sales_rows = list(CrmSalesOrder.objects.filter(organization=org, is_deleted=False, created_at__date=today).order_by("-created_at")[:50])
        if scope["can_access_crm"] and not scope["is_admin"] and not _crm_has_unrestricted_row_access(user, org):
            sales_rows = [row for row in sales_rows if _crm_can_view_row(user, org, row)]
        today_sales_total = sum((Decimal(str(getattr(row, "total_amount", 0) or 0)) for row in sales_rows), Decimal("0"))
        invoice_status_counts = {}
        total_invoice_amount = Decimal("0")
        paid_invoices_count = 0
        pending_invoices_count = 0
        this_month_invoices_count = 0
        for row in invoice_rows:
            if not isinstance(row, dict):
                continue
            status_name = str(
                row.get("status")
                or row.get("paymentStatus")
                or row.get("invoiceStatus")
                or "Unknown"
            ).strip() or "Unknown"
            invoice_status_counts[status_name] = invoice_status_counts.get(status_name, 0) + 1
            amount_value = _to_decimal(
                row.get("grandTotal")
                or row.get("grand_total")
                or row.get("totalAmount")
                or row.get("total_amount")
                or row.get("amount")
            )
            total_invoice_amount += amount_value
            normalized_status = status_name.strip().lower()
            if normalized_status in {"paid", "completed", "settled"}:
                paid_invoices_count += 1
            elif normalized_status in {"pending", "unpaid", "partial", "partially paid", "overdue", "due"}:
                pending_invoices_count += 1
            issue_date = _crm_report_parse_date(
                row.get("issueDate")
                or row.get("issue_date")
                or row.get("invoiceDate")
                or row.get("invoice_date")
                or row.get("date")
                or row.get("createdDate")
                or row.get("created_date")
            ) or _crm_report_issue_date_from_doc_no(
                row.get("docNo")
                or row.get("doc_no")
                or row.get("invoiceNo")
                or row.get("invoice_no")
            )
            if issue_date and issue_date.year == today.year and issue_date.month == today.month:
                this_month_invoices_count += 1
        context["accounts"] = {
            "currency": str(getattr(org, "currency", "") or "INR").strip().upper() or "INR",
            "country": str(getattr(org, "country", "") or "India").strip(),
            "today_sales_orders_count": len(sales_rows),
            "today_sales_total": _format_decimal_text(today_sales_total),
            "total_invoices_count": len([row for row in invoice_rows if isinstance(row, dict)]),
            "total_invoices_amount": _format_decimal_text(total_invoice_amount),
            "paid_invoices_count": paid_invoices_count,
            "pending_invoices_count": pending_invoices_count,
            "this_month_invoices_count": this_month_invoices_count,
            "invoice_status_summary": [
                {"status": name, "count": count}
                for name, count in sorted(invoice_status_counts.items(), key=lambda item: (-item[1], item[0].lower()))
            ][:20],
        }

    return context


def _build_ba_assistant_system_prompt(agent_name: str, scope, product_context=None):
    allowed_areas = []
    if scope["can_access_hr"]:
        allowed_areas.append("HR attendance, salary increments, overtime, and payroll-ready summaries")
    if scope["can_access_crm"]:
        allowed_areas.append("CRM leads, deals, meetings, and sales pipeline summaries")
    if scope["can_access_accounts"] or scope["can_access_billing"]:
        allowed_areas.append("accounts and billing summaries")
    if not allowed_areas:
        allowed_areas.append("no business data")
    product_context = product_context if isinstance(product_context, dict) else {}
    org_name = str(product_context.get("organization_name") or "").strip()
    enabled_modules = [
        str(row.get("name") or row.get("slug") or "").strip()
        for row in (product_context.get("enabled_modules") or [])
        if isinstance(row, dict)
    ]
    product_training_context = build_assistant_training_context(
        "business_autopilot",
        org_name=org_name,
        enabled_modules=enabled_modules,
        extra_notes=(
            [
                f"Current user type label: {scope['label']}",
                f"Current allowed sections: {', '.join(scope.get('allowed_sections') or []) or 'none'}",
            ]
        ),
    )
    return (
        (f"{product_training_context}\n\n" if product_training_context else "")
        +
        f"You are {agent_name}, a role-aware business assistant for Work Zilla Business Autopilot. "
        f"This user is a {scope['label']}. Only answer within these allowed areas: {'; '.join(allowed_areas)}. "
        "Use only the provided organization context. "
        "If the user asks about a restricted module, clearly say that their current access is limited. "
        "If data is missing, say it is not available in the current assistant data. "
        "Do not confuse 'feature not enabled' with 'no records today'. If the application has a feature but the current data has zero rows, clearly say the feature exists and there is no data right now. "
        "Behave like a capable internal employee of this organization: practical, context-aware, and responsible. "
        "Remember the recent discussion that is provided and continue the conversation naturally instead of answering in isolation. "
        "Mirror the user's language, script, and tone. "
        "If the user writes in casual English, reply in casual English. "
        "If the user writes in Tanglish or mixed Tamil-English, reply in the same natural mixed style and do not convert it into pure formal Tamil. "
        "If the user writes in Tamil script, keep it conversational and modern, not overly literary or stiff. "
        "If the user writes in Telugu, reply in Telugu. If the user writes in Kannada, reply in Kannada. "
        "Avoid sounding like a formal notice unless the user explicitly asks for a formal draft. "
        "Use the current module, recent conversation, and visible organization data together before concluding that data is unavailable. "
        "If the answer can be derived by totaling, grouping, comparing, or filtering the visible records, do that reasoning and answer directly. "
        "If the user's phrasing is referential like 'antha', 'that one', 'same user', or 'those leads', resolve it from recent conversation whenever possible. "
        "Do not invent names, dates, counts, meetings, overtime, attendance, or billing figures. "
        "If the question is simple like total users, total leads, top owner, today attendance, or sales total, answer directly in one or two natural lines. "
        "Do not answer with a bare number alone when the user asks by voice or chat. Always respond as a complete natural sentence with enough context to feel human. "
        "If the user asks for a very large list, many names, or huge detailed records, do not enumerate everything. Briefly explain that the dataset is too large to read out fully and ask them to check it manually in the application. "
        "Reply in the same language or mixed-language style used by the user question. "
        "Keep answers short, helpful, human, and business-focused."
    )


def _contains_tamil_chars(text: str):
    return bool(re.search(r"[\u0B80-\u0BFF]", str(text or "")))


def _contains_telugu_chars(text: str):
    return bool(re.search(r"[\u0C00-\u0C7F]", str(text or "")))


def _contains_kannada_chars(text: str):
    return bool(re.search(r"[\u0C80-\u0CFF]", str(text or "")))


def _contains_any(text: str, terms):
    haystack = str(text or "").strip().lower()
    for term in (terms or []):
        needle = str(term or "").strip().lower()
        if not needle:
            continue
        if len(re.sub(r"[^a-z0-9]", "", needle)) <= 2 and re.fullmatch(r"[a-z0-9]+", needle):
            if re.search(rf"(?<![a-z0-9]){re.escape(needle)}(?![a-z0-9])", haystack):
                return True
            continue
        if needle in haystack:
            return True
    return False


def _ba_looks_like_tanglish(message: str):
    text = str(message or "").strip().lower()
    if not text or not re.search(r"[a-z]", text):
        return False
    tanglish_markers = [
        "enna", "epdi", "eppo", "inga", "anga", "irukku", "venum", "venuma", "solu", "solunga",
        "pa", "la", "ku", "motham", "yevalo", "evlo", "adhigama", "yaru", "yaaru", "kitta",
        "innaiku", "inniku", "panna", "pananum", "theriyala", "illaya", "illa", "pesa",
    ]
    return any(marker in text for marker in tanglish_markers)


def _ba_detect_user_style(message: str):
    raw = str(message or "")
    has_tamil = _contains_tamil_chars(raw)
    has_telugu = _contains_telugu_chars(raw)
    has_kannada = _contains_kannada_chars(raw)
    has_english = bool(re.search(r"[A-Za-z]", raw))
    if not has_tamil and has_english and _ba_looks_like_tanglish(raw):
        return "tanglish"
    if has_tamil and has_english:
        return "tanglish"
    if has_telugu and has_english:
        return "telugu_mixed"
    if has_kannada and has_english:
        return "kannada_mixed"
    if has_tamil:
        return "tamil"
    if has_telugu:
        return "telugu"
    if has_kannada:
        return "kannada"
    return "english"


def _ba_language_instruction_from_text(text: str):
    style = _ba_detect_user_style(text)
    if style == "tanglish":
        return "If the text mixes Tamil and English, use natural office-style Tanglish. Keep English business words in English and pronounce Tamil casually."
    if style == "tamil":
        return "If the text contains Tamil, pronounce it in a smooth colloquial Tamil style."
    if style in {"telugu", "telugu_mixed"}:
        return "If the text contains Telugu, pronounce it naturally in a conversational Telugu style."
    if style in {"kannada", "kannada_mixed"}:
        return "If the text contains Kannada, pronounce it naturally in a conversational Kannada style."
    return "If the text is English or mixed business English, keep the delivery relaxed and natural."


def _ba_style_reply(style: str, english: str, tanglish: str = "", tamil: str = "", telugu: str = "", kannada: str = ""):
    normalized = str(style or "english").strip().lower()
    if normalized == "tamil":
        return str(tamil or tanglish or english)
    if normalized == "tanglish":
        return str(tanglish or english)
    if normalized == "telugu":
        return str(telugu or english)
    if normalized == "telugu_mixed":
        return str(telugu or english)
    if normalized == "kannada":
        return str(kannada or english)
    if normalized == "kannada_mixed":
        return str(kannada or english)
    return str(english)


def _ba_is_large_list_request(message: str):
    text = str(message or "").strip().lower()
    if not text:
        return False
    return _contains_any(
        text,
        [
            "list all",
            "show all",
            "all users",
            "all employees",
            "all leads",
            "all meetings",
            "all invoices",
            "all names",
            "all details",
            "full list",
            "complete list",
            "entire list",
            "user list",
            "employee list",
            "lead list",
            "meetings list",
            "meeting list",
            "invoice list",
            "members list",
            "member list",
            "names list",
            "ella per",
            "ella peru",
            "ellarayum",
            "motha list",
            "motham list",
            "list sollu",
            "show pannunga",
            "show me all",
            "tell all",
            "give all",
        ],
    )


def _ba_large_dataset_response(style: str, subject_english: str, count: int = 0, threshold: int = 5):
    safe_count = max(0, int(count or 0))
    safe_threshold = max(1, int(threshold or 5))
    if safe_count and safe_count <= safe_threshold:
        return ""
    if safe_count >= 100:
        english = f"{subject_english.capitalize()} {safe_count}-plus is too large to read out here. Please check it manually in the application."
        tanglish = f"{subject_english.capitalize()} {safe_count}-ku mela irukku. Ella details-um inga solla mudiyadhu. Application-la manual-a check pannunga."
        tamil = f"{subject_english.capitalize()} {safe_count}-ku mela இருக்கு. எல்லா details-um இங்க சொல்ல முடியாது. Application-la manual-a check pannunga."
        telugu = f"{subject_english.capitalize()} {safe_count}-ku paiga unnayi. Anni details ikkada cheppadam kashtam. Application-lo manual-ga check cheyyandi."
        kannada = f"{subject_english.capitalize()} {safe_count}-kkinta hecchu ide. Ella details-anna illi helakke agalla. Application-nalli manual-a check madi."
    else:
        english = f"There are {safe_count} {subject_english}. Listing everything here will be too long. Please check manually in the application."
        tanglish = f"{subject_english.capitalize()} total {safe_count} irukku. Ella details-um inga solla romba long aagum. Application-la manual-a check pannunga."
        tamil = f"{subject_english.capitalize()} total {safe_count} இருக்கு. எல்லா details-um இங்க சொல்ல ரொம்ப long ஆகும். Application-la manual-a check pannunga."
        telugu = f"Total {safe_count} {subject_english} unnayi. Anni details ikkada chepthe chala long avuthundi. Application-lo manual-ga check cheyyandi."
        kannada = f"Total {safe_count} {subject_english} ive. Ella details illi helidre tumba long agutte. Application-nalli manual-a check madi."
    return _ba_style_reply(
        style,
        english=english,
        tanglish=tanglish,
        tamil=tamil,
        telugu=telugu,
        kannada=kannada,
    )


def _ba_normalize_recent_messages(rows, limit=12):
    if not isinstance(rows, list):
        return []
    normalized = []
    for row in rows[-max(1, int(limit or 8)):]:
        if not isinstance(row, dict):
            continue
        role = str(row.get("role") or "").strip().lower()
        text = str(row.get("text") or row.get("content") or "").strip()
        if role not in {"user", "assistant"} or not text:
            continue
        normalized.append({
            "role": role,
            "text": text[:800],
        })
    return normalized


def _ba_normalize_history_messages(rows, limit=200):
    if not isinstance(rows, list):
        return []
    normalized = []
    for row in rows[-max(1, int(limit or 200)):]:
        if not isinstance(row, dict):
            continue
        role = str(row.get("role") or "").strip().lower()
        text = str(row.get("text") or row.get("content") or "").strip()
        item_id = str(row.get("id") or "").strip()[:120]
        if role not in {"user", "assistant"} or not text:
            continue
        normalized.append({
            "id": item_id or f"{role}-{len(normalized) + 1}",
            "role": role,
            "text": text[:4000],
        })
    return normalized


def _ba_assistant_question_route(message: str, scope):
    text = str(message or "").strip().lower()
    if not text:
        return "general"
    if _contains_any(
        text,
        [
            "application la",
            "app la",
            "app-level",
            "feature iruka",
            "option iruka",
            "module iruka",
            "system la",
            "namma application",
            "our application",
            "meeting schedule iruka",
            "meeting schedules iruka",
            "meetings option",
            "meetings feature",
        ],
    ):
        return "general"
    if scope.get("can_access_crm"):
        if _contains_any(
            text,
            [
                "adhigama lead",
                "most lead",
                "highest lead",
                "lead assign",
                "assigned lead",
                "open lead summary",
                "meetings",
                "meeting schedule",
                "today meeting",
                "total leads",
                "motham leads",
                "how many leads",
                "lead count",
                "leads total amount",
                "lead total amount",
                "total amount",
                "lead amount",
            ],
        ):
            return "crm"
    if scope.get("can_access_hr"):
        if _contains_any(text, ["overtime", "ot", "increment", "attendance", "leave"]):
            return "hr"
    if scope.get("can_access_accounts") or scope.get("can_access_billing"):
        if _contains_any(text, ["billing", "sales summary", "today sales", "invoice", "amount sold"]):
            return "accounts"
    if scope.get("can_access_ticketing"):
        if _contains_any(text, ["ticket", "tickets", "open ticket", "pending ticket", "ticket count", "ticket summary"]):
            return "ticketing"
    if _contains_any(
        text,
        [
            "how many users",
            "total users",
            "user count",
            "employee count",
            "how many employees",
            "motham users",
            "motham employee",
            "organization users",
            "org users",
            "restrictions",
            "access level",
            "what can i access",
            "what data can i access",
        ],
    ):
        return "general"
    return "general"


def _ba_merge_local_assistant_context(context, local_context):
    if not isinstance(context, dict) or not isinstance(local_context, dict):
        return context
    ticketing_context = local_context.get("ticketing") if isinstance(local_context.get("ticketing"), dict) else {}
    open_tickets = ticketing_context.get("openTickets") if isinstance(ticketing_context.get("openTickets"), list) else []
    total_tickets_count = int(ticketing_context.get("totalTicketsCount") or 0)
    normalized_open_tickets = []
    status_counts = {}
    for row in open_tickets[:50]:
        if not isinstance(row, dict):
            continue
        title = str(row.get("title") or row.get("subject") or "Ticket").strip() or "Ticket"
        status = str(row.get("status") or "Unknown").strip() or "Unknown"
        priority = str(row.get("priority") or "").strip()
        ticket_id = str(row.get("ticketId") or row.get("id") or "").strip()
        normalized_open_tickets.append({
            "ticket_id": ticket_id,
            "title": title,
            "status": status,
            "priority": priority,
        })
        status_counts[status] = status_counts.get(status, 0) + 1
    if normalized_open_tickets:
        context["ticketing"] = {
            "total_tickets_count": total_tickets_count or len(normalized_open_tickets),
            "open_tickets_count": len(normalized_open_tickets),
            "open_tickets": normalized_open_tickets[:12],
            "ticket_status_summary": [
                {"status": name, "count": count}
                for name, count in sorted(status_counts.items(), key=lambda item: (-item[1], item[0].lower()))
            ][:12],
        }
    return context


def _ba_assistant_direct_crm_answer(message: str, crm_context):
    text = str(message or "").strip().lower()
    style = _ba_detect_user_style(message)
    if not isinstance(crm_context, dict):
        return ""
    assignment_rows = crm_context.get("lead_assignment_summary") if isinstance(crm_context.get("lead_assignment_summary"), list) else []
    top_owner = crm_context.get("top_lead_owner") if isinstance(crm_context.get("top_lead_owner"), dict) else None
    today_meetings = crm_context.get("today_meetings") if isinstance(crm_context.get("today_meetings"), list) else []
    recent_leads = crm_context.get("recent_leads") if isinstance(crm_context.get("recent_leads"), list) else []
    visible_leads_count = int(crm_context.get("visible_leads_count") or 0)
    status_rows = crm_context.get("lead_status_summary") if isinstance(crm_context.get("lead_status_summary"), list) else []
    currency_code = str(crm_context.get("currency") or "INR").strip().upper() or "INR"
    country_name = str(crm_context.get("country") or "India").strip()
    if _ba_is_large_list_request(message) and _contains_any(text, ["lead", "leads", "meeting", "meetings", "crm"]):
        list_count = visible_leads_count
        subject = "crm records"
        if _contains_any(text, ["meeting", "meetings"]):
            list_count = len(today_meetings)
            subject = "meetings"
        elif _contains_any(text, ["lead", "leads"]):
            subject = "leads"
        large_reply = _ba_large_dataset_response(style, subject, list_count, threshold=5)
        if large_reply:
            return large_reply
    amount_question = _contains_any(
        text,
        [
            "total amount",
            "amount evlo",
            "amount yevalo",
            "amount yevlo",
            "evlo amount",
            "yevalo amount",
            "how much amount",
            "sum of",
            "total value",
            "lead amount",
        ],
    )
    matched_owner = None
    matched_count = None

    for row in assignment_rows:
        owner_name = str(row.get("name") or "").strip()
        if owner_name and owner_name.lower() in text:
            matched_owner = row
            break

    if assignment_rows:
        number_tokens = [int(value) for value in re.findall(r"\d+", text)]
        if number_tokens:
            unique_count_matches = []
            for candidate in number_tokens:
                rows = [row for row in assignment_rows if int(row.get("lead_count") or 0) == candidate]
                if len(rows) == 1:
                    unique_count_matches.append(rows[0])
            if unique_count_matches:
                matched_count = unique_count_matches[0]

    if amount_question:
        target_row = matched_owner or matched_count
        if not target_row and _contains_any(text, ["antha", "that", "those"]) and top_owner:
            target_row = top_owner
        if target_row:
            amount_display = _ba_format_currency_amount_display(currency_code, target_row.get("total_amount") or "0.00", country_name)
            amount_spoken = _ba_currency_amount_spoken(currency_code, target_row.get("total_amount") or "0.00", country_name)
            reply_text = _ba_style_reply(
                style,
                english=(
                    f"{target_row.get('name') or 'Unknown'} has {target_row.get('lead_count') or 0} leads. "
                    f"Total amount is {amount_display}."
                ),
                tanglish=(
                    f"{target_row.get('name') or 'Unknown'} kitta {target_row.get('lead_count') or 0} leads irukku. "
                    f"Total amount {amount_display}."
                ),
                tamil=(
                    f"{target_row.get('name') or 'Unknown'} கிட்ட {target_row.get('lead_count') or 0} leads இருக்கு. "
                    f"Total amount {amount_display}."
                ),
                telugu=(
                    f"{target_row.get('name') or 'Unknown'} దగ్గర {target_row.get('lead_count') or 0} leads ఉన్నాయి. "
                    f"Total amount {amount_display}."
                ),
                kannada=(
                    f"{target_row.get('name') or 'Unknown'} ಬಳಿ {target_row.get('lead_count') or 0} leads ಇವೆ. "
                    f"Total amount {amount_display}."
                ),
            )
            tts_text = _ba_style_reply(
                style,
                english=(
                    f"{target_row.get('name') or 'Unknown'} has {target_row.get('lead_count') or 0} leads. "
                    f"Total amount is {amount_spoken}."
                ),
                tanglish=(
                    f"{target_row.get('name') or 'Unknown'} kitta {target_row.get('lead_count') or 0} leads irukku. "
                    f"Total amount {amount_spoken}."
                ),
                tamil=(
                    f"{target_row.get('name') or 'Unknown'} kitta {target_row.get('lead_count') or 0} leads irukku. "
                    f"Total amount {amount_spoken}."
                ),
                telugu=(
                    f"{target_row.get('name') or 'Unknown'} daggara {target_row.get('lead_count') or 0} leads unnayi. "
                    f"Total amount {amount_spoken}."
                ),
                kannada=(
                    f"{target_row.get('name') or 'Unknown'} bali {target_row.get('lead_count') or 0} leads ive. "
                    f"Total amount {amount_spoken}."
                ),
            )
            return _ba_direct_response(reply_text, tts_text)
        # Fall back to the full assistant context search when the amount question
        # mentions a person/name we could not confidently map in the direct shortcut.
        return ""

    if _contains_any(text, ["adhigama lead", "most lead", "highest lead", "lead assign", "assigned lead"]):
        if not assignment_rows:
            return _ba_style_reply(
                style,
                english="Lead assignment summary is not available in the current visible CRM data.",
                tanglish="Ippo visible CRM data-la lead assignment summary illa.",
                tamil="இப்போ visible CRM data-la lead assignment summary illa.",
                telugu="ప్రస్తుతం కనిపిస్తున్న CRM dataలో lead assignment summary లేదు.",
                kannada="ಈಗ ಕಾಣಿಸುತ್ತಿರುವ CRM dataನಲ್ಲಿ lead assignment summary ಇಲ್ಲ.",
            )
        if top_owner:
            return _ba_style_reply(
                style,
                english=f"{top_owner.get('name') or 'Unknown'} has the highest leads. Count: {top_owner.get('lead_count') or 0}.",
                tanglish=f"{top_owner.get('name') or 'Unknown'} kitta than adhigama leads irukku. Count: {top_owner.get('lead_count') or 0}.",
                tamil=f"{top_owner.get('name') or 'Unknown'} கிட்ட தான் அதிகமான leads இருக்கு. Count: {top_owner.get('lead_count') or 0}.",
                telugu=f"{top_owner.get('name') or 'Unknown'} దగ్గరే ఎక్కువ leads ఉన్నాయి. Count: {top_owner.get('lead_count') or 0}.",
                kannada=f"{top_owner.get('name') or 'Unknown'} ಬಳಿ ಹೆಚ್ಚು leads ಇವೆ. Count: {top_owner.get('lead_count') or 0}.",
            )
        return ""

    if _contains_any(text, ["total leads", "motham leads", "how many leads", "lead count"]):
        return _ba_style_reply(
            style,
            english=f"Right now your CRM has {visible_leads_count} visible leads in total.",
            tanglish=f"Namma CRM-la ippo total-a {visible_leads_count} leads irukku.",
            tamil=f"நம்ம CRM-la இப்போ total-a {visible_leads_count} leads இருக்கு.",
            telugu=f"మీ CRMలో ప్రస్తుతం మొత్తం {visible_leads_count} visible leads ఉన్నాయి.",
            kannada=f"Nimma CRMnalli iga ottu {visible_leads_count} visible leads ive.",
        )

    if _contains_any(text, [" leads", "lead ", "crm leads", "leads summary", "lead summary"]) or text in {"leads", "lead", "crm"}:
        open_count = int(crm_context.get("open_leads_count") or 0)
        converted_count = 0
        for row in status_rows:
            if str(row.get("status") or "").strip().lower() == "converted":
                converted_count = int(row.get("count") or 0)
                break
        top_owner_name = str((top_owner or {}).get("name") or "").strip() or "nobody yet"
        top_owner_count = int((top_owner or {}).get("lead_count") or 0)
        return _ba_style_reply(
            style,
            english=(
                f"Right now CRM has {visible_leads_count} visible leads. "
                f"{open_count} are open and {converted_count} are converted. "
                f"Top allocation is with {top_owner_name} at {top_owner_count} leads."
            ),
            tanglish=(
                f"Ippo CRM-la {visible_leads_count} visible leads irukku. "
                f"Adhula {open_count} open, {converted_count} converted. "
                f"Top allocation {top_owner_name} kitta {top_owner_count} leads."
            ),
            tamil=(
                f"இப்போ CRM-la {visible_leads_count} visible leads இருக்கு. "
                f"அதுல {open_count} open, {converted_count} converted. "
                f"Top allocation {top_owner_name} கிட்ட {top_owner_count} leads."
            ),
            telugu=(
                f"ఇప్పుడే CRMలో {visible_leads_count} visible leads ఉన్నాయి. "
                f"అందులో {open_count} open, {converted_count} converted. "
                f"Top allocation {top_owner_name} దగ్గర {top_owner_count} leads ఉన్నాయి."
            ),
            kannada=(
                f"ಈಗ CRMನಲ್ಲಿ {visible_leads_count} visible leads ಇವೆ. "
                f"ಅದರಲ್ಲಿ {open_count} open, {converted_count} converted. "
                f"Top allocation {top_owner_name} ಬಳಿ {top_owner_count} leads ಇವೆ."
            ),
        )

    if _contains_any(text, ["open lead summary", "open leads", "my open leads"]):
        open_count = crm_context.get("open_leads_count")
        if recent_leads:
            top_rows = ", ".join(
                f"{row.get('lead_name') or 'Lead'} ({row.get('status') or '-'})"
                for row in recent_leads[:3]
            )
            return _ba_style_reply(
                style,
                english=f"Open leads count is {open_count or 0}. Recent leads: {top_rows}.",
                tanglish=f"Open leads count {open_count or 0}. Recent leads: {top_rows}.",
                tamil=f"Open leads count {open_count or 0}. Recent leads: {top_rows}.",
                telugu=f"Open leads count {open_count or 0}. Recent leads: {top_rows}.",
                kannada=f"Open leads count {open_count or 0}. Recent leads: {top_rows}.",
            )
        return _ba_style_reply(
            style,
            english=f"Open leads count is {open_count or 0}.",
            tanglish=f"Open leads count {open_count or 0}.",
            tamil=f"Open leads count {open_count or 0}.",
            telugu=f"Open leads count {open_count or 0}.",
            kannada=f"Open leads count {open_count or 0}.",
        )

    if _contains_any(text, ["today meeting", "meetings", "meeting schedule"]):
        if not today_meetings:
            return _ba_style_reply(
                style,
                english="There are no meetings scheduled for today.",
                tanglish="Innaikku scheduled meetings illa.",
                tamil="இன்னைக்கு scheduled meetings illa.",
                telugu="ఈ రోజు scheduled meetings లేవు.",
                kannada="ಇವತ್ತು scheduled meetings ಇಲ್ಲ.",
            )
        meeting_text = "; ".join(
            f"{row.get('title') or 'Meeting'} - {row.get('meeting_time') or 'time not set'}"
            for row in today_meetings[:4]
        )
        return _ba_style_reply(
            style,
            english=f"There are {len(today_meetings)} meetings today: {meeting_text}.",
            tanglish=f"Innaikku {len(today_meetings)} meeting irukku: {meeting_text}.",
            tamil=f"இன்னைக்கு {len(today_meetings)} meeting இருக்கு: {meeting_text}.",
            telugu=f"ఈ రోజు {len(today_meetings)} meetings ఉన్నాయి: {meeting_text}.",
            kannada=f"ಇವತ್ತು {len(today_meetings)} meetings ಇವೆ: {meeting_text}.",
        )

    return ""


def _ba_assistant_direct_hr_answer(message: str, hr_context):
    text = str(message or "").strip().lower()
    style = _ba_detect_user_style(message)
    if not isinstance(hr_context, dict):
        return ""
    overtime_rows = hr_context.get("overtime_this_month") if isinstance(hr_context.get("overtime_this_month"), list) else []
    increments = hr_context.get("recent_increments") if isinstance(hr_context.get("recent_increments"), list) else []
    checked_in = hr_context.get("employees_checked_in_today") if isinstance(hr_context.get("employees_checked_in_today"), list) else []
    if _ba_is_large_list_request(message) and _contains_any(text, ["employee", "employees", "attendance", "checked in", "increment", "increments", "overtime", "ot", "leave"]):
        list_count = 0
        subject = "employee records"
        if _contains_any(text, ["increment", "increments"]):
            list_count = len(increments)
            subject = "increment records"
        elif _contains_any(text, ["attendance", "checked in", "leave"]):
            list_count = len(checked_in)
            subject = "attendance records"
        elif _contains_any(text, ["overtime", "ot"]):
            list_count = len(overtime_rows)
            subject = "overtime records"
        large_reply = _ba_large_dataset_response(style, subject, list_count, threshold=5)
        if large_reply:
            return large_reply

    if _contains_any(text, ["overtime", "ot"]):
        if not overtime_rows:
            return _ba_style_reply(
                style,
                english="There is no overtime data for this month.",
                tanglish="Indha month-ku overtime data illa.",
                tamil="இந்த month-ku overtime data illa.",
                telugu="ఈ నెలకి overtime data లేదు.",
                kannada="ಈ ತಿಂಗಳಿಗೆ overtime data ಇಲ್ಲ.",
            )
        top_row = sorted(overtime_rows, key=lambda row: int(row.get("overtime_minutes") or 0), reverse=True)[0]
        return _ba_style_reply(
            style,
            english=f"{top_row.get('employee_name') or 'Unknown'} has the highest overtime with {top_row.get('overtime_minutes') or 0} minutes.",
            tanglish=f"{top_row.get('employee_name') or 'Unknown'} kitta adhigama overtime irukku. {top_row.get('overtime_minutes') or 0} minutes.",
            tamil=f"{top_row.get('employee_name') or 'Unknown'} கிட்ட அதிகமான overtime இருக்கு. {top_row.get('overtime_minutes') or 0} minutes.",
            telugu=f"{top_row.get('employee_name') or 'Unknown'} దగ్గర అత్యధిక overtime ఉంది. {top_row.get('overtime_minutes') or 0} minutes.",
            kannada=f"{top_row.get('employee_name') or 'Unknown'} ಬಳಿ ಹೆಚ್ಚು overtime ಇದೆ. {top_row.get('overtime_minutes') or 0} minutes.",
        )

    if _contains_any(text, ["increment"]):
        if not increments:
            return _ba_style_reply(
                style,
                english="Recent increment data is not available.",
                tanglish="Recent increment data illa.",
                tamil="Recent increment data illa.",
                telugu="Recent increment data అందుబాటులో లేదు.",
                kannada="Recent increment data ಲಭ್ಯ ಇಲ್ಲ.",
            )
        names = ", ".join(str(row.get("employee_name") or "").strip() for row in increments[:5] if str(row.get("employee_name") or "").strip())
        return _ba_style_reply(
            style,
            english=f"Recent increments list: {names}.",
            tanglish=f"Recent increments list-la: {names}.",
            tamil=f"Recent increments list-la: {names}.",
            telugu=f"Recent increments list: {names}.",
            kannada=f"Recent increments list: {names}.",
        )

    if _contains_any(text, ["attendance", "checked in", "leave"]):
        if checked_in:
            return _ba_style_reply(
                style,
                english=f"Employees checked in today: {', '.join(checked_in[:8])}.",
                tanglish=f"Innaikku checked-in employees: {', '.join(checked_in[:8])}.",
                tamil=f"இன்னைக்கு checked-in employees: {', '.join(checked_in[:8])}.",
                telugu=f"ఈ రోజు checked-in employees: {', '.join(checked_in[:8])}.",
                kannada=f"ಇವತ್ತು checked-in employees: {', '.join(checked_in[:8])}.",
            )
        return _ba_style_reply(
            style,
            english="There is no attendance check-in data for today.",
            tanglish="Innaikku attendance check-in data illa.",
            tamil="இன்னைக்கு attendance check-in data illa.",
            telugu="ఈ రోజు attendance check-in data లేదు.",
            kannada="ಇವತ್ತು attendance check-in data ಇಲ್ಲ.",
        )

    return ""


def _ba_assistant_direct_accounts_answer(message: str, accounts_context):
    text = str(message or "").strip().lower()
    style = _ba_detect_user_style(message)
    if not isinstance(accounts_context, dict):
        return ""
    currency_code = str(accounts_context.get("currency") or "INR").strip().upper() or "INR"
    country_name = str(accounts_context.get("country") or "India").strip()
    total_invoices_count = int(accounts_context.get("total_invoices_count") or 0)
    total_invoices_amount = str(accounts_context.get("total_invoices_amount") or "0.00")
    paid_invoices_count = int(accounts_context.get("paid_invoices_count") or 0)
    pending_invoices_count = int(accounts_context.get("pending_invoices_count") or 0)
    this_month_invoices_count = int(accounts_context.get("this_month_invoices_count") or 0)
    invoice_status_summary = accounts_context.get("invoice_status_summary") if isinstance(accounts_context.get("invoice_status_summary"), list) else []
    if _ba_is_large_list_request(message) and _contains_any(text, ["invoice", "invoices", "billing", "sales"]):
        subject = "invoices" if _contains_any(text, ["invoice", "invoices"]) else "billing records"
        large_reply = _ba_large_dataset_response(style, subject, total_invoices_count, threshold=5)
        if large_reply:
            return large_reply
    if _contains_any(
        text,
        [
            "total invoice",
            "invoice total",
            "how many invoice",
            "how many invoices",
            "invoice count",
            "motham invoice",
            "motham invoices",
            "evlo invoice",
            "yevalo invoice",
            "total-a invoice",
        ],
    ):
        status_text = ", ".join(
            f"{row.get('status')}: {int(row.get('count') or 0)}"
            for row in invoice_status_summary[:4]
        )
        if status_text:
            return _ba_style_reply(
                style,
                english=f"There are {total_invoices_count} invoices in your application right now. Status split: {status_text}.",
                tanglish=f"Namma application-la ippo total {total_invoices_count} invoices irukku. Status split: {status_text}.",
                tamil=f"நம்ம application-la இப்போ total {total_invoices_count} invoices இருக்கு. Status split: {status_text}.",
                telugu=f"మీ application-lo ippudu total {total_invoices_count} invoices ఉన్నాయి. Status split: {status_text}.",
                kannada=f"Nimma application-nalli iga total {total_invoices_count} invoices ive. Status split: {status_text}.",
            )
        return _ba_style_reply(
            style,
            english=f"There are {total_invoices_count} invoices in your application right now.",
            tanglish=f"Namma application-la ippo total {total_invoices_count} invoices irukku.",
            tamil=f"நம்ம application-la இப்போ total {total_invoices_count} invoices இருக்கு.",
            telugu=f"మీ application-lo ippudu total {total_invoices_count} invoices ఉన్నాయి.",
            kannada=f"Nimma application-nalli iga total {total_invoices_count} invoices ive.",
        )
    if _contains_any(
        text,
        [
            "invoice total amount",
            "total invoice value",
            "invoice amount total",
            "all invoice amount",
            "motham invoice amount",
            "invoice values yevlo",
            "invoice values yevalo",
            "how much invoice amount",
        ],
    ):
        amount_display = _ba_format_currency_amount_display(currency_code, total_invoices_amount, country_name)
        amount_spoken = _ba_currency_amount_spoken(currency_code, total_invoices_amount, country_name)
        reply_text = _ba_style_reply(
            style,
            english=f"Total invoice amount in your application is {amount_display}.",
            tanglish=f"Namma application-la total invoice amount {amount_display}.",
            tamil=f"நம்ம application-la total invoice amount {amount_display}.",
            telugu=f"మీ application-lo total invoice amount {amount_display}.",
            kannada=f"Nimma application-nalli total invoice amount {amount_display}.",
        )
        tts_text = _ba_style_reply(
            style,
            english=f"Total invoice amount in your application is {amount_spoken}.",
            tanglish=f"Namma application-la total invoice amount {amount_spoken}.",
            tamil=f"Namma application-la total invoice amount {amount_spoken}.",
            telugu=f"Mee application-lo total invoice amount {amount_spoken}.",
            kannada=f"Nimma application-nalli total invoice amount {amount_spoken}.",
        )
        return _ba_direct_response(reply_text, tts_text)
    if _contains_any(
        text,
        [
            "paid invoices",
            "how many paid invoices",
            "paid invoice count",
            "motham paid invoice",
        ],
    ):
        return _ba_style_reply(
            style,
            english=f"There are {paid_invoices_count} paid invoices right now.",
            tanglish=f"Ippo {paid_invoices_count} paid invoices irukku.",
            tamil=f"இப்போ {paid_invoices_count} paid invoices இருக்கு.",
            telugu=f"ప్రస్తుతం {paid_invoices_count} paid invoices ఉన్నాయి.",
            kannada=f"Iga {paid_invoices_count} paid invoices ive.",
        )
    if _contains_any(
        text,
        [
            "pending invoices",
            "how many pending invoices",
            "pending invoice count",
            "unpaid invoices",
            "due invoices",
            "motham pending invoice",
        ],
    ):
        return _ba_style_reply(
            style,
            english=f"There are {pending_invoices_count} pending invoices right now.",
            tanglish=f"Ippo {pending_invoices_count} pending invoices irukku.",
            tamil=f"இப்போ {pending_invoices_count} pending invoices இருக்கு.",
            telugu=f"ప్రస్తుతం {pending_invoices_count} pending invoices ఉన్నాయి.",
            kannada=f"Iga {pending_invoices_count} pending invoices ive.",
        )
    if _contains_any(
        text,
        [
            "this month invoices",
            "how many invoices this month",
            "invoice this month",
            "indha month invoice",
            "this month invoice count",
        ],
    ):
        return _ba_style_reply(
            style,
            english=f"There are {this_month_invoices_count} invoices for this month.",
            tanglish=f"Indha month-ku {this_month_invoices_count} invoices irukku.",
            tamil=f"இந்த month-ku {this_month_invoices_count} invoices இருக்கு.",
            telugu=f"ఈ నెలకి {this_month_invoices_count} invoices ఉన్నాయి.",
            kannada=f"Ii tingalige {this_month_invoices_count} invoices ive.",
        )
    if _contains_any(text, ["billing", "sales summary", "today sales", "invoice", "amount sold"]):
        return _ba_style_reply(
            style,
            english=(
                f"Today's sales orders count is {accounts_context.get('today_sales_orders_count') or 0}. "
                f"Total value is {accounts_context.get('today_sales_total') or '0.00'}."
            ),
            tanglish=(
                f"Today's sales orders count {accounts_context.get('today_sales_orders_count') or 0}. "
                f"Total value {accounts_context.get('today_sales_total') or '0.00'}."
            ),
            tamil=(
                f"Today's sales orders count {accounts_context.get('today_sales_orders_count') or 0}. "
                f"Total value {accounts_context.get('today_sales_total') or '0.00'}."
            ),
            telugu=(
                f"Today's sales orders count {accounts_context.get('today_sales_orders_count') or 0}. "
                f"Total value {accounts_context.get('today_sales_total') or '0.00'}."
            ),
            kannada=(
                f"Today's sales orders count {accounts_context.get('today_sales_orders_count') or 0}. "
                f"Total value {accounts_context.get('today_sales_total') or '0.00'}."
            ),
        )
    return ""


def _ba_assistant_direct_ticketing_answer(message: str, ticketing_context):
    text = str(message or "").strip().lower()
    style = _ba_detect_user_style(message)
    if not isinstance(ticketing_context, dict):
        return ""
    open_tickets = ticketing_context.get("open_tickets") if isinstance(ticketing_context.get("open_tickets"), list) else []
    total_tickets_count = int(ticketing_context.get("total_tickets_count") or len(open_tickets) or 0)
    open_tickets_count = int(ticketing_context.get("open_tickets_count") or len(open_tickets) or 0)
    status_summary = ticketing_context.get("ticket_status_summary") if isinstance(ticketing_context.get("ticket_status_summary"), list) else []
    if _ba_is_large_list_request(message) and _contains_any(text, ["ticket", "tickets"]):
        large_reply = _ba_large_dataset_response(style, "tickets", open_tickets_count, threshold=5)
        if large_reply:
            return large_reply
    if _contains_any(text, ["how many tickets", "ticket count", "total tickets", "open tickets", "pending tickets", "ticket summary", "motham ticket", "motham ticketing", "evlo ticket", "yevalo ticket", "ticketing count", "ticketing la motham"]):
        is_total_question = _contains_any(text, ["how many", "total", "motham", "evlo", "yevalo", "count"])
        if is_total_question and not _contains_any(text, ["open ticket", "pending ticket"]):
            return _ba_style_reply(
                style,
                english=f"Right now your ticketing has {total_tickets_count} tickets in total. Out of that, {open_tickets_count} are currently open.",
                tanglish=f"Namma ticketing-la ippo total-a {total_tickets_count} tickets irukku. Adhula {open_tickets_count} tickets open-a irukku.",
                tamil=f"நம்ம ticketing-la இப்போ total-a {total_tickets_count} tickets இருக்கு. அதுல {open_tickets_count} tickets open-a இருக்கு.",
                telugu=f"మీ ticketing-lo ippudu total-ga {total_tickets_count} tickets ఉన్నాయి. వాటిలో {open_tickets_count} tickets open-ga ఉన్నాయి.",
                kannada=f"Nimma ticketing-nalli iga total {total_tickets_count} tickets ive. Adaralli {open_tickets_count} tickets open ide.",
            )
        status_text = ", ".join(
            f"{row.get('status')}: {int(row.get('count') or 0)}"
            for row in status_summary[:4]
        )
        if status_text:
            return _ba_style_reply(
                style,
                english=f"Right now you have {open_tickets_count} open tickets. Status split: {status_text}.",
                tanglish=f"Ippo unga kitta {open_tickets_count} open tickets irukku. Status split: {status_text}.",
                tamil=f"இப்போ உங்களிடம் {open_tickets_count} open tickets இருக்கு. Status split: {status_text}.",
                telugu=f"ప్రస్తుతం మీ దగ్గర {open_tickets_count} open tickets ఉన్నాయి. Status split: {status_text}.",
                kannada=f"Iga nimma hattira {open_tickets_count} open tickets ive. Status split: {status_text}.",
            )
        return _ba_style_reply(
            style,
            english=f"Right now you have {open_tickets_count} open tickets.",
            tanglish=f"Ippo unga kitta {open_tickets_count} open tickets irukku.",
            tamil=f"இப்போ உங்களிடம் {open_tickets_count} open tickets இருக்கு.",
            telugu=f"ప్రస్తుతం మీ దగ్గర {open_tickets_count} open tickets ఉన్నాయి.",
            kannada=f"Iga nimma hattira {open_tickets_count} open tickets ive.",
        )
    if _contains_any(text, ["ticket list", "ticket details", "which tickets", "show tickets"]):
        if open_tickets_count <= 5 and open_tickets:
            ticket_text = ", ".join(
                f"{row.get('title') or 'Ticket'} ({row.get('status') or 'Unknown'})"
                for row in open_tickets[:5]
            )
            return _ba_style_reply(
                style,
                english=f"Your open tickets are: {ticket_text}.",
                tanglish=f"Unga open tickets: {ticket_text}.",
                tamil=f"உங்களோட open tickets: {ticket_text}.",
                telugu=f"మీ open tickets ఇవి: {ticket_text}.",
                kannada=f"Nimma open tickets ivu: {ticket_text}.",
            )
        large_reply = _ba_large_dataset_response(style, "tickets", open_tickets_count, threshold=5)
        if large_reply:
            return large_reply
    return ""


def _ba_assistant_direct_general_answer(message: str, scoped_context):
    text = str(message or "").strip().lower()
    style = _ba_detect_user_style(message)
    if not isinstance(scoped_context, dict):
        return ""
    org_summary = scoped_context.get("org_summary") if isinstance(scoped_context.get("org_summary"), dict) else {}
    scope = scoped_context.get("scope") if isinstance(scoped_context.get("scope"), dict) else {}
    product_capabilities = scoped_context.get("product_capabilities") if isinstance(scoped_context.get("product_capabilities"), dict) else {}
    total_users = int(org_summary.get("total_users") or 0)
    employee_users = int(org_summary.get("employee_users") or 0)
    if _ba_is_large_list_request(message) and _contains_any(text, ["user", "users", "employee", "employees", "member", "members", "name", "names"]):
        large_reply = _ba_large_dataset_response(style, "employee users", employee_users or total_users, threshold=5)
        if large_reply:
            return large_reply

    if _contains_any(
        text,
        [
            "how many users",
            "total users",
            "user count",
            "employee count",
            "how many employees",
            "motham users",
            "motham employee",
            "organization users",
            "org users",
        ],
    ):
        admin_users = int(org_summary.get("admin_users") or 0)
        org_name = str(org_summary.get("organization_name") or "your organization").strip()
        return _ba_style_reply(
            style,
            english=(
                f"{org_name} currently has {total_users} active users. "
                f"That includes {admin_users} admin users and {employee_users} employee users."
            ),
            tanglish=(
                f"{org_name}-la ippo {total_users} active users irukanga. "
                f"Adhula {admin_users} admin users, {employee_users} employee users."
            ),
            tamil=(
                f"{org_name}-la இப்போ {total_users} active users இருக்காங்க. "
                f"அதுல {admin_users} admin users, {employee_users} employee users."
            ),
            telugu=(
                f"{org_name}లో ప్రస్తుతం {total_users} active users ఉన్నారు. "
                f"అందులో {admin_users} admin users, {employee_users} employee users."
            ),
            kannada=(
                f"{org_name}ನಲ್ಲಿ ಈಗ {total_users} active users ಇದ್ದಾರೆ. "
                f"ಅದರಲ್ಲ {admin_users} admin users, {employee_users} employee users."
            ),
        )

    if _contains_any(
        text,
        [
            "total user list",
            "user list",
            "employee list",
            "list all users",
            "show users list",
            "show employee list",
            "members list",
            "member list",
            "users oda list",
            "employee oda list",
            "total user peru",
            "employee peru",
        ],
    ):
        org_name = str(org_summary.get("organization_name") or "your organization").strip()
        sample_user_names = [
            str(item or "").strip()
            for item in (org_summary.get("sample_user_names") or [])
            if str(item or "").strip()
        ]
        if total_users <= 5 and sample_user_names:
            names_text = ", ".join(sample_user_names[:5])
            return _ba_style_reply(
                style,
                english=f"{org_name} has {total_users} users. Their names are: {names_text}.",
                tanglish=f"{org_name}-la {total_users} users irukanga. Avanga peru: {names_text}.",
                tamil=f"{org_name}-la {total_users} users இருக்காங்க. அவங்க பேரு: {names_text}.",
                telugu=f"{org_name}లో {total_users} users ఉన్నారు. వారి పేర్లు: {names_text}.",
                kannada=f"{org_name}ನಲ್ಲಿ {total_users} users ಇದ್ದಾರೆ. Avra hesaru: {names_text}.",
            )
        if employee_users >= 100:
            return _ba_style_reply(
                style,
                english=f"There are more than 100 employee users. I cannot list all names here. Please check the application manually.",
                tanglish=f"Employee users 100-ku mela irukanga. Ella peraiyum inga solla mudiyadhu. Application-la manual-a check pannunga.",
                tamil=f"Employee users 100-ku mela இருக்காங்க. எல்லா பேரையும் இங்க சொல்ல முடியாது. Application-la manual-a check pannunga.",
                telugu=f"Employee users 100 కంటే ఎక్కువ మంది ఉన్నారు. అన్ని పేర్లు ఇక్కడ చెప్పలేను. Applicationలో manual-ga check చేయండి.",
                kannada=f"Employee users 100ಕ್ಕಿಂತ ಹೆಚ್ಚು ಇದ್ದಾರೆ. Ella hesarugalan illi helakke agalla. Application-nalli manual-a check madi.",
            )
        if total_users > 5:
            return _ba_style_reply(
                style,
                english=f"{org_name} has {total_users} users. Showing every name here will be too long. Please check the application manually.",
                tanglish=f"{org_name}-la {total_users} users irukanga. Ella peraiyum inga solla romba long aagum. Application-la manual-a check pannunga.",
                tamil=f"{org_name}-la {total_users} users இருக்காங்க. எல்லா பேரையும் இங்க சொல்ல ரொம்ப long ஆகும். Application-la manual-a check pannunga.",
                telugu=f"{org_name}లో {total_users} users ఉన్నారు. అన్ని పేర్లు ఇక్కడ చెప్పితే చాలా long అవుతుంది. Applicationలో manual-ga check చేయండి.",
                kannada=f"{org_name}ನಲ್ಲಿ {total_users} users ಇದ್ದಾರೆ. Ella hesarugalan illi helidre tumba long agutte. Application-nalli manual-a check madi.",
            )
        return _ba_style_reply(
            style,
            english="User list is not available right now.",
            tanglish="User list ippo available illa.",
            tamil="User list இப்போ available illa.",
            telugu="User list ఇప్పుడు available లేదు.",
            kannada="User list iga available illa.",
        )

    if _contains_any(
        text,
        [
            "meeting schedule iruka",
            "meeting schedules iruka",
            "meetings option",
            "meetings feature",
            "application la meeting",
            "app la meeting",
            "namma application la meeting",
            "our application has meetings",
            "meeting shedule option",
        ],
    ):
        crm_meetings_available = bool(product_capabilities.get("crm_meetings_available"))
        plan_name = str(product_capabilities.get("active_plan_name") or "").strip() or "current plan"
        meeting_user_types = product_capabilities.get("meeting_enabled_user_types") if isinstance(product_capabilities.get("meeting_enabled_user_types"), list) else []
        meeting_user_type_labels = ", ".join(
            str(row.get("label") or "").strip()
            for row in meeting_user_types
            if str(row.get("label") or "").strip()
        )
        availability_line = (
            "For this organization, it is currently enabled inside CRM."
            if crm_meetings_available
            else "For this organization, it is not currently enabled in the active modules."
        )
        user_type_line = (
            f"Plan-wise, meeting access is available for: {meeting_user_type_labels}."
            if meeting_user_type_labels
            else "Plan-wise user type meeting access details are not available right now."
        )
        return _ba_style_reply(
            style,
            english=(
                f"Yes, the application has a Meetings feature inside the CRM module. "
                f"{availability_line} Active plan: {plan_name}. {user_type_line}"
            ),
            tanglish=(
                f"Yes, namma application-la CRM module ullae Meetings feature irukku. "
                f"{'Indha organization-ku idhu ippo CRM-la enabled-a irukku.' if crm_meetings_available else 'Indha organization-ku idhu active modules-la ippo enabled illa.'} "
                f"Current plan {plan_name}. "
                f"{f'Plan-wise paatha {meeting_user_type_labels} users-ku meeting access irukku.' if meeting_user_type_labels else 'Plan-wise meeting access details ippo available illa.'}"
            ),
            tamil=(
                f"Yes, நம்ம application-la CRM module உள்ளே Meetings feature இருக்கு. "
                f"{'இந்த organization-ku இது இப்போ CRM-la enabled-a இருக்கு.' if crm_meetings_available else 'இந்த organization-ku இது active modules-la இப்போ enabled இல்லை.'} "
                f"Current plan {plan_name}. "
                f"{f'Plan-wise பார்த்தா {meeting_user_type_labels} users-ku meeting access இருக்கு.' if meeting_user_type_labels else 'Plan-wise meeting access details இப்போ available இல்லை.'}"
            ),
            telugu=(
                f"Yes, mana application-lo CRM module lopala Meetings feature undi. "
                f"{'Ee organization-ki idi ippudu CRM-lo enabled ga undi.' if crm_meetings_available else 'Ee organization-ki idi active modules-lo ippudu enabled ledu.'} "
                f"Current plan {plan_name}. "
                f"{f'Plan-wise chusthe {meeting_user_type_labels} users-ki meeting access undi.' if meeting_user_type_labels else 'Plan-wise meeting access details ippudu available levu.'}"
            ),
            kannada=(
                f"Yes, namma application-nalli CRM module olage Meetings feature ide. "
                f"{'Ee organization-ge idu iga CRM-nalli enabled ide.' if crm_meetings_available else 'Ee organization-ge idu active modules-nalli iga enabled illa.'} "
                f"Current plan {plan_name}. "
                f"{f'Plan-wise nodidre {meeting_user_type_labels} users-ge meeting access ide.' if meeting_user_type_labels else 'Plan-wise meeting access details iga available illa.'}"
            ),
        )

    if _contains_any(
        text,
        [
            "plan wise",
            "who has meeting access",
            "which user type has meeting access",
            "meeting access yaarukku",
            "meeting features eruku",
            "entha org user",
            "crm user ku meeting",
            "hrm user ku meeting",
            "full access user ku meeting",
        ],
    ):
        plan_name = str(product_capabilities.get("active_plan_name") or "").strip() or "current plan"
        meeting_user_types = product_capabilities.get("meeting_enabled_user_types") if isinstance(product_capabilities.get("meeting_enabled_user_types"), list) else []
        labels = [
            str(row.get("label") or "").strip()
            for row in meeting_user_types
            if str(row.get("label") or "").strip()
        ]
        if labels:
            labels_text = ", ".join(labels)
            return _ba_style_reply(
                style,
                english=f"Under the active plan {plan_name}, meeting access is available for these user types: {labels_text}.",
                tanglish=f"Active plan {plan_name} basis-la paatha, meeting access {labels_text} users-ku irukku.",
                tamil=f"Active plan {plan_name} basis-la பார்த்தா, meeting access {labels_text} users-ku இருக்கு.",
                telugu=f"Active plan {plan_name} ప్రకారం, meeting access {labels_text} users-ki ఉంది.",
                kannada=f"Active plan {plan_name} prakara, meeting access {labels_text} users-ge ide.",
            )
        return _ba_style_reply(
            style,
            english=f"I could not find any user-type meeting access mapping under the active plan {plan_name} right now.",
            tanglish=f"Active plan {plan_name}-ku user-type meeting access mapping ippo kidaikala.",
            tamil=f"Active plan {plan_name}-ku user-type meeting access mapping இப்போ கிடைக்கல.",
            telugu=f"Active plan {plan_name}కి user-type meeting access mapping ఇప్పుడు దొరకలేదు.",
            kannada=f"Active plan {plan_name}-ge user-type meeting access mapping iga sigalilla.",
        )

    if _contains_any(
        text,
        [
            "crm la enna iruku",
            "crm module features",
            "what is in crm",
            "crm options",
        ],
    ):
        return _ba_style_reply(
            style,
            english="CRM includes leads, contacts, teams, deals, sales orders, follow-ups, meetings, and reports.",
            tanglish="CRM-la leads, contacts, teams, deals, sales orders, follow-ups, meetings, reports ellam irukku.",
            tamil="CRM-la leads, contacts, teams, deals, sales orders, follow-ups, meetings, reports ellam irukku.",
            telugu="CRMలో leads, contacts, teams, deals, sales orders, follow-ups, meetings, reports ఉన్నాయి.",
            kannada="CRMನಲ್ಲಿ leads, contacts, teams, deals, sales orders, follow-ups, meetings, reports ಇವೆ.",
        )

    if _contains_any(text, ["restrictions", "access level", "what can i access", "what data can i access"]):
        allowed_sections = ", ".join(scope.get("allowed_sections") or []) or "no modules"
        label = str(scope.get("label") or "Business User").strip()
        return _ba_style_reply(
            style,
            english=f"Your current access is {label}. You can access these modules: {allowed_sections}.",
            tanglish=f"Ungaloda current access {label}. Neenga use panna mudiyura modules: {allowed_sections}.",
            tamil=f"உங்களோட current access {label}. நீங்க use பண்ண முடிஞ்ச modules: {allowed_sections}.",
            telugu=f"మీ current access {label}. మీరు access చేయగల modules: {allowed_sections}.",
            kannada=f"ನಿಮ್ಮ current access {label}. ನೀವು access ಮಾಡಬಹುದು ಅನ್ನೋ modules: {allowed_sections}.",
        )

    return ""


def _ba_assistant_direct_answer(message: str, scoped_context):
    if not isinstance(scoped_context, dict):
        return ""
    route = _ba_assistant_question_route(message, scoped_context.get("scope") or {})
    if route == "crm":
        return _ba_assistant_direct_crm_answer(message, scoped_context.get("crm"))
    if route == "hr":
        return _ba_assistant_direct_hr_answer(message, scoped_context.get("hr"))
    if route == "accounts":
        return _ba_assistant_direct_accounts_answer(message, scoped_context.get("accounts"))
    if route == "ticketing":
        return _ba_assistant_direct_ticketing_answer(message, scoped_context.get("ticketing"))
    return _ba_assistant_direct_general_answer(message, scoped_context)


def _ba_normalize_direct_response_payload(value):
    if isinstance(value, dict):
        reply_text = str(value.get("reply") or "").strip()
        tts_text = str(value.get("tts_text") or reply_text).strip() or reply_text
        if reply_text:
            return {"reply": reply_text, "tts_text": tts_text}
        return None
    reply_text = str(value or "").strip()
    if not reply_text:
        return None
    return {"reply": reply_text, "tts_text": reply_text}


def _get_org_membership(user: User, org: Organization):
    if not user or not user.is_authenticated or not org:
        return None
    return (
        OrganizationUser.objects
        .filter(organization=org, user=user, is_active=True, is_deleted=False)
        .only("id", "role", "department", "employee_role", "user_type", "is_active", "is_deleted")
        .first()
    )


def _can_manage_payroll(user: User, org: Organization = None):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if org and _hr_section_access_level(user, org, "hr") == "Full Access":
        return True
    profile = UserProfile.objects.filter(user=user).only("role").first()
    role = str(profile.role or "").strip().lower() if profile else ""
    if role in {"company_admin", "org_admin", "superadmin", "super_admin"}:
        return True
    membership = _get_org_membership(user, org)
    if not membership:
        return False
    employee_role = str(membership.employee_role or "").strip().lower()
    department = str(membership.department or "").strip().lower()
    return employee_role == "hr manager" or department == "hr"


def _can_view_salary_history(user: User, org: Organization = None):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if org and _hr_section_access_level(user, org, "hr") == "Full Access":
        return True
    profile = UserProfile.objects.filter(user=user).only("role").first()
    role = str(profile.role or "").strip().lower() if profile else ""
    if role in {"company_admin", "org_admin", "superadmin", "super_admin"}:
        return True
    membership = _get_org_membership(user, org)
    if not membership:
        return False
    membership_role = str(membership.role or "").strip().lower()
    employee_role = str(membership.employee_role or "").strip().lower()
    return membership_role == "company_admin" or employee_role == "hr manager"


def _can_manage_attendance_geo_settings(user: User, org: Organization = None):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if org and _hr_section_access_level(user, org, "hr") == "Full Access":
        return True
    membership = _get_org_membership(user, org)
    return bool(membership and _normalize_admin_role(getattr(membership, "role", "")) in {"company_admin", "org_admin", "owner"})


def _serialize_attendance_geo_setting(setting):
    return {
        "enabled": bool(getattr(setting, "enabled", False)),
        "location_name": str(getattr(setting, "location_name", "") or "").strip(),
        "latitude": float(setting.latitude) if getattr(setting, "latitude", None) is not None else None,
        "longitude": float(setting.longitude) if getattr(setting, "longitude", None) is not None else None,
        "radius_meters": int(getattr(setting, "radius_meters", 0) or 0),
        "allow_outside_fence": bool(getattr(setting, "allow_outside_fence", False)),
        "require_gps": bool(getattr(setting, "require_gps", True)),
        "google_maps_url": (
            f"https://www.google.com/maps?q={float(setting.latitude)},{float(setting.longitude)}"
            if getattr(setting, "latitude", None) is not None and getattr(setting, "longitude", None) is not None
            else ""
        ),
        "created_at": getattr(setting, "created_at", None).isoformat() if getattr(setting, "created_at", None) else None,
        "updated_at": getattr(setting, "updated_at", None).isoformat() if getattr(setting, "updated_at", None) else None,
    }


def _attendance_float(payload, key):
    raw = payload.get(key, None)
    if raw in (None, ""):
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        raise ValueError(f"invalid_{key}")


def _attendance_haversine_distance_meters(lat1, lon1, lat2, lon2):
    radius = 6371000.0
    phi1 = math.radians(float(lat1))
    phi2 = math.radians(float(lat2))
    delta_phi = math.radians(float(lat2) - float(lat1))
    delta_lambda = math.radians(float(lon2) - float(lon1))
    a = math.sin(delta_phi / 2.0) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2.0) ** 2
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return round(radius * c, 2)


def _attendance_validate_geo_payload(payload, *, require_reason=False):
    latitude = _attendance_float(payload, "latitude")
    longitude = _attendance_float(payload, "longitude")
    accuracy = _attendance_float(payload, "accuracy")
    outside_reason = str(payload.get("outside_reason") or "").strip()

    if latitude is None or longitude is None:
        raise ValueError("gps_coordinates_required")
    if latitude < -90 or latitude > 90:
        raise ValueError("invalid_latitude")
    if longitude < -180 or longitude > 180:
        raise ValueError("invalid_longitude")
    if accuracy is None:
        raise ValueError("gps_accuracy_required")
    if accuracy < 0:
        raise ValueError("invalid_accuracy")
    if require_reason and not outside_reason:
        raise ValueError("outside_reason_required")
    return latitude, longitude, accuracy, outside_reason


def _attendance_geo_response(entry, *, action, message):
    return {
        "ok": True,
        "action": action,
        "message": message,
        "attendance": _serialize_attendance_entry(entry),
    }


def _serialize_attendance_entry(entry):
    latest_proof = getattr(entry, "_prefetched_latest_proof", None)
    if latest_proof is None:
        latest_proof = entry.photo_proofs.order_by("-created_at", "-id").select_related("verified_by").first() if getattr(entry, "pk", None) else None

    return {
        "id": entry.id,
        "employee_name": entry.employee_name,
        "attendance_date": entry.attendance_date.isoformat() if entry.attendance_date else "",
        "checkin_time": entry.checkin_time.isoformat() if entry.checkin_time else None,
        "checkout_time": entry.checkout_time.isoformat() if entry.checkout_time else None,
        "checkin_latitude": float(entry.checkin_latitude) if entry.checkin_latitude is not None else None,
        "checkin_longitude": float(entry.checkin_longitude) if entry.checkin_longitude is not None else None,
        "checkin_accuracy": float(entry.checkin_accuracy) if entry.checkin_accuracy is not None else None,
        "checkout_latitude": float(entry.checkout_latitude) if entry.checkout_latitude is not None else None,
        "checkout_longitude": float(entry.checkout_longitude) if entry.checkout_longitude is not None else None,
        "checkout_accuracy": float(entry.checkout_accuracy) if entry.checkout_accuracy is not None else None,
        "checkin_distance_meters": float(entry.checkin_distance_meters) if entry.checkin_distance_meters is not None else None,
        "checkout_distance_meters": float(entry.checkout_distance_meters) if entry.checkout_distance_meters is not None else None,
        "checkin_inside_geofence": entry.checkin_inside_geofence,
        "checkout_inside_geofence": entry.checkout_inside_geofence,
        "geo_status": entry.geo_status,
        "attendance_mode": entry.attendance_mode,
        "face_verified": bool(entry.face_verified),
        "face_match_score": float(entry.face_match_score) if entry.face_match_score is not None else None,
        "external_verification_status": str(entry.external_verification_status or "").strip(),
        "verified_by": str(getattr(getattr(entry, "verified_by", None), "username", "") or "").strip(),
        "verified_at": entry.verified_at.isoformat() if entry.verified_at else None,
        "admin_notes": str(entry.admin_notes or "").strip(),
        "outside_reason": entry.outside_reason,
        "device_info": entry.device_info,
        "photo_proof": _serialize_attendance_photo_proof(latest_proof) if latest_proof else None,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
    }


def _serialize_face_recognition_setting(setting):
    return {
        "enabled": bool(getattr(setting, "enabled", False)),
        "require_internal_face": bool(getattr(setting, "require_internal_face", False)),
        "require_external_face": bool(getattr(setting, "require_external_face", False)),
        "min_match_score": float(getattr(setting, "min_match_score", 0.90) or 0.90),
        "photo_retention_days": int(getattr(setting, "photo_retention_days", 60) or 60),
        "allow_external_photo_proof": bool(getattr(setting, "allow_external_photo_proof", True)),
        "created_at": getattr(setting, "created_at", None).isoformat() if getattr(setting, "created_at", None) else None,
        "updated_at": getattr(setting, "updated_at", None).isoformat() if getattr(setting, "updated_at", None) else None,
    }


def _can_view_attendance_proof(user: User, org: Organization, proof: AttendancePhotoProof) -> bool:
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if _can_manage_attendance_geo_settings(user, org):
        return True
    membership = _get_org_membership(user, org)
    return bool(membership and membership.id == proof.employee_id)


def _serialize_attendance_photo_proof(proof: AttendancePhotoProof, user: Optional[User] = None, org: Optional[Organization] = None):
    can_view = bool(proof and user and org and _can_view_attendance_proof(user, org, proof))
    photo_deleted = bool(proof and not getattr(proof, "image", None) and getattr(proof, "expires_at", None) and getattr(proof, "expires_at") <= timezone.now())
    return {
        "id": proof.id,
        "attendance_mode": proof.attendance_mode,
        "face_verified": bool(proof.face_verified),
        "face_match_score": float(proof.face_match_score) if proof.face_match_score is not None else None,
        "gps_latitude": float(proof.gps_latitude) if proof.gps_latitude is not None else None,
        "gps_longitude": float(proof.gps_longitude) if proof.gps_longitude is not None else None,
        "gps_accuracy": float(proof.gps_accuracy) if proof.gps_accuracy is not None else None,
        "location_status": str(proof.location_status or "").strip(),
        "external_verification_status": str(proof.external_verification_status or "").strip(),
        "verified_by": str(getattr(getattr(proof, "verified_by", None), "username", "") or "").strip(),
        "verified_at": proof.verified_at.isoformat() if proof.verified_at else None,
        "admin_notes": str(proof.admin_notes or "").strip(),
        "expires_at": proof.expires_at.isoformat() if proof.expires_at else None,
        "created_at": proof.created_at.isoformat() if proof.created_at else None,
        "photo_url": proof.image.url if can_view and getattr(proof, "image", None) else None,
        "photo_deleted": photo_deleted,
        "photo_deleted_message": "Photo deleted after retention period." if photo_deleted else "",
        "can_view_photo": can_view,
    }


def cleanup_expired_attendance_photos(*, limit: int = 500) -> int:
    now = timezone.now()
    deleted_count = 0
    rows = list(
        AttendancePhotoProof.objects
        .filter(expires_at__isnull=False, expires_at__lte=now)
        .exclude(image="")
        .exclude(image__isnull=True)
        .order_by("expires_at", "id")[: max(1, int(limit or 500))]
    )
    for row in rows:
        if row.image:
            storage_name = row.image.name
            row.image.delete(save=False)
            if storage_name and default_storage.exists(storage_name):
                default_storage.delete(storage_name)
            row.save(update_fields=["image"])
            deleted_count += 1
    return deleted_count


def _ensure_default_module_catalog():
    # Keeps local/stale DBs usable even if seed migrations were not applied yet.
    # Avoid frequent write-lock contention on sqlite by doing minimal writes.
    defaults_by_slug = {row["slug"]: row for row in DEFAULT_ERP_MODULES}
    existing = {row.slug: row for row in Module.objects.filter(slug__in=defaults_by_slug.keys())}

    to_create = []
    to_update = []
    for slug, cfg in defaults_by_slug.items():
        row = existing.get(slug)
        if not row:
            to_create.append(
                Module(
                    slug=slug,
                    name=cfg["name"],
                    is_active=True,
                    sort_order=cfg["sort_order"],
                )
            )
            continue
        changed = False
        if row.name != cfg["name"]:
            row.name = cfg["name"]
            changed = True
        if row.sort_order != cfg["sort_order"]:
            row.sort_order = cfg["sort_order"]
            changed = True
        if not row.is_active:
            row.is_active = True
            changed = True
        if changed:
            to_update.append(row)

    if to_create:
        Module.objects.bulk_create(to_create, ignore_conflicts=True)
    if to_update:
        Module.objects.bulk_update(to_update, ["name", "is_active", "sort_order"])


def _ensure_org_modules(org):
    try:
        _ensure_default_module_catalog()
    except (OperationalError, DatabaseError):
        # If sqlite is temporarily locked, proceed with existing module catalog.
        pass
    active_modules = list(Module.objects.filter(is_active=True).order_by("sort_order", "name"))
    if not active_modules:
        return []
    existing = set(
        OrganizationModule.objects.filter(organization=org).values_list("module_id", flat=True)
    )
    pending = [
        OrganizationModule(organization=org, module=module, enabled=True)
        for module in active_modules
        if module.id not in existing
    ]
    if pending:
        OrganizationModule.objects.bulk_create(pending)
    return active_modules


def _default_plan_module_slugs(plan_name):
    key = str(plan_name or "").strip().lower()
    if "free" in key:
        return ["crm", "hrm", "projects", "accounts", "subscriptions"]
    if "starter" in key:
        return ["crm", "hrm", "projects", "accounts", "subscriptions"]
    if "growth" in key:
        return ["crm", "hrm", "projects", "accounts", "subscriptions", "ticketing", "stocks"]
    if "pro" in key:
        return ["crm", "hrm", "projects", "accounts", "subscriptions", "ticketing", "stocks"]
    return list(ERP_MODULE_SLUG_SET)


def _is_free_plan(plan):
    if not plan:
        return False
    prices = [
        plan.monthly_price or 0,
        plan.yearly_price or 0,
        plan.usd_monthly_price or 0,
        plan.usd_yearly_price or 0,
    ]
    return all(price <= 0 for price in prices)


def _is_business_autopilot_free_plan(plan):
    if not plan:
        return False
    product_slug = str(getattr(getattr(plan, "product", None), "slug", "") or "").strip().lower()
    if product_slug not in BUSINESS_AUTOPILOT_PRODUCT_SLUG_ALIASES:
        return False
    if _is_free_plan(plan):
        return True
    plan_name = str(getattr(plan, "name", "") or "").strip().lower()
    if "free" in plan_name:
        return True
    limits = getattr(plan, "limits", {}) or {}
    plan_tier = str(limits.get("plan_tier") or limits.get("tier") or "").strip().lower()
    return plan_tier == "free"


def _get_effective_employee_limit(plan):
    if not plan:
        return 0
    try:
        base_limit = int(getattr(plan, "employee_limit", 0) or 0)
    except (TypeError, ValueError):
        base_limit = 0
    if _is_business_autopilot_free_plan(plan):
        # Business Autopilot free trial: base 1 + extra 2 => total 3 users.
        return 3
    product_slug = str(getattr(getattr(plan, "product", None), "slug", "") or "").strip().lower()
    if product_slug in BUSINESS_AUTOPILOT_PRODUCT_SLUG_ALIASES:
        # Business Autopilot paid plans: base 1 user; additional users require paid add-ons.
        return 1
    return base_limit


def _trial_tier_module_slugs(trial_tier):
    tier = str(trial_tier or "").strip().lower()
    if tier in {"pro", "all", "full"}:
        return set(ERP_MODULE_SLUG_SET)
    if tier == "growth":
        return set(_default_plan_module_slugs("growth"))
    if tier == "starter":
        return set(_default_plan_module_slugs("starter"))
    if tier == "free":
        return set(_default_plan_module_slugs("free"))
    return None


def _get_active_erp_subscription(org):
    if not org:
        return None
    now = timezone.now()
    rows = (
        OrgSubscription.objects
        .filter(organization=org, plan__product__slug__in=BUSINESS_AUTOPILOT_PRODUCT_SLUG_ALIASES)
        .select_related("plan", "plan__product")
        .order_by("-start_date", "-id")
    )
    for row in rows:
        status = str(row.status or "").lower()
        if status not in {"active", "trialing"}:
            continue
        if row.end_date and row.end_date < now:
            continue
        return row
    return None


def _get_plan_erp_module_slugs(plan, subscription=None):
    if not plan:
        return set(ERP_MODULE_SLUG_SET)
    features = plan.features or {}
    limits = plan.limits or {}

    # Free trial plans can expose higher-tier modules for evaluation.
    trial_tier = (
        features.get("trial_features")
        or features.get("trial_features_tier")
        or limits.get("trial_features")
        or limits.get("trial_features_tier")
    )
    if _is_free_plan(plan):
        trial_modules = _trial_tier_module_slugs(trial_tier)
        if trial_modules:
            return trial_modules

    if str(getattr(subscription, "status", "") or "").strip().lower() == "trialing":
        trial_modules = _trial_tier_module_slugs(trial_tier)
        if trial_modules:
            return trial_modules

    configured = features.get("erp_enabled_modules")
    if isinstance(configured, list):
        normalized = {
            str(item or "").strip().lower()
            for item in configured
            if str(item or "").strip().lower() in ERP_MODULE_SLUG_SET
        }
        # Backward compatibility: subscriptions previously lived under Accounts.
        if "accounts" in normalized:
            normalized.add("subscriptions")
        if normalized:
            return normalized
    return set(_default_plan_module_slugs(plan.name))


def _serialize_employee_roles(org):
    rows = (
        OrganizationEmployeeRole.objects
        .filter(organization=org, is_active=True)
        .order_by("name")
    )
    return [{"id": row.id, "name": row.name} for row in rows]


def _serialize_departments(org):
    rows = (
        OrganizationDepartment.objects
        .filter(organization=org, is_active=True)
        .order_by("name")
    )
    return [{"id": row.id, "name": row.name} for row in rows]


def _normalize_membership_status(member: OrganizationUser):
    if not member:
        return OrganizationUser.STATUS_DELETED
    if getattr(member, "is_deleted", False):
        return OrganizationUser.STATUS_DELETED
    raw_status = str(getattr(member, "status", "") or "").strip().lower()
    if raw_status == OrganizationUser.STATUS_RESIGNED or getattr(member, "resigned_at", None):
        return OrganizationUser.STATUS_RESIGNED
    if not getattr(member, "is_active", False) or not getattr(getattr(member, "user", None), "is_active", True):
        return OrganizationUser.STATUS_INACTIVE
    if raw_status in {
        OrganizationUser.STATUS_ACTIVE,
        OrganizationUser.STATUS_INACTIVE,
        OrganizationUser.STATUS_RESIGNED,
        OrganizationUser.STATUS_DELETED,
    }:
        return raw_status
    return OrganizationUser.STATUS_ACTIVE if getattr(member, "is_active", False) else OrganizationUser.STATUS_INACTIVE


def _set_membership_status(member: OrganizationUser, status: str, *, changed_by: Optional[User] = None, commit: bool = True):
    normalized_status = str(status or "").strip().lower()
    if normalized_status not in {
        OrganizationUser.STATUS_ACTIVE,
        OrganizationUser.STATUS_INACTIVE,
        OrganizationUser.STATUS_RESIGNED,
        OrganizationUser.STATUS_DELETED,
    }:
        normalized_status = OrganizationUser.STATUS_INACTIVE
    now = timezone.now()
    member.status = normalized_status
    member.status_changed_at = now
    member.is_deleted = normalized_status == OrganizationUser.STATUS_DELETED
    member.deleted_at = now if normalized_status == OrganizationUser.STATUS_DELETED else None
    member.is_active = normalized_status == OrganizationUser.STATUS_ACTIVE
    if normalized_status == OrganizationUser.STATUS_RESIGNED:
        member.resigned_at = member.resigned_at or now
        member.resigned_by = changed_by
    elif normalized_status != OrganizationUser.STATUS_DELETED:
        member.resigned_at = None
        member.resigned_by = None
    member.user.is_active = normalized_status == OrganizationUser.STATUS_ACTIVE
    if commit:
        member.save(
            update_fields=[
                "status",
                "status_changed_at",
                "is_deleted",
                "deleted_at",
                "is_active",
                "resigned_at",
                "resigned_by",
                "updated_at",
            ]
        )
        member.user.save(update_fields=["is_active"])
    return member


def _active_billable_memberships(memberships):
    return [
        row for row in (memberships or [])
        if _normalize_membership_status(row) == OrganizationUser.STATUS_ACTIVE
    ]


def _normalize_org_user_taxonomy_assignments(org):
    active_role_names = {
        str(name or "").strip().lower()
        for name in OrganizationEmployeeRole.objects.filter(organization=org, is_active=True).values_list("name", flat=True)
        if str(name or "").strip()
    }
    active_department_names = {
        str(name or "").strip().lower()
        for name in OrganizationDepartment.objects.filter(organization=org, is_active=True).values_list("name", flat=True)
        if str(name or "").strip()
    }
    memberships = list(
        OrganizationUser.objects
        .filter(organization=org, role__in=ERP_EMPLOYEE_ROLES, is_deleted=False)
        .only("id", "employee_role", "department", "updated_at")
    )
    for membership in memberships:
        update_fields = []
        employee_role = str(membership.employee_role or "").strip()
        department = str(membership.department or "").strip()
        if employee_role and employee_role.lower() not in active_role_names:
            membership.employee_role = ""
            update_fields.append("employee_role")
        if department and department.lower() not in active_department_names:
            membership.department = ""
            update_fields.append("department")
        if update_fields:
            membership.save(update_fields=[*update_fields, "updated_at"])


def _serialize_org_users(org, *, include_deleted=False):
    _normalize_org_user_taxonomy_assignments(org)
    _ensure_org_admin_memberships(org)
    _sync_business_autopilot_membership_access(org)
    queryset = (
        OrganizationUser.objects
        .filter(organization=org, role__in=ERP_EMPLOYEE_ROLES)
        .select_related("user", "user__userprofile", "resigned_by")
        .order_by("-id")
    )
    memberships = queryset.filter(is_deleted=bool(include_deleted))
    return [_serialize_org_user_member(org, member) for member in memberships]


def _is_org_admin_account_member(org: Organization, member: OrganizationUser) -> bool:
    if not member:
        return False
    profile_role = _normalize_admin_role(getattr(getattr(member.user, "userprofile", None), "role", ""))
    membership_role = _normalize_admin_role(getattr(member, "role", ""))
    is_owner = bool(getattr(org, "owner_id", None) and member.user_id == org.owner_id)
    return (
        is_owner
        or membership_role == "company_admin"
        or profile_role in DELETE_PROTECTED_PROFILE_ROLES
    )


def _serialize_org_user_member(org: Organization, member: OrganizationUser):
    profile_role = _normalize_admin_role(getattr(getattr(member.user, "userprofile", None), "role", ""))
    is_org_admin_account = _is_org_admin_account_member(org, member)
    last_login = getattr(member.user, "last_login", None)
    status = _normalize_membership_status(member)
    return {
        "id": member.user_id,
        "membership_id": member.id,
        "name": _get_org_user_display_name(member.user),
        "first_name": str(getattr(member.user, "first_name", "") or "").strip(),
        "last_name": str(getattr(member.user, "last_name", "") or "").strip(),
        "employeeId": _format_employee_code(member.user_id),
        "email": member.user.email or "",
        "email_verified": bool(getattr(member.user, "email_verified", False)),
        "phone_number": str(
            getattr(getattr(member.user, "userprofile", None), "phone_number", "") or ""
        ).strip(),
        "role": member.role or "org_user",
        "user_type": _normalize_user_type_key(getattr(member, "user_type", "")),
        "profile_role": profile_role,
        "is_org_admin_account": is_org_admin_account,
        "can_delete": not is_org_admin_account,
        "can_toggle_status": not is_org_admin_account,
        "department": member.department or "",
        "employee_role": member.employee_role or "",
        "status": status,
        "is_active": status == OrganizationUser.STATUS_ACTIVE,
        "is_deleted": bool(member.is_deleted),
        "can_mark_resigned": (not is_org_admin_account) and status not in {OrganizationUser.STATUS_RESIGNED, OrganizationUser.STATUS_DELETED},
        "can_restore_active": (not is_org_admin_account) and status in {OrganizationUser.STATUS_INACTIVE, OrganizationUser.STATUS_RESIGNED},
        "last_login": last_login.isoformat() if last_login else "",
        "created_at": member.created_at.isoformat() if member.created_at else "",
        "deleted_at": member.deleted_at.isoformat() if member.deleted_at else "",
        "resigned_at": member.resigned_at.isoformat() if member.resigned_at else "",
        "resigned_by": _get_org_user_display_name(member.resigned_by) if getattr(member, "resigned_by", None) else "",
        "status_changed_at": member.status_changed_at.isoformat() if member.status_changed_at else "",
    }


def _safe_serialize_org_users(org, *, include_deleted=False):
    try:
        return _serialize_org_users(org, include_deleted=include_deleted)
    except (DatabaseError, OperationalError, IntegrityError):
        logger.exception("Failed to serialize Business Autopilot users for org_id=%s", getattr(org, "id", None))
        return []


def _build_membership_assignment_summary(memberships):
    rows = []
    for membership in memberships:
        display_name = _get_org_user_display_name(getattr(membership, "user", None))
        email = str(getattr(getattr(membership, "user", None), "email", "") or "").strip()
        label = display_name or email or f"User {membership.id}"
        rows.append(
            {
                "membership_id": membership.id,
                "user_id": membership.user_id,
                "name": display_name,
                "email": email,
                "label": label,
            }
        )
    return rows


def _safe_serialize_employee_roles(org):
    try:
        return _serialize_employee_roles(org)
    except (DatabaseError, OperationalError, IntegrityError):
        logger.exception("Failed to serialize employee roles for org_id=%s", getattr(org, "id", None))
        return []


def _safe_serialize_departments(org):
    try:
        return _serialize_departments(org)
    except (DatabaseError, OperationalError, IntegrityError):
        logger.exception("Failed to serialize departments for org_id=%s", getattr(org, "id", None))
        return []


def _list_org_user_memberships(org):
    return list(
        OrganizationUser.objects
        .filter(organization=org, role__in=ERP_EMPLOYEE_ROLES, is_deleted=False)
        .select_related("user")
        .order_by("id")
    )


def _compute_org_user_lock_ids(employee_limit, memberships=None, org: Organization = None):
    rows = memberships if isinstance(memberships, list) else []
    safe_limit = max(0, int(employee_limit or 0))
    resolved_org = org if org else None
    protected_ids = []
    regular_rows = []
    for row in rows:
        if _normalize_membership_status(row) != OrganizationUser.STATUS_ACTIVE:
            continue
        current_org = resolved_org if resolved_org else getattr(row, "organization", None)
        if current_org and _is_org_admin_account_member(current_org, row):
            protected_ids.append(row.id)
        else:
            regular_rows.append(row)
    remaining_capacity = max(0, safe_limit - len(protected_ids))
    unlocked_ids = [*protected_ids, *[row.id for row in regular_rows[:remaining_capacity]]]
    locked_ids = [row.id for row in regular_rows[remaining_capacity:]]
    return {
        "unlocked_ids": set(unlocked_ids),
        "locked_ids": set(locked_ids),
        "total_users": len([row for row in rows if _normalize_membership_status(row) == OrganizationUser.STATUS_ACTIVE]),
    }


def _attach_locked_state(users, locked_ids):
    safe_locked_ids = set(locked_ids or [])
    normalized = []
    for row in (users or []):
        membership_id = row.get("membership_id")
        is_locked = membership_id in safe_locked_ids
        next_row = dict(row)
        next_row["is_locked"] = is_locked
        if is_locked:
            next_row["status"] = OrganizationUser.STATUS_INACTIVE
            next_row["is_active"] = False
        normalized.append(next_row)
    return normalized


def _build_org_user_meta(org, users=None):
    active_sub = _get_active_erp_subscription(org)
    has_subscription = bool(active_sub and active_sub.plan)
    addon_count = int(active_sub.addon_count or 0) if has_subscription else 0
    allow_addons = bool(active_sub.plan.allow_addons) if has_subscription else False
    is_ba_trial = False
    if has_subscription:
        product_slug = str(getattr(getattr(active_sub.plan, "product", None), "slug", "") or "").strip().lower()
        status = str(getattr(active_sub, "status", "") or "").strip().lower()
        is_ba_trial = product_slug in BUSINESS_AUTOPILOT_PRODUCT_SLUG_ALIASES and status == "trialing"

    if has_subscription:
        base_limit = _get_effective_employee_limit(active_sub.plan)
        # For paid/tiered plans, minimum 1 user is included by default.
        if base_limit <= 0 and not _is_free_plan(active_sub.plan):
            base_limit = 1
        if is_ba_trial:
            base_limit = max(base_limit, 3)
    else:
        base_limit = 0

    employee_limit = max(0, base_limit + max(0, addon_count))
    memberships = _list_org_user_memberships(org)
    seat_summary = _get_org_subscription_user_type_seats(org)
    active_user_type_counts = _get_org_user_type_usage(org, memberships=memberships)
    lock_state = _compute_org_user_lock_ids(employee_limit, memberships=memberships, org=org)
    active_users = len(_active_billable_memberships(memberships))
    used_users = min(active_users, employee_limit)
    remaining_users = max(0, employee_limit - active_users)
    can_add_users = bool(has_subscription and active_users < employee_limit)
    limit_message = ""
    if not has_subscription:
        limit_message = "Free trial ended. Upgrade your plan to add users."
    elif not can_add_users:
        limit_message = "User limit reached. Please increase user limit or mark inactive/resigned users."

    if has_subscription and (_is_business_autopilot_free_plan(active_sub.plan) or is_ba_trial):
        base_included_users = 1
        extra_included_users = max(0, employee_limit - base_included_users)
    else:
        base_included_users = max(0, employee_limit - max(0, addon_count))
        extra_included_users = max(0, employee_limit - base_included_users)

    return {
        "employee_limit": employee_limit,
        "used_users": used_users,
        "remaining_users": remaining_users,
        "addon_count": max(0, addon_count),
        "allow_addons": allow_addons,
        "has_unlimited_users": False,
        "can_add_users": can_add_users,
        "has_subscription": has_subscription,
        "limit_message": limit_message,
        "base_included_users": base_included_users,
        "extra_included_users": extra_included_users,
        "active_users": active_users,
        "inactive_users": len([row for row in memberships if _normalize_membership_status(row) == OrganizationUser.STATUS_INACTIVE]),
        "resigned_users": len([row for row in memberships if _normalize_membership_status(row) == OrganizationUser.STATUS_RESIGNED]),
        "deleted_users": OrganizationUser.objects.filter(organization=org, role__in=ERP_EMPLOYEE_ROLES, is_deleted=True).count(),
        "user_types": _serialize_user_type_pricing_rows(
            seat_summary.get("config", []),
            seat_counts=seat_summary.get("counts"),
            active_counts=active_user_type_counts,
        ),
        "user_type_seat_counts": seat_summary.get("counts", {}),
        "user_type_active_counts": active_user_type_counts,
    }


def _build_org_users_response_payload(org, can_manage_users, *, created_user_credentials=None, credential_delivery=None, message=""):
    users = _safe_serialize_org_users(org)
    deleted_users = _safe_serialize_org_users(org, include_deleted=True)
    meta = _build_org_user_meta(org, users=users)
    lock_state = _compute_org_user_lock_ids(
        meta.get("employee_limit"),
        memberships=_list_org_user_memberships(org),
        org=org,
    )
    users = _attach_locked_state(users, lock_state["locked_ids"])
    user_counts = {
        "all": len(users),
        "active": len([row for row in users if str(row.get("status") or "").strip().lower() == OrganizationUser.STATUS_ACTIVE]),
        "inactive": len([row for row in users if str(row.get("status") or "").strip().lower() == OrganizationUser.STATUS_INACTIVE]),
        "resigned": len([row for row in users if str(row.get("status") or "").strip().lower() == OrganizationUser.STATUS_RESIGNED]),
        "deleted": len(deleted_users),
    }
    return {
        "authenticated": True,
        "organization_id": getattr(org, "id", None),
        "users": users,
        "deleted_users": deleted_users,
        "counts": user_counts,
        "employee_roles": _safe_serialize_employee_roles(org),
        "departments": _safe_serialize_departments(org),
        "can_manage_users": can_manage_users,
        "meta": meta,
        "created_user_credentials": created_user_credentials,
        "credential_delivery": credential_delivery,
        "message": message,
    }


def _sync_org_users_to_plan_limit(org, requested_by=None):
    memberships = _list_org_user_memberships(org)
    if not memberships:
        return {"auto_disabled_ids": [], "auto_enabled_ids": []}

    meta = _build_org_user_meta(org, users=None)
    lock_state = _compute_org_user_lock_ids(meta.get("employee_limit"), memberships=memberships, org=org)
    locked_ids = lock_state["locked_ids"]
    auto_disabled = []
    auto_enabled = []

    # ORG admin account must always remain active.
    for row in memberships:
        if _is_org_admin_account_member(org, row) and not row.is_active:
            row.is_active = True
            row.save(update_fields=["is_active", "updated_at"])
            _grant_business_autopilot_access(row.user, requested_by or row.user, row.role or "company_admin")
            auto_enabled.append(row.id)

    # Extra users are lock-managed by plan limit and must stay inactive.
    for row in memberships:
        if row.id in locked_ids and row.is_active:
            row.is_active = False
            row.save(update_fields=["is_active", "updated_at"])
            _revoke_business_autopilot_access(row.user)
            auto_disabled.append(row.id)

    return {"auto_disabled_ids": auto_disabled, "auto_enabled_ids": auto_enabled}


def _decimal_to_string(value, default="0.00"):
    try:
      amount = Decimal(str(value if value is not None else default))
    except (InvalidOperation, TypeError, ValueError):
      amount = Decimal(str(default or "0"))
    return f"{amount.quantize(Decimal('0.01'))}"


def _format_indian_currency_value(value, default="0.00"):
    normalized = _decimal_to_string(value, default)
    sign = ""
    if normalized.startswith("-"):
        sign = "-"
        normalized = normalized[1:]
    integer_part, _, decimal_part = normalized.partition(".")
    if len(integer_part) > 3:
        last_three = integer_part[-3:]
        remaining = integer_part[:-3]
        grouped_parts = []
        while len(remaining) > 2:
            grouped_parts.insert(0, remaining[-2:])
            remaining = remaining[:-2]
        if remaining:
            grouped_parts.insert(0, remaining)
        integer_part = ",".join(grouped_parts + [last_three])
    return f"{sign}{integer_part}.{decimal_part or '00'}"


def _format_pdf_inr(value, default="0.00"):
    return f"INR {_format_indian_currency_value(value, default)}"


def _to_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _coerce_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _format_employee_code(user_id):
    numeric_user_id = _coerce_int(user_id)
    if not numeric_user_id or numeric_user_id <= 0:
        return ""
    return f"EMP{numeric_user_id:03d}"


def _get_org_user_display_name(user):
    if not user:
        return ""
    full_name = " ".join(
        [
            str(getattr(user, "first_name", "") or "").strip(),
            str(getattr(user, "last_name", "") or "").strip(),
        ]
    ).strip()
    fallback = str(
        getattr(user, "first_name", "")
        or getattr(user, "username", "")
        or getattr(user, "email", "")
        or ""
    ).strip()
    return full_name or fallback


def _extract_person_name(payload):
    first_name = str(payload.get("first_name") or "").strip()
    last_name = str(payload.get("last_name") or "").strip()
    legacy_name = str(payload.get("name") or "").strip()

    if not first_name and legacy_name:
        parts = legacy_name.split()
        if parts:
            first_name = parts[0]
            last_name = " ".join(parts[1:]).strip()

    return first_name, last_name


def _resolve_org_department_from_payload(org, payload):
    department = str(payload.get("department") or "").strip()
    department_id = payload.get("department_id")
    if department_id not in (None, ""):
        try:
            department_id = int(department_id)
        except (TypeError, ValueError):
            return None, JsonResponse({"detail": "invalid_department"}, status=400)
        selected_department = OrganizationDepartment.objects.filter(
            organization=org, id=department_id, is_active=True
        ).first()
        if not selected_department:
            return None, JsonResponse({"detail": "department_not_found"}, status=404)
        return selected_department.name, None
    if not department:
        return None, JsonResponse({"detail": "department_required"}, status=400)
    department_row, _ = OrganizationDepartment.objects.get_or_create(
        organization=org,
        name=department,
        defaults={"is_active": True},
    )
    if not department_row.is_active:
        department_row.is_active = True
        department_row.save(update_fields=["is_active", "updated_at"])
    return department_row.name, None


def _resolve_org_employee_role_from_payload(org, payload):
    employee_role = str(payload.get("employee_role") or "").strip()
    employee_role_id = payload.get("employee_role_id")
    if employee_role_id not in (None, ""):
        try:
            employee_role_id = int(employee_role_id)
        except (TypeError, ValueError):
            return None, JsonResponse({"detail": "invalid_employee_role"}, status=400)
        selected_role = OrganizationEmployeeRole.objects.filter(
            organization=org,
            id=employee_role_id,
            is_active=True,
        ).first()
        if not selected_role:
            return None, JsonResponse({"detail": "employee_role_not_found"}, status=404)
        return selected_role.name, None
    if not employee_role:
        return None, JsonResponse({"detail": "employee_role_required"}, status=400)
    role_row, _ = OrganizationEmployeeRole.objects.get_or_create(
        organization=org,
        name=employee_role,
        defaults={"is_active": True},
    )
    if not role_row.is_active:
        role_row.is_active = True
        role_row.save(update_fields=["is_active", "updated_at"])
    return role_row.name, None


def _resolve_org_user_type_from_payload(org, payload):
    active_sub = _get_active_erp_subscription(org)
    config_rows = _get_subscription_user_types(active_sub)
    allowed_keys = {row["key"] for row in config_rows}
    requested_key = _normalize_user_type_key(payload.get("user_type"))
    if requested_key not in allowed_keys:
        if "full_access_user" in allowed_keys:
            requested_key = "full_access_user"
        elif allowed_keys:
            requested_key = next(iter(allowed_keys))
    return requested_key, config_rows


def _user_type_has_available_seat(org, user_type, *, memberships=None, exclude_membership_id=None):
    seat_summary = _get_org_subscription_user_type_seats(org)
    config_rows = seat_summary.get("config", [])
    seat_counts = seat_summary.get("counts", {})
    rows = memberships if isinstance(memberships, list) else _list_org_user_memberships(org)
    active_counts = _empty_user_type_counts(config_rows)
    for membership in rows:
        if exclude_membership_id and int(getattr(membership, "id", 0) or 0) == int(exclude_membership_id):
            continue
        if _normalize_membership_status(membership) != OrganizationUser.STATUS_ACTIVE:
            continue
        key = _normalize_user_type_key(getattr(membership, "user_type", ""))
        active_counts[key] = active_counts.get(key, 0) + 1
    return int(active_counts.get(user_type, 0)) < int(seat_counts.get(user_type, 0))


def _create_or_attach_org_user(org, payload, *, requested_by):
    first_name, last_name = _extract_person_name(payload)
    name = " ".join([first_name, last_name]).strip()
    email = (payload.get("email") or "").strip().lower()
    password = payload.get("password") or ""
    phone_number = str(payload.get("phone_number") or "").strip()
    confirm_existing_user = bool(payload.get("confirm_existing_user"))
    role = (payload.get("role") or "org_user").strip().lower()
    if role not in ERP_EMPLOYEE_ROLES:
        role = "org_user"
    if not phone_number:
        return {"ok": False, "response": JsonResponse({"detail": "phone_required"}, status=400)}
    department, department_error = _resolve_org_department_from_payload(org, payload)
    if department_error:
        return {"ok": False, "response": department_error}
    employee_role, employee_role_error = _resolve_org_employee_role_from_payload(org, payload)
    if employee_role_error:
        return {"ok": False, "response": employee_role_error}
    user_type, _ = _resolve_org_user_type_from_payload(org, payload)
    if not _user_type_has_available_seat(org, user_type):
        return {
            "ok": False,
            "response": JsonResponse(
                {
                    "detail": "user_type_limit_reached",
                    "message": "Selected user type limit reached. Increase seats or choose another available user type.",
                },
                status=403,
            ),
        }
    existing_user = User.objects.filter(email__iexact=email).first()
    if not name or not email:
        return {"ok": False, "response": JsonResponse({"detail": "name_email_required"}, status=400)}
    if not existing_user and not password:
        return {"ok": False, "response": JsonResponse({"detail": "password_required"}, status=400)}
    if password and len(password) < 6:
        return {"ok": False, "response": JsonResponse({"detail": "password_too_short"}, status=400)}

    newly_created_user = False
    plain_password_for_share = ""
    with transaction.atomic():
        if existing_user:
            existing_profile = UserProfile.objects.filter(user=existing_user).first()
            if existing_profile and existing_profile.organization_id and existing_profile.organization_id != org.id:
                return {"ok": False, "response": JsonResponse({"detail": "email_belongs_to_another_organization"}, status=409)}
            existing_products = _get_user_granted_products(existing_user)
            already_has_business_autopilot = any(
                product["slug"] == BUSINESS_AUTOPILOT_PRODUCT_SLUG for product in existing_products
            )
            if not already_has_business_autopilot and not confirm_existing_user:
                return {
                    "ok": False,
                    "response": JsonResponse(
                        {
                            "detail": "existing_org_user_requires_confirmation",
                            "message": "This user is already assigned to another product in this organization. The same password will continue to work.",
                            "same_password_allowed": True,
                            "existing_products": existing_products,
                        },
                        status=409,
                    ),
                }
            if already_has_business_autopilot:
                return {"ok": False, "response": JsonResponse({"detail": "user_already_assigned_to_business_autopilot"}, status=409)}
            user = existing_user
            if first_name or last_name:
                user.first_name = first_name
                user.last_name = last_name
                user.save(update_fields=["first_name", "last_name"])
            if existing_profile:
                existing_profile.organization = org
                existing_profile.role = role
                existing_profile.phone_number = phone_number
                existing_profile.save(update_fields=["organization", "role", "phone_number"])
            else:
                UserProfile.objects.create(
                    user=user,
                    organization=org,
                    role=role,
                    phone_number=phone_number,
                )
        else:
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_active=True,
            )
            UserProfile.objects.update_or_create(
                user=user,
                defaults={
                    "organization": org,
                    "role": role,
                    "phone_number": phone_number,
                },
            )
            newly_created_user = True
            plain_password_for_share = str(password or "")

        OrganizationUser.objects.update_or_create(
            organization=org,
            user=user,
            defaults={
                "role": role,
                "employee_role": employee_role,
                "department": department,
                "user_type": user_type,
                "status": OrganizationUser.STATUS_ACTIVE,
                "is_active": True,
                "is_deleted": False,
                "deleted_at": None,
                "resigned_at": None,
                "resigned_by": None,
                "status_changed_at": timezone.now(),
            },
        )
        _grant_business_autopilot_access(user, requested_by, role)
        _sync_org_users_to_plan_limit(org, requested_by=requested_by)

    is_existing_user_added = bool(existing_user) and not newly_created_user
    created_user_credentials = None
    credential_delivery = {
        "is_new_user": False,
        "email_sent": False,
        "status": "not_applicable",
    }
    if newly_created_user or is_existing_user_added:
        login_url = PUBLIC_LOGIN_URL
        password_for_share = plain_password_for_share if newly_created_user else "Use your existing password"
        created_user_credentials = {
            "name": _get_org_user_display_name(user),
            "email": str(user.email or "").strip(),
            "password": password_for_share,
            "login_url": login_url,
        }
        mail_sent = send_templated_email(
            user.email,
            "Your Work Zilla login credentials",
            "emails/business_autopilot_user_credentials.txt",
            {
                "name": created_user_credentials["name"] or "User",
                "email": created_user_credentials["email"],
                "password": password_for_share,
                "login_url": login_url,
                "organization_name": str(org.name or "").strip(),
            },
        )
        credential_delivery = {
            "is_new_user": bool(newly_created_user),
            "email_sent": bool(mail_sent),
            "status": "sent" if mail_sent else "failed",
        }
    return {
        "ok": True,
        "user": user,
        "created_user_credentials": created_user_credentials,
        "credential_delivery": credential_delivery,
        "newly_created_user": newly_created_user,
        "is_existing_user_added": is_existing_user_added,
    }


def _normalize_payroll_month(value):
    month = str(value or "").strip()
    if len(month) == 7 and month[4] == "-":
        return month
    return timezone.localdate().strftime("%Y-%m")


def _default_payroll_settings_payload():
    return {
        "enablePf": True,
        "enableEsi": True,
        "pfEmployeePercent": "12.00",
        "pfEmployerPercent": "12.00",
        "esiEmployeePercent": "0.75",
        "esiEmployerPercent": "3.25",
    }


def _serialize_payroll_settings(row):
    if not row:
        return _default_payroll_settings_payload()
    return {
        "enablePf": bool(row.enable_pf),
        "enableEsi": bool(row.enable_esi),
        "pfEmployeePercent": _decimal_to_string(row.pf_employee_percent, "12.00"),
        "pfEmployerPercent": _decimal_to_string(row.pf_employer_percent, "12.00"),
        "esiEmployeePercent": _decimal_to_string(row.esi_employee_percent, "0.75"),
        "esiEmployerPercent": _decimal_to_string(row.esi_employer_percent, "3.25"),
    }


def _serialize_salary_structure(row):
    return {
        "id": row.id,
        "name": row.name,
        "isDefault": bool(row.is_default),
        "basicSalaryPercent": _decimal_to_string(row.basic_salary_percent, "40.00"),
        "hraPercent": _decimal_to_string(row.hra_percent, "20.00"),
        "conveyanceFixed": _decimal_to_string(row.conveyance_fixed, "1600.00"),
        "autoSpecialAllowance": bool(row.auto_special_allowance),
        "basicSalary": _decimal_to_string(row.basic_salary),
        "hra": _decimal_to_string(row.hra),
        "conveyance": _decimal_to_string(row.conveyance),
        "specialAllowance": _decimal_to_string(row.special_allowance),
        "bonus": _decimal_to_string(row.bonus),
        "otherAllowances": _decimal_to_string(row.other_allowances),
        "applyPf": bool(row.apply_pf),
        "applyEsi": bool(row.apply_esi),
        "professionalTax": _decimal_to_string(row.professional_tax),
        "otherDeduction": _decimal_to_string(row.other_deduction),
        "notes": row.notes or "",
        "createdAt": row.created_at.isoformat() if row.created_at else "",
        "updatedAt": row.updated_at.isoformat() if row.updated_at else "",
    }


def _serialize_salary_history(row):
    return {
        "id": row.id,
        "employeeName": row.employee_name,
        "sourceUserId": row.source_user_id,
        "employeeId": _format_employee_code(row.source_user_id),
        "salaryStructureId": row.salary_structure_id,
        "salaryStructureName": row.salary_structure.name if row.salary_structure_id and row.salary_structure else "",
        "currentSalary": _decimal_to_string(row.current_salary),
        "monthlySalaryAmount": _decimal_to_string(row.monthly_salary_amount),
        "incrementType": row.increment_type or "percentage",
        "incrementValue": _decimal_to_string(row.increment_value),
        "effectiveFrom": row.effective_from.isoformat() if row.effective_from else "",
        "incrementAmount": _decimal_to_string(row.increment_amount),
        "newSalary": _decimal_to_string(row.new_salary),
        "notes": row.notes or "",
        "createdAt": row.created_at.isoformat() if row.created_at else "",
        "updatedAt": row.updated_at.isoformat() if row.updated_at else "",
    }


def _ensure_standard_salary_template(org: Organization):
    row, _ = SalaryStructure.objects.get_or_create(
        organization=org,
        name="Standard Template",
        defaults={
            "is_default": True,
            "basic_salary_percent": Decimal("40"),
            "hra_percent": Decimal("20"),
            "conveyance_fixed": Decimal("1600"),
            "auto_special_allowance": True,
            "apply_pf": True,
            "apply_esi": True,
        },
    )
    updates = []
    if not row.is_default:
        row.is_default = True
        updates.append("is_default")
    if row.basic_salary_percent != Decimal("40"):
        row.basic_salary_percent = Decimal("40")
        updates.append("basic_salary_percent")
    if row.hra_percent != Decimal("20"):
        row.hra_percent = Decimal("20")
        updates.append("hra_percent")
    if row.conveyance_fixed != Decimal("1600"):
        row.conveyance_fixed = Decimal("1600")
        updates.append("conveyance_fixed")
    if not row.auto_special_allowance:
        row.auto_special_allowance = True
        updates.append("auto_special_allowance")
    if updates:
        row.save(update_fields=updates)
    SalaryStructure.objects.filter(organization=org).exclude(id=row.id).filter(is_default=True).update(is_default=False)
    return row


def _serialize_payroll_entry(row):
    return {
        "id": row.id,
        "employeeName": row.employee_name,
        "sourceUserId": row.source_user_id,
        "month": row.payroll_month,
        "currency": row.currency or "INR",
        "salaryStructureId": row.salary_structure_id,
        "salaryStructureName": row.salary_structure.name if row.salary_structure_id and row.salary_structure else "",
        "salaryHistoryId": row.salary_history_id,
        "grossSalary": _decimal_to_string(row.gross_salary),
        "pfEmployeeAmount": _decimal_to_string(row.pf_employee_amount),
        "pfEmployerAmount": _decimal_to_string(row.pf_employer_amount),
        "esiEmployeeAmount": _decimal_to_string(row.esi_employee_amount),
        "esiEmployerAmount": _decimal_to_string(row.esi_employer_amount),
        "professionalTaxAmount": _decimal_to_string(row.professional_tax_amount),
        "otherDeductionAmount": _decimal_to_string(row.other_deduction_amount),
        "totalDeductions": _decimal_to_string(row.total_deductions),
        "netSalary": _decimal_to_string(row.net_salary),
        "earnings": row.earnings if isinstance(row.earnings, dict) else {},
        "deductions": row.deductions if isinstance(row.deductions, dict) else {},
        "status": row.status or "processed",
        "processedAt": row.processed_at.isoformat() if row.processed_at else "",
        "updatedAt": row.updated_at.isoformat() if row.updated_at else "",
        "slipNumber": row.payslip.slip_number if hasattr(row, "payslip") and row.payslip else "",
    }


def _serialize_payslip(row):
    return {
        "id": row.id,
        "payrollEntryId": row.payroll_entry_id,
        "slipNumber": row.slip_number,
        "generatedForMonth": row.generated_for_month,
        "employeeName": row.employee_name,
        "sourceUserId": row.source_user_id,
        "currency": row.currency or "INR",
        "generatedAt": row.generated_at.isoformat() if row.generated_at else "",
        "updatedAt": row.updated_at.isoformat() if row.updated_at else "",
    }


def _get_or_create_payroll_settings(org):
    row, _ = PayrollSettings.objects.get_or_create(organization=org)
    return row


def _serialize_org_payroll_profile(org):
    settings_obj, _ = OrganizationSettings.objects.get_or_create(organization=org)
    return {
        "organizationName": org.name,
        "country": org.country or "India",
        "currency": org.currency or "INR",
        "timezone": settings_obj.org_timezone or "UTC",
    }


def _format_currency_value(currency, amount):
    return f"{currency or 'INR'} {_decimal_to_string(amount)}"


def _serialize_salary_history_detail(row):
    return {
        "id": row.id,
        "effectiveDate": row.effective_from.isoformat() if row.effective_from else "",
        "previousSalary": _decimal_to_string(row.current_salary),
        "incrementType": row.increment_type or "percentage",
        "incrementValue": _decimal_to_string(row.increment_value),
        "incrementAmount": _decimal_to_string(row.increment_amount),
        "newSalary": _decimal_to_string(row.new_salary),
    }


def _serialize_modules(org):
    subscription = _get_active_erp_subscription(org)
    allowed_slugs = _get_plan_erp_module_slugs(
        subscription.plan if subscription and subscription.plan else None,
        subscription=subscription,
    )
    rows = (
        OrganizationModule.objects
        .filter(organization=org, module__is_active=True)
        .select_related("module")
        .order_by("module__sort_order", "module__name")
    )
    modules = [
        {
            "id": row.module_id,
            "name": row.module.name,
            "slug": row.module.slug,
            "enabled": bool(row.enabled and row.module.slug in allowed_slugs),
            "eligible": bool(row.module.slug in allowed_slugs),
            "path": MODULE_PATHS.get(row.module.slug, f"/{row.module.slug}"),
        }
        for row in rows
    ]
    enabled_modules = [module for module in modules if module["enabled"]]
    return modules, enabled_modules


def _default_accounts_workspace():
    return {
        "customers": [],
        "vendors": [],
        "itemMasters": [],
        "gstTemplates": [],
        "billingTemplates": [],
        "estimates": [],
        "invoices": [],
        "quickEstimateContacts": [],
        "quickEstimateSettings": {
            "headerText": "",
            "templateSize": "4in",
            "paymentProofRetentionDays": "45",
        },
    }


def _normalize_quick_estimate_header_html(value):
    return str(value or "").strip()


def _quick_estimate_header_text_length(value):
    return len(strip_tags(str(value or "")).strip())


def _default_india_gst_templates():
    return [
        {
            "id": "gst_default_india_igst",
            "name": "IGST",
            "taxScope": "Inter State",
            "cgst": "",
            "sgst": "",
            "igst": "18",
            "cess": "",
            "status": "Active",
            "notes": "",
        },
        {
            "id": "gst_default_india_cgst_sgst",
            "name": "CGST & SGST",
            "taxScope": "Intra State",
            "cgst": "9",
            "sgst": "9",
            "igst": "",
            "cess": "",
            "status": "Active",
            "notes": "",
        },
    ]


def _normalize_gst_template_text(value):
    return " ".join(str(value or "").strip().lower().split())


def _normalize_gst_template_number(value):
    return str(value or "").strip()


def _looks_like_legacy_india_gst_template(template):
    if not isinstance(template, dict):
        return False
    template_id = _normalize_gst_template_text(template.get("id"))
    template_name = _normalize_gst_template_text(template.get("name"))
    template_scope = _normalize_gst_template_text(template.get("taxScope") or template.get("scope"))
    cgst = _normalize_gst_template_number(template.get("cgst"))
    sgst = _normalize_gst_template_number(template.get("sgst"))
    igst = _normalize_gst_template_number(template.get("igst"))
    cess = _normalize_gst_template_number(template.get("cess"))

    if template_id in {
        "gst_india_igst",
        "gst_india_cgst_sgst",
        "gst_default_india_igst",
        "gst_default_india_cgst_sgst",
    }:
        return True

    if template_name in {"india gst", "igst", "cgst & sgst"} and template_scope in {"inter state", "intra state"}:
        return True

    return (
        template_name == "india gst"
        and cgst in {"9", "9.0"}
        and sgst in {"9", "9.0"}
        and igst in {"18", "18.0"}
        and cess in {"0", "0.0", ""}
    )


def _normalize_india_gst_templates(templates):
    canonical = _default_india_gst_templates()
    if not isinstance(templates, list):
        return canonical

    normalized = []
    for row in templates:
        row_id = _normalize_gst_template_text(row.get("id")) if isinstance(row, dict) else ""
        if row_id in {
            "gst_default_india_igst",
            "gst_default_india_cgst_sgst",
        } or _looks_like_legacy_india_gst_template(row):
            continue
        if isinstance(row, dict):
            normalized.append(row)

    return [*canonical, *normalized]


def _seed_accounts_workspace_defaults_for_org(payload, org):
    base = _normalize_accounts_workspace(payload)
    billing_profile = BillingProfile.objects.filter(organization=org).only("country").first()
    country = str(
        (getattr(billing_profile, "country", "") or getattr(org, "country", "") or "India")
    ).strip().lower()
    if country == "india":
        base["gstTemplates"] = _normalize_india_gst_templates(base.get("gstTemplates") or [])
    return base


def _ensure_accounts_workspace_defaults_for_org(org):
    workspace = AccountsWorkspace.objects.filter(organization=org).first()
    billing_profile = BillingProfile.objects.filter(organization=org).only("country").first()
    country = str(
        (getattr(billing_profile, "country", "") or getattr(org, "country", "") or "India")
    ).strip().lower()
    if not workspace:
        if country != "india":
            return None
        workspace = AccountsWorkspace.objects.create(
            organization=org,
            data=_default_accounts_workspace(),
        )
    seeded_data = _seed_accounts_workspace_defaults_for_org(workspace.data, org)
    if workspace.data != seeded_data:
        workspace.data = seeded_data
        workspace.save(update_fields=["data", "updated_at"])
    return workspace


def _normalize_accounts_workspace(payload):
    base = _default_accounts_workspace()
    if not isinstance(payload, dict):
        return base
    for key in ACCOUNTS_ALLOWED_ROOT_KEYS:
        value = payload.get(key)
        base[key] = value if isinstance(value, list) else []
    quick_estimate_settings = payload.get("quickEstimateSettings")
    if isinstance(quick_estimate_settings, dict):
        template_size = str(quick_estimate_settings.get("templateSize") or "4in").strip().lower()
        if template_size not in {"3in", "4in"}:
            template_size = "4in"
        payment_proof_retention_days = str(
            quick_estimate_settings.get("paymentProofRetentionDays") or "45"
        ).strip()
        if payment_proof_retention_days not in {"45", "60"}:
            payment_proof_retention_days = "45"
        base["quickEstimateSettings"] = {
            "headerText": _normalize_quick_estimate_header_html(quick_estimate_settings.get("headerText")),
            "templateSize": template_size,
            "paymentProofRetentionDays": payment_proof_retention_days,
        }
    return base


def _merge_accounts_workspace(existing_payload, incoming_payload):
    existing = _normalize_accounts_workspace(existing_payload)
    if not isinstance(incoming_payload, dict):
        return existing
    merged = dict(existing)
    for key in ACCOUNTS_ALLOWED_ROOT_KEYS:
        if key not in incoming_payload:
            continue
        value = incoming_payload.get(key)
        if isinstance(value, list):
            merged[key] = value
    if "quickEstimateSettings" in incoming_payload and isinstance(incoming_payload.get("quickEstimateSettings"), dict):
        template_size = str(incoming_payload.get("quickEstimateSettings", {}).get("templateSize") or "4in").strip().lower()
        if template_size not in {"3in", "4in"}:
            template_size = "4in"
        payment_proof_retention_days = str(
            incoming_payload.get("quickEstimateSettings", {}).get("paymentProofRetentionDays") or "45"
        ).strip()
        if payment_proof_retention_days not in {"45", "60"}:
            payment_proof_retention_days = "45"
        merged["quickEstimateSettings"] = {
            "headerText": _normalize_quick_estimate_header_html(incoming_payload.get("quickEstimateSettings", {}).get("headerText")),
            "templateSize": template_size,
            "paymentProofRetentionDays": payment_proof_retention_days,
        }
    return _normalize_accounts_workspace(merged)


def _parse_iso_date(value):
    normalized = str(value or "").strip()
    if not normalized:
        return None
    try:
        parsed = date.fromisoformat(normalized)
        return parsed
    except ValueError:
        return None


def _normalize_subscription_status(value):
    normalized = str(value or "").strip().lower()
    mapping = {
        "active": "Active",
        "expired": "Expired",
        "cancelled": "Cancelled",
        "canceled": "Cancelled",
    }
    return mapping.get(normalized, "Active")


def _get_accounts_customer_display_name(row):
    company_name = str(row.get("companyName") or row.get("name") or "").strip()
    client_name = str(row.get("clientName") or "").strip()
    if company_name and client_name:
        return f"{company_name} / {client_name}"
    return company_name or client_name


def _calculate_next_billing_date(start_date):
    if not isinstance(start_date, date):
        return None
    next_month = start_date.month + 1
    year = start_date.year
    while next_month > 12:
        year += 1
        next_month -= 12
    next_day = min(
        start_date.day,
        calendar.monthrange(year, next_month)[1]
    )
    return date(year, next_month, next_day)


def _effective_subscription_status(status, end_date):
    normalized_status = _normalize_subscription_status(status)
    if normalized_status == "Cancelled":
        return "Cancelled"
    today = timezone.localdate()
    if end_date and end_date < today:
        return "Expired"
    return "Active"


def _get_accounts_customer_lookup(org):
    workspace = _get_accounts_workspace(org)
    rows = workspace.data.get("customers") if isinstance(workspace.data, dict) else []
    if not isinstance(rows, list):
        return {}
    lookup = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        customer_id = str(row.get("id") or "").strip()
        if not customer_id:
            continue
        customer_name = _get_accounts_customer_display_name(row)
        if not customer_name:
            continue
        lookup[customer_id] = customer_name
    return lookup


def _serialize_subscription_category(row):
    return {
        "id": row.id,
        "name": row.name,
        "description": row.description or "",
        "createdAt": row.created_at.isoformat() if row.created_at else "",
    }


def _serialize_subscription_sub_category(row):
    return {
        "id": row.id,
        "name": row.name,
        "description": row.description or "",
        "categoryId": row.category_id,
        "categoryName": row.category.name if row.category else "",
        "createdAt": row.created_at.isoformat() if row.created_at else "",
    }


def _serialize_subscription(row, customer_lookup=None):
    if customer_lookup is None:
        customer_lookup = {}
    customer_id = row.customer_id
    customer_name = ""
    if customer_id is not None:
        customer_name = customer_lookup.get(str(customer_id), "")
    return {
        "id": row.id,
        "subscriptionTitle": row.subscription_title,
        "categoryId": row.category_id,
        "categoryName": row.category.name if row.category else "",
        "subCategoryId": row.sub_category_id,
        "subCategoryName": row.sub_category.name if row.sub_category else "",
        "customerId": customer_id,
        "customerName": customer_name,
        "emailAlertDays": _normalize_subscription_alert_days(row.email_alert_days),
        "whatsappAlertDays": _normalize_subscription_alert_days(row.whatsapp_alert_days),
        "emailAlertAssignTo": _serialize_subscription_alert_assignees(row.email_alert_assign_to),
        "whatsappAlertAssignTo": _serialize_subscription_alert_assignees(row.whatsapp_alert_assign_to),
        "planDurationDays": row.plan_duration_days,
        "paymentDescription": row.payment_description or "",
        "amount": _decimal_to_string(row.amount),
        "currency": row.currency or "INR",
        "startDate": row.start_date.isoformat() if row.start_date else "",
        "endDate": row.end_date.isoformat() if row.end_date else "",
        "nextBillingDate": row.next_billing_date.isoformat() if row.next_billing_date else "",
        "status": _effective_subscription_status(row.status, row.end_date),
        "createdBy": str(row.created_by_id or ""),
        "createdAt": row.created_at.isoformat() if row.created_at else "",
        "updatedAt": row.updated_at.isoformat() if row.updated_at else "",
    }


def _get_accounts_workspace(org):
    workspace, created = AccountsWorkspace.objects.get_or_create(
        organization=org,
        defaults={"data": _default_accounts_workspace()},
    )
    original_data = workspace.data
    seeded_data = _seed_accounts_workspace_defaults_for_org(workspace.data, org)
    workspace.data = seeded_data
    if created or original_data != seeded_data:
        workspace.save(update_fields=["data", "updated_at"])
    return workspace


def _normalize_site_admin_mobile(value):
    digits = re.sub(r"\D+", "", str(value or ""))
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def _normalize_site_admin_email(value):
    return str(value or "").strip().lower()


def _site_admin_is_reset_command(message):
    normalized = " ".join(str(message or "").strip().lower().split())
    return normalized in SITE_ADMIN_RESET_COMMANDS


def _site_admin_supported_module_labels():
    return ", ".join(module.module_name for module in get_site_admin_enabled_modules())


def _site_admin_instruction_text(module_key="quick_estimate"):
    return build_site_admin_instruction_context(module_key)


def _site_admin_detect_quick_estimate_intent(message):
    normalized = " ".join(str(message or "").strip().lower().split())
    if normalized in SITE_ADMIN_QUICK_ESTIMATE_HINTS:
        return True
    if "quick estimate" in normalized:
        return True
    if normalized.startswith("qe "):
        return True
    return False


def _site_admin_find_mobile(text):
    content = str(text or "")
    candidates = re.findall(r"(?:\+?91[\s-]*)?[6-9]\d[\d\s-]{8,12}", content)
    if not candidates:
        candidates = re.findall(r"(?:\+?91[\s-]*)?\d[\d\s-]{8,12}", content)
    for candidate in candidates:
        normalized = _normalize_site_admin_mobile(candidate)
        if len(normalized) == 10:
            return normalized
    return ""


def _site_admin_find_amount(text):
    content = str(text or "")
    preferred = re.findall(r"(?:rs\.?|inr|rupees|₹)\s*([0-9][0-9,]*(?:\.\d{1,2})?)", content, flags=re.IGNORECASE)
    fallback = re.findall(r"(?<!\d)([0-9]{2,}[0-9,]*(?:\.\d{1,2})?)(?!\d)", content)
    candidates = preferred or fallback
    for candidate in reversed(candidates):
        try:
            cleaned = candidate.replace(",", "").strip()
            amount = Decimal(cleaned)
        except (InvalidOperation, AttributeError):
            continue
        if amount > 0:
            return amount.quantize(Decimal("0.01"))
    return None


def _site_admin_find_email(text):
    match = re.search(r"([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})", str(text or ""), flags=re.IGNORECASE)
    return _normalize_site_admin_email(match.group(1)) if match else ""


def _site_admin_find_gst_number(text):
    match = re.search(r"\b\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z0-9]\b", str(text or "").upper())
    return match.group(0) if match else ""


def _site_admin_find_quantity_unit(text):
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*(nos?|pcs?|pieces?|qty|units?|kg|gms?|grams?|boxes?|packs?|sets?|hours?|days?)\b", str(text or ""), flags=re.IGNORECASE)
    if not matches:
        return None, ""
    quantity_raw, unit = matches[-1]
    try:
        quantity = Decimal(str(quantity_raw)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        return None, ""
    return quantity, str(unit or "").strip()


def _site_admin_parse_name_candidate(text):
    cleaned = " ".join(str(text or "").strip().split())
    if not cleaned:
        return ""
    lowered = cleaned.lower()
    if lowered in {"hi", "hello", "hey", "hai", "hii", "helo", "vanakkam", "start", "qe", "quick estimate"}:
        return ""
    if _site_admin_find_mobile(cleaned) or _site_admin_find_amount(cleaned) is not None:
        return ""
    if _site_admin_find_quantity_unit(cleaned)[0] is not None:
        return ""
    if any(token in lowered for token in ["rs.", " rs", "₹", "nos", "pcs", "pieces", "qty", "quantity", "gsm", "sheet", "printing"]):
        return ""
    if len(cleaned) > 80:
        return ""
    if sum(1 for ch in cleaned if ch.isalpha()) < 2:
        return ""
    return cleaned[:180]


def _site_admin_parse_item_text(text):
    content = " ".join(str(text or "").strip().split())
    if not content:
        return {
            "service_name": "",
            "description": "",
            "quantity": None,
            "unit": "",
        }
    amount = _site_admin_find_amount(content)
    if amount is not None:
        amount_pattern = re.compile(r"(?:rs\.?|inr|rupees|₹)\s*[0-9][0-9,]*(?:\.\d{1,2})?|[0-9][0-9,]*(?:\.\d{1,2})?$", flags=re.IGNORECASE)
        content = amount_pattern.sub("", content).strip(" -,:")
    quantity, unit = _site_admin_find_quantity_unit(content)
    service_name = content
    description = content
    split_match = re.search(r"\b\d+(?:\.\d+)?\s*(?:nos?|pcs?|pieces?|qty|units?|kg|gms?|grams?|boxes?|packs?|sets?|hours?|days?)\b", content, flags=re.IGNORECASE)
    if split_match:
        before_qty = content[:split_match.start()].strip(" -,")
        if before_qty:
            chunks = before_qty.split()
            service_name = " ".join(chunks[: min(4, len(chunks))]).strip(" -,")
            description = before_qty
    return {
        "service_name": service_name[:180],
        "description": description[:500],
        "quantity": quantity,
        "unit": unit[:40],
    }


def _site_admin_split_item_blocks(text):
    raw = str(text or "").strip()
    if not raw:
        return []
    normalized = raw.replace("\r\n", "\n").replace("\r", "\n")
    numbered_matches = list(re.finditer(r"(?<!\w)(\d+)[\.\)]\s+", normalized))
    if len(numbered_matches) >= 2:
        blocks = []
        for index, match in enumerate(numbered_matches):
            start = match.end()
            end = numbered_matches[index + 1].start() if index + 1 < len(numbered_matches) else len(normalized)
            block = normalized[start:end].strip(" \n\t-")
            if block:
                blocks.append(block)
        if blocks:
            return blocks
    line_blocks = []
    for line in normalized.split("\n"):
        if re.match(r"^\s*(?:[-*]|\d+[\.\)])\s+", line):
            block = re.sub(r"^\s*(?:[-*]|\d+[\.\)])\s+", "", line).strip()
            if block:
                line_blocks.append(block)
    if len(line_blocks) >= 2:
        return line_blocks
    return [normalized]


def _site_admin_build_item_entry(text, fallback_amount=None):
    item_meta = _site_admin_parse_item_text(text)
    amount = _site_admin_find_amount(text)
    if amount is None and fallback_amount is not None:
        amount = _to_decimal(fallback_amount).quantize(Decimal("0.01"))
    quantity = item_meta.get("quantity")
    rate = amount
    if amount is not None and quantity is not None and quantity > 0:
        try:
            rate = (amount / quantity).quantize(Decimal("0.01"))
        except (InvalidOperation, ZeroDivisionError):
            rate = amount
    return {
        "service_name": str(item_meta.get("service_name") or "").strip()[:180],
        "description": str(item_meta.get("description") or text or "").strip()[:2000],
        "quantity": quantity,
        "unit": str(item_meta.get("unit") or "").strip()[:40],
        "rate": rate if rate is not None else None,
        "amount": amount if amount is not None else None,
    }


def _site_admin_parse_item_entries(text, fallback_total=None):
    blocks = _site_admin_split_item_blocks(text)
    if not blocks:
        return [], None
    fallback_total_decimal = None
    if fallback_total not in (None, ""):
        fallback_total_decimal = _to_decimal(fallback_total).quantize(Decimal("0.01"))
    entries = [_site_admin_build_item_entry(block) for block in blocks]
    missing_amount_entries = [entry for entry in entries if entry.get("amount") is None]
    known_total = sum((entry.get("amount") or Decimal("0.00")) for entry in entries)
    if len(entries) == 1 and missing_amount_entries and fallback_total_decimal is not None and fallback_total_decimal > 0:
        entries[0] = _site_admin_build_item_entry(blocks[0], fallback_total_decimal)
    elif (
        len(entries) > 1
        and len(missing_amount_entries) == 1
        and fallback_total_decimal is not None
        and fallback_total_decimal > known_total
    ):
        missing_entry = missing_amount_entries[0]
        missing_entry["amount"] = (fallback_total_decimal - known_total).quantize(Decimal("0.01"))
        if missing_entry.get("quantity") is not None and missing_entry["quantity"] > 0:
            try:
                missing_entry["rate"] = (missing_entry["amount"] / missing_entry["quantity"]).quantize(Decimal("0.01"))
            except (InvalidOperation, ZeroDivisionError):
                missing_entry["rate"] = missing_entry["amount"]
        else:
            missing_entry["rate"] = missing_entry["amount"]
    total = sum((entry.get("amount") or Decimal("0.00")) for entry in entries).quantize(Decimal("0.01"))
    return entries, total if total > 0 else None


def _site_admin_parse_message_fields(message):
    lines = [line.strip() for line in str(message or "").splitlines() if str(line or "").strip()]
    combined = " ".join(lines)
    parsed_mobile = _site_admin_find_mobile(combined)
    parsed_amount = _site_admin_find_amount(combined)
    if (
        parsed_mobile
        and len(lines) == 1
        and re.sub(r"\s+", "", combined) == parsed_mobile
    ):
        parsed_amount = None
    parsed = {
        "mobile": parsed_mobile,
        "client_name": "",
        "email": _site_admin_find_email(combined),
        "address": "",
        "gst_number": _site_admin_find_gst_number(combined),
        "item_text": "",
        "amount": parsed_amount,
    }
    remaining_lines = list(lines)
    if parsed["mobile"]:
        remaining_lines = [line for line in remaining_lines if _normalize_site_admin_mobile(line) != parsed["mobile"]]
    for index, line in enumerate(remaining_lines):
        candidate = _site_admin_parse_name_candidate(line)
        if candidate:
            parsed["client_name"] = candidate
            remaining_lines.pop(index)
            break
    if remaining_lines:
        parsed["item_text"] = " ".join(remaining_lines).strip()
    elif not parsed["client_name"] and len(lines) == 1:
        parsed["item_text"] = ""
    return parsed


def _site_admin_customer_phone_matches(row, mobile):
    target = _normalize_site_admin_mobile(mobile)
    if not target:
        return False
    candidates = [
        row.get("phone"),
        *(item.get("number") for item in (row.get("phoneList") or []) if isinstance(item, dict)),
        *(item.get("number") for item in (row.get("additionalPhones") or []) if isinstance(item, dict)),
    ]
    return any(_normalize_site_admin_mobile(value) == target for value in candidates)


def _site_admin_get_or_create_customer(org, user, *, mobile, client_name="", email="", address="", gst_number=""):
    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    customers = data.get("customers") if isinstance(data.get("customers"), list) else []
    normalized_mobile = _normalize_site_admin_mobile(mobile)
    normalized_email = _normalize_site_admin_email(email)
    existing_row = next(
        (
            row for row in customers
            if isinstance(row, dict) and _site_admin_customer_phone_matches(row, normalized_mobile)
        ),
        None,
    )
    changed = False
    if existing_row is None:
        customer_id = f"cust_{secrets.token_hex(6)}"
        existing_row = {
            "id": customer_id,
            "companyName": client_name[:180],
            "clientName": client_name[:180],
            "name": client_name[:180],
            "gstin": gst_number[:32],
            "phoneCountryCode": "+91",
            "phone": normalized_mobile,
            "additionalPhones": [],
            "phoneList": [{"countryCode": "+91", "number": normalized_mobile}],
            "email": normalized_email,
            "additionalEmails": [],
            "emailList": [normalized_email] if normalized_email else [],
            "billingAddress": address[:500],
            "shippingAddress": address[:500],
            "billingCountry": str(org.country or "India").strip() or "India",
            "shippingCountry": str(org.country or "India").strip() or "India",
            "billingState": "",
            "shippingState": "",
            "billingPincode": "",
            "shippingPincode": "",
            "createdAt": timezone.now().isoformat(),
            "updatedAt": timezone.now().isoformat(),
        }
        customers.append(existing_row)
        changed = True
    else:
        if client_name and not str(existing_row.get("clientName") or "").strip():
            existing_row["clientName"] = client_name[:180]
            changed = True
        if client_name and not str(existing_row.get("companyName") or existing_row.get("name") or "").strip():
            existing_row["companyName"] = client_name[:180]
            existing_row["name"] = client_name[:180]
            changed = True
        if normalized_email and not str(existing_row.get("email") or "").strip():
            existing_row["email"] = normalized_email
            changed = True
        if normalized_email:
            email_list = existing_row.get("emailList") if isinstance(existing_row.get("emailList"), list) else []
            if normalized_email not in {_normalize_site_admin_email(item) for item in email_list}:
                email_list.append(normalized_email)
                existing_row["emailList"] = [item for item in email_list if str(item or "").strip()]
                changed = True
        if address and not str(existing_row.get("billingAddress") or "").strip():
            existing_row["billingAddress"] = address[:500]
            existing_row["shippingAddress"] = address[:500]
            changed = True
        if gst_number and not str(existing_row.get("gstin") or "").strip():
            existing_row["gstin"] = gst_number[:32]
            changed = True
        if not _site_admin_customer_phone_matches(existing_row, normalized_mobile):
            phone_list = existing_row.get("phoneList") if isinstance(existing_row.get("phoneList"), list) else []
            phone_list.append({"countryCode": "+91", "number": normalized_mobile})
            existing_row["phoneList"] = phone_list
            changed = True
        existing_row["updatedAt"] = timezone.now().isoformat()
    if changed:
        data["customers"] = customers
        workspace.data = data
        workspace.updated_by = user
        workspace.save(update_fields=["data", "updated_by", "updated_at"])
    customer_name = _get_accounts_customer_display_name(existing_row) or client_name or normalized_mobile
    return existing_row, customer_name


def _site_admin_update_customer_for_estimate(org, user, estimate, *, mobile="", client_name=""):
    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    customers = data.get("customers") if isinstance(data.get("customers"), list) else []
    target_mobile = _normalize_site_admin_mobile(mobile or estimate.mobile)
    target_client_name = str(client_name or estimate.client_name or "").strip()[:180]
    estimate_customer_id = str(getattr(estimate, "customer_id", "") or "").strip()

    existing_row = None
    if estimate_customer_id:
        existing_row = next(
            (
                row for row in customers
                if isinstance(row, dict) and str(row.get("id") or "").strip() == estimate_customer_id
            ),
            None,
        )
    if existing_row is None:
        existing_row = next(
            (
                row for row in customers
                if isinstance(row, dict) and _site_admin_customer_phone_matches(row, estimate.mobile)
            ),
            None,
        )

    if existing_row is None:
        existing_row, _ = _site_admin_get_or_create_customer(
            org,
            user,
            mobile=target_mobile,
            client_name=target_client_name,
            email=estimate.email or "",
            address=estimate.address or "",
            gst_number=estimate.gst_number or "",
        )
        return existing_row

    conflicting_row = next(
        (
            row for row in customers
            if isinstance(row, dict)
            and row is not existing_row
            and _site_admin_customer_phone_matches(row, target_mobile)
        ),
        None,
    )
    if conflicting_row is not None:
        conflict_name = _get_accounts_customer_display_name(conflicting_row) or target_mobile
        raise ValueError(f"Mobile number {target_mobile} already exists for {conflict_name}.")

    changed = False
    if target_client_name:
        if str(existing_row.get("clientName") or "").strip() != target_client_name:
            existing_row["clientName"] = target_client_name
            changed = True
        if str(existing_row.get("companyName") or "").strip() != target_client_name:
            existing_row["companyName"] = target_client_name
            changed = True
        if str(existing_row.get("name") or "").strip() != target_client_name:
            existing_row["name"] = target_client_name
            changed = True

    if target_mobile:
        if _normalize_site_admin_mobile(existing_row.get("phone")) != target_mobile:
            existing_row["phone"] = target_mobile
            changed = True
        if str(existing_row.get("phoneCountryCode") or "").strip() != "+91":
            existing_row["phoneCountryCode"] = "+91"
            changed = True
        phone_list = existing_row.get("phoneList") if isinstance(existing_row.get("phoneList"), list) else []
        if not any(_normalize_site_admin_mobile(item.get("number")) == target_mobile for item in phone_list if isinstance(item, dict)):
            phone_list.append({"countryCode": "+91", "number": target_mobile})
            existing_row["phoneList"] = phone_list
            changed = True

    existing_row["updatedAt"] = timezone.now().isoformat()
    if changed:
        data["customers"] = customers
        workspace.data = data
        workspace.updated_by = user
        workspace.save(update_fields=["data", "updated_by", "updated_at"])
    return existing_row


def _get_quick_estimate_contact_store(data):
    rows = data.get("quickEstimateContacts")
    return rows if isinstance(rows, list) else []


def _upsert_quick_estimate_contact(
    org,
    user,
    *,
    contact_id="",
    mobile="",
    client_name="",
    email="",
    address="",
    gst_number="",
):
    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    contacts = _get_quick_estimate_contact_store(data)
    normalized_mobile = _normalize_site_admin_mobile(mobile)
    normalized_name = str(client_name or "").strip()[:180]
    normalized_email = _normalize_site_admin_email(email)
    normalized_address = str(address or "").strip()[:500]
    normalized_gst = str(gst_number or "").strip()[:32]
    target_id = str(contact_id or "").strip()
    row = None
    if target_id:
        row = next((item for item in contacts if isinstance(item, dict) and str(item.get("id") or "").strip() == target_id), None)
    if row is None and normalized_mobile:
        row = next((item for item in contacts if isinstance(item, dict) and _normalize_site_admin_mobile(item.get("phone")) == normalized_mobile), None)
    if row is None:
        target_id = target_id or f"qe_contact_{secrets.token_hex(6)}"
        row = {"id": target_id, "createdAt": timezone.now().isoformat()}
        contacts.append(row)
    row["id"] = target_id or str(row.get("id") or "").strip()
    row["clientName"] = normalized_name
    row["phone"] = normalized_mobile
    row["email"] = normalized_email
    row["address"] = normalized_address
    row["gstin"] = normalized_gst
    row["updatedAt"] = timezone.now().isoformat()
    data["quickEstimateContacts"] = contacts
    workspace.data = data
    workspace.updated_by = user
    workspace.save(update_fields=["data", "updated_by", "updated_at"])
    return row


def _next_quick_estimate_number(org):
    with transaction.atomic():
        sequence, _ = QuickEstimateSequence.objects.select_for_update().get_or_create(
            organization=org,
            defaults={"next_number": 1},
        )
        current_number = max(1, int(sequence.next_number or 1))
        sequence.next_number = current_number + 1
        sequence.save(update_fields=["next_number", "updated_at"])
    return current_number, f"QE-{current_number:04d}"


def _format_quick_estimate_amount(value):
    amount = _to_decimal(value).quantize(Decimal("0.01"))
    return f"{amount:.2f}".rstrip("0").rstrip(".") if amount % 1 else f"{int(amount)}"


def _serialize_quick_estimate_item(row):
    return {
        "id": row.id,
        "service_name": row.service_name or "",
        "description": row.description or "",
        "quantity": _decimal_to_string(row.quantity) if row.quantity is not None else "",
        "unit": row.unit or "",
        "rate": _decimal_to_string(row.rate) if row.rate is not None else "",
        "amount": _decimal_to_string(row.amount),
    }


def _serialize_quick_estimate_item_snapshot(entry):
    if not isinstance(entry, dict):
        return {}
    quantity = entry.get("quantity")
    rate = entry.get("rate")
    amount = entry.get("amount")
    return {
        "service_name": str(entry.get("service_name") or "").strip(),
        "description": str(entry.get("description") or "").strip(),
        "quantity": _decimal_to_string(quantity) if quantity is not None else "",
        "unit": str(entry.get("unit") or "").strip(),
        "rate": _decimal_to_string(rate) if rate is not None else "",
        "amount": _decimal_to_string(amount) if amount is not None else "",
    }


def _build_quick_estimate_whatsapp_url(row):
    public_preview_url = _build_quick_estimate_public_preview_url(row)
    items = list(row.items.all().order_by("id"))
    item_lines = []
    for index, item in enumerate(items, start=1):
        left = " ".join(part for part in [item.service_name, item.description] if str(part or "").strip()).strip()
        qty_text = ""
        if item.quantity is not None:
            qty_value = _decimal_to_string(item.quantity)
            qty_text = f" {qty_value}{(' ' + item.unit) if item.unit else ''}".strip()
        item_lines.append(
            f"{index}. {left}{(' ' + qty_text) if qty_text else ''} - Rs.{_format_quick_estimate_amount(item.amount)}".strip()
        )
    message = "\n".join([
        "Ultra HD Prints / Work Zilla",
        f"Quick Estimate: {row.estimate_number}",
        "",
        f"Client: {row.client_name}",
        f"Mobile: {row.mobile}",
        "",
        "Items:",
        *item_lines,
        "",
        f"Total: Rs.{_format_quick_estimate_amount(row.total_amount)}",
        "",
        f"Preview: {public_preview_url}",
        "",
        "Thank you.",
    ])
    return f"https://wa.me/91{row.mobile}?text={quote(message)}"


def _get_public_site_base_url():
    configured = str(getattr(settings, "SITE_BASE_URL", "") or "").strip()
    if configured:
        return configured.rstrip("/")
    return "https://getworkzilla.com"


def _base36_encode(value: int) -> str:
    digits = "0123456789abcdefghijklmnopqrstuvwxyz"
    value = int(value or 0)
    if value <= 0:
        return "0"
    encoded = ""
    while value:
        value, remainder = divmod(value, 36)
        encoded = digits[remainder] + encoded
    return encoded


def _base36_decode(value: str) -> int:
    return int(str(value or "0").strip().lower(), 36)


def _build_quick_estimate_public_token(row) -> str:
    estimate_id = int(getattr(row, "id", 0) or 0)
    org_id = int(getattr(row, "organization_id", 0) or 0)
    id_part = _base36_encode(estimate_id)
    checksum_source = f"{estimate_id}:{org_id}:{settings.SECRET_KEY}".encode("utf-8")
    checksum = hashlib.sha256(checksum_source).hexdigest()[:8].lower()
    return f"qe{id_part}{checksum}"


def _resolve_quick_estimate_public_token(token: str) -> Optional[int]:
    raw = str(token or "").strip().lower()
    if not raw.startswith("qe") or len(raw) <= 10:
        return None
    id_part = raw[2:-8]
    checksum = raw[-8:]
    if not id_part:
        return None
    try:
        estimate_id = _base36_decode(id_part)
    except ValueError:
        return None
    row = QuickEstimate.objects.filter(id=estimate_id).only("id", "organization_id").first()
    if not row:
        return None
    expected = _build_quick_estimate_public_token(row)
    return estimate_id if expected == raw else None


def _build_quick_estimate_public_preview_url(row):
    token = _build_quick_estimate_public_token(row)
    return f"{_get_public_site_base_url()}/api/business-autopilot/qe/{token}/"


def _get_quick_estimate_settings(org):
    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    settings_data = data.get("quickEstimateSettings") if isinstance(data.get("quickEstimateSettings"), dict) else {}
    header_text = _normalize_quick_estimate_header_html(settings_data.get("headerText"))
    template_size = str(settings_data.get("templateSize") or "4in").strip().lower()
    if template_size not in {"3in", "4in"}:
        template_size = "4in"
    retention_days = str(settings_data.get("paymentProofRetentionDays") or "45").strip()
    if retention_days not in {"45", "60"}:
        retention_days = "45"
    return {
        "headerText": header_text,
        "templateSize": template_size,
        "paymentProofRetentionDays": retention_days,
    }


def _purge_expired_quick_estimate_payment_proof(row):
    if not row or not getattr(row, "pk", None) or not str(getattr(row, "payment_proof_image", "") or "").strip():
        return row
    settings_data = _get_quick_estimate_settings(getattr(row, "organization", None))
    retention_days = int(str(settings_data.get("paymentProofRetentionDays") or "45").strip() or "45")
    updated_at = getattr(row, "updated_at", None) or getattr(row, "created_at", None)
    if not updated_at:
        return row
    if timezone.now() < updated_at + timedelta(days=retention_days):
        return row
    row.payment_proof_image = ""
    row.save(update_fields=["payment_proof_image", "updated_at"])
    return row


def _render_quick_estimate_thermal_preview(row, org):
    seller = InvoiceSellerProfile.objects.order_by("-updated_at").first()
    company_name = str(getattr(seller, "company_name", "") or getattr(org, "name", "") or "Work Zilla").strip() or "Work Zilla"
    settings_data = _get_quick_estimate_settings(org)
    header_text = _normalize_quick_estimate_header_html(settings_data.get("headerText"))
    template_size = str(settings_data.get("templateSize") or "4in").strip().lower()
    assigned_user = getattr(row, "assigned_user", None)
    if template_size not in {"3in", "4in"}:
        template_size = "4in"
    return render_to_string(
        "business_autopilot/quick_estimate_thermal_preview.html",
        {
            "company_name": company_name,
            "header_text": header_text,
            "template_size": template_size,
            "thermal_width_px": 288 if template_size == "3in" else 384,
            "estimate": row,
            "items": list(row.items.all().order_by("id")),
            "formatted_total": _format_quick_estimate_amount(row.total_amount),
            "formatted_subtotal": _format_quick_estimate_amount(row.subtotal),
            "created_at_label": timezone.localtime(row.created_at).strftime("%d/%m/%Y, %I:%M %p") if row.created_at else "",
            "assigned_user_name": _get_org_user_display_name(assigned_user),
        },
    )


def _wrap_quick_estimate_pdf_lines(text, max_width, font_name, font_size):
    cleaned = re.sub(r"\s+", " ", str(text or "").strip())
    if not cleaned:
        return []
    words = cleaned.split(" ")
    lines = []
    current = []
    for word in words:
        candidate = " ".join(current + [word]).strip()
        if current and pdfmetrics.stringWidth(candidate, font_name, font_size) > max_width:
            lines.append(" ".join(current).strip())
            current = [word]
        else:
            current.append(word)
    if current:
        lines.append(" ".join(current).strip())
    return lines


def _quick_estimate_pdf_header_segments(header_html, company_name, content_width):
    raw_html = str(header_html or "").strip()
    if not raw_html:
        return [{"text": line, "font_name": "Helvetica-Bold", "font_size": 13} for line in _wrap_quick_estimate_pdf_lines(company_name, content_width, "Helvetica-Bold", 13)]

    normalized = raw_html
    normalized = re.sub(r"<\s*br\s*/?\s*>", "\n", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"<\s*/?\s*p[^>]*>", "\n", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"<\s*/?\s*div[^>]*>", "\n", normalized, flags=re.IGNORECASE)

    segments = []
    for raw_line in [part.strip() for part in normalized.splitlines() if part.strip()]:
        is_bold = bool(re.search(r"<\s*(strong|b)\b", raw_line, flags=re.IGNORECASE))
        font_size_match = re.search(r"<font[^>]*size=[\"']?([1-7])[\"']?[^>]*>", raw_line, flags=re.IGNORECASE)
        html_font_size = font_size_match.group(1) if font_size_match else ""
        text = strip_tags(raw_line).strip()
        if not text:
            continue
        if html_font_size == "4":
            font_name = "Helvetica-Bold"
            font_size = 13.8
        elif html_font_size == "3":
            font_name = "Helvetica-Bold" if is_bold else "Helvetica"
            font_size = 10.6
        elif html_font_size == "2":
            font_name = "Helvetica-Bold" if is_bold else "Helvetica"
            font_size = 9.3
        else:
            font_name = "Helvetica-Bold" if is_bold else "Helvetica"
            font_size = 13 if is_bold else 9.4
        wrapped = _wrap_quick_estimate_pdf_lines(text, content_width, font_name, font_size)
        for line in wrapped:
            segments.append({
                "text": line,
                "font_name": font_name,
                "font_size": font_size,
            })
    return segments or [{"text": line, "font_name": "Courier-Bold", "font_size": 13} for line in _wrap_quick_estimate_pdf_lines(company_name, content_width, "Courier-Bold", 13)]


def _quick_estimate_pdf_header_images(header_html):
    raw_html = str(header_html or "").strip()
    if not raw_html:
        return []
    images = []
    for match in re.finditer(r"<img[^>]+src=[\"']([^\"']+)[\"'][^>]*>", raw_html, flags=re.IGNORECASE):
        image = _ba_pdf_image_from_data_url(match.group(1))
        if image:
            images.append(image)
    return images


def _quick_estimate_pdf_response(row, org):
    settings_data = _get_quick_estimate_settings(org)
    template_size = str(settings_data.get("templateSize") or "4in").strip().lower()
    if template_size not in {"3in", "4in"}:
        template_size = "4in"
    page_width = 3 * inch if template_size == "3in" else 4 * inch
    margin_x = 0.18 * inch if template_size == "3in" else 0.22 * inch
    content_width = page_width - (margin_x * 2)
    seller = InvoiceSellerProfile.objects.order_by("-updated_at").first()
    company_name = str(getattr(seller, "company_name", "") or getattr(org, "name", "") or "Work Zilla").strip() or "Work Zilla"
    header_text = _normalize_quick_estimate_header_html(settings_data.get("headerText"))
    header_segments = _quick_estimate_pdf_header_segments(header_text, company_name, content_width)
    header_images = _quick_estimate_pdf_header_images(header_text)
    created_at_label = timezone.localtime(row.created_at).strftime("%d/%m/%Y, %I:%M %p") if row.created_at else ""
    assigned_user_name = _get_org_user_display_name(getattr(row, "assigned_user", None)) or "-"
    items = list(row.items.all().order_by("id"))
    item_blocks = []
    total_height = 0
    amount_font = _ba_get_unicode_pdf_font() or "Helvetica"
    item_text_width = content_width - 60
    for item in items:
        title = str(item.service_name or "").strip()
        description = str(item.description or "").strip()
        qty_line = ""
        if item.quantity:
            qty_line = f"Qty: {item.quantity}{f' {item.unit}' if item.unit else ''}"
        lines = []
        if title:
            lines.extend(_wrap_quick_estimate_pdf_lines(title, item_text_width, "Helvetica-Bold", 10.2))
        if description:
            lines.extend(_wrap_quick_estimate_pdf_lines(description, item_text_width, "Helvetica", 9.4))
        if qty_line:
            lines.extend(_wrap_quick_estimate_pdf_lines(qty_line, item_text_width, "Helvetica", 9))
        if not lines:
            lines = ["-"]
        block_height = max(20, len(lines) * 13 + 8)
        total_height += block_height
        item_blocks.append({
            "lines": lines,
            "amount": f"Rs { _format_quick_estimate_amount(item.amount) }",
            "title_lines": len(_wrap_quick_estimate_pdf_lines(title, item_text_width, "Helvetica-Bold", 10.2)) if title else 0,
        })
    header_height = sum(14 if segment["font_size"] >= 13 else 10 for segment in header_segments)
    header_image_height = len(header_images) * 58
    meta_height = 11  # title separator gap
    for label, value in [
        ("Estimate No", row.estimate_number or "-"),
        ("Date", created_at_label or "-"),
        ("Client", row.client_name or "-"),
        ("Mobile", row.mobile or "-"),
        ("EMP Name", assigned_user_name),
    ]:
        meta_width = content_width * 0.62 if str(label).lower() == "date" else content_width * 0.54
        value_lines = _wrap_quick_estimate_pdf_lines(value, meta_width, "Helvetica-Bold", 10) or ["-"]
        meta_height += 10 + (max(0, len(value_lines) - 1) * 9)

    items_section_height = 30 + total_height
    total_section_height = 28
    footer_height = 10
    top_padding = 18
    bottom_padding = 4
    page_height = (
        top_padding
        + header_height
        + header_image_height
        + 20  # quick estimate title block
        + meta_height
        + items_section_height
        + total_section_height
        + footer_height
        + bottom_padding
    )
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(page_width, page_height))
    pdf.setTitle(str(row.estimate_number or "quick-estimate"))
    left = margin_x
    right = page_width - margin_x
    y = page_height - 18

    for segment in header_segments:
        pdf.setFillColorRGB(0.07, 0.09, 0.13)
        pdf.setFont(segment["font_name"], segment["font_size"])
        pdf.drawCentredString(page_width / 2, y, segment["text"])
        y -= 14 if segment["font_size"] >= 13 else 10

    if header_images:
        y -= 2
        image_size = min(48, max(34, content_width * 0.28))
        for image in header_images:
            pdf.drawImage(
                image,
                (page_width - image_size) / 2,
                y - image_size,
                width=image_size,
                height=image_size,
                preserveAspectRatio=True,
                mask="auto",
            )
            y -= image_size + 10

    y -= 6
    pdf.setFillColorRGB(0, 0, 0)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawCentredString(page_width / 2, y - 1, "QUICK ESTIMATE")
    y -= 14

    pdf.setStrokeColorRGB(0, 0, 0)
    pdf.setDash(3, 2)
    pdf.line(left, y, right, y)
    y -= 11
    pdf.setDash()

    meta_rows = [
        ("Estimate No", row.estimate_number or "-"),
        ("Date", created_at_label or "-"),
        ("Client", row.client_name or "-"),
        ("Mobile", row.mobile or "-"),
        ("EMP Name", assigned_user_name),
    ]
    for label, value in meta_rows:
        pdf.setFont("Helvetica", 9.2)
        pdf.setFillColorRGB(0, 0, 0)
        pdf.drawString(left, y, str(label))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.setFillColorRGB(0, 0, 0)
        meta_width = content_width * 0.62 if str(label).lower() == "date" else content_width * 0.54
        value_lines = _wrap_quick_estimate_pdf_lines(value, meta_width, "Helvetica-Bold", 10) or ["-"]
        pdf.drawRightString(right, y, value_lines[0])
        y -= 10
        for continuation in value_lines[1:]:
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawRightString(right, y, continuation)
            y -= 9

    pdf.setStrokeColorRGB(0, 0, 0)
    pdf.setDash(3, 2)
    pdf.line(left, y + 2, right, y + 2)
    pdf.setDash()
    y -= 10

    pdf.setFont("Helvetica-Bold", 9.4)
    pdf.setFillColorRGB(0, 0, 0)
    pdf.drawString(left, y, "Items")
    pdf.drawRightString(right, y, "Amt")
    y -= 8
    pdf.setStrokeColorRGB(0, 0, 0)
    pdf.line(left, y, right, y)
    y -= 12

    for block in item_blocks:
        line_y = y
        for idx, line in enumerate(block["lines"]):
            if idx < block["title_lines"]:
                pdf.setFont("Helvetica-Bold", 10.2)
                pdf.setFillColorRGB(0, 0, 0)
            elif idx == len(block["lines"]) - 1 and line.startswith("Qty:"):
                pdf.setFont("Helvetica", 9)
                pdf.setFillColorRGB(0, 0, 0)
            else:
                pdf.setFont("Helvetica", 9.4)
                pdf.setFillColorRGB(0, 0, 0)
            pdf.drawString(left, line_y, line)
            line_y -= 12.6
        pdf.setFont(amount_font, 10.2)
        pdf.setFillColorRGB(0, 0, 0)
        pdf.drawRightString(right, y, block["amount"])
        y = line_y - 3.5
        pdf.setStrokeColorRGB(0, 0, 0)
        pdf.setDash(3, 2)
        pdf.line(left, y, right, y)
        pdf.setDash()
        y -= 8

    y -= 6
    pdf.setFont("Helvetica-Bold", 11.5)
    pdf.setFillColorRGB(0, 0, 0)
    pdf.drawString(left, y, "Total")
    pdf.setFont(amount_font, 11.5)
    pdf.drawRightString(right, y, f"Rs { _format_quick_estimate_amount(row.total_amount) }")
    y -= 6

    pdf.setStrokeColorRGB(0, 0, 0)
    pdf.setDash(3, 2)
    pdf.line(left, y, right, y)
    pdf.setDash()
    y -= 6
    pdf.setFont("Helvetica", 9)
    pdf.setFillColorRGB(0, 0, 0)
    pdf.drawCentredString(page_width / 2, y, "Thank you.")
    pdf.showPage()
    pdf.save()
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    filename = f"{str(row.estimate_number or 'quick-estimate').replace(' ', '_')}_{template_size}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _safe_render_quick_estimate_thermal_preview(row, org):
    try:
        return _render_quick_estimate_thermal_preview(row, org)
    except Exception:
        logger.exception("Quick estimate thermal preview render failed for estimate_id=%s", getattr(row, "id", None))
        return ""


def _serialize_quick_estimate(row, include_preview=False):
    row = _purge_expired_quick_estimate_payment_proof(row)
    created_by = getattr(row, "created_by", None) or getattr(getattr(row, "organization", None), "owner", None)
    assigned_user = getattr(row, "assigned_user", None)
    assigned_by = getattr(row, "assigned_by", None)
    payment_verified_by = getattr(row, "payment_verified_by", None)
    job_verified_by = getattr(row, "job_verified_by", None)
    delivery_verified_by = getattr(row, "delivery_verified_by", None)
    payload = {
        "id": row.id,
        "estimate_number": row.estimate_number,
        "mobile": row.mobile,
        "client_name": row.client_name,
        "notes": row.notes or "",
        "email": row.email or "",
        "address": row.address or "",
        "gst_number": row.gst_number or "",
        "subtotal": _decimal_to_string(row.subtotal),
        "tax_amount": _decimal_to_string(row.tax_amount),
        "total_amount": _decimal_to_string(row.total_amount),
        "status": row.status,
        "payment_status": str(getattr(row, "payment_status", "") or ""),
        "payment_mode": str(getattr(row, "payment_mode", "") or ""),
        "job_status": str(getattr(row, "job_status", "") or ""),
        "delivery_status": str(getattr(row, "delivery_status", "") or ""),
        "payment_proof_image": str(getattr(row, "payment_proof_image", "") or ""),
        "payment_verified_by_name": _get_org_user_display_name(payment_verified_by),
        "job_verified_by_name": _get_org_user_display_name(job_verified_by),
        "delivery_verified_by_name": _get_org_user_display_name(delivery_verified_by),
        "customer_id": row.customer_id or "",
        "created_by_id": getattr(row, "created_by_id", None) or getattr(getattr(row, "organization", None), "owner_id", None),
        "created_by_name": _get_org_user_display_name(created_by),
        "created_by_username": str(getattr(created_by, "username", "") or "").strip(),
        "created_by_email": str(getattr(created_by, "email", "") or "").strip(),
        "assigned_user_id": getattr(row, "assigned_user_id", None),
        "assigned_membership_id": (
            OrganizationUser.objects
            .filter(organization=row.organization, user_id=getattr(row, "assigned_user_id", None), is_deleted=False)
            .values_list("id", flat=True)
            .first()
            if getattr(row, "assigned_user_id", None)
            else None
        ),
        "assigned_user_name": _get_org_user_display_name(assigned_user),
        "assigned_by_id": getattr(row, "assigned_by_id", None),
        "assigned_by_name": _get_org_user_display_name(assigned_by),
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "updated_at": row.updated_at.isoformat() if row.updated_at else "",
        "whatsapp_url": _build_quick_estimate_whatsapp_url(row),
        "thermal_preview_url": f"/api/business-autopilot/quick-estimates/{row.id}/thermal-preview/",
        "thermal_preview_pdf_url": f"/api/business-autopilot/quick-estimates/{row.id}/thermal-preview/?format=pdf",
        "items": [_serialize_quick_estimate_item(item) for item in row.items.all().order_by("id")],
    }
    if include_preview:
        payload["thermal_preview_html"] = _safe_render_quick_estimate_thermal_preview(row, row.organization)
    return payload


def _serialize_quick_estimate_history(row):
    snapshot = row.snapshot if isinstance(row.snapshot, dict) else {}
    return {
        "id": row.id,
        "action": row.action,
        "details": row.note or "",
        "edit_by": _get_org_user_display_name(getattr(row, "actor", None)),
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "snapshot": snapshot,
    }


def _record_quick_estimate_history(estimate, *, action="updated", actor=None, note="", snapshot=None):
    if not estimate:
        return
    QuickEstimateHistory.objects.create(
        quick_estimate=estimate,
        action=str(action or QuickEstimateHistory.ACTION_UPDATED).strip() or QuickEstimateHistory.ACTION_UPDATED,
        actor=actor,
        note=str(note or "").strip(),
        snapshot=snapshot if isinstance(snapshot, dict) else {},
    )


def _serialize_quick_estimate_contact(row, *, linked_estimate_count=0):
    if not isinstance(row, dict):
        row = {}
    phone = _normalize_site_admin_mobile(row.get("phone"))
    client_name = str(row.get("clientName") or row.get("companyName") or row.get("name") or "").strip()
    return {
        "id": str(row.get("id") or "").strip(),
        "client_name": client_name,
        "mobile": phone,
        "email": _normalize_site_admin_email(row.get("email")),
        "address": str(row.get("address") or row.get("billingAddress") or row.get("shippingAddress") or "").strip(),
        "gst_number": str(row.get("gstin") or "").strip(),
        "linked_estimate_count": int(linked_estimate_count or 0),
        "created_at": str(row.get("createdAt") or "").strip(),
        "updated_at": str(row.get("updatedAt") or "").strip(),
    }


def _quick_estimate_contact_rows_with_counts(org):
    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    contacts = _get_quick_estimate_contact_store(data)
    estimate_counts = {}
    for customer_id in (
        QuickEstimate.objects
        .filter(organization=org)
        .exclude(customer_id="")
        .values_list("customer_id", flat=True)
    ):
        key = str(customer_id or "").strip()
        if key:
            estimate_counts[key] = estimate_counts.get(key, 0) + 1
    rows = []
    for row in contacts:
        if not isinstance(row, dict):
            continue
        contact_id = str(row.get("id") or "").strip()
        if not contact_id:
            continue
        rows.append(_serialize_quick_estimate_contact(row, linked_estimate_count=estimate_counts.get(contact_id, 0)))
    rows.sort(key=lambda item: (str(item.get("client_name") or "").lower(), str(item.get("mobile") or "")))
    return workspace, data, contacts, rows


def _sync_quick_estimate_contacts_from_estimates(org, user=None):
    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    contacts = _get_quick_estimate_contact_store(data)
    known_ids = {
        str(row.get("id") or "").strip()
        for row in contacts
        if isinstance(row, dict) and str(row.get("id") or "").strip()
    }
    changed = False
    estimates = (
        QuickEstimate.objects
        .filter(organization=org)
        .exclude(customer_id="")
        .order_by("created_at", "id")
    )
    for estimate in estimates:
        contact_id = str(getattr(estimate, "customer_id", "") or "").strip()
        if not contact_id or contact_id in known_ids:
            continue
        normalized_mobile = _normalize_site_admin_mobile(getattr(estimate, "mobile", ""))
        normalized_name = str(getattr(estimate, "client_name", "") or "").strip()[:180]
        if not normalized_mobile or not normalized_name:
            continue
        contacts.append({
            "id": contact_id,
            "clientName": normalized_name,
            "phone": normalized_mobile,
            "email": _normalize_site_admin_email(getattr(estimate, "email", "")),
            "address": str(getattr(estimate, "address", "") or "").strip()[:500],
            "gstin": str(getattr(estimate, "gst_number", "") or "").strip()[:32],
            "createdAt": (estimate.created_at.isoformat() if getattr(estimate, "created_at", None) else timezone.now().isoformat()),
            "updatedAt": (estimate.updated_at.isoformat() if getattr(estimate, "updated_at", None) else timezone.now().isoformat()),
        })
        known_ids.add(contact_id)
        changed = True
    if changed:
        data["quickEstimateContacts"] = contacts
        workspace.data = data
        workspace.updated_by = user
        workspace.save(update_fields=["data", "updated_by", "updated_at"])
    return workspace, data, contacts


def _ensure_quick_estimate_contact_name_mobile(*, mobile="", client_name="", require_client_name=True):
    normalized_mobile = _normalize_site_admin_mobile(mobile)
    normalized_name = str(client_name or "").strip()
    if not normalized_mobile:
        raise ValueError("Please enter the mobile number.")
    if require_client_name and not normalized_name:
        raise ValueError("Please enter the client name.")
    return normalized_mobile, normalized_name[:180]


def _to_decimal(value):
    try:
        return Decimal(str(value or 0))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def _normalize_subscription_alert_days(value):
    if value is None:
        return None

    if isinstance(value, str):
        source = value.strip()
        if not source:
            return []
        items = [part.strip() for part in source.replace(";", ",").split(",") if part.strip()]
        if len(items) == 1 and "," not in source and ";" not in source and source.isdigit():
            items = [source]
        candidates = items
    elif isinstance(value, (list, tuple, set)):
        candidates = value
    else:
        candidates = [value]

    def _coerce_item(item):
        try:
            number = int(str(item).strip())
        except (TypeError, ValueError):
            return None
        if number < 0:
            return None
        return str(number)

    normalized = []
    for item in candidates:
        if str(item).strip() == "":
            continue
        normalized_item = _coerce_item(item)
        if normalized_item is None:
            return None
        if normalized_item not in normalized:
            normalized.append(normalized_item)
    return normalized


def _is_valid_alert_email(value: str):
    text = str(value or "").strip()
    if not text:
        return False
    if "@" not in text:
        return False
    local, domain = text.rsplit("@", 1)
    if not local or not domain:
        return False
    if "." not in domain:
        return False
    return True


def _normalize_subscription_alert_assignees(value):
    if value is None:
        return []
    if not isinstance(value, (list, tuple, set)):
        return None
    normalized = []
    seen = set()
    for row in value:
        if isinstance(row, dict):
            raw_type = str(row.get("type") or "").strip().lower()
            raw_value = str(row.get("value") or row.get("name") or row.get("label") or "").strip()
            raw_label = str(row.get("label") or row.get("name") or raw_value).strip()
            if raw_type == "user":
                user_id = _coerce_positive_int(row.get("id") or row.get("value"))
                if not user_id or not raw_value:
                    return None
                recipient_key = f"user:{user_id}"
                if recipient_key in seen:
                    continue;
                normalized.append({
                    "type": "user",
                    "value": str(user_id),
                    "label": raw_label
                })
                seen.add(recipient_key)
            elif raw_type == "department":
                if not raw_value:
                    return None
                recipient_key = f"department:{raw_value.lower()}"
                if recipient_key in seen:
                    continue;
                normalized.append({
                    "type": "department",
                    "value": raw_value,
                    "label": raw_value
                })
                seen.add(recipient_key)
            elif raw_type == "email":
                email_value = raw_value.lower()
                if not _is_valid_alert_email(email_value):
                    return None
                recipient_key = f"email:{email_value}"
                if recipient_key in seen:
                    continue
                normalized.append({
                    "type": "email",
                    "value": email_value,
                    "label": raw_label or email_value,
                })
                seen.add(recipient_key)
            else:
                return None
        elif isinstance(row, str):
            raw_value = row.strip()
            if not raw_value:
                continue
            if raw_value.lower().startswith("user:"):
                raw_text = raw_value.split(":", 1)[1].strip()
                user_id = _coerce_positive_int(raw_text)
                if not user_id:
                    return None
                recipient_key = f"user:{user_id}"
                if recipient_key in seen:
                    continue;
                normalized.append({
                    "type": "user",
                    "value": str(user_id),
                    "label": raw_text
                })
                seen.add(recipient_key)
            elif raw_value.lower().startswith("department:"):
                raw_text = raw_value.split(":", 1)[1].strip()
                if not raw_text:
                    return None
                recipient_key = f"department:{raw_text.lower()}"
                if recipient_key in seen:
                    continue;
                normalized.append({
                    "type": "department",
                    "value": raw_text,
                    "label": raw_text
                })
                seen.add(recipient_key)
            elif raw_value.lower().startswith("email:"):
                raw_text = raw_value.split(":", 1)[1].strip().lower()
                if not _is_valid_alert_email(raw_text):
                    return None
                recipient_key = f"email:{raw_text}"
                if recipient_key in seen:
                    continue
                normalized.append({
                    "type": "email",
                    "value": raw_text,
                    "label": raw_text
                })
                seen.add(recipient_key)
            else:
                return None
        else:
            return None
    return normalized


def _serialize_subscription_alert_assignees(value):
    return _normalize_subscription_alert_assignees(value) or []


def _coerce_subscription_alert_assignees(value):
    normalized = _normalize_subscription_alert_assignees(value)
    if normalized is None:
        return None
    return normalized


def _get_org_admin_alert_recipients(org):
    if not org:
        return []
    rows = (
        OrganizationUser.objects
        .filter(organization=org, role="company_admin", is_active=True, user__is_active=True)
        .select_related("user")
    )
    recipients = []
    seen = set()
    for row in rows:
        user_id = _coerce_positive_int(row.user_id)
        if not user_id:
            continue
        recipient_key = f"user:{user_id}"
        if recipient_key in seen:
            continue
        label = _get_org_user_display_name(row.user)
        recipients.append({
            "type": "user",
            "value": str(user_id),
            "label": label
        })
        seen.add(recipient_key)
    return recipients


def _merge_subscription_alert_recipients(primary, org):
    merged = []
    seen = set()
    for entry in list(primary or []) + list(_get_org_admin_alert_recipients(org) or []):
        if not isinstance(entry, dict):
            continue
        entry_type = str(entry.get("type") or "").strip().lower()
        entry_value = str(entry.get("value") or "").strip()
        if not entry_type or not entry_value:
            continue
        key = f"{entry_type}:{entry_value.lower()}"
        if key in seen:
            continue
        merged.append(entry)
        seen.add(key)
    return merged


_LEGACY_AUTO_TAX_VALUES = {"0", "0.0", "0.00", "0.000"}
_GST_TEMPLATE_ID_ALIASES = {
    "gst_default_india_igst": "gst_default_india_igst",
    "gst_default_india_cgst_sgst": "gst_default_india_cgst_sgst",
    "gst_india_igst": "gst_default_india_igst",
    "gst_india_cgst_sgst": "gst_default_india_cgst_sgst",
}


def _document_template_components(gst_template):
    if not isinstance(gst_template, dict):
        return {
            "cgst": Decimal("0"),
            "sgst": Decimal("0"),
            "igst": Decimal("0"),
            "cess": Decimal("0"),
        }
    return {
        "cgst": _to_decimal(gst_template.get("cgst")),
        "sgst": _to_decimal(gst_template.get("sgst")),
        "igst": _to_decimal(gst_template.get("igst")),
        "cess": _to_decimal(gst_template.get("cess")),
    }


def _document_template_total_percent(gst_template):
    return sum(_document_template_components(gst_template).values(), Decimal("0"))


def _resolve_gst_template_by_id(gst_templates_by_id, template_id):
    raw_template_id = str(template_id or "").strip()
    if not raw_template_id:
        return None
    normalized_template_id = raw_template_id.lower().replace("-", "_").replace(" ", "_")
    canonical_template_id = _GST_TEMPLATE_ID_ALIASES.get(normalized_template_id, normalized_template_id)
    candidate_ids = []
    for value in (raw_template_id, normalized_template_id, canonical_template_id):
        if value and value not in candidate_ids:
            candidate_ids.append(value)
    if "cgst" in normalized_template_id and "sgst" in normalized_template_id:
        inferred = "gst_default_india_cgst_sgst"
        if inferred not in candidate_ids:
            candidate_ids.append(inferred)
    elif "igst" in normalized_template_id:
        inferred = "gst_default_india_igst"
        if inferred not in candidate_ids:
            candidate_ids.append(inferred)

    if isinstance(gst_templates_by_id, dict):
        for candidate_id in candidate_ids:
            direct = gst_templates_by_id.get(candidate_id)
            if isinstance(direct, dict):
                return direct

    for default_template in _default_india_gst_templates():
        default_template_id = str(default_template.get("id") or "").strip()
        if default_template_id in candidate_ids:
            return default_template

    return None


def _resolve_document_line_tax_override(row):
    if not isinstance(row, dict):
        return {
            "has_override": False,
            "tax_percent": Decimal("0"),
            "tax_percent_source": "auto",
        }

    raw_tax_value = row.get("taxPercent")
    if raw_tax_value is None:
        raw_tax_value = row.get("tax_percent")
    raw_tax_percent = str(raw_tax_value if raw_tax_value is not None else "").strip()
    raw_tax_source = row.get("taxPercentSource")
    if raw_tax_source is None:
        raw_tax_source = row.get("tax_percent_source")
    tax_source = str(raw_tax_source or "").strip().lower()

    if tax_source == "manual":
        has_override = bool(raw_tax_percent)
        return {
            "has_override": has_override,
            "tax_percent": _to_decimal(raw_tax_percent),
            "tax_percent_source": "manual" if has_override else "auto",
        }

    if tax_source in {"auto", "template"}:
        return {
            "has_override": False,
            "tax_percent": Decimal("0"),
            "tax_percent_source": "auto",
        }

    if (not raw_tax_percent) or (raw_tax_percent in _LEGACY_AUTO_TAX_VALUES):
        return {
            "has_override": False,
            "tax_percent": Decimal("0"),
            "tax_percent_source": "auto",
        }

    return {
        "has_override": True,
        "tax_percent": _to_decimal(raw_tax_percent),
        "tax_percent_source": "manual",
    }


def _document_totals(document, gst_templates_by_id):
    items = document.get("items") if isinstance(document, dict) else []
    if not isinstance(items, list):
        items = []
    gst_template_id = (
        str((document or {}).get("gstTemplateId") or (document or {}).get("gst_template_id") or "").strip()
        if isinstance(document, dict)
        else ""
    )
    gst_template = _resolve_gst_template_by_id(gst_templates_by_id, gst_template_id)
    template_components = _document_template_components(gst_template)
    default_tax = _document_template_total_percent(gst_template)
    subtotal = Decimal("0")
    tax_total = Decimal("0")
    tax_breakdown = {
        "cgst": Decimal("0"),
        "sgst": Decimal("0"),
        "igst": Decimal("0"),
        "cess": Decimal("0"),
    }
    for row in items:
        if not isinstance(row, dict):
            continue
        qty = _to_decimal(row.get("qty"))
        rate = _to_decimal(row.get("rate"))
        line_total = qty * rate
        tax_override = _resolve_document_line_tax_override(row)
        tax_pct = tax_override["tax_percent"] if tax_override["has_override"] else default_tax
        component_total = sum(template_components.values(), Decimal("0"))
        if component_total > 0:
            effective_components = {
                key: (tax_pct * value) / component_total
                for key, value in template_components.items()
            }
        else:
            effective_components = {
                "cgst": Decimal("0"),
                "sgst": Decimal("0"),
                "igst": tax_pct,
                "cess": Decimal("0"),
            }
        subtotal += line_total
        tax_total += (line_total * tax_pct) / Decimal("100")
        for key, value in effective_components.items():
            tax_breakdown[key] += (line_total * value) / Decimal("100")
    breakdown_percent = {
        key: ((value / subtotal) * Decimal("100")) if subtotal > 0 else Decimal("0")
        for key, value in tax_breakdown.items()
    }
    return {
        "subtotal": float(subtotal),
        "tax_total": float(tax_total),
        "grand_total": float(subtotal + tax_total),
        "tax_breakdown": {key: float(value) for key, value in tax_breakdown.items()},
        "breakdown_percent": {key: float(value) for key, value in breakdown_percent.items()},
    }


@require_http_methods(["GET", "POST"])
def org_enabled_modules(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False, "modules": []}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "modules": []})

    _ensure_org_modules(org)
    can_manage = _can_manage_modules(request.user)

    if request.method == "POST":
        if not can_manage:
            return JsonResponse({"detail": "forbidden"}, status=403)
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        module_slug = (payload.get("module_slug") or "").strip().lower()
        enabled = payload.get("enabled")
        if not module_slug or not isinstance(enabled, bool):
            return JsonResponse({"detail": "invalid_payload"}, status=400)
        module = Module.objects.filter(slug=module_slug, is_active=True).first()
        if not module:
            return JsonResponse({"detail": "module_not_found"}, status=404)
        active_sub = _get_active_erp_subscription(org)
        allowed_slugs = _get_plan_erp_module_slugs(
            active_sub.plan if active_sub and active_sub.plan else None,
            subscription=active_sub,
        )
        if enabled and module_slug not in allowed_slugs:
            return JsonResponse({"detail": "module_not_allowed_for_plan"}, status=400)
        org_module, _ = OrganizationModule.objects.get_or_create(
            organization=org,
            module=module,
            defaults={"enabled": enabled},
        )
        if org_module.enabled != enabled:
            org_module.enabled = enabled
            org_module.save(update_fields=["enabled", "updated_at"])

    modules, enabled_modules = _serialize_modules(org)
    return JsonResponse(
        {
            "authenticated": True,
            "organization": {
                "id": org.id,
                "name": org.name,
                "company_key": org.company_key,
            },
            "modules": enabled_modules,
            "enabled_modules": enabled_modules,
            "catalog": modules,
            "can_manage_modules": can_manage,
            "can_manage_users": _can_manage_users(request.user, org),
        }
    )


@require_http_methods(["GET", "POST"])
def org_users(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False, "users": [], "meta": {}}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "users": [], "meta": {}})

    can_manage_users = _can_manage_users(request.user, org)

    created_user_credentials = None
    credential_delivery = {
        "is_new_user": False,
        "email_sent": False,
        "status": "not_applicable",
    }

    if request.method == "POST":
        if not can_manage_users:
            return JsonResponse({"detail": "forbidden"}, status=403)
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        if payload.get("bulk_import") is True:
            rows = payload.get("rows")
            if not isinstance(rows, list) or not rows:
                return JsonResponse({"detail": "rows_required"}, status=400)
            results = []
            imported_count = 0
            skipped_count = 0
            failed_count = 0
            for index, row in enumerate(rows, start=1):
                if not isinstance(row, dict):
                    failed_count += 1
                    results.append({
                        "row_number": index,
                        "email": "",
                        "status": "failed",
                        "detail": "invalid_row",
                        "message": "Each imported row must be an object.",
                    })
                    continue
                _sync_org_users_to_plan_limit(org, requested_by=request.user)
                current_users = _safe_serialize_org_users(org)
                current_meta = _build_org_user_meta(org, users=current_users)
                if not current_meta.get("can_add_users"):
                    failed_count += 1
                    results.append({
                        "row_number": index,
                        "email": str(row.get("email") or "").strip().lower(),
                        "status": "failed",
                        "detail": "employee_limit_reached",
                        "message": current_meta.get("limit_message") or "User limit reached. Add-on users required.",
                    })
                    continue
                import_payload = dict(row)
                import_payload["confirm_existing_user"] = True
                result = _create_or_attach_org_user(org, import_payload, requested_by=request.user)
                if not result.get("ok"):
                    response = result.get("response")
                    detail_payload = {}
                    try:
                        detail_payload = json.loads(response.content.decode("utf-8") or "{}")
                    except Exception:
                        detail_payload = {}
                    detail = str(detail_payload.get("detail") or "import_failed").strip() or "import_failed"
                    failed_count += 1
                    results.append({
                        "row_number": index,
                        "email": str(row.get("email") or "").strip().lower(),
                        "status": "failed",
                        "detail": detail,
                        "message": str(detail_payload.get("message") or detail).strip(),
                    })
                    continue
                imported_count += 1
                status = "created" if result.get("newly_created_user") else "attached"
                if result.get("is_existing_user_added"):
                    skipped_count += 0
                results.append({
                    "row_number": index,
                    "email": str(getattr(result.get("user"), "email", "") or "").strip().lower(),
                    "status": status,
                    "detail": status,
                    "message": "Imported successfully." if status == "created" else "Existing user linked successfully.",
                })
            _sync_org_users_to_plan_limit(org, requested_by=request.user)
            payload = _build_org_users_response_payload(
                org,
                can_manage_users,
                message=f"Imported {imported_count} user(s).",
            )
            payload["organization"] = {
                "id": org.id,
                "name": org.name,
                "company_key": org.company_key,
            }
            payload["import_summary"] = {
                "total_rows": len(rows),
                "imported_count": imported_count,
                "skipped_count": skipped_count,
                "failed_count": failed_count,
                "results": results,
            }
            return JsonResponse(payload)
        _sync_org_users_to_plan_limit(org, requested_by=request.user)
        current_users = _safe_serialize_org_users(org)
        current_meta = _build_org_user_meta(org, users=current_users)
        if not current_meta.get("can_add_users"):
            return JsonResponse(
                {
                    "detail": "employee_limit_reached",
                    "message": current_meta.get("limit_message") or "User limit reached. Please increase user limit or mark inactive/resigned users.",
                    "meta": current_meta,
                },
                status=403,
            )
        result = _create_or_attach_org_user(org, payload, requested_by=request.user)
        if not result.get("ok"):
            return result["response"]
        created_user_credentials = result.get("created_user_credentials")
        credential_delivery = result.get("credential_delivery") or credential_delivery

    _sync_org_users_to_plan_limit(org, requested_by=request.user)
    payload = _build_org_users_response_payload(
        org,
        can_manage_users,
        created_user_credentials=created_user_credentials,
        credential_delivery=credential_delivery,
    )
    requested_status = str(request.GET.get("status") or "all").strip().lower()
    if requested_status in {OrganizationUser.STATUS_ACTIVE, OrganizationUser.STATUS_INACTIVE, OrganizationUser.STATUS_RESIGNED}:
        payload["users"] = [
            row for row in payload.get("users", [])
            if str(row.get("status") or "").strip().lower() == requested_status
        ]
    elif requested_status == OrganizationUser.STATUS_DELETED:
        payload["users"] = []
        payload["deleted_users"] = payload.get("deleted_users", [])
    payload["selected_status"] = requested_status or "all"
    payload["organization"] = {
        "id": org.id,
        "name": org.name,
        "company_key": org.company_key,
    }
    return JsonResponse(payload)


@require_http_methods(["GET"])
def org_user_email_check(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "available": False}, status=404)

    can_manage_users = _can_manage_users(request.user, org)
    if not can_manage_users:
        return JsonResponse({"detail": "forbidden"}, status=403)

    email = str(request.GET.get("email") or "").strip().lower()
    if not email:
        return JsonResponse({"ok": False, "available": False, "message": "Email is required."}, status=400)
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return JsonResponse({"ok": False, "available": False, "message": "Enter a valid email."}, status=400)

    existing_user = User.objects.filter(email__iexact=email).first()
    if not existing_user:
        return JsonResponse(
            {
                "ok": True,
                "available": True,
                "existing_user": False,
                "same_password_allowed": False,
                "password_required": True,
                "message": "Email is available.",
            }
        )

    existing_profile = UserProfile.objects.filter(user=existing_user).first()
    if existing_profile and existing_profile.organization_id and existing_profile.organization_id != org.id:
        return JsonResponse(
            {
                "ok": True,
                "available": False,
                "existing_user": True,
                "same_password_allowed": False,
                "password_required": False,
                "belongs_to_another_organization": True,
                "message": "This email is already assigned to another organization.",
            },
            status=409,
        )

    existing_products = _get_user_granted_products(existing_user)
    already_has_business_autopilot = any(
        product["slug"] == BUSINESS_AUTOPILOT_PRODUCT_SLUG for product in existing_products
    )
    if already_has_business_autopilot:
        return JsonResponse(
            {
                "ok": True,
                "available": False,
                "existing_user": True,
                "same_password_allowed": True,
                "password_required": False,
                "already_assigned_to_business_autopilot": True,
                "existing_products": existing_products,
                "message": "This user is already created in Business Autopilot.",
            },
            status=409,
        )

    return JsonResponse(
        {
            "ok": True,
            "available": False,
            "existing_user": True,
            "same_password_allowed": True,
            "password_required": False,
            "existing_products": existing_products,
            "message": "This user is already created in another product. The same password will continue to work.",
        }
    )


@require_http_methods(["PUT", "DELETE", "POST"])
def org_user_detail(request, membership_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "users": [], "meta": {}})

    can_manage_users = _can_manage_users(request.user, org)
    if not can_manage_users:
        return JsonResponse({"detail": "forbidden"}, status=403)

    membership = OrganizationUser.objects.filter(organization=org, id=membership_id).select_related("user").first()
    if not membership:
        return JsonResponse({"detail": "user_not_found"}, status=404)

    payload = {}
    resolved_method = request.method
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        action = str(payload.get("action") or "").strip().lower()
        if action == "restore":
            resolved_method = "RESTORE"
        elif action in {"delete", "remove"}:
            resolved_method = "DELETE"
        elif action in {"update", "edit", "save"} or any(
            key in payload
            for key in {
                "first_name",
                "last_name",
                "name",
                "email",
                "password",
                "phone_number",
                "role",
                "user_type",
                "department",
                "department_id",
                "employee_role",
                "employee_role_id",
                "is_active",
            }
        ):
            resolved_method = "PUT"
        else:
            return JsonResponse({"detail": "invalid_action"}, status=400)

    if resolved_method == "DELETE":
        try:
            if _is_org_admin_account_member(org, membership):
                return JsonResponse(
                    {
                        "detail": "org_admin_delete_forbidden",
                        "message": "ORG admin account cannot be deleted from Users.",
                    },
                    status=403,
                )
            permanent = str(request.GET.get("permanent") or "").strip().lower() in {"1", "true", "yes"}
            if request.method == "POST":
                permanent_raw = str(payload.get("permanent") or "").strip().lower()
                permanent = permanent or permanent_raw in {"1", "true", "yes"}
            if membership.is_deleted and not permanent:
                return JsonResponse({"detail": "user_already_deleted"}, status=400)

            linked_leads = _crm_collect_leads_linked_to_user(org, membership.user_id)
            linked_deals = _crm_collect_deals_linked_to_user(org, membership.user_id)

            reassign_membership_ids = payload.get("reassign_to_membership_ids") if isinstance(payload, dict) else None
            target_user_ids = _crm_resolve_target_user_ids(org, reassign_membership_ids, exclude_user_id=membership.user_id)
            if not target_user_ids:
                admin_user_id = _crm_resolve_org_admin_user_id(org, fallback_user_id=getattr(request.user, "id", None))
                target_user_ids = [admin_user_id] if admin_user_id else []

            crm_reassign_result = None
            if target_user_ids:
                try:
                    with transaction.atomic():
                        crm_reassign_result = _crm_snapshot_and_reassign_user_records(
                            org,
                            membership=membership,
                            target_user_ids=target_user_ids,
                            performed_by=request.user,
                        )
                except Exception:
                    logger.exception("Failed to reassign CRM records for deleted membership_id=%s", membership_id)
                    crm_reassign_result = None

            if permanent:
                _revoke_business_autopilot_access(membership.user)
                membership.delete()
                message = "User permanently deleted."
            else:
                _set_membership_status(membership, OrganizationUser.STATUS_DELETED, changed_by=request.user)
                _revoke_business_autopilot_access(membership.user)
                message = "User moved to deleted items."

            _sync_org_users_to_plan_limit(org, requested_by=request.user)
            payload = _build_org_users_response_payload(org, can_manage_users, message=message)
            payload["affected_leads"] = [_serialize_crm_lead(row) for row in linked_leads]
            payload["affected_deals"] = [_serialize_crm_deal(row) for row in linked_deals]
            payload["crm_reassignment"] = crm_reassign_result.get("summary") if crm_reassign_result else {}

            reassignment_summary = payload.get("crm_reassignment") or {}
            target_names = []
            if crm_reassign_result and isinstance(crm_reassign_result, dict):
                target_names = [name for name in (crm_reassign_result.get("target_names") or []) if str(name or "").strip()]
            if not target_names and target_user_ids:
                target_display = _crm_build_user_display_map(target_user_ids)
                target_names = [target_display.get(user_id, "") for user_id in target_user_ids if target_display.get(user_id, "")]
            target_label = ", ".join([name for name in target_names if name]) or "Org Admin"
            total_linked = sum(int(reassignment_summary.get(key) or 0) for key in ("leads", "deals", "sales_orders", "meetings"))
            payload["linked_records_message"] = (
                f"{total_linked} CRM record(s) were reassigned to {target_label}."
                if total_linked
                else ""
            )
            return JsonResponse(payload)
        except Exception as exc:
            logger.exception("Business Autopilot user delete failed membership_id=%s", membership_id)
            message = "Unable to delete this user right now. Please try again."
            debug_error = str(exc) if getattr(settings, "DEBUG", False) else ""
            return JsonResponse(
                {
                    "detail": "delete_failed",
                    "message": message,
                    "debug_error": debug_error,
                },
                status=500,
            )

    if resolved_method == "RESTORE":
        if not membership.is_deleted:
            return JsonResponse({"detail": "user_not_deleted"}, status=400)
        _sync_org_users_to_plan_limit(org, requested_by=request.user)
        preview_meta = _build_org_user_meta(org, users=None)
        if not preview_meta.get("can_add_users"):
            return JsonResponse(
                {
                    "detail": "employee_limit_reached",
                    "message": preview_meta.get("limit_message") or "User limit reached. Add-on users required to restore this user.",
                    "meta": preview_meta,
                },
                status=403,
            )
        should_restore_crm = False
        if isinstance(payload, dict):
            should_restore_crm = str(payload.get("restore_crm_assignments") or "").strip().lower() in {"1", "true", "yes"}

        _set_membership_status(membership, OrganizationUser.STATUS_ACTIVE, changed_by=request.user)
        _sync_org_users_to_plan_limit(org, requested_by=request.user)
        membership.refresh_from_db()
        if membership.is_active:
            _grant_business_autopilot_access(membership.user, request.user, membership.role or "org_user")
        restore_payload = _build_org_users_response_payload(org, can_manage_users, message="User restored successfully.")
        if should_restore_crm:
            try:
                restore_result = _crm_restore_user_records_from_snapshot(org, membership=membership, performed_by=request.user)
            except DatabaseError:
                logger.exception("Failed to restore CRM records from snapshot for membership_id=%s", membership_id)
                restore_result = None
            restore_payload["crm_restore"] = restore_result.get("summary") if restore_result else {}
        return JsonResponse(restore_payload)

    if request.method != "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
    if membership.is_deleted:
        return JsonResponse({"detail": "user_not_found"}, status=404)

    role = (payload.get("role") or membership.role or "org_user").strip().lower()
    if role not in ERP_EMPLOYEE_ROLES:
        role = "org_user"

    first_name, last_name = _extract_person_name(payload)
    name = " ".join([first_name, last_name]).strip()
    email = str(payload.get("email") or membership.user.email or "").strip().lower()
    password = str(payload.get("password") or "")
    phone_number = str(payload.get("phone_number") or "").strip()
    employee_role = (payload.get("employee_role") or "").strip()
    employee_role_id = payload.get("employee_role_id")
    user_type, _ = _resolve_org_user_type_from_payload(org, {"user_type": payload.get("user_type") or getattr(membership, "user_type", "")})
    department = (payload.get("department") or membership.department or "").strip()
    department_id = payload.get("department_id")
    is_active = payload.get("is_active")

    if not email:
        return JsonResponse({"detail": "email_required"}, status=400)
    if password and len(password) < 6:
        return JsonResponse({"detail": "password_too_short"}, status=400)

    existing_user = User.objects.filter(email__iexact=email).exclude(id=membership.user_id).first()
    if existing_user:
        return JsonResponse({"detail": "email_already_exists"}, status=409)

    if department_id not in (None, ""):
        try:
            department_id = int(department_id)
        except (TypeError, ValueError):
            return JsonResponse({"detail": "invalid_department"}, status=400)
        selected_department = OrganizationDepartment.objects.filter(
            organization=org,
            id=department_id,
            is_active=True,
        ).first()
        if not selected_department:
            return JsonResponse({"detail": "department_not_found"}, status=404)
        department = selected_department.name

    if employee_role_id not in (None, ""):
        try:
            employee_role_id = int(employee_role_id)
        except (TypeError, ValueError):
            return JsonResponse({"detail": "invalid_employee_role"}, status=400)
        selected_role = OrganizationEmployeeRole.objects.filter(
            organization=org,
            id=employee_role_id,
            is_active=True,
        ).first()
        if not selected_role:
            return JsonResponse({"detail": "employee_role_not_found"}, status=404)
        employee_role = selected_role.name

    with transaction.atomic():
        previous_user_type = _normalize_user_type_key(getattr(membership, "user_type", ""))
        if name:
            membership.user.first_name = first_name
            membership.user.last_name = last_name
        membership.user.email = email
        membership.user.username = email
        if password:
            membership.user.set_password(password)
            membership.user.save(update_fields=["first_name", "last_name", "email", "username", "password"])
        else:
            membership.user.save(update_fields=["first_name", "last_name", "email", "username"])

        profile = UserProfile.objects.filter(user=membership.user).first()
        if profile:
            profile.organization = org
            profile.role = role
            profile.phone_number = phone_number
            profile.save(update_fields=["organization", "role", "phone_number"])
        else:
            UserProfile.objects.create(
                user=membership.user,
                organization=org,
                role=role,
                phone_number=phone_number,
            )

        membership.role = role
        membership.department = department
        membership.employee_role = employee_role
        membership.user_type = user_type
        if isinstance(is_active, bool) and is_active and not membership.is_active:
            _sync_org_users_to_plan_limit(org, requested_by=request.user)
            preview_meta = _build_org_user_meta(org, users=None)
            preview_lock_state = _compute_org_user_lock_ids(
                preview_meta.get("employee_limit"),
                memberships=_list_org_user_memberships(org),
                org=org,
            )
            if membership.id in preview_lock_state["locked_ids"]:
                return JsonResponse(
                    {
                        "detail": "employee_limit_reached",
                        "message": preview_meta.get("limit_message") or "User limit reached. Add-on users required.",
                        "meta": preview_meta,
                    },
                    status=403,
                )
            if not _user_type_has_available_seat(org, user_type, exclude_membership_id=membership.id):
                return JsonResponse(
                    {
                        "detail": "user_type_limit_reached",
                        "message": "Selected user type limit reached. Increase seats or choose another available user type.",
                        "meta": preview_meta,
                    },
                    status=403,
                )
        elif user_type != previous_user_type and _normalize_membership_status(membership) == OrganizationUser.STATUS_ACTIVE:
            if not _user_type_has_available_seat(org, user_type, exclude_membership_id=membership.id):
                return JsonResponse(
                    {
                        "detail": "user_type_limit_reached",
                        "message": "Selected user type limit reached. Increase seats or choose another available user type.",
                    },
                    status=403,
                )

        if isinstance(is_active, bool):
            if not is_active and _is_org_admin_account_member(org, membership):
                return JsonResponse(
                    {
                        "detail": "org_admin_deactivate_forbidden",
                        "message": "ORG admin account cannot be deactivated.",
                    },
                    status=403,
                )
            _set_membership_status(
                membership,
                OrganizationUser.STATUS_ACTIVE if is_active else OrganizationUser.STATUS_INACTIVE,
                changed_by=request.user,
                commit=False,
            )
        elif _is_org_admin_account_member(org, membership):
            _set_membership_status(membership, OrganizationUser.STATUS_ACTIVE, changed_by=request.user, commit=False)
        membership.save(
            update_fields=[
                "role",
                "department",
                "employee_role",
                "user_type",
                "status",
                "status_changed_at",
                "is_active",
                "is_deleted",
                "deleted_at",
                "resigned_at",
                "resigned_by",
                "updated_at",
            ]
        )
        membership.user.is_active = _normalize_membership_status(membership) == OrganizationUser.STATUS_ACTIVE
        membership.user.save(update_fields=["is_active"])
        if _normalize_membership_status(membership) == OrganizationUser.STATUS_ACTIVE:
            _grant_business_autopilot_access(membership.user, request.user, role)
        else:
            _revoke_business_autopilot_access(membership.user)

    _sync_org_users_to_plan_limit(org, requested_by=request.user)
    return JsonResponse(_build_org_users_response_payload(org, can_manage_users))


@require_http_methods(["POST"])
def org_user_toggle_status(request, membership_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "users": [], "meta": {}})

    can_manage_users = _can_manage_users(request.user, org)
    if not can_manage_users:
        return JsonResponse({"detail": "forbidden"}, status=403)

    membership = OrganizationUser.objects.filter(organization=org, id=membership_id, is_deleted=False).select_related("user").first()
    if not membership or not membership.user:
        return JsonResponse({"detail": "user_not_found"}, status=404)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    enabled = payload.get("enabled")
    if not isinstance(enabled, bool):
        return JsonResponse({"detail": "enabled_required"}, status=400)
    if not enabled and _is_org_admin_account_member(org, membership):
        return JsonResponse(
            {
                "detail": "org_admin_deactivate_forbidden",
                "message": "ORG admin account cannot be deactivated.",
            },
            status=403,
        )

    _sync_org_users_to_plan_limit(org, requested_by=request.user)
    if enabled and _normalize_membership_status(membership) != OrganizationUser.STATUS_ACTIVE:
        preview_meta = _build_org_user_meta(org, users=None)
        preview_lock_state = _compute_org_user_lock_ids(
            preview_meta.get("employee_limit"),
            memberships=_list_org_user_memberships(org),
            org=org,
        )
        if membership.id in preview_lock_state["locked_ids"]:
            return JsonResponse(
                {
                    "detail": "employee_limit_reached",
                    "message": preview_meta.get("limit_message") or "User limit reached. Add-on users required.",
                    "meta": preview_meta,
                },
                status=403,
            )
        if not _user_type_has_available_seat(org, _normalize_user_type_key(getattr(membership, "user_type", "")), exclude_membership_id=membership.id):
            return JsonResponse(
                {
                    "detail": "user_type_limit_reached",
                    "message": "Selected user type limit reached. Increase seats or choose another available user type.",
                    "meta": preview_meta,
                },
                status=403,
            )

    next_status = OrganizationUser.STATUS_ACTIVE if enabled else OrganizationUser.STATUS_INACTIVE
    _set_membership_status(membership, next_status, changed_by=request.user)
    if enabled:
        _grant_business_autopilot_access(membership.user, request.user, membership.role or "org_user")
    else:
        _revoke_business_autopilot_access(membership.user)

    _sync_org_users_to_plan_limit(org, requested_by=request.user)
    users = _safe_serialize_org_users(org)
    meta = _build_org_user_meta(org, users=users)
    lock_state = _compute_org_user_lock_ids(
        meta.get("employee_limit"),
        memberships=_list_org_user_memberships(org),
        org=org,
    )
    users = _attach_locked_state(users, lock_state["locked_ids"])
    message = "User activated." if enabled else "User marked inactive."

    return JsonResponse(
        {
            "authenticated": True,
            "organization_id": getattr(org, "id", None),
            "users": users,
            "employee_roles": _safe_serialize_employee_roles(org),
            "departments": _safe_serialize_departments(org),
            "can_manage_users": can_manage_users,
            "meta": meta,
            "message": message,
        }
    )


@require_http_methods(["POST"])
def org_user_mark_resigned(request, membership_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "users": [], "meta": {}})
    can_manage_users = _can_manage_users(request.user, org)
    if not can_manage_users:
        return JsonResponse({"detail": "forbidden"}, status=403)
    membership = OrganizationUser.objects.filter(organization=org, id=membership_id, is_deleted=False).select_related("user").first()
    if not membership or not membership.user:
        return JsonResponse({"detail": "user_not_found"}, status=404)
    if _is_org_admin_account_member(org, membership):
        return JsonResponse({"detail": "org_admin_resign_forbidden", "message": "ORG admin account cannot be marked as resigned."}, status=403)
    _set_membership_status(membership, OrganizationUser.STATUS_RESIGNED, changed_by=request.user)
    _revoke_business_autopilot_access(membership.user)
    _sync_org_users_to_plan_limit(org, requested_by=request.user)
    return JsonResponse(_build_org_users_response_payload(org, can_manage_users, message="User marked as resigned."))


@require_http_methods(["POST"])
def org_user_mark_inactive(request, membership_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "users": [], "meta": {}})
    can_manage_users = _can_manage_users(request.user, org)
    if not can_manage_users:
        return JsonResponse({"detail": "forbidden"}, status=403)
    membership = OrganizationUser.objects.filter(organization=org, id=membership_id, is_deleted=False).select_related("user").first()
    if not membership or not membership.user:
        return JsonResponse({"detail": "user_not_found"}, status=404)
    if _is_org_admin_account_member(org, membership):
        return JsonResponse({"detail": "org_admin_deactivate_forbidden", "message": "ORG admin account cannot be marked inactive."}, status=403)
    _set_membership_status(membership, OrganizationUser.STATUS_INACTIVE, changed_by=request.user)
    _revoke_business_autopilot_access(membership.user)
    _sync_org_users_to_plan_limit(org, requested_by=request.user)
    return JsonResponse(_build_org_users_response_payload(org, can_manage_users, message="User marked as inactive."))


@require_http_methods(["POST"])
def org_user_restore_active(request, membership_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "users": [], "meta": {}})
    can_manage_users = _can_manage_users(request.user, org)
    if not can_manage_users:
        return JsonResponse({"detail": "forbidden"}, status=403)
    membership = OrganizationUser.objects.filter(organization=org, id=membership_id, is_deleted=False).select_related("user").first()
    if not membership or not membership.user:
        return JsonResponse({"detail": "user_not_found"}, status=404)
    preview_meta = _build_org_user_meta(org, users=None)
    if not preview_meta.get("can_add_users"):
        return JsonResponse(
            {
                "detail": "employee_limit_reached",
                "message": preview_meta.get("limit_message") or "User limit reached. Please increase user limit or mark inactive/resigned users.",
                "meta": preview_meta,
            },
            status=403,
        )
    if not _user_type_has_available_seat(org, _normalize_user_type_key(getattr(membership, "user_type", "")), exclude_membership_id=membership.id):
        return JsonResponse(
            {
                "detail": "user_type_limit_reached",
                "message": "Selected user type limit reached. Increase seats or choose another available user type.",
                "meta": preview_meta,
            },
            status=403,
        )
    _set_membership_status(membership, OrganizationUser.STATUS_ACTIVE, changed_by=request.user)
    _grant_business_autopilot_access(membership.user, request.user, membership.role or "org_user")
    _sync_org_users_to_plan_limit(org, requested_by=request.user)
    return JsonResponse(_build_org_users_response_payload(org, can_manage_users, message="User restored to active status."))


@require_http_methods(["POST"])
def org_user_resend_credentials(request, membership_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "users": []})

    can_manage_users = _can_manage_users(request.user, org)
    if not can_manage_users:
        return JsonResponse({"detail": "forbidden"}, status=403)

    membership = OrganizationUser.objects.filter(organization=org, id=membership_id, is_deleted=False).select_related("user").first()
    if not membership or not membership.user:
        return JsonResponse({"detail": "user_not_found"}, status=404)

    user = membership.user
    email = str(user.email or "").strip()
    if not email:
        return JsonResponse({"detail": "email_required"}, status=400)

    temp_password = _generate_temp_login_password()
    user.set_password(temp_password)
    user.save(update_fields=["password"])

    login_url = PUBLIC_LOGIN_URL
    credentials = {
        "name": _get_org_user_display_name(user),
        "email": email,
        "password": temp_password,
        "login_url": login_url,
    }
    mail_sent = send_templated_email(
        email,
        "Your Work Zilla login credentials",
        "emails/business_autopilot_user_credentials.txt",
        {
            "name": credentials["name"] or "User",
            "email": credentials["email"],
            "password": temp_password,
            "login_url": login_url,
            "organization_name": str(org.name or "").strip(),
        },
    )

    return JsonResponse(
        {
            "authenticated": True,
            "credentials": credentials,
            "email_sent": bool(mail_sent),
            "status": "sent" if mail_sent else "failed",
        }
    )


@require_http_methods(["POST"])
def org_user_verify_email(request, membership_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "users": [], "meta": {}})

    can_manage_users = _can_manage_users(request.user, org)
    if not can_manage_users:
        return JsonResponse({"detail": "forbidden"}, status=403)

    membership = OrganizationUser.objects.filter(organization=org, id=membership_id).select_related("user").first()
    if not membership or not membership.user:
        return JsonResponse({"detail": "user_not_found"}, status=404)

    user = membership.user
    email = str(user.email or "").strip()
    if not email:
        return JsonResponse({"detail": "email_required"}, status=400)

    if not bool(getattr(user, "email_verified", False)):
        mark_email_verified(user)
        user.email_verification_sent_at = None
        user.save(update_fields=["email_verification_sent_at"])

    _sync_org_users_to_plan_limit(org, requested_by=request.user)
    users = _safe_serialize_org_users(org)
    meta = _build_org_user_meta(org, users=users)
    lock_state = _compute_org_user_lock_ids(
        meta.get("employee_limit"),
        memberships=_list_org_user_memberships(org),
        org=org,
    )
    users = _attach_locked_state(users, lock_state["locked_ids"])

    if membership.is_active and not (membership.id in lock_state["locked_ids"]):
        _grant_business_autopilot_access(user, request.user, membership.role or "org_user")

    return JsonResponse(
        {
            "authenticated": True,
            "users": users,
            "employee_roles": _safe_serialize_employee_roles(org),
            "departments": _safe_serialize_departments(org),
            "can_manage_users": can_manage_users,
            "meta": meta,
            "message": "User email verified successfully.",
        }
    )


@require_http_methods(["GET", "POST"])
def org_role_access(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False, "role_access_map": {}}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "role_access_map": {}})

    can_manage_users = _can_manage_users(request.user, org)
    settings_obj, _ = OrganizationSettings.objects.get_or_create(organization=org)

    if request.method == "POST":
        if not can_manage_users:
            return JsonResponse({"detail": "forbidden"}, status=403)
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        role_access_blob = str(payload.get("role_access_blob") or "").strip()
        if role_access_blob:
            try:
                decoded = base64.b64decode(role_access_blob.encode("ascii"), validate=True).decode("utf-8")
                role_access_map = json.loads(decoded or "{}")
            except (ValueError, binascii.Error, UnicodeDecodeError, json.JSONDecodeError):
                return JsonResponse({"detail": "invalid_role_access_blob"}, status=400)
        else:
            role_access_map = payload.get("role_access_map", payload)
        normalized_role_access_map = _normalize_role_access_map(role_access_map)
        settings_obj.business_autopilot_role_access_map = normalized_role_access_map
        settings_obj.save(update_fields=["business_autopilot_role_access_map"])
    else:
        normalized_role_access_map = _normalize_role_access_map(
            settings_obj.business_autopilot_role_access_map or {}
        )

    return JsonResponse(
        {
            "authenticated": True,
            "organization": {
                "id": org.id,
                "name": org.name,
                "company_key": org.company_key,
            },
            "can_manage_users": can_manage_users,
            "role_access_map": normalized_role_access_map,
        }
    )


@require_http_methods(["GET", "POST"])
def org_employee_roles(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False, "employee_roles": []}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "employee_roles": []})

    can_manage_users = _can_manage_users(request.user, org)

    if request.method == "POST":
        if not can_manage_users:
            return JsonResponse({"detail": "forbidden"}, status=403)
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        name = (payload.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name_required"}, status=400)
        conflicting_department_exists = (
            OrganizationDepartment.objects
            .filter(organization=org, name__iexact=name, is_active=True)
            .exists()
        )
        if conflicting_department_exists:
            return JsonResponse({"detail": "employee_role_matches_department_name"}, status=400)
        existing_role = (
            OrganizationEmployeeRole.objects
            .filter(organization=org, name__iexact=name)
            .order_by("-is_active", "id")
            .first()
        )
        if existing_role:
            update_fields = []
            if not existing_role.is_active:
                existing_role.is_active = True
                update_fields.append("is_active")
            if existing_role.name != name:
                existing_role.name = name
                update_fields.append("name")
            if update_fields:
                try:
                    existing_role.save(update_fields=[*update_fields, "updated_at"])
                except IntegrityError:
                    return JsonResponse({"detail": "employee_role_exists"}, status=400)
        else:
            try:
                OrganizationEmployeeRole.objects.create(
                    organization=org,
                    name=name,
                    is_active=True,
                )
            except IntegrityError:
                return JsonResponse({"detail": "employee_role_exists"}, status=400)

    return JsonResponse(
        {
            "authenticated": True,
            "organization": {
                "id": org.id,
                "name": org.name,
                "company_key": org.company_key,
            },
            "employee_roles": _serialize_employee_roles(org),
            "departments": _serialize_departments(org),
            "can_manage_users": can_manage_users,
        }
    )


@require_http_methods(["PUT", "DELETE", "POST"])
def org_employee_role_detail(request, role_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None}, status=404)
    if not _can_manage_users(request.user, org):
        return JsonResponse({"detail": "forbidden"}, status=403)

    role = OrganizationEmployeeRole.objects.filter(
        organization=org,
        id=role_id,
    ).first()
    if not role:
        return JsonResponse({"detail": "employee_role_not_found"}, status=404)

    assigned_memberships = list(
        OrganizationUser.objects
        .filter(organization=org, employee_role__iexact=role.name, role__in=ERP_EMPLOYEE_ROLES, is_deleted=False)
        .select_related("user")
        .order_by("id")
    )
    affected_users = _build_membership_assignment_summary(assigned_memberships)

    payload = {}
    resolved_method = request.method
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        action = str(payload.get("action") or "").strip().lower()
        if action in {"delete", "remove"}:
            resolved_method = "DELETE"
        elif action in {"update", "edit", "save"} or "name" in payload:
            resolved_method = "PUT"
        else:
            return JsonResponse({"detail": "invalid_action"}, status=400)

    if resolved_method == "PUT":
        if request.method != "POST":
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        name = (payload.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name_required"}, status=400)
        conflicting_department_exists = (
            OrganizationDepartment.objects
            .filter(organization=org, name__iexact=name, is_active=True)
            .exists()
        )
        if conflicting_department_exists:
            return JsonResponse({"detail": "employee_role_matches_department_name"}, status=400)
        duplicate = (
            OrganizationEmployeeRole.objects
            .filter(organization=org, name__iexact=name, is_active=True)
            .exclude(id=role.id)
            .exists()
        )
        if duplicate:
            return JsonResponse({"detail": "employee_role_exists"}, status=400)
        role.name = name
        role.is_active = True
        try:
            role.save(update_fields=["name", "is_active", "updated_at"])
        except IntegrityError:
            return JsonResponse({"detail": "employee_role_exists"}, status=400)
    else:
        if role.is_active:
            if assigned_memberships:
                for membership in assigned_memberships:
                    if membership.employee_role:
                        membership.employee_role = ""
                        membership.save(update_fields=["employee_role", "updated_at"])
            role.is_active = False
            role.save(update_fields=["is_active", "updated_at"])

    return JsonResponse(
        {
            "authenticated": True,
            "employee_roles": _serialize_employee_roles(org),
            "departments": _serialize_departments(org),
            "users": _safe_serialize_org_users(org),
            "can_manage_users": _can_manage_users(request.user, org),
            "affected_users": affected_users,
        }
    )


@require_http_methods(["GET", "POST"])
def org_departments(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False, "departments": []}, status=401)

    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "departments": []})

    can_manage_users = _can_manage_users(request.user, org)

    if request.method == "POST":
        if not can_manage_users:
            return JsonResponse({"detail": "forbidden"}, status=403)
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        name = (payload.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name_required"}, status=400)
        conflicting_role_exists = (
            OrganizationEmployeeRole.objects
            .filter(organization=org, name__iexact=name, is_active=True)
            .exists()
        )
        if conflicting_role_exists:
            return JsonResponse({"detail": "department_matches_employee_role_name"}, status=400)
        existing_department = (
            OrganizationDepartment.objects
            .filter(organization=org, name__iexact=name)
            .order_by("-is_active", "id")
            .first()
        )
        if existing_department:
            update_fields = []
            if not existing_department.is_active:
                existing_department.is_active = True
                update_fields.append("is_active")
            if existing_department.name != name:
                existing_department.name = name
                update_fields.append("name")
            if update_fields:
                try:
                    existing_department.save(update_fields=[*update_fields, "updated_at"])
                except IntegrityError:
                    return JsonResponse({"detail": "department_exists"}, status=400)
        else:
            try:
                OrganizationDepartment.objects.create(
                    organization=org,
                    name=name,
                    is_active=True,
                )
            except IntegrityError:
                return JsonResponse({"detail": "department_exists"}, status=400)

    return JsonResponse(
        {
            "authenticated": True,
            "organization": {
                "id": org.id,
                "name": org.name,
                "company_key": org.company_key,
            },
            "departments": _serialize_departments(org),
            "can_manage_users": can_manage_users,
        }
    )


@require_http_methods(["PUT", "DELETE", "POST"])
def org_department_detail(request, department_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None}, status=404)
    if not _can_manage_users(request.user, org):
        return JsonResponse({"detail": "forbidden"}, status=403)

    department = OrganizationDepartment.objects.filter(
        organization=org,
        id=department_id,
    ).first()
    if not department:
        return JsonResponse({"detail": "department_not_found"}, status=404)

    assigned_memberships = list(
        OrganizationUser.objects
        .filter(organization=org, department__iexact=department.name, role__in=ERP_EMPLOYEE_ROLES, is_deleted=False)
        .select_related("user")
        .order_by("id")
    )
    affected_users = _build_membership_assignment_summary(assigned_memberships)

    payload = {}
    resolved_method = request.method
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        action = str(payload.get("action") or "").strip().lower()
        if action in {"delete", "remove"}:
            resolved_method = "DELETE"
        elif action in {"update", "edit", "save"} or "name" in payload:
            resolved_method = "PUT"
        else:
            return JsonResponse({"detail": "invalid_action"}, status=400)

    if resolved_method == "PUT":
        if request.method != "POST":
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        name = (payload.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name_required"}, status=400)
        previous_name = str(department.name or "").strip()
        conflicting_role_exists = (
            OrganizationEmployeeRole.objects
            .filter(organization=org, name__iexact=name, is_active=True)
            .exists()
        )
        if conflicting_role_exists:
            return JsonResponse({"detail": "department_matches_employee_role_name"}, status=400)
        duplicate = (
            OrganizationDepartment.objects
            .filter(organization=org, name__iexact=name, is_active=True)
            .exclude(id=department.id)
            .exists()
        )
        if duplicate:
            return JsonResponse({"detail": "department_exists"}, status=400)
        department.name = name
        department.is_active = True
        try:
            department.save(update_fields=["name", "is_active", "updated_at"])
        except IntegrityError:
            return JsonResponse({"detail": "department_exists"}, status=400)
        if previous_name and previous_name.lower() != name.lower():
            OrganizationUser.objects.filter(
                organization=org,
                department__iexact=previous_name,
                role__in=ERP_EMPLOYEE_ROLES,
            ).update(department=name, updated_at=timezone.now())
    else:
        if department.is_active:
            if assigned_memberships:
                for membership in assigned_memberships:
                    if membership.department:
                        membership.department = ""
                        membership.save(update_fields=["department", "updated_at"])
            department.is_active = False
            department.save(update_fields=["is_active", "updated_at"])

    return JsonResponse(
        {
            "authenticated": True,
            "departments": _serialize_departments(org),
            "employee_roles": _serialize_employee_roles(org),
            "users": _safe_serialize_org_users(org),
            "can_manage_users": _can_manage_users(request.user, org),
            "affected_users": affected_users,
        }
    )


@require_http_methods(["GET", "PUT"])
def payroll_workspace(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse(
            {
                "authenticated": True,
                "organization": None,
                "organization_profile": {
                    "organizationName": "",
                    "country": "India",
                    "currency": "INR",
                    "timezone": "UTC",
                },
                "payroll_settings": _default_payroll_settings_payload(),
                "salary_structures": [],
                "salary_history": [],
                "payroll_entries": [],
                "payslips": [],
                "employee_directory": [],
                "permissions": {
                    "can_manage_payroll": False,
                    "can_view_all_payroll": False,
                    "can_view_salary_history": False,
                },
            }
        )

    can_manage_payroll = _can_manage_payroll(request.user, org)
    can_view_salary_history = _can_view_salary_history(request.user, org)
    settings_obj, _ = OrganizationSettings.objects.get_or_create(organization=org)
    payroll_settings = _get_or_create_payroll_settings(org)
    standard_template = _ensure_standard_salary_template(org)

    if request.method == "PUT":
        if not can_manage_payroll:
            return JsonResponse({"detail": "forbidden"}, status=403)
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)

        org_profile = payload.get("organization_profile") if isinstance(payload, dict) else {}
        payroll_payload = payload.get("payroll_settings") if isinstance(payload, dict) else {}
        structure_payloads = payload.get("salary_structures") if isinstance(payload, dict) else []
        history_payloads = payload.get("salary_history") if isinstance(payload, dict) else []
        entry_payloads = payload.get("payroll_entries") if isinstance(payload, dict) else []
        payslip_payloads = payload.get("payslips") if isinstance(payload, dict) else []

        with transaction.atomic():
            org.name = str((org_profile or {}).get("organizationName") or org.name).strip() or org.name
            org.country = str((org_profile or {}).get("country") or org.country or "India").strip() or "India"
            org.currency = str((org_profile or {}).get("currency") or org.currency or "INR").strip().upper()[:10] or "INR"
            org.save(update_fields=["name", "country", "currency"])

            settings_obj.org_timezone = str((org_profile or {}).get("timezone") or settings_obj.org_timezone or "UTC").strip() or "UTC"
            settings_obj.save(update_fields=["org_timezone"])

            payroll_settings.enable_pf = _to_bool((payroll_payload or {}).get("enablePf"), True)
            payroll_settings.enable_esi = _to_bool((payroll_payload or {}).get("enableEsi"), True)
            payroll_settings.pf_employee_percent = _to_decimal((payroll_payload or {}).get("pfEmployeePercent") or "12")
            payroll_settings.pf_employer_percent = _to_decimal((payroll_payload or {}).get("pfEmployerPercent") or "12")
            payroll_settings.esi_employee_percent = _to_decimal((payroll_payload or {}).get("esiEmployeePercent") or "0.75")
            payroll_settings.esi_employer_percent = _to_decimal((payroll_payload or {}).get("esiEmployerPercent") or "3.25")
            payroll_settings.updated_by = request.user
            payroll_settings.save()

            existing_structures = {row.id: row for row in SalaryStructure.objects.filter(organization=org)}
            keep_structure_ids = set()
            structure_id_map = {}
            for item in structure_payloads if isinstance(structure_payloads, list) else []:
                if not isinstance(item, dict):
                    continue
                structure_id = _coerce_int(item.get("id"))
                row = existing_structures.get(structure_id) if structure_id else None
                if not row:
                    row = SalaryStructure(organization=org)
                row.name = str(item.get("name") or "").strip() or f"Structure {timezone.now().strftime('%H%M%S')}"
                row.is_default = _to_bool(item.get("isDefault"), False)
                row.basic_salary_percent = _to_decimal(item.get("basicSalaryPercent") or "40")
                row.hra_percent = _to_decimal(item.get("hraPercent") or "20")
                row.conveyance_fixed = _to_decimal(item.get("conveyanceFixed") or item.get("conveyance") or "1600")
                row.auto_special_allowance = _to_bool(item.get("autoSpecialAllowance"), True)
                row.basic_salary = _to_decimal(item.get("basicSalary"))
                row.hra = _to_decimal(item.get("hra"))
                row.conveyance = _to_decimal(item.get("conveyance"))
                row.special_allowance = _to_decimal(item.get("specialAllowance"))
                row.bonus = _to_decimal(item.get("bonus"))
                row.other_allowances = _to_decimal(item.get("otherAllowances"))
                row.apply_pf = _to_bool(item.get("applyPf"), True)
                row.apply_esi = _to_bool(item.get("applyEsi"), True)
                row.professional_tax = _to_decimal(item.get("professionalTax"))
                row.other_deduction = _to_decimal(item.get("otherDeduction"))
                row.notes = str(item.get("notes") or "").strip()
                row.save()
                keep_structure_ids.add(row.id)
                structure_id_map[str(item.get("id") or "")] = row.id
            SalaryStructure.objects.filter(organization=org).exclude(id__in=keep_structure_ids).delete()

            keep_history_ids = set()
            history_id_map = {}
            if can_view_salary_history:
                existing_history = {row.id: row for row in EmployeeSalaryHistory.objects.filter(organization=org)}
                for item in history_payloads if isinstance(history_payloads, list) else []:
                    if not isinstance(item, dict):
                        continue
                    history_id = _coerce_int(item.get("id"))
                    row = existing_history.get(history_id) if history_id else None
                    if not row:
                        row = EmployeeSalaryHistory(organization=org)
                    resolved_structure_id = _coerce_int(item.get("salaryStructureId"))
                    resolved_structure_id = structure_id_map.get(str(item.get("salaryStructureId")), resolved_structure_id)
                    row.employee_name = str(item.get("employeeName") or "").strip()
                    row.source_user_id = _coerce_int(item.get("sourceUserId"))
                    row.salary_structure_id = resolved_structure_id if resolved_structure_id in keep_structure_ids else resolved_structure_id
                    row.current_salary = _to_decimal(item.get("currentSalary") or item.get("monthlySalaryAmount"))
                    row.monthly_salary_amount = _to_decimal(item.get("monthlySalaryAmount"))
                    row.increment_type = str(item.get("incrementType") or "percentage").strip().lower() or "percentage"
                    if row.increment_type not in {"percentage", "fixed"}:
                        row.increment_type = "percentage"
                    row.increment_value = _to_decimal(item.get("incrementValue"))
                    effective_from = str(item.get("effectiveFrom") or "").strip()
                    row.effective_from = effective_from or timezone.localdate().isoformat()
                    row.increment_amount = _to_decimal(item.get("incrementAmount"))
                    row.new_salary = _to_decimal(item.get("newSalary") or item.get("monthlySalaryAmount"))
                    row.notes = str(item.get("notes") or "").strip()
                    row.save()
                    keep_history_ids.add(row.id)
                    history_id_map[str(item.get("id") or "")] = row.id
                EmployeeSalaryHistory.objects.filter(organization=org).exclude(id__in=keep_history_ids).delete()
            default_structure_id = next((row_id for row_id in keep_structure_ids if existing_structures.get(row_id) and existing_structures[row_id].is_default), None)
            explicit_default_ids = []
            for row in SalaryStructure.objects.filter(organization=org, id__in=keep_structure_ids):
                if row.is_default:
                    explicit_default_ids.append(row.id)
            if explicit_default_ids:
                SalaryStructure.objects.filter(organization=org).exclude(id=explicit_default_ids[0]).update(is_default=False)
            else:
                standard_template = _ensure_standard_salary_template(org)
                keep_structure_ids.add(standard_template.id)

            existing_entries = {row.id: row for row in PayrollEntry.objects.filter(organization=org)}
            keep_entry_ids = set()
            entry_id_map = {}
            for item in entry_payloads if isinstance(entry_payloads, list) else []:
                if not isinstance(item, dict):
                    continue
                entry_id = _coerce_int(item.get("id"))
                row = existing_entries.get(entry_id) if entry_id else None
                if not row:
                    row = PayrollEntry(organization=org)
                resolved_structure_id = _coerce_int(item.get("salaryStructureId"))
                resolved_structure_id = structure_id_map.get(str(item.get("salaryStructureId")), resolved_structure_id)
                resolved_history_id = _coerce_int(item.get("salaryHistoryId"))
                resolved_history_id = history_id_map.get(str(item.get("salaryHistoryId")), resolved_history_id)
                row.employee_name = str(item.get("employeeName") or "").strip()
                row.source_user_id = _coerce_int(item.get("sourceUserId"))
                row.payroll_month = _normalize_payroll_month(item.get("month"))
                row.currency = str(item.get("currency") or org.currency or "INR").strip().upper()[:10] or "INR"
                row.salary_structure_id = resolved_structure_id
                row.salary_history_id = resolved_history_id
                row.gross_salary = _to_decimal(item.get("grossSalary"))
                row.pf_employee_amount = _to_decimal(item.get("pfEmployeeAmount"))
                row.pf_employer_amount = _to_decimal(item.get("pfEmployerAmount"))
                row.esi_employee_amount = _to_decimal(item.get("esiEmployeeAmount"))
                row.esi_employer_amount = _to_decimal(item.get("esiEmployerAmount"))
                row.professional_tax_amount = _to_decimal(item.get("professionalTaxAmount"))
                row.other_deduction_amount = _to_decimal(item.get("otherDeductionAmount"))
                row.total_deductions = _to_decimal(item.get("totalDeductions"))
                row.net_salary = _to_decimal(item.get("netSalary"))
                row.earnings = item.get("earnings") if isinstance(item.get("earnings"), dict) else {}
                row.deductions = item.get("deductions") if isinstance(item.get("deductions"), dict) else {}
                row.status = str(item.get("status") or "processed").strip().lower() or "processed"
                row.save()
                keep_entry_ids.add(row.id)
                entry_id_map[str(item.get("id") or "")] = row.id
            PayrollEntry.objects.filter(organization=org).exclude(id__in=keep_entry_ids).delete()

            existing_payslips = {row.id: row for row in Payslip.objects.filter(organization=org)}
            keep_payslip_ids = set()
            for item in payslip_payloads if isinstance(payslip_payloads, list) else []:
                if not isinstance(item, dict):
                    continue
                payslip_id = _coerce_int(item.get("id"))
                row = existing_payslips.get(payslip_id) if payslip_id else None
                if not row:
                    row = Payslip(organization=org)
                resolved_entry_id = _coerce_int(item.get("payrollEntryId"))
                resolved_entry_id = entry_id_map.get(str(item.get("payrollEntryId")), resolved_entry_id)
                if not resolved_entry_id:
                    continue
                row.payroll_entry_id = resolved_entry_id
                row.slip_number = str(item.get("slipNumber") or f"SLIP-{resolved_entry_id}").strip()
                row.generated_for_month = _normalize_payroll_month(item.get("generatedForMonth") or item.get("month"))
                row.employee_name = str(item.get("employeeName") or "").strip()
                row.source_user_id = _coerce_int(item.get("sourceUserId"))
                row.currency = str(item.get("currency") or org.currency or "INR").strip().upper()[:10] or "INR"
                row.save()
                keep_payslip_ids.add(row.id)
            Payslip.objects.filter(organization=org).exclude(id__in=keep_payslip_ids).delete()

    salary_structures = list(
        SalaryStructure.objects.filter(organization=org).order_by("name", "id")
    )
    salary_history = list(
        EmployeeSalaryHistory.objects.filter(organization=org).select_related("salary_structure")
    ) if can_view_salary_history else []
    payroll_entries_qs = PayrollEntry.objects.filter(organization=org).select_related("salary_structure", "salary_history", "payslip")
    payslips_qs = Payslip.objects.filter(organization=org).select_related("payroll_entry")

    if not can_manage_payroll:
        payroll_entries_qs = payroll_entries_qs.filter(source_user_id=request.user.id)
        payslips_qs = payslips_qs.filter(source_user_id=request.user.id)
        salary_history = [row for row in salary_history if row.source_user_id == request.user.id]

    payroll_entries = list(payroll_entries_qs.order_by("-payroll_month", "employee_name", "-id"))
    payslips = list(payslips_qs.order_by("-generated_at", "-id"))

    return JsonResponse(
        {
            "authenticated": True,
            "organization": {
                "id": org.id,
                "name": org.name,
                "company_key": org.company_key,
            },
            "organization_profile": _serialize_org_payroll_profile(org),
            "payroll_settings": _serialize_payroll_settings(payroll_settings),
            "salary_structures": [_serialize_salary_structure(row) for row in salary_structures],
            "salary_history": [_serialize_salary_history(row) for row in salary_history],
            "payroll_entries": [_serialize_payroll_entry(row) for row in payroll_entries],
            "payslips": [_serialize_payslip(row) for row in payslips],
            "employee_directory": _serialize_org_users(org),
            "permissions": {
                "can_manage_payroll": can_manage_payroll,
                "can_view_all_payroll": can_manage_payroll,
                "can_view_salary_history": can_view_salary_history,
            },
        }
    )


@require_http_methods(["GET"])
def employees_search(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    if not _can_view_salary_history(request.user, org):
        return JsonResponse({"detail": "forbidden"}, status=403)

    query = str(request.GET.get("q") or "").strip().lower()
    if not query:
        return JsonResponse([], safe=False)

    memberships = (
        OrganizationUser.objects
        .filter(organization=org, is_active=True, user__is_active=True, role__in=ERP_EMPLOYEE_ROLES, is_deleted=False)
        .select_related("user")
    )
    matches = []
    for member in memberships:
        display_name = _get_org_user_display_name(member.user)
        haystack = " ".join(
            [
                display_name,
                str(member.user.username or ""),
                str(member.user.email or ""),
                str(member.department or ""),
                str(member.employee_role or ""),
            ]
        ).strip().lower()
        if query not in haystack:
            continue
        matches.append(
            {
                "id": member.user_id,
                "name": display_name,
                "employee_id": _format_employee_code(member.user_id),
            }
        )

    matches.sort(key=lambda row: (str(row.get("name") or "").lower(), row.get("id") or 0))
    return JsonResponse(matches[:20], safe=False)


@require_http_methods(["GET", "POST", "PATCH"])
def attendance_geo_settings(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    if not _get_active_erp_subscription(org):
        return JsonResponse({"detail": "active_subscription_required"}, status=403)

    setting, _ = AttendanceGeoSetting.objects.get_or_create(
        organization=org,
        defaults={"radius_meters": 100, "require_gps": True},
    )
    if request.method == "GET":
        return JsonResponse({"setting": _serialize_attendance_geo_setting(setting)})
    if not _can_manage_attendance_geo_settings(request.user, org):
        return JsonResponse({"detail": "forbidden"}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    location_name = str(payload.get("location_name", setting.location_name) or "").strip()
    enabled = bool(payload.get("enabled", setting.enabled))
    allow_outside_fence = bool(payload.get("allow_outside_fence", setting.allow_outside_fence))
    require_gps = bool(payload.get("require_gps", setting.require_gps))
    try:
        latitude = _attendance_float(payload, "latitude") if "latitude" in payload else (float(setting.latitude) if setting.latitude is not None else None)
        longitude = _attendance_float(payload, "longitude") if "longitude" in payload else (float(setting.longitude) if setting.longitude is not None else None)
        radius_meters = int(payload.get("radius_meters", setting.radius_meters or 0) or 0)
    except (TypeError, ValueError):
        return JsonResponse({"detail": "invalid_geo_settings"}, status=400)

    if latitude is not None and (latitude < -90 or latitude > 90):
        return JsonResponse({"detail": "Latitude must be -90 to 90"}, status=400)
    if longitude is not None and (longitude < -180 or longitude > 180):
        return JsonResponse({"detail": "Longitude must be -180 to 180"}, status=400)
    if radius_meters and radius_meters < MIN_GEO_RADIUS_METERS:
        return JsonResponse({"detail": "Radius minimum is 20 meters"}, status=400)
    if enabled and (latitude is None or longitude is None or radius_meters < MIN_GEO_RADIUS_METERS):
        return JsonResponse({"detail": "Geo attendance requires office location and valid radius"}, status=400)

    setting.location_name = location_name
    setting.enabled = enabled
    setting.latitude = latitude
    setting.longitude = longitude
    setting.radius_meters = radius_meters or setting.radius_meters or 100
    setting.allow_outside_fence = allow_outside_fence
    setting.require_gps = require_gps
    setting.save()
    return JsonResponse({"setting": _serialize_attendance_geo_setting(setting)})


@require_http_methods(["GET"])
def attendance_my_records(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    membership = _get_org_membership(request.user, org)
    if not membership:
        return JsonResponse({"detail": "employee_not_found"}, status=403)

    rows = list(
        AttendanceEntry.objects.filter(organization=org, employee_membership=membership)
        .select_related("employee_membership")
        .order_by("-attendance_date", "-updated_at", "-id")[:31]
    )
    return JsonResponse(
        {
            "records": [_serialize_attendance_entry(row) for row in rows],
            "employee_name": _get_org_user_display_name(request.user),
        }
    )


def _attendance_geo_punch(request, *, action):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    if not _get_active_erp_subscription(org):
        return JsonResponse({"detail": "active_subscription_required"}, status=403)
    membership = _get_org_membership(request.user, org)
    if not membership or _normalize_membership_status(membership) != OrganizationUser.STATUS_ACTIVE:
        return JsonResponse({"detail": "employee_not_found"}, status=403)

    setting = AttendanceGeoSetting.objects.filter(organization=org).first()
    if not setting or not setting.enabled:
        return JsonResponse({"detail": "geo_attendance_not_enabled"}, status=400)
    if setting.latitude is None or setting.longitude is None:
        return JsonResponse({"detail": "office_location_not_configured"}, status=400)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    try:
        latitude, longitude, accuracy, outside_reason = _attendance_validate_geo_payload(payload)
    except ValueError as exc:
        detail = str(exc)
        message = {
            "gps_coordinates_required": "Location permission denied. Please enable GPS permission.",
            "gps_accuracy_required": "GPS accuracy is too low. Please try again.",
            "invalid_accuracy": "GPS accuracy is too low. Please try again.",
            "invalid_latitude": "Latitude must be -90 to 90",
            "invalid_longitude": "Longitude must be -180 to 180",
            "outside_reason_required": "Outside reason is required.",
        }.get(detail, "Invalid geo payload")
        return JsonResponse({"detail": detail, "message": message}, status=400)

    if accuracy > MAX_GPS_ACCURACY_METERS:
        return JsonResponse({"detail": "low_gps_accuracy", "message": "GPS accuracy is too low. Please try again."}, status=400)

    distance_meters = _attendance_haversine_distance_meters(latitude, longitude, float(setting.latitude), float(setting.longitude))
    inside_geofence = distance_meters <= float(setting.radius_meters or 0)
    if action == "checkin" and not inside_geofence and not setting.allow_outside_fence:
        return JsonResponse({"detail": "outside_office_radius", "message": "You are outside the allowed office radius.", "distance_meters": distance_meters, "inside_geofence": False}, status=400)
    if action == "checkin" and not inside_geofence and setting.allow_outside_fence and not outside_reason:
        return JsonResponse({"detail": "outside_reason_required", "message": "Outside reason is required.", "distance_meters": distance_meters, "inside_geofence": False}, status=400)

    now = timezone.now()
    today = timezone.localdate()
    employee_name = _get_org_user_display_name(request.user)
    geo_status = AttendanceEntry.GEO_STATUS_INSIDE if inside_geofence else AttendanceEntry.GEO_STATUS_OUTSIDE
    device_info = str(request.META.get("HTTP_USER_AGENT") or "").strip()[:1000]

    with transaction.atomic():
        entry, _ = AttendanceEntry.objects.select_for_update().get_or_create(
            organization=org,
            employee_membership=membership,
            attendance_date=today,
            defaults={"employee_name": employee_name, "geo_status": geo_status, "device_info": device_info},
        )
        entry.employee_name = employee_name
        entry.geo_status = geo_status
        entry.device_info = device_info
        if action == "checkin":
            if entry.checkin_time:
                return JsonResponse({"detail": "duplicate_checkin", "message": "Check-in already exists for today."}, status=400)
            entry.checkin_time = now
            entry.checkin_latitude = latitude
            entry.checkin_longitude = longitude
            entry.checkin_accuracy = accuracy
            entry.checkin_distance_meters = distance_meters
            entry.checkin_inside_geofence = inside_geofence
            entry.outside_reason = outside_reason
        else:
            if not entry.checkin_time:
                return JsonResponse({"detail": "missing_checkin", "message": "Prevent checkout without check-in"}, status=400)
            if entry.checkout_time:
                return JsonResponse({"detail": "duplicate_checkout", "message": "Check-out already exists for today."}, status=400)
            entry.checkout_time = now
            entry.checkout_latitude = latitude
            entry.checkout_longitude = longitude
            entry.checkout_accuracy = accuracy
            entry.checkout_distance_meters = distance_meters
            entry.checkout_inside_geofence = inside_geofence
            if outside_reason:
                entry.outside_reason = outside_reason
        entry.save()

    if action == "checkin":
        message = "Check-in saved successfully. You are inside office location." if inside_geofence else "Check-in saved successfully."
    else:
        message = "Check-out saved successfully."
    response = _attendance_geo_response(entry, action=action, message=message)
    response["distance_meters"] = distance_meters
    response["inside_geofence"] = inside_geofence
    response["gps_accuracy_warning"] = accuracy > 100
    return JsonResponse(response)


@require_http_methods(["POST"])
def attendance_geo_checkin(request):
    return _attendance_geo_punch(request, action="checkin")


@require_http_methods(["POST"])
def attendance_geo_checkout(request):
    return _attendance_geo_punch(request, action="checkout")


def _get_face_setting(org: Organization) -> FaceRecognitionSetting:
    setting, _ = FaceRecognitionSetting.objects.get_or_create(
        organization=org,
        defaults={
            "enabled": False,
            "require_internal_face": False,
            "require_external_face": False,
            "min_match_score": Decimal("0.90"),
            "photo_retention_days": 60,
            "allow_external_photo_proof": True,
        },
    )
    return setting


@require_http_methods(["GET"])
def hrm_face_settings(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    setting = _get_face_setting(org)
    return JsonResponse({"setting": _serialize_face_recognition_setting(setting)})


@require_http_methods(["POST"])
def hrm_face_settings_update(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    if not _can_manage_attendance_geo_settings(request.user, org):
        return JsonResponse({"detail": "forbidden"}, status=403)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    setting = _get_face_setting(org)
    setting.enabled = bool(payload.get("enabled", setting.enabled))
    setting.require_internal_face = bool(payload.get("require_internal_face", setting.require_internal_face))
    setting.require_external_face = bool(payload.get("require_external_face", setting.require_external_face))
    setting.allow_external_photo_proof = bool(payload.get("allow_external_photo_proof", setting.allow_external_photo_proof))
    try:
        setting.min_match_score = Decimal(str(payload.get("min_match_score", setting.min_match_score)))
        setting.photo_retention_days = int(payload.get("photo_retention_days", setting.photo_retention_days))
    except (InvalidOperation, TypeError, ValueError):
        return JsonResponse({"detail": "invalid_face_settings"}, status=400)
    if setting.min_match_score < Decimal("0.50") or setting.min_match_score > Decimal("0.99"):
        return JsonResponse({"detail": "min_match_score_range"}, status=400)
    if setting.photo_retention_days < 1 or setting.photo_retention_days > 365:
        return JsonResponse({"detail": "photo_retention_days_range"}, status=400)
    setting.save()
    return JsonResponse({"setting": _serialize_face_recognition_setting(setting)})


@require_http_methods(["GET"])
def hrm_face_enrollment_status(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    setting = _get_face_setting(org)
    membership = _get_org_membership(request.user, org)
    if not membership:
        return JsonResponse({"detail": "employee_not_found"}, status=403)
    profile = EmployeeFaceProfile.objects.filter(organization=org, employee=membership, is_active=True).first()
    enrolled = bool(profile and membership.face_enrolled)
    return JsonResponse({
        "enabled": bool(setting.enabled),
        "employee_name": _get_org_user_display_name(request.user),
        "employee_membership_id": membership.id,
        "face_enrolled": enrolled,
        "requires_enrollment": bool(setting.enabled and not enrolled),
        "min_required_images": 3,
        "max_required_images": 5,
    })


@require_http_methods(["POST"])
def hrm_face_enrollment_capture(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    setting = _get_face_setting(org)
    membership = _get_org_membership(request.user, org)
    if not membership:
        return JsonResponse({"detail": "employee_not_found"}, status=403)
    if not setting.enabled:
        return JsonResponse({"detail": "face_recognition_not_enabled"}, status=400)
    images = request.FILES.getlist("images")
    if len(images) < 3 or len(images) > 5:
        return JsonResponse({"detail": "enrollment_images_3_to_5_required"}, status=400)

    embeddings = []
    try:
        for image in images:
            embeddings.append(generate_embedding(image))
    except FaceRecognitionUnavailable as exc:
        return JsonResponse({"detail": "face_library_unavailable", "message": str(exc)}, status=503)
    except FaceRecognitionValidationError as exc:
        return JsonResponse({"detail": "face_validation_failed", "message": str(exc)}, status=400)

    profile, _ = EmployeeFaceProfile.objects.update_or_create(
        organization=org,
        employee=membership,
        defaults={
            "face_embedding": encrypt_embeddings(embeddings),
            "embedding_model_name": "face_recognition/128d",
            "enrolled_at": timezone.now(),
            "is_active": True,
        },
    )
    membership.face_enrolled = True
    membership.save(update_fields=["face_enrolled", "updated_at"])
    return JsonResponse({
        "ok": True,
        "face_enrolled": True,
        "enrolled_at": profile.enrolled_at.isoformat() if profile.enrolled_at else None,
    })


def _face_upload_file(request):
    uploaded = request.FILES.get("image") or request.FILES.get("photo") or request.FILES.get("selfie")
    if not uploaded:
        raise FaceRecognitionValidationError("Face photo is required.")
    if uploaded.size > 8 * 1024 * 1024:
        raise FaceRecognitionValidationError("Photo size must be 8 MB or less.")
    content_type = str(getattr(uploaded, "content_type", "") or "").lower()
    if content_type and not content_type.startswith("image/"):
        raise FaceRecognitionValidationError("Only image uploads are allowed.")
    return uploaded


def _parse_attendance_mode(setting: AttendanceGeoSetting, latitude, longitude, accuracy):
    if setting and setting.enabled and setting.latitude is not None and setting.longitude is not None and latitude is not None and longitude is not None:
        distance_meters = _attendance_haversine_distance_meters(latitude, longitude, float(setting.latitude), float(setting.longitude))
        inside_geofence = distance_meters <= float(setting.radius_meters or 0)
        return (
            AttendanceEntry.MODE_INTERNAL if inside_geofence else AttendanceEntry.MODE_EXTERNAL,
            distance_meters,
            inside_geofence,
            AttendanceEntry.GEO_STATUS_INSIDE if inside_geofence else AttendanceEntry.GEO_STATUS_OUTSIDE,
        )
    return AttendanceEntry.MODE_SELF, None, None, AttendanceEntry.GEO_STATUS_MANUAL


@require_http_methods(["POST"])
def hrm_attendance_checkin_face(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    membership = _get_org_membership(request.user, org)
    if not membership or _normalize_membership_status(membership) != OrganizationUser.STATUS_ACTIVE:
        return JsonResponse({"detail": "employee_not_found"}, status=403)

    face_setting = _get_face_setting(org)
    if not face_setting.enabled:
        return JsonResponse({"detail": "face_recognition_not_enabled"}, status=400)
    face_profile = EmployeeFaceProfile.objects.filter(organization=org, employee=membership, is_active=True).first()
    if not face_profile or not membership.face_enrolled:
        return JsonResponse({"detail": "face_enrollment_required"}, status=400)

    try:
        uploaded = _face_upload_file(request)
        latitude = _attendance_float(request.POST, "latitude")
        longitude = _attendance_float(request.POST, "longitude")
        accuracy = _attendance_float(request.POST, "accuracy")
        outside_reason = str(request.POST.get("outside_reason") or "").strip()
    except FaceRecognitionValidationError as exc:
        return JsonResponse({"detail": "invalid_face_image", "message": str(exc)}, status=400)
    except ValueError as exc:
        return JsonResponse({"detail": str(exc)}, status=400)

    geo_setting = AttendanceGeoSetting.objects.filter(organization=org).first()
    attendance_mode, distance_meters, inside_geofence, geo_status = _parse_attendance_mode(geo_setting, latitude, longitude, accuracy)
    if attendance_mode == AttendanceEntry.MODE_EXTERNAL and geo_setting and not geo_setting.allow_outside_fence:
        return JsonResponse({"detail": "outside_office_radius", "message": "You are outside the allowed office radius."}, status=400)
    if attendance_mode == AttendanceEntry.MODE_EXTERNAL and not face_setting.allow_external_photo_proof:
        return JsonResponse({"detail": "external_photo_proof_disabled"}, status=400)
    if attendance_mode == AttendanceEntry.MODE_EXTERNAL and not outside_reason:
        return JsonResponse({"detail": "outside_reason_required", "message": "Outside reason is required."}, status=400)

    require_face = (
        attendance_mode == AttendanceEntry.MODE_INTERNAL and face_setting.require_internal_face
    ) or (
        attendance_mode == AttendanceEntry.MODE_EXTERNAL and face_setting.require_external_face
    ) or attendance_mode == AttendanceEntry.MODE_SELF
    if not require_face:
        return JsonResponse({"detail": "face_verification_not_required"}, status=400)

    try:
        verification = verify_employee_face(face_profile, uploaded, min_score=float(face_setting.min_match_score))
    except FaceRecognitionUnavailable as exc:
        return JsonResponse({"detail": "face_library_unavailable", "message": str(exc)}, status=503)
    except FaceRecognitionValidationError as exc:
        return JsonResponse({"detail": "face_verification_failed", "message": str(exc)}, status=400)
    if not verification.matched:
        return JsonResponse({"detail": "face_mismatch", "score": verification.score, "threshold": verification.threshold}, status=400)

    compressed = compress_uploaded_photo(uploaded)
    now = timezone.now()
    today = timezone.localdate()
    employee_name = _get_org_user_display_name(request.user)
    device_info = str(request.META.get("HTTP_USER_AGENT") or "").strip()[:1000]
    retention_days = int(face_setting.photo_retention_days or 60)

    with transaction.atomic():
        entry, _ = AttendanceEntry.objects.select_for_update().get_or_create(
            organization=org,
            employee_membership=membership,
            attendance_date=today,
            defaults={"employee_name": employee_name},
        )
        if entry.checkin_time:
            return JsonResponse({"detail": "duplicate_checkin", "message": "Check-in already exists for today."}, status=400)
        entry.employee_name = employee_name
        entry.checkin_time = now
        entry.checkin_latitude = latitude
        entry.checkin_longitude = longitude
        entry.checkin_accuracy = accuracy
        entry.checkin_distance_meters = distance_meters
        entry.checkin_inside_geofence = inside_geofence
        entry.geo_status = geo_status
        entry.attendance_mode = attendance_mode
        entry.face_verified = True
        entry.face_match_score = verification.score
        entry.device_info = device_info
        entry.outside_reason = outside_reason
        if attendance_mode == AttendanceEntry.MODE_EXTERNAL:
            entry.external_verification_status = AttendanceEntry.EXTERNAL_STATUS_PENDING
            entry.verified_by = None
            entry.verified_at = None
        entry.save()

        proof = AttendancePhotoProof(
            organization=org,
            employee=membership,
            attendance=entry,
            attendance_mode=attendance_mode,
            face_verified=True,
            face_match_score=verification.score,
            gps_latitude=latitude,
            gps_longitude=longitude,
            gps_accuracy=accuracy,
            location_status="Inside Fence" if attendance_mode == AttendanceEntry.MODE_INTERNAL else ("Outside Fence" if attendance_mode == AttendanceEntry.MODE_EXTERNAL else "Manual"),
            external_verification_status=AttendancePhotoProof.VERIFICATION_PENDING if attendance_mode == AttendanceEntry.MODE_EXTERNAL else AttendancePhotoProof.VERIFICATION_APPROVED,
            expires_at=now + timedelta(days=retention_days),
        )
        proof.image.save(
            f"attendance-proof-{membership.id}-{today.isoformat()}.jpg",
            ContentFile(compressed.read()),
            save=False,
        )
        proof.save()
        entry._prefetched_latest_proof = proof

    response = _attendance_geo_response(
        entry,
        action="checkin_face",
        message="Face verified and attendance check-in saved successfully.",
    )
    response["proof"] = _serialize_attendance_photo_proof(proof, request.user, org)
    return JsonResponse(response)


def _update_external_proof_status(request, *, next_status: str):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    if not _can_manage_attendance_geo_settings(request.user, org):
        return JsonResponse({"detail": "forbidden"}, status=403)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)
    proof_id = payload.get("proof_id")
    attendance_id = payload.get("attendance_id")
    proof = AttendancePhotoProof.objects.filter(organization=org).select_related("attendance", "verified_by").order_by("-created_at", "-id")
    if proof_id:
        proof = proof.filter(id=proof_id).first()
    elif attendance_id:
        proof = proof.filter(attendance_id=attendance_id).first()
    else:
        proof = None
    if not proof:
        return JsonResponse({"detail": "proof_not_found"}, status=404)

    notes = str(payload.get("admin_notes") or "").strip()
    proof.external_verification_status = next_status
    proof.verified_by = request.user
    proof.verified_at = timezone.now()
    proof.admin_notes = notes
    proof.save(update_fields=["external_verification_status", "verified_by", "verified_at", "admin_notes"])

    attendance = proof.attendance
    attendance.external_verification_status = next_status
    attendance.verified_by = request.user
    attendance.verified_at = proof.verified_at
    attendance.admin_notes = notes
    attendance.save(update_fields=["external_verification_status", "verified_by", "verified_at", "admin_notes", "updated_at"])
    attendance._prefetched_latest_proof = proof

    return JsonResponse({
        "ok": True,
        "attendance": _serialize_attendance_entry(attendance),
        "proof": _serialize_attendance_photo_proof(proof, request.user, org),
    })


@require_http_methods(["POST"])
def hrm_attendance_external_proof_verify(request):
    return _update_external_proof_status(request, next_status=AttendancePhotoProof.VERIFICATION_APPROVED)


@require_http_methods(["POST"])
def hrm_attendance_external_proof_reject(request):
    return _update_external_proof_status(request, next_status=AttendancePhotoProof.VERIFICATION_REJECTED)


@require_http_methods(["GET"])
def employee_salary_history(request, employee_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    if not _can_view_salary_history(request.user, org):
        return JsonResponse({"detail": "forbidden"}, status=403)

    membership = (
        OrganizationUser.objects
        .filter(organization=org, user_id=employee_id)
        .select_related("user")
        .first()
    )
    history_rows = list(
        EmployeeSalaryHistory.objects
        .filter(organization=org, source_user_id=employee_id)
        .select_related("salary_structure")
        .order_by("effective_from", "id")
    )
    if not membership and not history_rows:
        return JsonResponse({"detail": "employee_not_found"}, status=404)

    employee_name = ""
    if history_rows:
        employee_name = str(history_rows[0].employee_name or "").strip()
    if not employee_name and membership:
        employee_name = _get_org_user_display_name(membership.user)

    return JsonResponse(
        {
            "employee_id": employee_id,
            "employee_code": _format_employee_code(employee_id),
            "employee_name": employee_name,
            "history": [_serialize_salary_history_detail(row) for row in history_rows],
        }
    )


@require_http_methods(["GET"])
def payroll_payslip_pdf(request, payslip_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    payslip = (
        Payslip.objects
        .filter(organization=org, id=payslip_id)
        .select_related("payroll_entry", "payroll_entry__salary_structure")
        .first()
    )
    if not payslip:
        return JsonResponse({"detail": "payslip_not_found"}, status=404)

    can_manage_payroll = _can_manage_payroll(request.user, org)
    if not can_manage_payroll and payslip.source_user_id != request.user.id:
        return JsonResponse({"detail": "forbidden"}, status=403)

    payroll_entry = payslip.payroll_entry
    currency = payslip.currency or payroll_entry.currency or org.currency or "INR"
    earnings = payroll_entry.earnings if isinstance(payroll_entry.earnings, dict) else {}
    deductions = payroll_entry.deductions if isinstance(payroll_entry.deductions, dict) else {}

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    left = 18 * mm
    top = height - 18 * mm

    pdf.setTitle(f"Payslip {payslip.slip_number}")
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(left, top, org.name or "Organization")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(left, top - 14, f"Payslip for {payslip.generated_for_month}")
    pdf.drawRightString(width - left, top - 14, f"Slip No: {payslip.slip_number}")

    info_y = top - 38
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left, info_y, "Employee Name")
    pdf.drawString(left + 75 * mm, info_y, "Employee ID")
    pdf.drawString(left + 120 * mm, info_y, "Currency")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(left, info_y - 14, payslip.employee_name or "-")
    pdf.drawString(left + 75 * mm, info_y - 14, f"EMP-{payslip.source_user_id or payroll_entry.id}")
    pdf.drawString(left + 120 * mm, info_y - 14, currency)

    section_y = info_y - 42
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(left, section_y, "Earnings")
    pdf.drawString(left + 105 * mm, section_y, "Deductions")

    row_y = section_y - 16
    pdf.setFont("Helvetica", 10)
    for index in range(max(len(earnings), len(deductions), 1)):
        earning_items = list(earnings.items())
        deduction_items = list(deductions.items())
        if index < len(earning_items):
            label, amount = earning_items[index]
            pdf.drawString(left, row_y, str(label))
            pdf.drawRightString(left + 80 * mm, row_y, _format_currency_value(currency, amount))
        if index < len(deduction_items):
            label, amount = deduction_items[index]
            pdf.drawString(left + 105 * mm, row_y, str(label))
            pdf.drawRightString(width - left, row_y, _format_currency_value(currency, amount))
        row_y -= 14

    row_y -= 10
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(left, row_y, "Gross Salary")
    pdf.drawRightString(left + 80 * mm, row_y, _format_currency_value(currency, payroll_entry.gross_salary))
    pdf.drawString(left + 105 * mm, row_y, "Net Salary")
    pdf.drawRightString(width - left, row_y, _format_currency_value(currency, payroll_entry.net_salary))
    row_y -= 16
    pdf.setFont("Helvetica", 10)
    pdf.drawString(left, row_y, f"Total Deductions: {_format_currency_value(currency, payroll_entry.total_deductions)}")
    pdf.drawString(left + 105 * mm, row_y, f"Generated: {timezone.localtime(payslip.generated_at).strftime('%Y-%m-%d %H:%M') if payslip.generated_at else '-'}")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{payslip.slip_number}.pdf"'
    return response


@require_http_methods(["GET", "PUT", "POST"])
def accounts_workspace(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None, "data": _default_accounts_workspace()})

    workspace = _get_accounts_workspace(org)
    billing_profile = BillingProfile.objects.filter(organization=org).only("country").first()

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("__crm_action") or payload.get("action") or "").strip().upper()
        if body_action in {"PUT", "PATCH", "UPDATE"}:
            resolved_method = "PUT"
        elif override_method == "PUT":
            resolved_method = "PUT"

    if resolved_method == "PUT":
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        data = _merge_accounts_workspace(workspace.data, payload.get("data"))
        data = _seed_accounts_workspace_defaults_for_org(data, org)
        duplicate_error = _validate_accounts_workspace_customer_uniqueness(data)
        if duplicate_error:
            return JsonResponse(duplicate_error, status=409)
        workspace.data = data
        workspace.updated_by = request.user
        workspace.save(update_fields=["data", "updated_by", "updated_at"])

    return JsonResponse(
        {
            "authenticated": True,
            "organization": {
                "id": org.id,
                "name": org.name,
                "company_key": org.company_key,
            },
            "organization_profile": _serialize_org_payroll_profile(org),
            "billing_country": str(getattr(billing_profile, "country", "") or "").strip(),
            "data": _normalize_accounts_workspace(workspace.data),
            "updated_at": workspace.updated_at.isoformat() if workspace.updated_at else "",
        }
    )


def _coerce_positive_int(value):
    number = _coerce_int(value)
    return number if number and number > 0 else None


def _coerce_non_negative_int(value):
    number = _coerce_int(value)
    if number is None:
        return None
    return number if number >= 0 else None


def _coerce_subscription_alert_days(value):
    source_values = _normalize_subscription_alert_days(value)
    if source_values is None:
        return None
    return [int(item) for item in source_values]


def _serialize_accounts_customer_options(org):
    workspace = _get_accounts_workspace(org)
    rows = workspace.data.get("customers") if isinstance(workspace.data, dict) else []
    if not isinstance(rows, list):
        return []
    options = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        customer_id = str(row.get("id") or "").strip()
        if not customer_id:
            continue
        customer_name = _get_accounts_customer_display_name(row)
        if not customer_name:
            continue
        company_name = str(row.get("companyName") or row.get("name") or "").strip()
        client_name = str(row.get("clientName") or "").strip()
        phone_country_code = str(row.get("phoneCountryCode") or row.get("phone_country_code") or "+91").strip() or "+91"
        phone = str(row.get("phone") or "").strip()
        additional_phones = row.get("additionalPhones") if isinstance(row.get("additionalPhones"), list) else []
        phone_list = row.get("phoneList") if isinstance(row.get("phoneList"), list) else additional_phones
        email = str(row.get("email") or "").strip()
        additional_emails = row.get("additionalEmails") if isinstance(row.get("additionalEmails"), list) else []
        email_list = row.get("emailList") if isinstance(row.get("emailList"), list) else additional_emails
        billing_country = str(row.get("billingCountry") or row.get("country") or "").strip()
        billing_state = str(row.get("billingState") or row.get("state") or "").strip()
        billing_pincode = str(row.get("billingPincode") or row.get("pincode") or "").strip()
        shipping_country = str(row.get("shippingCountry") or row.get("country") or "").strip()
        shipping_state = str(row.get("shippingState") or row.get("state") or "").strip()
        shipping_pincode = str(row.get("shippingPincode") or row.get("pincode") or "").strip()
        options.append({
            "id": customer_id,
            "name": customer_name,
            "displayName": customer_name,
            "companyName": company_name,
            "clientName": client_name,
            "phoneCountryCode": phone_country_code,
            "phone": phone,
            "phoneList": phone_list,
            "additionalPhones": additional_phones,
            "email": email,
            "emailList": email_list,
            "additionalEmails": additional_emails,
            "billingCountry": billing_country,
            "billingState": billing_state,
            "billingPincode": billing_pincode,
            "shippingCountry": shipping_country,
            "shippingState": shipping_state,
            "shippingPincode": shipping_pincode,
        })
    return options


@require_http_methods(["GET", "POST"])
def accounts_subscription_categories(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None}, status=404)
    if request.method == "POST" and not _can_manage_modules(request.user):
        return JsonResponse({"detail": "forbidden"}, status=403)

    if request.method == "GET":
        rows = list(
            SubscriptionCategory.objects
            .filter(organization=org)
            .order_by("name", "id")
        )
        return JsonResponse({"categories": [_serialize_subscription_category(row) for row in rows]})

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    name = str(payload.get("name") or "").strip()
    if not name:
        return JsonResponse({"detail": "name_required"}, status=400)

    description = str(payload.get("description") or "").strip()
    try:
        row = SubscriptionCategory.objects.create(
            organization=org,
            name=name,
            description=description,
        )
    except IntegrityError:
        return JsonResponse({"detail": "category_exists"}, status=409)

    return JsonResponse(
        {
            "category": _serialize_subscription_category(row),
        },
        status=201,
    )


@require_http_methods(["GET", "PUT", "DELETE"])
def accounts_subscription_category_detail(request, category_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None}, status=404)
    if request.method in {"PUT", "DELETE"} and not _can_manage_modules(request.user):
        return JsonResponse({"detail": "forbidden"}, status=403)

    row = SubscriptionCategory.objects.filter(organization=org, id=category_id).first()
    if not row:
        return JsonResponse({"detail": "category_not_found"}, status=404)

    if request.method == "GET":
        return JsonResponse({"category": _serialize_subscription_category(row)})

    if request.method == "PUT":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)

        name = str(payload.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name_required"}, status=400)

        description = str(payload.get("description") or "").strip()
        row.name = name
        row.description = description
        try:
            row.save(update_fields=["name", "description"])
        except IntegrityError:
            return JsonResponse({"detail": "category_exists"}, status=409)
        return JsonResponse({"category": _serialize_subscription_category(row)})

    row.delete()
    return JsonResponse({"deleted": True})


@require_http_methods(["GET", "POST"])
def accounts_subscription_sub_categories(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None}, status=404)
    if request.method == "POST" and not _can_manage_modules(request.user):
        return JsonResponse({"detail": "forbidden"}, status=403)

    if request.method == "GET":
        rows = list(
            SubscriptionSubCategory.objects
            .filter(organization=org)
            .select_related("category")
            .order_by("category_id", "name", "id")
        )
        return JsonResponse({"subCategories": [_serialize_subscription_sub_category(row) for row in rows]})

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    category_id = _coerce_positive_int(payload.get("categoryId") or payload.get("category_id"))
    if not category_id:
        return JsonResponse({"detail": "category_required"}, status=400)
    category = SubscriptionCategory.objects.filter(organization=org, id=category_id).first()
    if not category:
        return JsonResponse({"detail": "category_not_found"}, status=404)

    name = str(payload.get("name") or "").strip()
    if not name:
        return JsonResponse({"detail": "name_required"}, status=400)

    description = str(payload.get("description") or "").strip()
    try:
        row = SubscriptionSubCategory.objects.create(
            organization=org,
            category=category,
            name=name,
            description=description,
        )
    except IntegrityError:
        return JsonResponse({"detail": "sub_category_exists"}, status=409)

    return JsonResponse(
        {
            "subCategory": _serialize_subscription_sub_category(row),
        },
        status=201,
    )


@require_http_methods(["GET", "PUT", "DELETE"])
def accounts_subscription_sub_category_detail(request, sub_category_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None}, status=404)
    if request.method in {"PUT", "DELETE"} and not _can_manage_modules(request.user):
        return JsonResponse({"detail": "forbidden"}, status=403)

    row = SubscriptionSubCategory.objects.filter(organization=org, id=sub_category_id).select_related("category").first()
    if not row:
        return JsonResponse({"detail": "sub_category_not_found"}, status=404)

    if request.method == "GET":
        return JsonResponse({"subCategory": _serialize_subscription_sub_category(row)})

    if request.method == "PUT":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)

        name = str(payload.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name_required"}, status=400)
        category_id = _coerce_positive_int(payload.get("categoryId") or payload.get("category_id"))
        if category_id:
            category = SubscriptionCategory.objects.filter(organization=org, id=category_id).first()
            if not category:
                return JsonResponse({"detail": "category_not_found"}, status=404)
            row.category = category

        row.name = name
        row.description = str(payload.get("description") or "").strip()
        try:
            row.save(update_fields=["category", "name", "description"])
        except IntegrityError:
            return JsonResponse({"detail": "sub_category_exists"}, status=409)
        return JsonResponse({"subCategory": _serialize_subscription_sub_category(row)})

    row.delete()
    return JsonResponse({"deleted": True})


@require_http_methods(["GET", "POST"])
def accounts_subscriptions(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None}, status=404)
    if request.method == "POST" and not _can_manage_modules(request.user):
        return JsonResponse({"detail": "forbidden"}, status=403)

    customer_lookup = _get_accounts_customer_lookup(org)

    if request.method == "GET":
        rows = list(
            Subscription.objects
            .filter(organization=org)
            .select_related("category", "sub_category", "created_by")
            .order_by("-created_at")
        )
        payload = {
            "subscriptions": [_serialize_subscription(row, customer_lookup=customer_lookup) for row in rows],
            "categoryOptions": [_serialize_subscription_category(row) for row in SubscriptionCategory.objects.filter(organization=org).order_by("name")],
            "subCategoryOptions": [_serialize_subscription_sub_category(row) for row in SubscriptionSubCategory.objects.filter(organization=org).select_related("category").order_by("category_id", "name")],
            "customerOptions": _serialize_accounts_customer_options(org),
        }
        return JsonResponse(payload)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    start_date = _parse_iso_date(payload.get("startDate") or payload.get("start_date"))

    end_date = _parse_iso_date(payload.get("endDate") or payload.get("end_date"))
    status = _normalize_subscription_status(payload.get("status"))
    category_id = _coerce_positive_int(payload.get("categoryId") or payload.get("category_id"))
    sub_category_id = _coerce_positive_int(payload.get("subCategoryId") or payload.get("sub_category_id"))

    category = None
    if category_id:
        category = SubscriptionCategory.objects.filter(organization=org, id=category_id).first()
        if not category:
            return JsonResponse({"detail": "category_not_found"}, status=404)

    sub_category = None
    if sub_category_id:
        sub_category = SubscriptionSubCategory.objects.filter(organization=org, id=sub_category_id).first()
        if not sub_category:
            return JsonResponse({"detail": "sub_category_not_found"}, status=404)
        if category and sub_category.category_id != category.id:
            return JsonResponse({"detail": "sub_category_category_mismatch"}, status=400)

    customer_id = _coerce_positive_int(payload.get("customerId") or payload.get("customer_id"))
    raw_plan_duration_days = payload.get("planDurationDays", payload.get("plan_duration_days", 30))
    plan_duration_days = _coerce_positive_int(raw_plan_duration_days)
    if raw_plan_duration_days is not None and plan_duration_days is None and str(raw_plan_duration_days).strip() not in ("", "None", "null"):
        return JsonResponse({"detail": "plan_duration_invalid"}, status=400)
    amount = _to_decimal(payload.get("amount"))
    currency = str(payload.get("currency") or org.currency or "INR").strip().upper() or "INR"
    payment_description = str(payload.get("paymentDescription") or payload.get("payment_description") or "").strip()
    subscription_title = str(payload.get("subscriptionTitle") or payload.get("subscription_title") or "").strip()
    if not subscription_title:
        return JsonResponse({"detail": "subscription_title_required"}, status=400)
    raw_email_alert_recipients = payload.get("emailAlertAssignTo") if "emailAlertAssignTo" in payload else payload.get("email_alert_assign_to")
    raw_whatsapp_alert_recipients = payload.get("whatsappAlertAssignTo") if "whatsappAlertAssignTo" in payload else payload.get("whatsapp_alert_assign_to")
    email_alert_recipients = _coerce_subscription_alert_assignees(raw_email_alert_recipients)
    if raw_email_alert_recipients is not None and email_alert_recipients is None:
        return JsonResponse({"detail": "email_alert_assign_to_invalid"}, status=400)
    whatsapp_alert_recipients = _coerce_subscription_alert_assignees(raw_whatsapp_alert_recipients)
    if raw_whatsapp_alert_recipients is not None and whatsapp_alert_recipients is None:
        return JsonResponse({"detail": "whatsapp_alert_assign_to_invalid"}, status=400)
    email_alert_recipients = _merge_subscription_alert_recipients(email_alert_recipients, org)
    whatsapp_alert_recipients = _merge_subscription_alert_recipients(whatsapp_alert_recipients, org)
    raw_email_alert_days = payload.get("emailAlertDays") if "emailAlertDays" in payload else payload.get("email_alert_days")
    raw_whatsapp_alert_days = payload.get("whatsappAlertDays") if "whatsappAlertDays" in payload else payload.get("whatsapp_alert_days")
    email_alert_days = _coerce_subscription_alert_days(raw_email_alert_days)
    if raw_email_alert_days is not None and raw_email_alert_days != "" and email_alert_days is None:
        return JsonResponse({"detail": "email_alert_days_invalid"}, status=400)
    whatsapp_alert_days = _coerce_subscription_alert_days(raw_whatsapp_alert_days)
    if raw_whatsapp_alert_days is not None and raw_whatsapp_alert_days != "" and whatsapp_alert_days is None:
        return JsonResponse({"detail": "whatsapp_alert_days_invalid"}, status=400)

    next_billing_date = _calculate_next_billing_date(start_date) if start_date else None

    row = Subscription.objects.create(
        organization=org,
        category=category,
        sub_category=sub_category,
        subscription_title=subscription_title,
        customer_id=customer_id,
        plan_duration_days=plan_duration_days,
        payment_description=payment_description,
        amount=amount,
        currency=currency,
        start_date=start_date,
        end_date=end_date,
        email_alert_days=email_alert_days,
        whatsapp_alert_days=whatsapp_alert_days,
        email_alert_assign_to=email_alert_recipients,
        whatsapp_alert_assign_to=whatsapp_alert_recipients,
        next_billing_date=next_billing_date,
        status=status,
        created_by=request.user,
    )
    return JsonResponse(
        {
            "subscription": _serialize_subscription(row, customer_lookup=_get_accounts_customer_lookup(org)),
        },
        status=201,
    )


@require_http_methods(["GET", "PUT", "DELETE"])
def accounts_subscription_detail(request, subscription_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"authenticated": True, "organization": None}, status=404)
    if request.method in {"PUT", "DELETE"} and not _can_manage_modules(request.user):
        return JsonResponse({"detail": "forbidden"}, status=403)

    row = (
        Subscription.objects
        .filter(organization=org, id=subscription_id)
        .select_related("category", "sub_category", "created_by")
        .first()
    )
    if not row:
        return JsonResponse({"detail": "subscription_not_found"}, status=404)

    if request.method == "GET":
        return JsonResponse({"subscription": _serialize_subscription(row, customer_lookup=_get_accounts_customer_lookup(org))})

    if request.method == "PUT":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)

        start_date = _parse_iso_date(payload.get("startDate") or payload.get("start_date"))
        if start_date is not None:
            row.start_date = start_date
        end_date = _parse_iso_date(payload.get("endDate") or payload.get("end_date"))

        row.subscription_title = str(payload.get("subscriptionTitle") or payload.get("subscription_title") or row.subscription_title).strip() or row.subscription_title
        row.payment_description = str(payload.get("paymentDescription") or payload.get("payment_description") or row.payment_description).strip()
        amount = payload.get("amount")
        if amount is not None:
            row.amount = _to_decimal(amount)
        if "planDurationDays" in payload or "plan_duration_days" in payload:
            raw_plan_duration_days = payload.get("planDurationDays", payload.get("plan_duration_days"))
            plan_duration_days = _coerce_positive_int(raw_plan_duration_days)
            if plan_duration_days is None and str(raw_plan_duration_days).strip() not in ("", "None", "null"):
                return JsonResponse({"detail": "plan_duration_invalid"}, status=400)
            row.plan_duration_days = plan_duration_days
        row.currency = str(payload.get("currency") or row.currency).strip().upper() or "INR"
        row.status = _normalize_subscription_status(payload.get("status") or row.status)
        row.customer_id = _coerce_positive_int(payload.get("customerId") or payload.get("customer_id")) or row.customer_id
        if "emailAlertDays" in payload or "email_alert_days" in payload:
            raw_email_alert_days = payload.get("emailAlertDays") if "emailAlertDays" in payload else payload.get("email_alert_days")
            email_alert_days = _coerce_subscription_alert_days(raw_email_alert_days)
            if raw_email_alert_days is not None and raw_email_alert_days != "" and email_alert_days is None:
                return JsonResponse({"detail": "email_alert_days_invalid"}, status=400)
            row.email_alert_days = email_alert_days
        if "whatsappAlertDays" in payload or "whatsapp_alert_days" in payload:
            raw_whatsapp_alert_days = payload.get("whatsappAlertDays") if "whatsappAlertDays" in payload else payload.get("whatsapp_alert_days")
            whatsapp_alert_days = _coerce_subscription_alert_days(raw_whatsapp_alert_days)
            if raw_whatsapp_alert_days is not None and raw_whatsapp_alert_days != "" and whatsapp_alert_days is None:
                return JsonResponse({"detail": "whatsapp_alert_days_invalid"}, status=400)
            row.whatsapp_alert_days = whatsapp_alert_days
        if "emailAlertAssignTo" in payload or "email_alert_assign_to" in payload:
            raw_email_alert_recipients = payload.get("emailAlertAssignTo") if "emailAlertAssignTo" in payload else payload.get("email_alert_assign_to")
            email_alert_recipients = _coerce_subscription_alert_assignees(raw_email_alert_recipients)
            if raw_email_alert_recipients is not None and email_alert_recipients is None:
                return JsonResponse({"detail": "email_alert_assign_to_invalid"}, status=400)
            row.email_alert_assign_to = _merge_subscription_alert_recipients(email_alert_recipients, org)
        if "whatsappAlertAssignTo" in payload or "whatsapp_alert_assign_to" in payload:
            raw_whatsapp_alert_recipients = payload.get("whatsappAlertAssignTo") if "whatsappAlertAssignTo" in payload else payload.get("whatsapp_alert_assign_to")
            whatsapp_alert_recipients = _coerce_subscription_alert_assignees(raw_whatsapp_alert_recipients)
            if raw_whatsapp_alert_recipients is not None and whatsapp_alert_recipients is None:
                return JsonResponse({"detail": "whatsapp_alert_assign_to_invalid"}, status=400)
            row.whatsapp_alert_assign_to = _merge_subscription_alert_recipients(whatsapp_alert_recipients, org)

        category_id_raw = "categoryId" in payload and payload.get("categoryId")
        category_id = _coerce_positive_int(payload.get("categoryId") if category_id_raw else payload.get("category_id"))
        if category_id_raw:
            if not category_id:
                return JsonResponse({"detail": "category_invalid"}, status=400)
            category = SubscriptionCategory.objects.filter(organization=org, id=category_id).first()
            if not category:
                return JsonResponse({"detail": "category_not_found"}, status=404)
            row.category = category

        sub_category_id_raw = "subCategoryId" in payload and payload.get("subCategoryId")
        sub_category_id = _coerce_positive_int(payload.get("subCategoryId") if sub_category_id_raw else payload.get("sub_category_id"))
        if sub_category_id_raw:
            if not sub_category_id:
                return JsonResponse({"detail": "sub_category_invalid"}, status=400)
            sub_category = SubscriptionSubCategory.objects.filter(organization=org, id=sub_category_id).first()
            if not sub_category:
                return JsonResponse({"detail": "sub_category_not_found"}, status=404)
            if row.category_id and sub_category.category_id != row.category_id:
                return JsonResponse({"detail": "sub_category_category_mismatch"}, status=400)
            row.sub_category = sub_category

        row.next_billing_date = _calculate_next_billing_date(row.start_date) if row.start_date else None
        if end_date is not None:
            row.end_date = end_date

        row.save(update_fields=[
            "category",
            "sub_category",
            "subscription_title",
            "email_alert_assign_to",
            "whatsapp_alert_assign_to",
            "plan_duration_days",
            "email_alert_days",
            "whatsapp_alert_days",
            "customer_id",
            "payment_description",
            "amount",
            "currency",
            "start_date",
            "end_date",
            "next_billing_date",
            "status",
        ])
        return JsonResponse({"subscription": _serialize_subscription(row, customer_lookup=_get_accounts_customer_lookup(org))})

    row.delete()
    return JsonResponse({"deleted": True})


@require_http_methods(["GET", "POST"])
def org_openai_settings(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    settings_obj, _ = OrganizationSettings.objects.get_or_create(organization=org)
    scope = _business_autopilot_ai_scope(request.user, org)
    if request.method == "POST":
        if not _can_manage_openai(request.user, org):
            return JsonResponse({"detail": "forbidden"}, status=403)
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        account_email = str(payload.get("account_email") or "").strip()
        model = str(payload.get("model") or "gpt-4o-mini").strip() or "gpt-4o-mini"
        api_key = str(payload.get("api_key") or "").strip()
        enabled = bool(payload.get("enabled"))
        agent_name = str(payload.get("agent_name") or DEFAULT_BA_OPENAI_AGENT_NAME).strip() or DEFAULT_BA_OPENAI_AGENT_NAME
        voice_gender = str(payload.get("voice_gender") or "female").strip().lower()
        if voice_gender not in {"male", "female"}:
            voice_gender = "female"
        wake_word_enabled = bool(payload.get("wake_word_enabled"))
        wake_phrase = str(payload.get("wake_phrase") or "").strip()
        update_fields = []
        if "api_key" in payload and api_key:
            settings_obj.business_autopilot_openai_api_key = api_key[:200]
            update_fields.append("business_autopilot_openai_api_key")
        settings_obj.business_autopilot_ai_agent_name = agent_name[:120]
        settings_obj.business_autopilot_openai_account_email = account_email[:60]
        settings_obj.business_autopilot_openai_model = model[:120]
        settings_obj.business_autopilot_openai_enabled = enabled
        settings_obj.business_autopilot_ai_voice_gender = voice_gender
        settings_obj.business_autopilot_ai_wake_word_enabled = wake_word_enabled
        settings_obj.business_autopilot_ai_wake_phrase = wake_phrase[:120]
        update_fields.extend([
            "business_autopilot_ai_agent_name",
            "business_autopilot_openai_account_email",
            "business_autopilot_openai_model",
            "business_autopilot_openai_enabled",
            "business_autopilot_ai_voice_gender",
            "business_autopilot_ai_wake_word_enabled",
            "business_autopilot_ai_wake_phrase",
        ])
        settings_obj.save(update_fields=list(dict.fromkeys(update_fields)))
        return JsonResponse({
            "saved": True,
            **_serialize_openai_settings(settings_obj),
        })

    if not (_can_manage_openai(request.user, org) or scope["can_chat"]):
        return JsonResponse({"detail": "forbidden"}, status=403)
    payload = _serialize_openai_settings(settings_obj)
    if not _can_manage_openai(request.user, org):
        payload["account_email"] = ""
        payload["masked_api_key"] = ""
    payload["scope"] = {
        "user_type": scope["user_type"],
        "label": scope["label"],
        "allowed_sections": scope["allowed_sections"],
        "can_access_crm": scope["can_access_crm"],
        "can_access_hr": scope["can_access_hr"],
        "can_access_accounts": scope["can_access_accounts"],
        "can_access_billing": scope["can_access_billing"],
    }
    return JsonResponse(payload)


@require_http_methods(["POST"])
def org_openai_test(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    if not _can_manage_openai(request.user, org):
        return JsonResponse({"detail": "forbidden"}, status=403)

    settings_obj, _ = OrganizationSettings.objects.get_or_create(organization=org)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        payload = {}
    api_key = str(payload.get("api_key") or settings_obj.business_autopilot_openai_api_key or "").strip()
    model = str(payload.get("model") or settings_obj.business_autopilot_openai_model or "gpt-4o-mini").strip() or "gpt-4o-mini"
    if not api_key:
        return JsonResponse({"detail": "openai_api_key_missing"}, status=400)

    try:
        response = requests.post(
            OPENAI_CHAT_COMPLETIONS_URL,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": model,
                "messages": [
                    {"role": "system", "content": "Reply with exactly: Connection OK"},
                    {"role": "user", "content": "Test connection."},
                ],
                "max_tokens": 16,
                "temperature": 0,
            },
            timeout=30,
        )
    except requests.RequestException as exc:
        return JsonResponse({"detail": f"openai_request_failed: {exc}"}, status=502)

    data = response.json() if response.content else {}
    if response.status_code >= 400:
        return JsonResponse({"detail": data.get("error", {}).get("message") or "openai_connection_failed"}, status=400)
    return JsonResponse({"ok": True, "model": model})


@require_http_methods(["POST"])
def org_openai_transcribe(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    scope = _business_autopilot_ai_scope(request.user, org)
    if not scope["can_chat"]:
        return JsonResponse({"detail": "forbidden"}, status=403)

    settings_obj, _ = OrganizationSettings.objects.get_or_create(organization=org)
    api_key = str(settings_obj.business_autopilot_openai_api_key or "").strip()
    if not api_key:
        return JsonResponse({"detail": "openai_api_key_missing"}, status=400)

    audio_file = request.FILES.get("audio")
    if not audio_file:
        return JsonResponse({"detail": "audio_required"}, status=400)
    current_section = str(request.POST.get("current_section") or "").strip().lower()
    preferred_input_language = str(request.POST.get("preferred_input_language") or "").strip().lower()
    speech_context = str(request.POST.get("speech_context") or "").strip()
    allow_empty = str(request.POST.get("allow_empty") or "").strip().lower() in {"1", "true", "yes", "on"}

    filename = str(getattr(audio_file, "name", "") or "voice-note.webm").strip() or "voice-note.webm"
    model = "gpt-4o-mini-transcribe"
    prompt_parts = [
        "This is a Work Zilla Business Autopilot organization voice command.",
        "Common business terms include CRM, HRM, lead, leads, deal, deals, meeting, meetings, attendance, payroll, billing, invoice, follow-up, task, overtime, sales, and subscription.",
        "Preserve English business words like CRM, leads, HRM, sales, billing, and the configured assistant name exactly when spoken.",
        "For Tamil-English speech, preserve romanized words such as motham, yevalo, evlo, thevaiya, lead, leads, meeting, billing, users, and CRM naturally instead of forcing unrelated words.",
        "Do not over-correct partial Tamil colloquial speech into different Tamil words when the audio is unclear; prefer the closest heard wording.",
        "If the speech sounds South Indian and ambiguous, prefer Tamil or Tanglish over Malayalam.",
        "Never rewrite spoken Tamil words into Malayalam script.",
    ]
    if current_section:
        prompt_parts.append(f"Current module is {current_section}.")
    if preferred_input_language in {"ta", "ta-in", "tamil"}:
        prompt_parts.append("Preferred spoken language is Tamil. Transcribe the speech in Tamil or natural Tanglish only.")
    if speech_context:
        prompt_parts.append(speech_context[:500])
    prompt_text = " ".join(prompt_parts)[:900]
    try:
        response = requests.post(
            OPENAI_AUDIO_TRANSCRIPTIONS_URL,
            headers={
                "Authorization": f"Bearer {api_key}",
            },
            data={
                "model": model,
                "prompt": prompt_text,
            },
            files={
                "file": (filename, audio_file.read(), getattr(audio_file, "content_type", "application/octet-stream")),
            },
            timeout=120,
        )
    except requests.RequestException as exc:
        return JsonResponse({"detail": f"openai_request_failed: {exc}"}, status=502)

    data = response.json() if response.content else {}
    if response.status_code >= 400:
        detail = data.get("error", {}).get("message") or "openai_transcription_failed"
        if allow_empty:
            return JsonResponse({"text": "", "detail": detail})
        return JsonResponse({"detail": detail}, status=400)
    transcript_text = str((data or {}).get("text") or "").strip()
    if not transcript_text:
        if allow_empty:
            return JsonResponse({"text": ""})
        return JsonResponse({"detail": "transcript_empty"}, status=400)
    return JsonResponse({"text": transcript_text})


@require_http_methods(["POST"])
def org_openai_tts(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    scope = _business_autopilot_ai_scope(request.user, org)
    if not scope["can_chat"]:
        return JsonResponse({"detail": "forbidden"}, status=403)

    settings_obj, _ = OrganizationSettings.objects.get_or_create(organization=org)
    api_key = str(settings_obj.business_autopilot_openai_api_key or "").strip()
    if not api_key:
        return JsonResponse({"detail": "openai_api_key_missing"}, status=400)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    input_text = str(payload.get("text") or "").strip()
    if not input_text:
        return JsonResponse({"detail": "text_required"}, status=400)

    configured_voice_gender = str(getattr(settings_obj, "business_autopilot_ai_voice_gender", "") or "female").strip().lower()
    if configured_voice_gender not in {"male", "female"}:
        configured_voice_gender = "female"
    selected_voice = "alloy" if configured_voice_gender == "male" else "marin"
    instructions = (
        "Speak like a friendly office assistant in a natural, casual, human conversational way. "
        "Avoid sounding robotic or like a formal announcement. "
        "Use smooth pacing, a warm tone, and natural pauses. "
        "If the text mixes languages, code-switch naturally instead of forcing one language style. "
        + ("Use a slightly deeper voice style. " if configured_voice_gender == "male" else "Use a slightly softer voice style. ")
        + _ba_language_instruction_from_text(input_text)
    )
    try:
        response = requests.post(
            OPENAI_AUDIO_SPEECH_URL,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": "gpt-4o-mini-tts",
                "voice": selected_voice,
                "input": input_text[:4096],
                "instructions": instructions,
                "response_format": "mp3",
            },
            timeout=120,
        )
    except requests.RequestException as exc:
        return JsonResponse({"detail": f"openai_request_failed: {exc}"}, status=502)

    if response.status_code >= 400:
        try:
            data = response.json() if response.content else {}
        except ValueError:
            data = {}
        return JsonResponse({"detail": data.get("error", {}).get("message") or "openai_tts_failed"}, status=400)

    audio_response = HttpResponse(response.content, content_type="audio/mpeg")
    audio_response["Content-Disposition"] = 'inline; filename="assistant-reply.mp3"'
    return audio_response


@require_http_methods(["GET", "PUT", "DELETE"])
def org_openai_chat_history(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    scope = _business_autopilot_ai_scope(request.user, org)
    if not scope["can_chat"]:
        return JsonResponse({"detail": "forbidden"}, status=403)

    raw_date = str(request.GET.get("date") or "").strip()
    if request.method == "PUT":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        raw_date = str(payload.get("date") or raw_date).strip()
    chat_date = parse_date(raw_date) if raw_date else timezone.localdate()
    if not chat_date:
        return JsonResponse({"detail": "invalid_date"}, status=400)

    history_row = BusinessAutopilotChatHistory.objects.filter(
        organization=org,
        user=request.user,
        chat_date=chat_date,
    ).first()

    if request.method == "GET":
        return JsonResponse({
            "date": chat_date.isoformat(),
            "messages": _ba_normalize_history_messages(getattr(history_row, "messages", []), limit=250),
            "updated_at": history_row.updated_at.isoformat() if history_row and history_row.updated_at else "",
        })

    if request.method == "DELETE":
        if history_row:
            history_row.delete()
        return JsonResponse({"ok": True, "date": chat_date.isoformat()})

    messages = _ba_normalize_history_messages(payload.get("messages"), limit=250)
    history_row, _ = BusinessAutopilotChatHistory.objects.update_or_create(
        organization=org,
        user=request.user,
        chat_date=chat_date,
        defaults={"messages": messages},
    )
    return JsonResponse({
        "ok": True,
        "date": chat_date.isoformat(),
        "messages": messages,
        "updated_at": history_row.updated_at.isoformat() if history_row.updated_at else "",
    })


def _site_admin_get_state(org, user):
    state, _ = SiteAdminChatState.objects.get_or_create(
        organization=org,
        user=user,
        defaults={"collected_data": {}},
    )
    return state


def _site_admin_clear_state(state):
    state.intent = ""
    state.current_step = ""
    state.collected_data = {}
    state.awaiting_whatsapp_share = False
    state.last_quick_estimate = None
    state.save(update_fields=["intent", "current_step", "collected_data", "awaiting_whatsapp_share", "last_quick_estimate", "updated_at"])


def _site_admin_existing_customer_by_mobile(org, mobile):
    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    customers = data.get("customers") if isinstance(data.get("customers"), list) else []
    normalized_mobile = _normalize_site_admin_mobile(mobile)
    return next(
        (
            row for row in customers
            if isinstance(row, dict) and _site_admin_customer_phone_matches(row, normalized_mobile)
        ),
        None,
    )


def _site_admin_next_quick_estimate_step(collected):
    if not str(collected.get("mobile") or "").strip():
        return "mobile", "Please share the mobile number."
    if not bool(collected.get("existing_client")) and not str(collected.get("client_name") or "").strip():
        return "client_name", "Please share the client name."
    if not str(collected.get("item_text") or "").strip():
        existing_name = str(collected.get("client_name") or "").strip()
        if bool(collected.get("existing_client")) and existing_name:
            return "item_text", f"Client name: {existing_name}. Please share the estimate item details."
        return "item_text", "Please share the estimate item details."
    if not str(collected.get("amount") or "").strip():
        return "amount", "Please share the estimate amount."
    return "", ""


def _site_admin_merge_quick_estimate_fields(state, message):
    collected = dict(state.collected_data or {})
    parsed = _site_admin_parse_message_fields(message)
    current_step = str(state.current_step or "").strip()
    raw_message = " ".join(str(message or "").strip().split())

    if parsed.get("mobile"):
        collected["mobile"] = parsed["mobile"]
    elif current_step == "mobile":
        explicit_mobile = _normalize_site_admin_mobile(raw_message)
        if len(explicit_mobile) == 10:
            collected["mobile"] = explicit_mobile

    if collected.get("mobile"):
        existing_customer = _site_admin_existing_customer_by_mobile(state.organization, collected["mobile"])
        if existing_customer:
            existing_name = str(existing_customer.get("clientName") or existing_customer.get("companyName") or existing_customer.get("name") or "").strip()
            collected["existing_client"] = True
            collected["customer_id"] = str(existing_customer.get("id") or "").strip()
            if existing_name:
                collected["client_name"] = existing_name[:180]
            if not collected.get("email"):
                collected["email"] = _normalize_site_admin_email(existing_customer.get("email"))
            if not collected.get("address"):
                collected["address"] = str(existing_customer.get("billingAddress") or existing_customer.get("shippingAddress") or "").strip()[:500]
            if not collected.get("gst_number"):
                collected["gst_number"] = str(existing_customer.get("gstin") or "").strip()[:32]
        else:
            collected["existing_client"] = False

    if parsed.get("email"):
        collected["email"] = parsed["email"]
    if parsed.get("gst_number"):
        collected["gst_number"] = parsed["gst_number"]

    if parsed.get("client_name"):
        collected["client_name"] = parsed["client_name"]
    elif current_step == "client_name":
        name_candidate = _site_admin_parse_name_candidate(raw_message)
        if name_candidate:
            collected["client_name"] = name_candidate

    item_text_candidate = str(parsed.get("item_text") or "").strip()
    if current_step == "item_text" and raw_message and not item_text_candidate:
        item_text_candidate = raw_message
    client_name_ready = bool(collected.get("existing_client")) or bool(str(collected.get("client_name") or "").strip())
    if (
        item_text_candidate
        and client_name_ready
        and not _site_admin_detect_quick_estimate_intent(item_text_candidate)
    ):
        collected["item_text"] = item_text_candidate[:500]

    amount_value = parsed.get("amount")
    if amount_value is not None:
        collected["amount"] = _decimal_to_string(amount_value)
    elif current_step == "amount":
        explicit_amount = _site_admin_find_amount(raw_message)
        if explicit_amount is not None:
            collected["amount"] = _decimal_to_string(explicit_amount)

    if current_step == "item_text" and str(collected.get("item_text") or "").strip() and not str(collected.get("amount") or "").strip():
        _, amount_from_item = _site_admin_parse_item_entries(
            str(collected.get("item_text") or ""),
            fallback_total=parsed.get("amount"),
        )
        if amount_from_item is not None:
            collected["amount"] = _decimal_to_string(amount_from_item)

    if str(collected.get("item_text") or "").strip():
        _, computed_total = _site_admin_parse_item_entries(
            str(collected.get("item_text") or ""),
            fallback_total=collected.get("amount") or parsed.get("amount"),
        )
        if computed_total is not None:
            collected["amount"] = _decimal_to_string(computed_total)

    return collected


def _site_admin_create_quick_estimate(org, user, collected):
    mobile, client_name = _ensure_quick_estimate_contact_name_mobile(
        mobile=collected.get("mobile"),
        client_name=collected.get("client_name"),
        require_client_name=not bool(collected.get("existing_client")),
    )
    customer_row, customer_name = _site_admin_get_or_create_customer(
        org,
        user,
        mobile=mobile,
        client_name=client_name,
        email=str(collected.get("email") or "").strip(),
        address=str(collected.get("address") or "").strip(),
        gst_number=str(collected.get("gst_number") or "").strip(),
    )
    estimate_sequence, estimate_number = _next_quick_estimate_number(org)
    item_entries, parsed_total = _site_admin_parse_item_entries(
        collected.get("item_text"),
        fallback_total=collected.get("amount"),
    )
    amount = (parsed_total or _to_decimal(collected.get("amount"))).quantize(Decimal("0.01"))
    with transaction.atomic():
        estimate = QuickEstimate.objects.create(
            organization=org,
            customer_id=str(customer_row.get("id") or "").strip(),
            estimate_sequence=estimate_sequence,
            estimate_number=estimate_number,
            mobile=mobile,
            client_name=str(client_name or customer_name or mobile).strip()[:180],
            notes="",
            email=str(collected.get("email") or "").strip()[:254],
            address=str(collected.get("address") or "").strip()[:1000],
            gst_number=str(collected.get("gst_number") or "").strip()[:32],
            subtotal=amount,
            tax_amount=Decimal("0.00"),
            total_amount=amount,
            status=QuickEstimate.STATUS_CREATED,
            created_by=user,
        )
        if not item_entries:
            item_entries = [_site_admin_build_item_entry(collected.get("item_text"), amount)]
        for entry in item_entries:
            QuickEstimateItem.objects.create(
                quick_estimate=estimate,
                service_name=str(entry.get("service_name") or "").strip()[:180],
                description=str(entry.get("description") or collected.get("item_text") or "").strip()[:2000],
                quantity=entry.get("quantity"),
                unit=str(entry.get("unit") or "").strip()[:40],
                rate=entry.get("rate"),
                amount=(entry.get("amount") or Decimal("0.00")).quantize(Decimal("0.01")),
            )
        _upsert_quick_estimate_contact(
            org,
            user,
            contact_id=str(customer_row.get("id") or "").strip(),
            mobile=mobile,
            client_name=str(client_name or customer_name or mobile).strip()[:180],
            email=str(collected.get("email") or "").strip(),
            address=str(collected.get("address") or "").strip(),
            gst_number=str(collected.get("gst_number") or "").strip(),
        )
    return estimate


def _site_admin_quick_estimate_response(
    estimate,
    reply,
    *,
    action="quick_estimate_created",
    whatsapp_share_pending=True,
    whatsapp_url=None,
):
    return {
        "reply": reply,
        "action": action,
        "quick_estimate_id": estimate.id,
        "estimate_number": estimate.estimate_number,
        "whatsapp_share_pending": bool(whatsapp_share_pending),
        "whatsapp_url": whatsapp_url,
        "thermal_preview_html": _safe_render_quick_estimate_thermal_preview(estimate, estimate.organization),
    }


def _normalize_quick_estimate_progress_status(value: str) -> str:
    normalized = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    return QuickEstimate.PROGRESS_COMPLETED if normalized == QuickEstimate.PROGRESS_COMPLETED else QuickEstimate.PROGRESS_NON_COMPLETED


def _site_admin_update_quick_estimate_items(
    estimate,
    item_text,
    *,
    mobile="",
    client_name="",
    notes="",
    payment_status="",
    payment_mode="",
    job_status="",
    delivery_status="",
    payment_proof_image="",
    user=None,
):
    item_entries, parsed_total = _site_admin_parse_item_entries(item_text)
    if not item_entries:
        return None
    next_mobile, next_client_name = _ensure_quick_estimate_contact_name_mobile(
        mobile=mobile or estimate.mobile,
        client_name=client_name or estimate.client_name,
        require_client_name=True,
    )
    amount = (parsed_total or Decimal("0.00")).quantize(Decimal("0.01"))
    previous_snapshot = {
        "mobile": estimate.mobile,
        "client_name": estimate.client_name,
        "notes": str(getattr(estimate, "notes", "") or ""),
        "payment_status": str(getattr(estimate, "payment_status", "") or ""),
        "payment_mode": str(getattr(estimate, "payment_mode", "") or ""),
        "job_status": str(getattr(estimate, "job_status", "") or ""),
        "delivery_status": str(getattr(estimate, "delivery_status", "") or ""),
        "payment_proof_image": str(getattr(estimate, "payment_proof_image", "") or ""),
        "total_amount": _decimal_to_string(estimate.total_amount),
        "items": [_serialize_quick_estimate_item(item) for item in estimate.items.all().order_by("id")],
    }
    with transaction.atomic():
        estimate.items.all().delete()
        for entry in item_entries:
            QuickEstimateItem.objects.create(
                quick_estimate=estimate,
                service_name=str(entry.get("service_name") or "").strip()[:180],
                description=str(entry.get("description") or item_text or "").strip()[:2000],
                quantity=entry.get("quantity"),
                unit=str(entry.get("unit") or "").strip()[:40],
                rate=entry.get("rate"),
                amount=(entry.get("amount") or Decimal("0.00")).quantize(Decimal("0.01")),
            )
        if user is not None:
            customer_row = _site_admin_update_customer_for_estimate(
                estimate.organization,
                user,
                estimate,
                mobile=next_mobile,
                client_name=next_client_name,
            )
            customer_id = str(customer_row.get("id") or estimate.customer_id or "").strip()
        else:
            customer_id = str(estimate.customer_id or "").strip()
        estimate.customer_id = customer_id
        estimate.mobile = next_mobile
        estimate.client_name = next_client_name or estimate.client_name
        estimate.notes = str(notes or getattr(estimate, "notes", "") or "").strip()[:120]
        estimate.subtotal = amount
        estimate.tax_amount = Decimal("0.00")
        estimate.total_amount = amount
        estimate.status = QuickEstimate.STATUS_CREATED
        estimate.payment_status = _normalize_quick_estimate_progress_status(payment_status or getattr(estimate, "payment_status", ""))
        estimate.payment_mode = str(payment_mode or getattr(estimate, "payment_mode", "") or "").strip().lower()[:20]
        estimate.job_status = _normalize_quick_estimate_progress_status(job_status or getattr(estimate, "job_status", ""))
        estimate.delivery_status = _normalize_quick_estimate_progress_status(delivery_status or getattr(estimate, "delivery_status", ""))
        if estimate.payment_status == QuickEstimate.PROGRESS_COMPLETED and estimate.payment_mode == "online":
            estimate.payment_proof_image = str(payment_proof_image or getattr(estimate, "payment_proof_image", "") or "")
            if user is not None:
                estimate.payment_verified_by = user
        elif estimate.payment_status == QuickEstimate.PROGRESS_COMPLETED and estimate.payment_mode == "cash":
            estimate.payment_proof_image = ""
            if user is not None:
                estimate.payment_verified_by = user
        else:
            estimate.payment_proof_image = ""
            estimate.payment_mode = ""
            estimate.payment_verified_by = None
        if estimate.job_status == QuickEstimate.PROGRESS_COMPLETED and user is not None:
            estimate.job_verified_by = user
        elif estimate.job_status != QuickEstimate.PROGRESS_COMPLETED:
            estimate.job_verified_by = None
        if estimate.delivery_status == QuickEstimate.PROGRESS_COMPLETED and user is not None:
            estimate.delivery_verified_by = user
        elif estimate.delivery_status != QuickEstimate.PROGRESS_COMPLETED:
            estimate.delivery_verified_by = None
        estimate.save(update_fields=[
            "customer_id",
            "mobile",
            "client_name",
            "notes",
            "subtotal",
            "tax_amount",
            "total_amount",
            "status",
            "payment_status",
            "payment_mode",
            "job_status",
            "delivery_status",
            "payment_proof_image",
            "payment_verified_by",
            "job_verified_by",
            "delivery_verified_by",
            "updated_at",
        ])
        if user is not None:
            _upsert_quick_estimate_contact(
                estimate.organization,
                user,
                contact_id=customer_id,
                mobile=estimate.mobile,
                client_name=estimate.client_name,
                email=estimate.email,
                address=estimate.address,
                gst_number=estimate.gst_number,
            )
        _record_quick_estimate_history(
            estimate,
            action=QuickEstimateHistory.ACTION_UPDATED,
            actor=user,
            note=f"Estimate updated for {estimate.client_name}.",
            snapshot={
                "before": previous_snapshot,
                "after": {
                    "mobile": estimate.mobile,
                    "client_name": estimate.client_name,
                    "notes": estimate.notes,
                    "payment_status": estimate.payment_status,
                    "payment_mode": estimate.payment_mode,
                    "job_status": estimate.job_status,
                    "delivery_status": estimate.delivery_status,
                    "payment_proof_image": estimate.payment_proof_image,
                    "total_amount": _decimal_to_string(estimate.total_amount),
                    "items": [_serialize_quick_estimate_item_snapshot(entry) for entry in item_entries],
                },
            },
        )
    return estimate


def _extract_accounts_customer_phone_keys(row):
    if not isinstance(row, dict):
        return set()
    keys = set()
    primary_phone = _normalize_site_admin_mobile(row.get("phone"))
    if primary_phone:
        keys.add(primary_phone)
    for item in (row.get("phoneList") or []):
        if isinstance(item, dict):
            number = _normalize_site_admin_mobile(item.get("number"))
            if number:
                keys.add(number)
    for item in (row.get("additionalPhones") or []):
        if isinstance(item, dict):
            number = _normalize_site_admin_mobile(item.get("number"))
            if number:
                keys.add(number)
    return keys


def _extract_accounts_customer_email_keys(row):
    if not isinstance(row, dict):
        return set()
    keys = set()
    primary_email = _normalize_site_admin_email(row.get("email"))
    if primary_email:
        keys.add(primary_email)
    for item in (row.get("emailList") or []):
        normalized = _normalize_site_admin_email(item)
        if normalized:
            keys.add(normalized)
    for item in (row.get("additionalEmails") or []):
        normalized = _normalize_site_admin_email(item)
        if normalized:
            keys.add(normalized)
    return keys


def _validate_accounts_workspace_customer_uniqueness(data):
    rows = data.get("customers") if isinstance(data.get("customers"), list) else []
    phone_index = {}
    email_index = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        customer_id = str(row.get("id") or "").strip() or f"customer_{len(phone_index) + len(email_index)}"
        for phone in _extract_accounts_customer_phone_keys(row):
            if phone in phone_index and phone_index[phone] != customer_id:
                return {
                    "detail": "duplicate_customer",
                    "message": "A client with the same mobile number already exists.",
                    "duplicate_fields": ["phone"],
                }
            phone_index[phone] = customer_id
        for email in _extract_accounts_customer_email_keys(row):
            if email in email_index and email_index[email] != customer_id:
                return {
                    "detail": "duplicate_customer",
                    "message": "A client with the same email ID already exists.",
                    "duplicate_fields": ["email"],
                }
            email_index[email] = customer_id
    return None


@require_http_methods(["POST"])
def site_admin_chat(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)
    message = str(payload.get("message") or "").strip()
    if not message:
        return JsonResponse({"detail": "message_required"}, status=400)

    state = _site_admin_get_state(org, request.user)
    normalized_message = " ".join(message.lower().split())

    if _site_admin_is_reset_command(normalized_message):
        _site_admin_clear_state(state)
        return JsonResponse({
            "reply": "Quick Estimate draft cleared.",
            "action": "state_cleared",
            "quick_estimate_id": None,
            "estimate_number": "",
            "whatsapp_share_pending": False,
            "whatsapp_url": None,
            "thermal_preview_html": "",
        })

    if state.awaiting_whatsapp_share and state.last_quick_estimate_id:
        estimate = QuickEstimate.objects.filter(organization=org, id=state.last_quick_estimate_id).prefetch_related("items").first()
        if estimate:
            if normalized_message in {"yes", "y", "share", "ok"}:
                estimate.status = QuickEstimate.STATUS_SHARED
                estimate.save(update_fields=["status", "updated_at"])
                state.intent = ""
                state.current_step = ""
                state.collected_data = {}
                state.awaiting_whatsapp_share = False
                state.save(update_fields=["intent", "current_step", "collected_data", "awaiting_whatsapp_share", "updated_at"])
                return JsonResponse({
                    "reply": "Opening WhatsApp Web.",
                    "action": "open_whatsapp",
                    "quick_estimate_id": estimate.id,
                    "estimate_number": estimate.estimate_number,
                    "whatsapp_share_pending": False,
                    "whatsapp_url": _build_quick_estimate_whatsapp_url(estimate),
                    "thermal_preview_html": _render_quick_estimate_thermal_preview(estimate, org),
                })
            if normalized_message in {"no", "n", "skip"}:
                state.intent = ""
                state.current_step = ""
                state.collected_data = {}
                state.awaiting_whatsapp_share = False
                state.save(update_fields=["intent", "current_step", "collected_data", "awaiting_whatsapp_share", "updated_at"])
                return JsonResponse({
                    "reply": "Okay, WhatsApp sharing was skipped.",
                    "action": "share_skipped",
                    "quick_estimate_id": estimate.id,
                    "estimate_number": estimate.estimate_number,
                    "whatsapp_share_pending": False,
                    "whatsapp_url": None,
                    "thermal_preview_html": _render_quick_estimate_thermal_preview(estimate, org),
                })

    should_start_qe = True

    if not should_start_qe:
        lowered = normalized_message
        if lowered in {"find client", "today estimates"}:
            return JsonResponse({
                "reply": "This operation is coming soon. Please use QE Create for now.",
                "action": "coming_soon",
                "quick_estimate_id": None,
                "estimate_number": "",
                "whatsapp_share_pending": False,
                "whatsapp_url": None,
                "thermal_preview_html": "",
            })
        return JsonResponse({
            "reply": f"Site Admin currently supports {_site_admin_supported_module_labels() or 'Quick Estimate'}. Click QE Create to open Quick Estimate.",
            "action": "unsupported",
            "quick_estimate_id": None,
            "estimate_number": "",
            "whatsapp_share_pending": False,
            "whatsapp_url": None,
            "thermal_preview_html": "",
        })

    if state.intent != SiteAdminChatState.INTENT_QUICK_ESTIMATE:
        state.intent = SiteAdminChatState.INTENT_QUICK_ESTIMATE
        state.collected_data = {}

    collected = _site_admin_merge_quick_estimate_fields(state, message)
    step, reply = _site_admin_next_quick_estimate_step(collected)
    if step:
        state.current_step = step
        state.collected_data = collected
        state.awaiting_whatsapp_share = False
        state.save(update_fields=["intent", "current_step", "collected_data", "awaiting_whatsapp_share", "updated_at"])
        return JsonResponse({
            "reply": reply,
            "action": "collecting_quick_estimate",
            "quick_estimate_id": None,
            "estimate_number": "",
            "whatsapp_share_pending": False,
            "whatsapp_url": None,
            "thermal_preview_html": "",
        })

    estimate = _site_admin_create_quick_estimate(org, request.user, collected)
    state.intent = ""
    state.current_step = ""
    state.collected_data = {}
    state.awaiting_whatsapp_share = False
    state.last_quick_estimate = estimate
    state.save(update_fields=["intent", "current_step", "collected_data", "awaiting_whatsapp_share", "last_quick_estimate", "updated_at"])
    reply = f"Quick Estimate {estimate.estimate_number} created for {estimate.client_name} - ₹{_format_quick_estimate_amount(estimate.total_amount)}."
    return JsonResponse(_site_admin_quick_estimate_response(estimate, reply, whatsapp_share_pending=False))


@require_http_methods(["GET", "POST", "PATCH", "DELETE"])
def quick_estimate_contacts(request, contact_id: str = ""):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("__action") or payload.get("action") or "").strip().upper()
        if body_action in {"PATCH", "DELETE"}:
            resolved_method = body_action
        elif override_method in {"PATCH", "DELETE"}:
            resolved_method = override_method

    _sync_quick_estimate_contacts_from_estimates(org, request.user)
    workspace, data, contacts, rows = _quick_estimate_contact_rows_with_counts(org)
    if resolved_method == "GET":
        if contact_id:
            payload = next((row for row in rows if row["id"] == str(contact_id or "").strip()), None)
            if not payload:
                return JsonResponse({"detail": "quick_estimate_contact_not_found"}, status=404)
            return JsonResponse({"contact": payload})
        return JsonResponse({"contacts": rows})

    if payload is None:
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)

    normalized_contact_id = str(contact_id or payload.get("contact_id") or payload.get("id") or "").strip()
    target_row = next(
        (row for row in contacts if isinstance(row, dict) and str(row.get("id") or "").strip() == normalized_contact_id),
        None,
    )
    if not target_row:
        return JsonResponse({"detail": "quick_estimate_contact_not_found"}, status=404)

    if resolved_method == "DELETE":
        contacts[:] = [row for row in contacts if not (isinstance(row, dict) and str(row.get("id") or "").strip() == normalized_contact_id)]
        data["quickEstimateContacts"] = contacts
        workspace.data = data
        workspace.updated_by = request.user
        workspace.save(update_fields=["data", "updated_by", "updated_at"])
        return JsonResponse({"message": "Contact deleted successfully."})

    try:
        normalized_mobile, normalized_name = _ensure_quick_estimate_contact_name_mobile(
            mobile=payload.get("mobile") or target_row.get("phone"),
            client_name=payload.get("client_name") or payload.get("clientName") or target_row.get("clientName"),
            require_client_name=True,
        )
    except ValueError as exc:
        return JsonResponse({"detail": "validation_error", "message": str(exc)}, status=400)

    conflicting_row = next(
        (
            row for row in contacts
            if isinstance(row, dict)
            and row is not target_row
            and _normalize_site_admin_mobile(row.get("phone")) == normalized_mobile
        ),
        None,
    )
    if conflicting_row is not None:
        return JsonResponse(
            {"detail": "duplicate_contact", "message": "A client with the same mobile number already exists."},
            status=400,
        )

    normalized_email = _normalize_site_admin_email(payload.get("email") or target_row.get("email"))
    normalized_address = str(payload.get("address") or target_row.get("address") or "").strip()[:500]
    normalized_gst = str(payload.get("gst_number") or payload.get("gstin") or target_row.get("gstin") or "").strip()[:32]
    target_row["clientName"] = normalized_name
    target_row["phone"] = normalized_mobile
    target_row["email"] = normalized_email
    target_row["address"] = normalized_address
    target_row["gstin"] = normalized_gst
    target_row["updatedAt"] = timezone.now().isoformat()
    data["quickEstimateContacts"] = contacts
    workspace.data = data
    workspace.updated_by = request.user
    workspace.save(update_fields=["data", "updated_by", "updated_at"])

    linked_estimates = QuickEstimate.objects.filter(organization=org, customer_id=normalized_contact_id)
    linked_estimates.update(
        mobile=normalized_mobile,
        client_name=normalized_name,
        email=normalized_email,
        address=normalized_address,
        gst_number=normalized_gst,
        updated_at=timezone.now(),
    )
    return JsonResponse(
        {
            "message": "Contact updated successfully.",
            "contact": _serialize_quick_estimate_contact(target_row, linked_estimate_count=linked_estimates.count()),
        }
    )


@require_http_methods(["GET", "POST", "PATCH", "DELETE"])
def quick_estimates(request, estimate_id: int = None):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("__action") or payload.get("action") or "").strip().upper()
        if body_action in {"PATCH", "DELETE"}:
            resolved_method = body_action
        elif override_method in {"PATCH", "DELETE"}:
            resolved_method = override_method

    qs = (
        QuickEstimate.objects
        .filter(organization=org)
        .select_related(
            "organization__owner",
            "created_by",
            "assigned_user",
            "assigned_by",
            "payment_verified_by",
            "job_verified_by",
            "delivery_verified_by",
        )
        .prefetch_related("items")
        .order_by("-created_at", "-id")
    )
    if resolved_method == "GET" and estimate_id is None:
        return JsonResponse({"quick_estimates": [_serialize_quick_estimate(row) for row in qs[:100]]})

    if resolved_method in {"PATCH", "DELETE"} and not estimate_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        estimate_id = _coerce_positive_int(payload.get("quick_estimate_id") or payload.get("estimate_id") or payload.get("id"))

    row = qs.filter(id=estimate_id).first()
    if not row:
        return JsonResponse({"detail": "quick_estimate_not_found"}, status=404)

    if resolved_method == "GET":
        return JsonResponse({"quick_estimate": _serialize_quick_estimate(row, include_preview=True)})

    if payload is None:
        if request.content_type and request.content_type.startswith("multipart/form-data"):
            payload = request.POST.dict()
        else:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)

    patch_action = str(payload.get("action") or payload.get("__action") or "").strip().lower()
    if resolved_method == "DELETE":
        patch_action = "cancel"

    if patch_action == "cancel":
        reason = str(payload.get("reason") or payload.get("cancel_reason") or "").strip()
        if not reason:
            return JsonResponse({"detail": "cancel_reason_required", "message": "Please enter the cancel reason."}, status=400)
        row.status = QuickEstimate.STATUS_CANCELLED
        row.save(update_fields=["status", "updated_at"])
        _record_quick_estimate_history(
            row,
            action=QuickEstimateHistory.ACTION_CANCELLED,
            actor=request.user,
            note=f"Estimate cancelled. Reason: {reason}",
            snapshot={"reason": reason},
        )
        row = (
            QuickEstimate.objects
            .filter(id=row.id)
            .select_related(
                "organization__owner",
                "created_by",
                "assigned_user",
                "assigned_by",
                "payment_verified_by",
                "job_verified_by",
                "delivery_verified_by",
            )
            .prefetch_related("items")
            .first()
        ) or row
        return JsonResponse({
            "message": f"{row.estimate_number} cancelled.",
            "quick_estimate": _serialize_quick_estimate(row, include_preview=True),
        })

    if patch_action == "reopen":
        row.status = QuickEstimate.STATUS_CREATED
        row.save(update_fields=["status", "updated_at"])
        _record_quick_estimate_history(
            row,
            action=QuickEstimateHistory.ACTION_REOPENED,
            actor=request.user,
            note="Estimate reopened.",
            snapshot={},
        )
        row = (
            QuickEstimate.objects
            .filter(id=row.id)
            .select_related(
                "organization__owner",
                "created_by",
                "assigned_user",
                "assigned_by",
                "payment_verified_by",
                "job_verified_by",
                "delivery_verified_by",
            )
            .prefetch_related("items")
            .first()
        ) or row
        return JsonResponse({
            "message": f"{row.estimate_number} reopened.",
            "quick_estimate": _serialize_quick_estimate(row, include_preview=True),
        })

    if patch_action == "assign":
        assigned_user_id = _coerce_positive_int(
            payload.get("assigned_user_id") or payload.get("assignedUserId") or payload.get("user_id")
        )
        assigned_membership_id = _coerce_positive_int(
            payload.get("membership_id") or payload.get("assigned_membership_id") or payload.get("assignedMembershipId")
        )
        if not assigned_user_id and not assigned_membership_id:
            row.assigned_user = None
            row.assigned_by = request.user
            row.save(update_fields=["assigned_user", "assigned_by", "updated_at"])
            row = (
                QuickEstimate.objects
                .filter(id=row.id)
                .select_related(
                    "organization__owner",
                    "created_by",
                    "assigned_user",
                    "assigned_by",
                    "payment_verified_by",
                    "job_verified_by",
                    "delivery_verified_by",
                )
                .prefetch_related("items")
                .first()
            ) or row
            _record_quick_estimate_history(
                row,
                action=QuickEstimateHistory.ACTION_UNASSIGNED,
                actor=request.user,
                note="Assigned user removed.",
                snapshot={},
            )
            return JsonResponse(
                {
                    "message": f"Assigned user removed from {row.estimate_number}.",
                    "quick_estimate": _serialize_quick_estimate(row, include_preview=True),
                }
            )
        membership_query = OrganizationUser.objects.filter(
            organization=org,
            role__in=ERP_EMPLOYEE_ROLES,
            is_deleted=False,
        ).select_related("user")
        membership = None
        if assigned_membership_id:
            membership = membership_query.filter(id=assigned_membership_id).first()
        if membership is None and assigned_user_id:
            membership = membership_query.filter(user_id=assigned_user_id).first()
        if not membership or _normalize_membership_status(membership) != OrganizationUser.STATUS_ACTIVE:
            return JsonResponse({"detail": "assigned_user_not_found", "message": "Selected org user is not active."}, status=404)
        created_by_was_empty = not row.created_by_id
        if created_by_was_empty:
            row.created_by = request.user
        row.assigned_user = membership.user
        row.assigned_by = request.user
        update_fields = ["assigned_user", "assigned_by", "updated_at"]
        if created_by_was_empty:
            update_fields.append("created_by")
        row.save(update_fields=update_fields)
        row = (
            QuickEstimate.objects
            .filter(id=row.id)
            .select_related(
                "organization__owner",
                "created_by",
                "assigned_user",
                "assigned_by",
                "payment_verified_by",
                "job_verified_by",
                "delivery_verified_by",
            )
            .prefetch_related("items")
            .first()
        ) or row
        _record_quick_estimate_history(
            row,
            action=QuickEstimateHistory.ACTION_ASSIGNED,
            actor=request.user,
            note=f"Assigned to {_get_org_user_display_name(membership.user)}.",
            snapshot={
                "assigned_user_id": membership.user_id,
                "assigned_user_name": _get_org_user_display_name(membership.user),
            },
        )
        return JsonResponse(
            {
                "message": f"{row.estimate_number} assigned to {_get_org_user_display_name(membership.user)}.",
                "quick_estimate": _serialize_quick_estimate(row, include_preview=True),
            }
        )

    if patch_action == "payment":
        next_payment_status = _normalize_quick_estimate_progress_status(
            payload.get("payment_status") or payload.get("paymentStatus") or row.payment_status
        )
        next_payment_mode = str(payload.get("payment_mode") or payload.get("paymentMode") or row.payment_mode or "").strip().lower()[:20]
        if next_payment_status == QuickEstimate.PROGRESS_COMPLETED and next_payment_mode not in {"cash", "online"}:
            return JsonResponse({"detail": "payment_mode_required", "message": "Please choose cash or online payment mode."}, status=400)
        next_payment_proof_image = str(
            payload.get("payment_proof_image")
            or payload.get("paymentProofImage")
            or getattr(row, "payment_proof_image", "")
            or ""
        ).strip()
        if next_payment_status == QuickEstimate.PROGRESS_COMPLETED and next_payment_mode == "online" and not next_payment_proof_image:
            return JsonResponse({"detail": "payment_proof_required", "message": "Please upload the payment proof image."}, status=400)

        previous_snapshot = {
            "payment_status": str(getattr(row, "payment_status", "") or ""),
            "payment_mode": str(getattr(row, "payment_mode", "") or ""),
            "payment_proof_image": str(getattr(row, "payment_proof_image", "") or ""),
        }
        row.payment_status = next_payment_status
        row.payment_mode = next_payment_mode if next_payment_status == QuickEstimate.PROGRESS_COMPLETED else ""
        if row.payment_status == QuickEstimate.PROGRESS_COMPLETED and row.payment_mode == "online":
            row.payment_proof_image = next_payment_proof_image
            row.payment_verified_by = request.user
        elif row.payment_status == QuickEstimate.PROGRESS_COMPLETED and row.payment_mode == "cash":
            row.payment_proof_image = ""
            row.payment_verified_by = request.user
        else:
            row.payment_proof_image = ""
            row.payment_mode = ""
            row.payment_verified_by = None
        row.save(update_fields=["payment_status", "payment_mode", "payment_proof_image", "payment_verified_by", "updated_at"])
        _record_quick_estimate_history(
            row,
            action=QuickEstimateHistory.ACTION_UPDATED,
            actor=request.user,
            note=f"Payment updated for {row.estimate_number}.",
            snapshot={
                "before": previous_snapshot,
                "after": {
                    "payment_status": row.payment_status,
                    "payment_mode": row.payment_mode,
                    "payment_proof_image": row.payment_proof_image,
                },
            },
        )
        row = (
            QuickEstimate.objects
            .filter(id=row.id)
            .select_related(
                "organization__owner",
                "created_by",
                "assigned_user",
                "assigned_by",
                "payment_verified_by",
                "job_verified_by",
                "delivery_verified_by",
            )
            .prefetch_related("items")
            .first()
        ) or row
        return JsonResponse(
            {
                "message": f"Payment updated for {row.estimate_number}.",
                "quick_estimate": _serialize_quick_estimate(row, include_preview=True),
            }
        )

    raw_item_text = str(payload.get("item_text") or payload.get("message") or "").strip()
    item_text = "\n".join(line.rstrip() for line in raw_item_text.splitlines() if line.strip())
    if not item_text:
        return JsonResponse({"detail": "item_text_required", "message": "Please share the estimate item details."}, status=400)
    next_mobile = _normalize_site_admin_mobile(payload.get("mobile") or row.mobile)
    next_client_name = str(payload.get("client_name") or payload.get("clientName") or row.client_name or "").strip()[:180]
    next_notes = str(payload.get("notes") or payload.get("note") or getattr(row, "notes", "") or "").strip()[:120]
    next_payment_status = _normalize_quick_estimate_progress_status(payload.get("payment_status") or payload.get("paymentStatus") or row.payment_status)
    next_payment_mode = str(payload.get("payment_mode") or payload.get("paymentMode") or getattr(row, "payment_mode", "") or "").strip().lower()[:20]
    next_job_status = _normalize_quick_estimate_progress_status(payload.get("job_status") or payload.get("jobStatus") or row.job_status)
    next_delivery_status = _normalize_quick_estimate_progress_status(payload.get("delivery_status") or payload.get("deliveryStatus") or row.delivery_status)
    next_payment_proof_image = str(
        payload.get("payment_proof_image")
        or payload.get("paymentProofImage")
        or getattr(row, "payment_proof_image", "")
        or ""
    ).strip()
    if len(next_mobile) != 10:
        return JsonResponse({"detail": "invalid_mobile", "message": "Please share a valid 10-digit mobile number."}, status=400)
    if not next_client_name:
        return JsonResponse({"detail": "client_name_required", "message": "Please share the client name."}, status=400)
    if next_payment_status == QuickEstimate.PROGRESS_COMPLETED and next_payment_mode not in {"cash", "online"}:
        return JsonResponse({"detail": "payment_mode_required", "message": "Please choose cash or online payment mode."}, status=400)
    if next_payment_status == QuickEstimate.PROGRESS_COMPLETED and not next_payment_proof_image:
        if next_payment_mode == "online":
            return JsonResponse({"detail": "payment_proof_required", "message": "Please upload the payment proof image."}, status=400)
        next_payment_proof_image = ""
    try:
        updated = _site_admin_update_quick_estimate_items(
            row,
            item_text,
            mobile=next_mobile,
            client_name=next_client_name,
            notes=next_notes,
            payment_status=next_payment_status,
            payment_mode=next_payment_mode,
            job_status=next_job_status,
            delivery_status=next_delivery_status,
            payment_proof_image=next_payment_proof_image,
            user=request.user,
        )
    except ValueError as exc:
        return JsonResponse({"detail": "duplicate_customer_mobile", "message": str(exc)}, status=409)
    if not updated:
        return JsonResponse({"detail": "invalid_item_text", "message": "Please share valid estimate item details."}, status=400)
    updated.refresh_from_db()
    updated = (
        QuickEstimate.objects
        .filter(organization=org, id=updated.id)
        .select_related(
            "organization__owner",
            "created_by",
            "assigned_user",
            "assigned_by",
            "payment_verified_by",
            "job_verified_by",
            "delivery_verified_by",
        )
        .prefetch_related("items")
        .first()
    ) or updated
    return JsonResponse(
        _site_admin_quick_estimate_response(
            updated,
            f"Quick Estimate {updated.estimate_number} updated.",
            action="quick_estimate_updated",
            whatsapp_share_pending=False,
        )
    )


@require_http_methods(["GET", "POST", "PATCH"])
def quick_estimate_settings(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("action") or payload.get("__action") or "").strip().upper()
        if body_action == "PATCH":
            resolved_method = "PATCH"
        elif override_method == "PATCH":
            resolved_method = "PATCH"

    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    if resolved_method == "GET":
        return JsonResponse({"settings": _get_quick_estimate_settings(org)})

    if payload is None:
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)

    header_text = _normalize_quick_estimate_header_html(payload.get("headerText") or payload.get("header_text"))
    template_size = str(payload.get("templateSize") or payload.get("template_size") or "4in").strip().lower()
    payment_proof_retention_days = str(
        payload.get("paymentProofRetentionDays")
        or payload.get("payment_proof_retention_days")
        or "45"
    ).strip()
    if template_size not in {"3in", "4in"}:
        template_size = "4in"
    if payment_proof_retention_days not in {"45", "60"}:
        payment_proof_retention_days = "45"
    if _quick_estimate_header_text_length(header_text) > QUICK_ESTIMATE_HEADER_TEXT_MAX_LENGTH:
        return JsonResponse(
            {
                "detail": "header_text_too_long",
                "message": f"Header text supports up to {QUICK_ESTIMATE_HEADER_TEXT_MAX_LENGTH} characters.",
            },
            status=400,
        )
    data["quickEstimateSettings"] = {
        "headerText": header_text,
        "templateSize": template_size,
        "paymentProofRetentionDays": payment_proof_retention_days,
    }
    workspace.data = data
    workspace.updated_by = request.user
    workspace.save(update_fields=["data", "updated_by", "updated_at"])
    return JsonResponse({
        "message": "Quick Estimate settings saved.",
        "settings": _get_quick_estimate_settings(org),
    })


@require_http_methods(["GET"])
def quick_estimate_thermal_preview(request, estimate_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    row = QuickEstimate.objects.filter(organization=org, id=estimate_id).prefetch_related("items").first()
    if not row:
        return JsonResponse({"detail": "quick_estimate_not_found"}, status=404)
    requested_format = str(request.GET.get("format") or "").strip().lower()
    if requested_format == "pdf":
        return _quick_estimate_pdf_response(row, org)
    return HttpResponse(_render_quick_estimate_thermal_preview(row, org), content_type="text/html; charset=utf-8")


@require_http_methods(["GET"])
def quick_estimate_public_preview(request, signed_token: str):
    estimate_id = _resolve_quick_estimate_public_token(signed_token)
    if estimate_id is None:
        signer = TimestampSigner(salt="business-autopilot.quick-estimate-preview")
        try:
            estimate_id = int(signer.unsign(signed_token, max_age=60 * 60 * 24 * 30))
        except (BadSignature, SignatureExpired, ValueError):
            return HttpResponse("Preview link is invalid or expired.", status=404, content_type="text/plain; charset=utf-8")

    row = (
        QuickEstimate.objects
        .filter(id=estimate_id)
        .select_related("organization")
        .prefetch_related("items")
        .first()
    )
    if not row or not getattr(row, "organization", None):
        return HttpResponse("Preview not found.", status=404, content_type="text/plain; charset=utf-8")
    return HttpResponse(_render_quick_estimate_thermal_preview(row, row.organization), content_type="text/html; charset=utf-8")


@require_http_methods(["GET"])
def quick_estimate_history(request, estimate_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    estimate = QuickEstimate.objects.filter(organization=org, id=estimate_id).first()
    if not estimate:
        return JsonResponse({"detail": "quick_estimate_not_found"}, status=404)
    rows = (
        QuickEstimateHistory.objects
        .filter(quick_estimate=estimate)
        .select_related("actor")
        .order_by("-created_at", "-id")
    )
    return JsonResponse({"history": [_serialize_quick_estimate_history(row) for row in rows[:100]]})


@require_http_methods(["POST"])
def org_openai_chat(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    scope = _business_autopilot_ai_scope(request.user, org)
    if not scope["can_chat"]:
        return JsonResponse({"detail": "forbidden"}, status=403)

    settings_obj, _ = OrganizationSettings.objects.get_or_create(organization=org)
    if not settings_obj.business_autopilot_openai_enabled and not str(settings_obj.business_autopilot_openai_api_key or "").strip():
        return JsonResponse({"detail": "assistant_not_enabled"}, status=400)
    api_key = str(settings_obj.business_autopilot_openai_api_key or "").strip()
    if not api_key:
        return JsonResponse({"detail": "openai_api_key_missing"}, status=400)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    message = str(payload.get("message") or "").strip()
    if not message:
        return JsonResponse({"detail": "message_required"}, status=400)
    recent_messages = _ba_normalize_recent_messages(payload.get("recent_messages"))
    current_section = str(payload.get("current_section") or "").strip().lower()
    local_context = payload.get("local_context") if isinstance(payload.get("local_context"), dict) else {}

    agent_name = str(settings_obj.business_autopilot_ai_agent_name or DEFAULT_BA_OPENAI_AGENT_NAME).strip() or DEFAULT_BA_OPENAI_AGENT_NAME
    model = str(settings_obj.business_autopilot_openai_model or "gpt-4o-mini").strip() or "gpt-4o-mini"
    context = _build_ba_assistant_scope_context(request.user, org, scope)
    context = _ba_merge_local_assistant_context(context, local_context)
    direct_answer = _ba_assistant_direct_answer(message, {
        **context,
        "scope": {
            "can_access_crm": scope["can_access_crm"],
            "can_access_hr": scope["can_access_hr"],
            "can_access_accounts": scope["can_access_accounts"],
            "can_access_billing": scope["can_access_billing"],
            "can_access_ticketing": "ticketing" in (scope.get("allowed_sections") or []),
        },
    })
    normalized_direct_answer = _ba_normalize_direct_response_payload(direct_answer)
    if normalized_direct_answer:
        return JsonResponse({
            "reply": normalized_direct_answer["reply"],
            "tts_text": normalized_direct_answer["tts_text"],
            "agent_name": agent_name,
            "model": "tool-direct",
            "scope": {
                "user_type": scope["user_type"],
                "label": scope["label"],
                "allowed_sections": scope["allowed_sections"],
                "can_access_crm": scope["can_access_crm"],
                "can_access_hr": scope["can_access_hr"],
                "can_access_accounts": scope["can_access_accounts"],
                "can_access_billing": scope["can_access_billing"],
            },
        })
    context_json = json.dumps(context, ensure_ascii=True)[:24000]
    system_prompt = _build_ba_assistant_system_prompt(
        agent_name,
        scope,
        {
            "organization_name": context.get("org_summary", {}).get("organization_name"),
            "enabled_modules": context.get("product_capabilities", {}).get("enabled_modules") or [],
        },
    )
    conversation_text = "\n".join(
        f"{row.get('role')}: {row.get('text')}"
        for row in recent_messages
    ).strip()
    recent_conversation_block = f"Recent conversation:\n{conversation_text}\n\n" if conversation_text else ""

    try:
        response = requests.post(
            OPENAI_CHAT_COMPLETIONS_URL,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": model,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {
                        "role": "user",
                        "content": (
                            f"Current module/page: {current_section or 'unknown'}\n\n"
                            "Organization context JSON:\n"
                            f"{context_json}\n\n"
                            f"{recent_conversation_block}"
                            f"User question: {message}"
                        ),
                    },
                ],
                "temperature": 0.2,
                "max_tokens": 500,
            },
            timeout=60,
        )
    except requests.RequestException as exc:
        return JsonResponse({"detail": f"openai_request_failed: {exc}"}, status=502)

    data = response.json() if response.content else {}
    if response.status_code >= 400:
        return JsonResponse({"detail": data.get("error", {}).get("message") or "openai_chat_failed"}, status=400)
    answer = (
        (((data.get("choices") or [{}])[0]).get("message") or {}).get("content")
        if isinstance(data, dict)
        else ""
    )
    answer_text = str(answer or "").strip() or "No response received."
    return JsonResponse({
        "reply": answer_text,
        "tts_text": answer_text,
        "agent_name": agent_name,
        "model": model,
        "scope": {
            "user_type": scope["user_type"],
            "label": scope["label"],
            "allowed_sections": scope["allowed_sections"],
            "can_access_crm": scope["can_access_crm"],
            "can_access_hr": scope["can_access_hr"],
            "can_access_accounts": scope["can_access_accounts"],
            "can_access_billing": scope["can_access_billing"],
        },
    })


@require_http_methods(["GET"])
def accounts_document_print(request, doc_type: str, doc_id: str):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    normalized_doc_type = (doc_type or "").strip().lower()
    doc_type_aliases = {
        "estimate": "estimate",
        "invoice": "invoice",
        "salesorder": "sales_order",
        "sales-order": "sales_order",
        "sales_order": "sales_order",
        "sales order": "sales_order",
    }
    normalized_doc_type = doc_type_aliases.get(normalized_doc_type, normalized_doc_type)
    if normalized_doc_type not in {"estimate", "invoice", "sales_order"}:
        return JsonResponse({"detail": "invalid_doc_type"}, status=400)

    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)
    if normalized_doc_type == "sales_order":
        sales_order = (
            CrmSalesOrder.objects
            .filter(organization=org, id=doc_id, is_deleted=False)
            .select_related("deal")
            .first()
        )
        if not sales_order:
            return JsonResponse({"detail": "document_not_found"}, status=404)
        payload = _serialize_crm_sales_order(sales_order)
        document = {
            "id": sales_order.id,
            "docNo": sales_order.order_id,
            "customerName": payload.get("company") or payload.get("customer_name") or "-",
            "customerGstin": payload.get("customer_gstin") or "",
            "billingAddress": payload.get("billing_address") or "",
            "issueDate": payload.get("issue_date") or "",
            "dueDate": payload.get("due_date") or "",
            "status": "",
            "gstTemplateId": payload.get("gst_template_id") or "",
            "billingTemplateId": payload.get("billing_template_id") or "",
            "items": payload.get("items") or [],
            "notes": payload.get("notes") or "",
            "termsText": payload.get("terms_text") or "",
            "paymentStatusNotes": payload.get("payment_status_notes") or "",
            "paymentStatus": payload.get("payment_status") or sales_order.payment_status or "pending",
            "paidAmount": payload.get("paid_amount") if payload.get("paid_amount") is not None else sales_order.paid_amount,
            "paymentMode": payload.get("payment_mode") or sales_order.payment_mode or "",
            "paymentDate": payload.get("payment_date") or (sales_order.payment_date.isoformat() if sales_order.payment_date else ""),
            "transactionId": payload.get("transaction_id") or sales_order.transaction_id or "",
            "balanceAmount": payload.get("balance_amount") if payload.get("balance_amount") is not None else float(max(Decimal("0"), (sales_order.total_amount or Decimal("0")) - Decimal(str(sales_order.paid_amount or 0)))),
            "salesperson": payload.get("salesperson") or "",
            "sourceDealId": payload.get("source_deal_id") or "",
            "subtotal": payload.get("subtotal") or 0,
            "taxTotal": payload.get("tax_total") or 0,
            "grandTotal": payload.get("grand_total") or 0,
        }
    else:
        list_key = "estimates" if normalized_doc_type == "estimate" else "invoices"
        document = next((row for row in data.get(list_key, []) if str(row.get("id")) == str(doc_id)), None)
    if not document:
        return JsonResponse({"detail": "document_not_found"}, status=404)

    gst_templates = {str(row.get("id")): row for row in data.get("gstTemplates", []) if isinstance(row, dict)}
    billing_templates = {str(row.get("id")): row for row in data.get("billingTemplates", []) if isinstance(row, dict)}
    totals = _document_totals(document, gst_templates)
    gst_template = _resolve_gst_template_by_id(gst_templates, str(document.get("gstTemplateId") or ""))
    billing_template = billing_templates.get(str(document.get("billingTemplateId") or ""))
    org_billing_profile = (
        BillingProfile.objects
        .filter(organization=org)
        .only("company_name", "address_line1", "address_line2", "city", "state", "country", "gstin")
        .first()
    )
    items = document.get("items") if isinstance(document.get("items"), list) else []
    gst_template_default_tax = _document_template_total_percent(gst_template)

    line_rows = []
    for row in items:
        if not isinstance(row, dict):
            continue
        qty = _to_decimal(row.get("qty"))
        rate = _to_decimal(row.get("rate"))
        amount = qty * rate
        raw_description = str(row.get("description") or "").replace("\r\n", "\n")
        raw_custom_text = str(row.get("customText") or "").strip()
        if raw_custom_text:
            description_main = raw_description.strip()
            description_custom = raw_custom_text
        else:
            first_line, _, remaining = raw_description.partition("\n")
            description_main = first_line.strip()
            description_custom = remaining.strip()
        line_tax_override = _resolve_document_line_tax_override(row)
        line_tax_percent = line_tax_override["tax_percent"] if line_tax_override["has_override"] else gst_template_default_tax
        line_rows.append(
            {
                "description": description_main,
                "description_custom": description_custom,
                "hsn_sac_type": str(row.get("hsnSacType") or row.get("hsn_sac_type") or "").strip(),
                "hsn_sac_code": str(row.get("hsnSacCode") or row.get("hsnCode") or row.get("sacCode") or "").strip(),
                "qty": str(row.get("qty") or ""),
                "rate": float(rate),
                "tax_percent": float(line_tax_percent),
                "amount": float(amount),
            }
        )

    paid_amount_value = _to_decimal(document.get("paidAmount"))
    balance_amount_value = max(Decimal("0"), _to_decimal(totals.get("grand_total")) - paid_amount_value)

    context = {
        "org": org,
        "doc_type": normalized_doc_type,
        "doc_type_label": "Estimate" if normalized_doc_type == "estimate" else "Sales Order" if normalized_doc_type == "sales_order" else "Invoice",
        "document": document,
        "items": line_rows,
        "gst_template": gst_template,
        "billing_template": billing_template,
        "totals": totals,
        "gst_total_percent": float(
            (_to_decimal(gst_template.get("cgst")) + _to_decimal(gst_template.get("sgst")) + _to_decimal(gst_template.get("igst")) + _to_decimal(gst_template.get("cess")))
            if isinstance(gst_template, dict)
            else Decimal("0")
        ),
        "theme_color": (billing_template or {}).get("themeColor") or "#22c55e",
        "company_logo_data_url": (billing_template or {}).get("companyLogoDataUrl") or "",
        "payment_status": str(document.get("paymentStatus") or "Pending").strip().title() or "Pending",
        "paid_amount": float(paid_amount_value),
        "balance_amount": float(balance_amount_value),
        "payment_mode": str(document.get("paymentMode") or "").strip(),
        "payment_date": str(document.get("paymentDate") or "").strip(),
        "transaction_id": str(document.get("transactionId") or "").strip(),
        "has_payment": paid_amount_value > 0,
    }
    if str(request.GET.get("format") or "").strip().lower() == "pdf":
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        left = 18 * mm
        right = width - 18 * mm
        top = height - 20 * mm
        logo_image = _ba_pdf_image_from_data_url(context["company_logo_data_url"])
        if not logo_image:
            logo_image = _ba_resolve_org_logo_image(org, request.user)
        money_font = _ba_get_unicode_pdf_font() or "Helvetica"
        tax_breakdown = totals.get("tax_breakdown") or {}
        breakdown_percent = totals.get("breakdown_percent") or {}
        gst_breakdown = []
        for key, label in (("cgst", "CGST"), ("sgst", "SGST"), ("igst", "IGST"), ("cess", "CESS")):
            amount = _to_decimal(tax_breakdown.get(key))
            if amount > 0:
                gst_breakdown.append((label, float(breakdown_percent.get(key) or 0), amount))

        pdf.setTitle(f"{context['doc_type_label']} {document.get('docNo') or doc_id}")
        pdf.setFillColorRGB(0, 0, 0)
        logo_height = _ba_draw_pdf_logo(pdf, logo_image, left, top + 1 * mm, 24 * mm, 14 * mm)
        right_section_x = right
        seller_x = left
        seller_header_y = top - (logo_height + 3 * mm if logo_height else 0)
        org_company_name = str(getattr(org_billing_profile, "company_name", "") or "").strip()
        seller_heading = org_company_name or str(org.name or "Organization")
        pdf.setFont("Helvetica-Bold", 18)
        pdf.drawString(seller_x, seller_header_y, seller_heading)
        pdf.setFont("Helvetica", 9)
        seller_lines = []
        org_address_line1 = str(getattr(org_billing_profile, "address_line1", "") or "").strip()
        org_address_line2 = str(getattr(org_billing_profile, "address_line2", "") or "").strip()
        org_city = str(getattr(org_billing_profile, "city", "") or "").strip()
        org_state = str(getattr(org_billing_profile, "state", "") or "").strip()
        org_country = str(getattr(org_billing_profile, "country", "") or "").strip()
        org_city_line = ", ".join([value for value in [org_city, org_state] if value])
        if org_address_line1:
            seller_lines.append(org_address_line1)
        if org_address_line2:
            seller_lines.append(org_address_line2)
        if org_city_line:
            seller_lines.append(org_city_line)
        elif org_country:
            seller_lines.append(org_country)
        seller_y = seller_header_y - 8 * mm
        for line in seller_lines[:3]:
            pdf.drawString(seller_x, seller_y, str(line))
            seller_y -= 4 * mm

        invoice_meta_y = top - 2 * mm
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawRightString(right_section_x, invoice_meta_y, context["doc_type_label"].upper())
        pdf.setFont("Helvetica", 9)
        pdf.drawRightString(right_section_x, invoice_meta_y - 6 * mm, f"# {document.get('docNo') or '-'}")
        pdf.drawRightString(right_section_x, invoice_meta_y - 10 * mm, f"Date {document.get('issueDate') or '-'}")
        pdf.setFont(money_font, 9)
        pdf.drawRightString(right_section_x, invoice_meta_y - 14 * mm, f"Amount {_format_pdf_inr(totals.get('grand_total') or '0')}")
        pdf.setFont("Helvetica", 9)
        pdf.drawRightString(right_section_x, invoice_meta_y - 18 * mm, f"Customer : {document.get('customerName') or '-'}")

        billing_address_lines = [line for line in str(document.get("billingAddress") or "").splitlines() if str(line).strip()]
        right_section_bottom = invoice_meta_y - 18 * mm
        left_section_bottom = seller_y if seller_lines else seller_header_y
        billed_y = min(left_section_bottom, right_section_bottom) - 8 * mm
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(left, billed_y, "BILLED TO")
        pdf.setFont("Helvetica", 9)
        billed_lines = [document.get("customerName") or "-"]
        customer_company = ""
        raw_customer = str(document.get("customerName") or "").strip()
        if raw_customer and raw_customer.lower() != str(org.name or "").strip().lower():
            customer_company = raw_customer
        if customer_company:
            billed_lines.append(customer_company)
        billed_lines.extend(billing_address_lines[:5])
        if document.get("customerGstin"):
            billed_lines.append(f"GSTIN: {document.get('customerGstin')}")
        billed_line_y = billed_y - 4 * mm
        for line in billed_lines[:7]:
            pdf.drawString(left, billed_line_y, str(line))
            billed_line_y -= 4 * mm

        details_y = billed_y
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawRightString(right_section_x, details_y, "DOCUMENT DETAILS")
        pdf.setFont("Helvetica", 9)
        details_rows = [
            f"Due Date {document.get('dueDate') or '-'}",
            f"Status {document.get('status') or '-'}",
        ]
        if isinstance(gst_template, dict) and str(gst_template.get("name") or "").strip():
            details_rows.append(f"GST Template {gst_template.get('name')}")
        for index, line in enumerate(details_rows, start=1):
            pdf.drawRightString(right_section_x, details_y - (index * 4 * mm), line)

        table_top = min(billed_line_y, details_y - (len(details_rows) + 1) * 4 * mm) - 8 * mm
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(left, table_top, "DESCRIPTION")
        pdf.drawString(left + 56 * mm, table_top, "GST")
        pdf.drawString(left + 90 * mm, table_top, "HSN / SAC")
        pdf.drawString(left + 112 * mm, table_top, "UNITS")
        pdf.drawString(left + 131 * mm, table_top, "UNIT PRICE")
        pdf.drawRightString(right, table_top, "AMOUNT")
        pdf.line(left, table_top - 2 * mm, right, table_top - 2 * mm)
        row_y = table_top - 8 * mm
        pdf.setFont("Helvetica", 9)
        for row in line_rows[:12]:
            line_tax = _to_decimal(row.get("tax_percent"))
            line_type = str(row.get("hsn_sac_type") or row.get("hsnSacType") or "HSN").strip().upper() or "HSN"
            line_type = "SAC" if line_type == "SAC" else "HSN"
            line_code = str(row.get("hsn_sac_code") or "").strip() or "-"
            line_code_display = f"{line_type} {line_code}" if line_code != "-" else line_type
            description_text = str(row.get("description_custom") or row.get("description") or "-")
            description_lines = _ba_wrap_pdf_text(pdf, description_text, 86 * mm, "Helvetica", 9)
            for idx, line in enumerate(description_lines[:2]):
                pdf.drawString(left, row_y - (idx * 4 * mm), line)
            text_bottom_y = row_y - ((len(description_lines[:2]) - 1) * 4 * mm)
            pdf.drawString(left + 90 * mm, text_bottom_y, line_code_display)
            pdf.drawString(left + 114 * mm, text_bottom_y, str(row.get("qty") or ""))
            pdf.setFont(money_font, 9)
            pdf.drawString(left + 131 * mm, text_bottom_y, _format_pdf_inr(row.get('rate') or '0'))
            pdf.drawRightString(right, text_bottom_y, _format_pdf_inr(row.get('amount') or '0'))
            pdf.setFont("Helvetica", 9)
            if line_tax > 0:
                pdf.setFont("Helvetica", 8)
                pdf.drawString(left + 56 * mm, text_bottom_y, f"GST {float(line_tax)}%")
                pdf.setFont("Helvetica", 9)
            pdf.line(left, text_bottom_y - 4 * mm, right, text_bottom_y - 4 * mm)
            row_y = text_bottom_y - 8 * mm

        summary_y = row_y - 8 * mm
        pdf.setFont(money_font, 9)
        pdf.drawRightString(right, summary_y, f"Sub Total {_format_pdf_inr(totals.get('subtotal') or '0')}")
        summary_y -= 5 * mm
        if gst_breakdown:
            for label, percent, tax_value in gst_breakdown:
                pdf.drawRightString(right, summary_y, f"{label} @ {percent:.2f}% {_format_pdf_inr(tax_value)}")
                summary_y -= 5 * mm
        else:
            pdf.drawRightString(right, summary_y, f"GST / Tax {_format_pdf_inr(totals.get('tax_total') or '0')}")
            summary_y -= 5 * mm
        pdf.setFont(money_font, 11)
        pdf.drawRightString(right, summary_y - 1 * mm, f"Total {_format_pdf_inr(totals.get('grand_total') or '0')}")

        payment_y = summary_y - 16 * mm
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(left, payment_y, "PAYMENT DETAILS")
        pdf.setFont("Helvetica", 9)
        payment_lines = [f"Payment Status: {context['payment_status']}"]
        if _to_decimal(context.get("paid_amount")) > 0:
            pdf.setFont(money_font, 9)
            payment_lines.append(f"Paid Amount: {_format_pdf_inr(context.get('paid_amount') or '0')}")
            payment_lines.append(f"Balance Amount: {_format_pdf_inr(context.get('balance_amount') or '0')}")
            pdf.setFont("Helvetica", 9)
            if context.get("payment_mode"):
                payment_lines.append(f"Payment Mode: {context['payment_mode']}")
            if context.get("payment_date"):
                payment_lines.append(f"Payment Date: {context['payment_date']}")
            if context.get("transaction_id"):
                payment_lines.append(f"Transaction ID: {context['transaction_id']}")
        payment_line_y = payment_y - 4 * mm
        for line in payment_lines[:6]:
            pdf.drawString(left, payment_line_y, str(line))
            payment_line_y -= 4 * mm

        pdf.showPage()
        pdf.save()
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        generated_at = timezone.localtime().strftime("%Y%m%d%H%M%S")
        filename = f"{normalized_doc_type}_{document.get('docNo') or doc_id}_{generated_at}.pdf".replace(" ", "_")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        return response
    return render(request, "business_autopilot/accounts/document_print.html", context)


@require_http_methods(["POST"])
def accounts_document_email(request, doc_type: str, doc_id: str):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    normalized_doc_type = (doc_type or "").strip().lower()
    doc_type_aliases = {
        "estimate": "estimate",
        "invoice": "invoice",
    }
    normalized_doc_type = doc_type_aliases.get(normalized_doc_type, normalized_doc_type)
    if normalized_doc_type not in {"estimate", "invoice"}:
        return JsonResponse({"detail": "invalid_doc_type"}, status=400)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    extra_message = str(payload.get("message") or "").strip()
    subject_override = str(payload.get("subject") or "").strip()
    recipient_override = payload.get("to")
    document_override = payload.get("document") if isinstance(payload.get("document"), dict) else None

    recipients = []
    if isinstance(recipient_override, (list, tuple)):
        recipients = [str(item or "").strip() for item in recipient_override]
    elif isinstance(recipient_override, str):
        recipients = [item.strip() for item in recipient_override.split(",")]
    recipients = [item for item in recipients if item]

    workspace = _get_accounts_workspace(org)
    data = _normalize_accounts_workspace(workspace.data)

    list_key = "estimates" if normalized_doc_type == "estimate" else "invoices"
    document = document_override or next(
        (row for row in data.get(list_key, []) if str(row.get("id")) == str(doc_id)),
        None,
    )
    if not document:
        return JsonResponse({"detail": "document_not_found"}, status=404)

    customer_id = str(document.get("customerId") or document.get("customer_id") or "").strip()
    customer_name = str(
        document.get("customerName")
        or document.get("customer_name")
        or document.get("companyOrClientName")
        or document.get("company_or_client_name")
        or document.get("company")
        or document.get("clients")
        or ""
    ).strip()

    if not recipients:
        customers = data.get("customers") if isinstance(data.get("customers"), list) else []
        customer = None
        if customer_id:
            customer = next((row for row in customers if str(row.get("id") or "").strip() == customer_id), None)
        if customer is None and customer_name:
            normalized_customer_name = customer_name.lower().strip()
            customer = next(
                (
                    row
                    for row in customers
                    if str(row.get("companyName") or row.get("name") or "").lower().strip() == normalized_customer_name
                ),
                None,
            )
        if customer:
            primary_email = str(customer.get("email") or "").strip()
            additional_emails = customer.get("additionalEmails") if isinstance(customer.get("additionalEmails"), list) else []
            email_list = customer.get("emailList") if isinstance(customer.get("emailList"), list) else additional_emails
            recipients = [primary_email, *[str(item or "").strip() for item in email_list]]
            recipients = [item for item in recipients if item]

    if not recipients:
        return JsonResponse({"detail": "client_email_missing"}, status=400)

    gst_templates = {str(row.get("id")): row for row in data.get("gstTemplates", []) if isinstance(row, dict)}
    totals = _document_totals(document, gst_templates)
    grand_total = totals.get("grand_total") if isinstance(totals, dict) else None

    doc_no = str(document.get("docNo") or document.get("doc_no") or "").strip()
    issue_date = str(document.get("issueDate") or document.get("issue_date") or document.get("date") or "").strip()
    due_date = str(document.get("dueDate") or document.get("due_date") or "").strip()

    doc_label = "Invoice" if normalized_doc_type == "invoice" else "Estimate"
    subject = subject_override or f"{doc_label} {doc_no or doc_id} from {org.name}"

    print_url = request.build_absolute_uri(
        f"/api/business-autopilot/accounts/documents/{normalized_doc_type}/{doc_id}/print?format=pdf"
    )

    context = {
        "org_name": org.name,
        "doc_type": normalized_doc_type,
        "doc_label": doc_label,
        "doc_no": doc_no or doc_id,
        "customer_name": customer_name or "-",
        "issue_date": issue_date or "-",
        "due_date": due_date or "-",
        "grand_total": grand_total if grand_total is not None else "-",
        "extra_message": extra_message,
        "print_url": print_url,
    }

    mail_sent = send_templated_email(recipients, subject, "emails/accounts_document_sent.txt", context)
    if not mail_sent:
        return JsonResponse({"detail": "email_send_failed"}, status=502)
    return JsonResponse({"sent": True, "recipients": recipients})


def _crm_is_admin(user: User, org: Organization):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    profile = UserProfile.objects.filter(user=user).only("role").first()
    profile_role = _normalize_admin_role(getattr(profile, "role", ""))
    if profile_role in {"company_admin", "org_admin", "owner", "superadmin", "super_admin"}:
        return True
    membership = _get_org_membership(user, org)
    membership_role = _normalize_admin_role(getattr(membership, "role", "")) if membership else ""
    return membership_role in {"company_admin", "org_admin", "owner"}


def _crm_to_decimal(value):
    try:
        normalized = str(value or "0").strip() or "0"
        normalized = normalized.replace(",", "")
        normalized = "".join(ch for ch in normalized if ch.isdigit() or ch in {".", "-"})
        return Decimal(normalized or "0")
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def _normalize_payment_status(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if normalized == "paid":
        return "paid"
    if normalized in {"partial", "partially paid"}:
        return "partial"
    return "pending"


def _crm_order_id(org: Organization):
    day_key = timezone.localdate().strftime("%d%m%Y")
    prefix = f"SO-{day_key}-"
    pattern = re.compile(rf"^{re.escape(prefix)}(\d{{3}})$")
    last_seq = 0
    existing_ids = (
        CrmSalesOrder.objects
        .filter(organization=org, order_id__startswith=prefix)
        .values_list("order_id", flat=True)
    )
    for raw_id in existing_ids:
        match = pattern.match(str(raw_id or "").strip())
        if not match:
            continue
        try:
            seq = int(match.group(1))
        except (TypeError, ValueError):
            continue
        if seq > last_seq:
            last_seq = seq
    return f"{prefix}{last_seq + 1:03d}"


def _crm_issue_date_from_order_id(order_id: str):
    raw_order_id = str(order_id or "").strip()
    if not raw_order_id:
        return ""
    compact_match = re.match(r"^SO-(\d{2})(\d{2})(\d{4})-\d{3}$", raw_order_id, re.IGNORECASE)
    if compact_match:
        day = int(compact_match.group(1))
        month = int(compact_match.group(2))
        year = int(compact_match.group(3))
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return ""
    dashed_match = re.match(r"^SO-(\d{2})-(\d{2})-(\d{4})-\d{3}$", raw_order_id, re.IGNORECASE)
    if dashed_match:
        day = int(dashed_match.group(1))
        month = int(dashed_match.group(2))
        year = int(dashed_match.group(3))
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return ""
    return ""


def _crm_resolve_unique_order_id(org: Organization, requested_order_id: str = ""):
    requested = str(requested_order_id or "").strip()[:30]
    if requested:
        if not CrmSalesOrder.objects.filter(organization=org, order_id=requested).exists():
            return requested
    for _ in range(25):
        candidate = _crm_order_id(org)
        if not CrmSalesOrder.objects.filter(organization=org, order_id=candidate).exists():
            return candidate
    return _crm_order_id(org)


def _crm_body_row_id(payload, *keys):
    if not isinstance(payload, dict):
        return None
    for key in keys:
        row_id = _coerce_positive_int(payload.get(key))
        if row_id:
            return row_id
    return None


def _crm_reference_id_from_related_to(org: Organization, related_to: str):
    related_value = str(related_to or "").strip()
    if not related_value:
        return ""
    lead_match = (
        CrmLead.objects
        .filter(organization=org, is_deleted=False, lead_name__iexact=related_value)
        .order_by("-updated_at", "-id")
        .values_list("crm_reference_id", flat=True)
        .first()
    )
    if lead_match:
        return str(lead_match).strip()
    lead_company_match = (
        CrmLead.objects
        .filter(organization=org, is_deleted=False, company__iexact=related_value)
        .order_by("-updated_at", "-id")
        .values_list("crm_reference_id", flat=True)
        .first()
    )
    if lead_company_match:
        return str(lead_company_match).strip()
    deal_match = (
        CrmDeal.objects
        .filter(organization=org, is_deleted=False, deal_name__iexact=related_value)
        .order_by("-updated_at", "-id")
        .values_list("crm_reference_id", flat=True)
        .first()
    )
    if deal_match:
        return str(deal_match).strip()
    deal_company_match = (
        CrmDeal.objects
        .filter(organization=org, is_deleted=False, company__iexact=related_value)
        .order_by("-updated_at", "-id")
        .values_list("crm_reference_id", flat=True)
        .first()
    )
    if deal_company_match:
        return str(deal_company_match).strip()
    return ""


def _crm_clean_user_id_list(value):
    raw = value if isinstance(value, list) else []
    user_ids = []
    for item in raw:
        parsed = _coerce_positive_int(item)
        if parsed:
            user_ids.append(parsed)
    return list(dict.fromkeys(user_ids))


def _crm_extract_assigned_user_names_from_payload(payload) -> list:
    if not isinstance(payload, dict):
        return []
    if isinstance(payload.get("assigned_user_names"), list):
        return [str(item or "").strip() for item in payload.get("assigned_user_names") if str(item or "").strip()]
    if isinstance(payload.get("assignedUserNames"), list):
        return [str(item or "").strip() for item in payload.get("assignedUserNames") if str(item or "").strip()]
    if isinstance(payload.get("assignedUser"), list):
        return [str(item or "").strip() for item in payload.get("assignedUser") if str(item or "").strip()]
    if isinstance(payload.get("assigned_user"), list):
        return [str(item or "").strip() for item in payload.get("assigned_user") if str(item or "").strip()]
    if payload.get("assignedUser"):
        return [str(item or "").strip() for item in str(payload.get("assignedUser") or "").split(",") if str(item or "").strip()]
    if payload.get("assignedTo"):
        return [str(item or "").strip() for item in str(payload.get("assignedTo") or "").split(",") if str(item or "").strip()]
    return []


def _crm_resolve_user_ids_from_names(org: Organization, value):
    raw_names = value if isinstance(value, list) else []
    normalized_names = [str(item or "").strip().lower() for item in raw_names if str(item or "").strip()]
    if not normalized_names:
        return []
    member_users = [
        membership.user
        for membership in OrganizationUser.objects.filter(organization=org, is_deleted=False).select_related("user")
        if membership.user_id
    ]
    resolved_ids = []
    for user in member_users:
        full_name = str(getattr(user, "get_full_name", lambda: "")() or "").strip()
        display_name = _get_org_user_display_name(user)
        candidates = {
            str(getattr(user, "email", "") or "").strip().lower(),
            str(getattr(user, "username", "") or "").strip().lower(),
            full_name.lower(),
            str(display_name or "").strip().lower(),
        }
        if any(candidate and candidate in normalized_names for candidate in candidates):
            resolved_ids.append(int(user.id))
    return list(dict.fromkeys(resolved_ids))


def _crm_clean_int_list(value):
    raw = value if isinstance(value, list) else []
    numbers = []
    for item in raw:
        try:
            parsed = int(str(item or "").strip())
        except (TypeError, ValueError):
            continue
        if parsed < 0:
            continue
        numbers.append(parsed)
    return list(dict.fromkeys(numbers))


def _crm_sales_order_payload_dict(row: CrmSalesOrder):
    raw_payload = row.products if isinstance(row.products, dict) else {}
    if not isinstance(raw_payload, dict):
        raw_payload = {}
    items = raw_payload.get("items")
    if not isinstance(items, list):
        items = row.products if isinstance(row.products, list) else []
    return {
        **raw_payload,
        "items": [item for item in items if isinstance(item, dict)],
    }


def _crm_normalize_sales_order_items(items):
    normalized_items = []
    if not isinstance(items, list):
        return normalized_items
    for index, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        description = str(item.get("description") or item.get("customText") or "").strip()
        hsn_sac_type = str(item.get("hsnSacType") or item.get("hsn_sac_type") or "HSN").strip().upper() or "HSN"
        hsn_sac_code = str(item.get("hsnSacCode") or item.get("hsn_sac_code") or item.get("hsnCode") or item.get("sacCode") or "").strip()
        qty_value = _crm_to_decimal(item.get("qty"))
        qty = qty_value if qty_value > 0 else Decimal("1")
        rate = _crm_to_decimal(item.get("rate"))
        tax_override = _resolve_document_line_tax_override(item)
        qty_text = format(qty.quantize(Decimal("1")), "f") if qty == qty.to_integral() else format(qty.normalize(), "f")
        normalized_items.append({
            "id": str(item.get("id") or f"crm_so_line_{index + 1}").strip(),
            "itemMasterId": str(item.get("itemMasterId") or "").strip(),
            "inventoryItemId": str(item.get("inventoryItemId") or "").strip(),
            "description": description,
            "customText": str(item.get("customText") or "").strip(),
            "hsnSacType": "SAC" if hsn_sac_type == "SAC" else "HSN",
            "hsnSacCode": hsn_sac_code,
            "qty": qty_text,
            "rate": str(rate),
            "taxPercent": str(tax_override["tax_percent"]) if tax_override["has_override"] else "",
            "taxPercentSource": tax_override["tax_percent_source"],
        })
    return normalized_items


def _crm_sales_order_totals(items, default_tax_percent=Decimal("0")):
    subtotal = Decimal("0")
    tax_total = Decimal("0")
    total_qty = Decimal("0")
    first_rate = Decimal("0")
    first_tax = Decimal("0")
    for index, item in enumerate(items):
        qty = _crm_to_decimal(item.get("qty"))
        rate = _crm_to_decimal(item.get("rate"))
        tax_override = _resolve_document_line_tax_override(item)
        tax_percent = tax_override["tax_percent"] if tax_override["has_override"] else _to_decimal(default_tax_percent)
        line_subtotal = qty * rate
        subtotal += line_subtotal
        tax_total += line_subtotal * (tax_percent / Decimal("100"))
        total_qty += qty
        if index == 0:
            first_rate = rate
            first_tax = tax_percent
    return {
        "subtotal": subtotal,
        "tax_total": tax_total,
        "grand_total": subtotal + tax_total,
        "total_qty": total_qty,
        "first_rate": first_rate,
        "first_tax": first_tax,
    }


def _crm_can_access_row(user: User, org: Organization, row):
    return _crm_can_view_row(user, org, row)


def _crm_assigned_user_names(primary_user, raw_user_ids):
    ordered_ids = []
    primary_user_id = getattr(primary_user, "id", None)
    if primary_user_id:
        ordered_ids.append(int(primary_user_id))
    for user_id in _crm_clean_user_id_list(raw_user_ids):
        if user_id not in ordered_ids:
            ordered_ids.append(int(user_id))
    if not ordered_ids:
        return []
    users_by_id = {user.id: user for user in User.objects.filter(id__in=ordered_ids)}
    resolved_names = []
    for user_id in ordered_ids:
        matched_user = users_by_id.get(user_id)
        if not matched_user:
            continue
        display_name = _get_org_user_display_name(matched_user)
        if display_name:
            resolved_names.append(display_name)
    return resolved_names


def _serialize_crm_lead(row: CrmLead):
    assigned_user_ids = _crm_clean_user_id_list(row.assigned_user_ids)
    assigned_user_names = _crm_assigned_user_names(row.assigned_user, assigned_user_ids)
    assigned_user_name = ""
    if row.assigned_user_id and row.assigned_user:
        assigned_user_name = str(row.assigned_user.get_full_name() or row.assigned_user.email or "").strip()
    completed_by_user_name = _get_org_user_display_name(row.completed_by_user) if getattr(row, "completed_by_user_id", None) else ""
    completed_by_team = str(getattr(row, "completed_by_team", "") or "").strip()
    completed_by_type = str(getattr(row, "completed_by_type", "") or "").strip()
    completed_by_user_ids = _crm_clean_user_id_list(getattr(row, "completed_by_user_ids", []) or [])
    completed_by_user_names = []
    if completed_by_user_ids:
        users_by_id = {user.id: user for user in User.objects.filter(id__in=completed_by_user_ids)}
        for user_id in completed_by_user_ids:
            matched_user = users_by_id.get(int(user_id))
            if not matched_user:
                continue
            display_name = _get_org_user_display_name(matched_user)
            if display_name:
                completed_by_user_names.append(display_name)
    completed_by_team_names = [
        str(name or "").strip()
        for name in (getattr(row, "completed_by_team_names", None) or [])
        if str(name or "").strip()
    ]
    completed_by_name_parts = []
    completed_by_name_parts.extend(completed_by_user_names)
    completed_by_name_parts.extend(completed_by_team_names)
    if not completed_by_name_parts:
        if completed_by_type.lower() == "team" and completed_by_team:
            completed_by_name_parts = [completed_by_team]
        elif completed_by_user_name:
            completed_by_name_parts = [completed_by_user_name]
    completed_by_name = ", ".join(completed_by_name_parts)
    return {
        "id": row.id,
        "crm_reference_id": str(row.crm_reference_id or "").strip(),
        "lead_name": row.lead_name,
        "company": row.company,
        "phone": row.phone,
        "lead_amount": float(row.lead_amount or 0),
        "lead_source": row.lead_source,
        "assign_type": row.assign_type,
        "assigned_user_id": row.assigned_user_id,
        "assigned_user_name": assigned_user_name,
        "assigned_user_ids": assigned_user_ids,
        "assigned_user_names": assigned_user_names,
        "assigned_team": row.assigned_team,
        "stage": row.stage,
        "priority": row.priority or "Medium",
        "status": row.status,
        "is_deleted": bool(row.is_deleted),
        "deleted_at": row.deleted_at.isoformat() if row.deleted_at else None,
        "final_proposal_amount": float(getattr(row, "final_proposal_amount", 0) or 0),
        "proposal_finalized_at": row.proposal_finalized_at.isoformat() if getattr(row, "proposal_finalized_at", None) else None,
        "proposal_finalized_by_id": row.proposal_finalized_by_id if hasattr(row, "proposal_finalized_by_id") else None,
        "proposal_finalized_by_name": _get_org_user_display_name(row.proposal_finalized_by) if getattr(row, "proposal_finalized_by_id", None) else "",
        "completed_by_type": completed_by_type,
        "completed_by_user_id": getattr(row, "completed_by_user_id", None),
        "completed_by_user_name": completed_by_user_name,
        "completed_by_team": completed_by_team,
        "completed_by_user_ids": completed_by_user_ids,
        "completed_by_user_names": completed_by_user_names,
        "completed_by_team_names": completed_by_team_names,
        "completed_by_name": completed_by_name,
        "created_by_id": row.created_by_id,
        "created_by_name": _get_org_user_display_name(row.created_by) if row.created_by_id else "",
        "updated_by_id": row.updated_by_id,
        "updated_by_name": _get_org_user_display_name(row.updated_by) if row.updated_by_id else "",
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "updated_at": row.updated_at.isoformat() if row.updated_at else "",
    }


def _crm_lead_snapshot(row: CrmLead) -> dict:
    if not row:
        return {}
    return {
        "crm_reference_id": str(getattr(row, "crm_reference_id", "") or "").strip(),
        "lead_name": str(getattr(row, "lead_name", "") or "").strip(),
        "company": str(getattr(row, "company", "") or "").strip(),
        "phone": str(getattr(row, "phone", "") or "").strip(),
        "lead_amount": float(getattr(row, "lead_amount", 0) or 0),
        "lead_source": str(getattr(row, "lead_source", "") or "").strip(),
        "assign_type": str(getattr(row, "assign_type", "") or "").strip(),
        "assigned_user_id": getattr(row, "assigned_user_id", None),
        "assigned_user_ids": _crm_clean_user_id_list(getattr(row, "assigned_user_ids", []) or []),
        "assigned_team": str(getattr(row, "assigned_team", "") or "").strip(),
        "stage": str(getattr(row, "stage", "") or "").strip(),
        "priority": str(getattr(row, "priority", "") or "").strip(),
        "status": str(getattr(row, "status", "") or "").strip(),
        "is_deleted": bool(getattr(row, "is_deleted", False)),
        "final_proposal_amount": float(getattr(row, "final_proposal_amount", 0) or 0),
        "proposal_finalized_at": getattr(row, "proposal_finalized_at", None).isoformat() if getattr(row, "proposal_finalized_at", None) else None,
        "completed_by_type": str(getattr(row, "completed_by_type", "") or "").strip(),
        "completed_by_user_id": getattr(row, "completed_by_user_id", None),
        "completed_by_team": str(getattr(row, "completed_by_team", "") or "").strip(),
        "completed_by_user_ids": _crm_clean_user_id_list(getattr(row, "completed_by_user_ids", []) or []),
        "completed_by_team_names": [
            str(item or "").strip()
            for item in (getattr(row, "completed_by_team_names", None) or [])
            if str(item or "").strip()
        ],
    }


def _crm_diff_snapshots(before: dict, after: dict) -> list:
    safe_before = before if isinstance(before, dict) else {}
    safe_after = after if isinstance(after, dict) else {}
    keys = sorted(set(safe_before.keys()) | set(safe_after.keys()))
    changes = []
    for key in keys:
        old_value = safe_before.get(key)
        new_value = safe_after.get(key)
        if old_value == new_value:
            continue
        changes.append({"field": key, "old": old_value, "new": new_value})
    return changes


def _crm_log_lead_modification(org: Organization, lead: CrmLead, *, changed_by: User = None, action: str = "update", changes: list = None):
    try:
        CrmLeadModification.objects.create(
            organization=org,
            lead=lead,
            lead_reference_id=str(getattr(lead, "crm_reference_id", "") or "").strip(),
            lead_name=str(getattr(lead, "lead_name", "") or "").strip(),
            changed_by=changed_by if getattr(changed_by, "is_authenticated", False) else None,
            action=str(action or "update").strip()[:40] or "update",
            changes=changes if isinstance(changes, list) else [],
        )
    except (DatabaseError, OperationalError, IntegrityError):
        logger.exception("Failed to record CRM lead modification org_id=%s lead_id=%s", getattr(org, "id", None), getattr(lead, "id", None))


def _serialize_crm_lead_modification(row: CrmLeadModification):
    if not row:
        return {}
    return {
        "id": row.id,
        "lead_id": row.lead_id,
        "lead_reference_id": str(row.lead_reference_id or "").strip(),
        "lead_name": str(row.lead_name or "").strip(),
        "action": str(row.action or "").strip(),
        "changed_by_id": row.changed_by_id,
        "changed_by_name": _get_org_user_display_name(row.changed_by) if row.changed_by_id else "",
        "changes": row.changes if isinstance(row.changes, list) else [],
        "created_at": timezone.localtime(row.created_at).isoformat() if row.created_at else "",
    }


def _serialize_crm_contact(row: CrmContact):
    return {
        "id": row.id,
        "name": row.name,
        "company": row.company,
        "email": row.email,
        "phone_country_code": row.phone_country_code or "+91",
        "phone": row.phone,
        "tag": row.tag or "Client",
        "is_deleted": bool(row.is_deleted),
        "deleted_at": row.deleted_at.isoformat() if row.deleted_at else None,
        "created_by_id": row.created_by_id,
        "created_by_name": _get_org_user_display_name(row.created_by) if row.created_by_id else "",
        "updated_by_id": row.updated_by_id,
        "updated_by_name": _get_org_user_display_name(row.updated_by) if row.updated_by_id else "",
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "updated_at": row.updated_at.isoformat() if row.updated_at else "",
    }


def _crm_normalize_phone_digits(value: str) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def _crm_find_duplicate_contact(
    org: Organization,
    *,
    company: str = "",
    email: str = "",
    phone_country_code: str = "+91",
    phone: str = "",
    exclude_contact_id: Optional[int] = None,
):
    normalized_company = str(company or "").strip()
    normalized_company_lower = normalized_company.lower()
    normalized_email = str(email or "").strip()
    normalized_email_lower = normalized_email.lower()
    normalized_phone_country_code = str(phone_country_code or "+91").strip() or "+91"
    normalized_phone_country_code_lower = normalized_phone_country_code.lower()
    normalized_phone_digits = _crm_normalize_phone_digits(phone)

    duplicate_q = Q()
    has_filters = False
    if normalized_company_lower:
        duplicate_q |= Q(company__iexact=normalized_company)
        has_filters = True
    if normalized_email_lower:
        duplicate_q |= Q(email__iexact=normalized_email)
        has_filters = True
    if normalized_phone_digits:
        duplicate_q |= Q(phone_country_code__iexact=normalized_phone_country_code)
        has_filters = True
    if not has_filters:
        return None, []

    queryset = CrmContact.objects.filter(organization=org, is_deleted=False)
    if exclude_contact_id:
        queryset = queryset.exclude(id=exclude_contact_id)
    candidates = queryset.filter(duplicate_q).select_related("created_by").order_by("-created_at", "-id")

    for candidate in candidates:
        matched_fields = []
        candidate_company = str(getattr(candidate, "company", "") or "").strip().lower()
        candidate_email = str(getattr(candidate, "email", "") or "").strip().lower()
        candidate_phone_country_code = str(getattr(candidate, "phone_country_code", "") or "+91").strip().lower()
        candidate_phone_digits = _crm_normalize_phone_digits(getattr(candidate, "phone", ""))
        if normalized_company_lower and candidate_company and candidate_company == normalized_company_lower:
            matched_fields.append("company")
        if normalized_email_lower and candidate_email and candidate_email == normalized_email_lower:
            matched_fields.append("email")
        if (
            normalized_phone_digits
            and candidate_phone_digits
            and candidate_phone_country_code == normalized_phone_country_code_lower
            and candidate_phone_digits == normalized_phone_digits
        ):
            matched_fields.append("phone")
        if matched_fields:
            return candidate, matched_fields

    return None, []


def _crm_collect_leads_linked_to_user(org: Organization, user_id: int):
    safe_user_id = _coerce_positive_int(user_id)
    if not safe_user_id:
        return []
    rows = (
        CrmLead.objects
        .filter(organization=org, is_deleted=False)
        .select_related("assigned_user", "created_by")
        .order_by("-created_at", "-id")
    )
    linked = []
    for row in rows:
        assigned_ids = set(_crm_clean_user_id_list(row.assigned_user_ids))
        assigned_user_id = _coerce_positive_int(getattr(row, "assigned_user_id", None))
        if assigned_user_id:
            assigned_ids.add(int(assigned_user_id))
        if safe_user_id in assigned_ids:
            linked.append(row)
    return linked


def _crm_collect_deals_linked_to_user(org: Organization, user_id: int):
    safe_user_id = _coerce_positive_int(user_id)
    if not safe_user_id:
        return []
    rows = (
        CrmDeal.objects
        .filter(organization=org, is_deleted=False)
        .select_related("assigned_user", "lead", "created_by")
        .order_by("-created_at", "-id")
    )
    linked = []
    for row in rows:
        assigned_ids = set(_crm_clean_user_id_list(row.assigned_user_ids))
        assigned_user_id = _coerce_positive_int(getattr(row, "assigned_user_id", None))
        if assigned_user_id:
            assigned_ids.add(int(assigned_user_id))
        if safe_user_id in assigned_ids:
            linked.append(row)
    return linked


def _crm_resolve_org_admin_user_id(org: Organization, fallback_user_id: int = None) -> int:
    if getattr(org, "owner_id", None):
        return int(org.owner_id)
    admin_membership = (
        OrganizationUser.objects
        .filter(organization=org, is_deleted=False, role__iexact="company_admin")
        .order_by("id")
        .first()
    )
    if admin_membership and admin_membership.user_id:
        return int(admin_membership.user_id)
    if fallback_user_id:
        return int(fallback_user_id)
    return 0


def _crm_resolve_target_user_ids(org: Organization, membership_ids, *, exclude_user_id: int = None) -> list:
    safe_ids = [
        _coerce_positive_int(value)
        for value in (membership_ids or [])
    ]
    safe_ids = [value for value in safe_ids if value]
    if not safe_ids:
        return []
    memberships = (
        OrganizationUser.objects
        .filter(organization=org, is_deleted=False, id__in=safe_ids)
        .select_related("user")
    )
    resolved = []
    exclude_id = _coerce_positive_int(exclude_user_id)
    for membership in memberships:
        user_id = _coerce_positive_int(membership.user_id)
        if not user_id or (exclude_id and user_id == exclude_id):
            continue
        if membership.user and not membership.user.is_active:
            continue
        resolved.append(user_id)
    # Keep stable ordering + uniqueness.
    seen = set()
    ordered = []
    for user_id in resolved:
        if user_id in seen:
            continue
        seen.add(user_id)
        ordered.append(user_id)
    return ordered


def _crm_build_user_display_map(user_ids):
    ids = [_coerce_positive_int(value) for value in (user_ids or [])]
    ids = [value for value in ids if value]
    if not ids:
        return {}
    users_by_id = {user.id: user for user in User.objects.filter(id__in=ids)}
    display = {}
    for user_id in ids:
        user = users_by_id.get(user_id)
        if not user:
            continue
        name = str(user.get_full_name() or "").strip()
        display[user_id] = name or str(user.email or "").strip() or f"User {user_id}"
    return display


def _crm_snapshot_and_reassign_user_records(
    org: Organization,
    *,
    membership: OrganizationUser,
    target_user_ids,
    performed_by: User = None,
):
    source_user_id = _coerce_positive_int(getattr(membership, "user_id", None))
    if not source_user_id:
        return {"snapshot": None, "summary": {"leads": 0, "deals": 0, "sales_orders": 0, "meetings": 0}}

    leads = _crm_collect_leads_linked_to_user(org, source_user_id)
    deals = _crm_collect_deals_linked_to_user(org, source_user_id)
    sales_orders = list(
        CrmSalesOrder.objects
        .filter(organization=org, is_deleted=False, assigned_user_id=source_user_id)
        .select_related("assigned_user")
        .order_by("-created_at", "-id")
    )
    meetings = list(
        CrmMeeting.objects
        .filter(organization=org, is_deleted=False)
        .order_by("-created_at", "-id")
    )
    linked_meetings = []
    for meeting in meetings:
        owner_ids = set(_crm_clean_user_id_list(meeting.owner_user_ids))
        if source_user_id in owner_ids:
            linked_meetings.append(meeting)

    target_ids = [int(value) for value in (target_user_ids or []) if _coerce_positive_int(value)]
    if not target_ids:
        target_ids = []
    display_map = _crm_build_user_display_map([source_user_id, *target_ids])

    snapshot_payload = {
        "source_user_id": source_user_id,
        "source_user_name": display_map.get(source_user_id, ""),
        "targets_user_ids": target_ids,
        "targets_user_names": [display_map.get(user_id, "") for user_id in target_ids],
        "leads": [
            {
                "id": row.id,
                "assign_type": str(row.assign_type or ""),
                "assigned_team": str(row.assigned_team or ""),
                "assigned_user_id": row.assigned_user_id,
                "assigned_user_ids": _crm_clean_user_id_list(row.assigned_user_ids),
            }
            for row in leads
        ],
        "deals": [
            {
                "id": row.id,
                "assigned_team": str(row.assigned_team or ""),
                "assigned_user_id": row.assigned_user_id,
                "assigned_user_ids": _crm_clean_user_id_list(row.assigned_user_ids),
            }
            for row in deals
        ],
        "sales_orders": [
            {
                "id": row.id,
                "assigned_user_id": row.assigned_user_id,
            }
            for row in sales_orders
        ],
        "meetings": [
            {
                "id": row.id,
                "owner_user_ids": _crm_clean_user_id_list(row.owner_user_ids),
                "owner_names": str(row.owner_names or ""),
            }
            for row in linked_meetings
        ],
    }

    snapshot = None
    try:
        snapshot = BusinessAutopilotUserCrmReassignmentSnapshot.objects.create(
            organization=org,
            membership=membership,
            source_user_id=source_user_id,
            created_by=performed_by if performed_by and getattr(performed_by, "id", None) else None,
            reassigned_to_user_ids=target_ids,
            snapshot=snapshot_payload,
        )
    except DatabaseError:
        # Fallback for environments where migrations have not been applied yet.
        logger.exception("Failed to persist CRM reassignment snapshot for membership_id=%s", getattr(membership, "id", None))
        snapshot = None

    if not target_ids:
        return {"snapshot": snapshot, "summary": {"leads": len(leads), "deals": len(deals), "sales_orders": len(sales_orders), "meetings": len(linked_meetings)}}

    target_display_names = [display_map.get(user_id, "") for user_id in target_ids]
    fallback_user_id = target_ids[0]
    fallback_user = User.objects.filter(id=fallback_user_id).first()

    def _merge_reassignment(existing_ids: set[int]) -> list[int]:
        merged = set(int(value) for value in existing_ids if _coerce_positive_int(value))
        merged.discard(source_user_id)
        for target_id in target_ids:
            merged.add(int(target_id))
        if not merged:
            merged.update(target_ids)
        return list(merged)

    # Leads
    for row in leads:
        existing_ids = set(_crm_clean_user_id_list(row.assigned_user_ids))
        if row.assigned_user_id:
            existing_ids.add(int(row.assigned_user_id))
        if source_user_id not in existing_ids and row.assigned_user_id != source_user_id:
            continue
        next_ids = _merge_reassignment(existing_ids)
        row.assign_type = "Users"
        row.assigned_team = ""
        row.assigned_user_ids = next_ids
        row.assigned_user_id = next_ids[0] if next_ids else None
        if fallback_user and row.assigned_user_id:
            row.assigned_user = fallback_user
        row.updated_by = performed_by if performed_by else row.updated_by
        row.save(update_fields=["assign_type", "assigned_team", "assigned_user_ids", "assigned_user", "updated_by", "updated_at"])

    # Deals
    for row in deals:
        existing_ids = set(_crm_clean_user_id_list(row.assigned_user_ids))
        if row.assigned_user_id:
            existing_ids.add(int(row.assigned_user_id))
        if source_user_id not in existing_ids and row.assigned_user_id != source_user_id:
            continue
        next_ids = _merge_reassignment(existing_ids)
        row.assigned_team = ""
        row.assigned_user_ids = next_ids
        row.assigned_user_id = next_ids[0] if next_ids else None
        if fallback_user and row.assigned_user_id:
            row.assigned_user = fallback_user
        row.updated_by = performed_by if performed_by else row.updated_by
        row.save(update_fields=["assigned_team", "assigned_user_ids", "assigned_user", "updated_by", "updated_at"])

    # Sales orders
    for row in sales_orders:
        row.assigned_user_id = target_ids[0]
        if fallback_user:
            row.assigned_user = fallback_user
        row.updated_by = performed_by if performed_by else row.updated_by
        row.save(update_fields=["assigned_user", "updated_by", "updated_at"])

    # Meetings
    for row in linked_meetings:
        existing_ids = set(_crm_clean_user_id_list(row.owner_user_ids))
        next_ids = _merge_reassignment(existing_ids)
        row.owner_user_ids = next_ids
        # Refresh owner names for UI.
        owners_display = _crm_build_user_display_map(next_ids)
        row.owner_names = ", ".join([owners_display.get(user_id, "") for user_id in next_ids if owners_display.get(user_id, "")]).strip()
        row.updated_by = performed_by if performed_by else row.updated_by
        row.save(update_fields=["owner_user_ids", "owner_names", "updated_by", "updated_at"])

    return {
        "snapshot": snapshot,
        "summary": {
            "leads": len(leads),
            "deals": len(deals),
            "sales_orders": len(sales_orders),
            "meetings": len(linked_meetings),
        },
        "target_names": target_display_names,
        "target_user_ids": target_ids,
    }


def _crm_restore_user_records_from_snapshot(org: Organization, *, membership: OrganizationUser, performed_by: User = None) -> dict:
    source_user_id = _coerce_positive_int(getattr(membership, "user_id", None))
    if not source_user_id:
        return {"restored": False, "summary": {"leads": 0, "deals": 0, "sales_orders": 0, "meetings": 0}}
    snapshot = (
        BusinessAutopilotUserCrmReassignmentSnapshot.objects
        .filter(organization=org, membership=membership, source_user_id=source_user_id, reverted_at__isnull=True)
        .order_by("-created_at", "-id")
        .first()
    )
    if not snapshot:
        return {"restored": False, "summary": {"leads": 0, "deals": 0, "sales_orders": 0, "meetings": 0}}

    payload = snapshot.snapshot or {}
    restored_counts = {"leads": 0, "deals": 0, "sales_orders": 0, "meetings": 0}

    lead_rows = {row.id: row for row in CrmLead.objects.filter(organization=org, is_deleted=False, id__in=[item.get("id") for item in (payload.get("leads") or [])]).select_related("assigned_user")}
    for item in payload.get("leads") or []:
        row = lead_rows.get(item.get("id"))
        if not row:
            continue
        row.assign_type = str(item.get("assign_type") or row.assign_type or "Users")
        row.assigned_team = str(item.get("assigned_team") or "")
        row.assigned_user_ids = _crm_clean_user_id_list(item.get("assigned_user_ids"))
        row.assigned_user_id = _coerce_positive_int(item.get("assigned_user_id")) or None
        row.updated_by = performed_by if performed_by else row.updated_by
        row.save(update_fields=["assign_type", "assigned_team", "assigned_user_ids", "assigned_user", "updated_by", "updated_at"])
        restored_counts["leads"] += 1

    deal_rows = {row.id: row for row in CrmDeal.objects.filter(organization=org, is_deleted=False, id__in=[item.get("id") for item in (payload.get("deals") or [])]).select_related("assigned_user")}
    for item in payload.get("deals") or []:
        row = deal_rows.get(item.get("id"))
        if not row:
            continue
        row.assigned_team = str(item.get("assigned_team") or "")
        row.assigned_user_ids = _crm_clean_user_id_list(item.get("assigned_user_ids"))
        row.assigned_user_id = _coerce_positive_int(item.get("assigned_user_id")) or None
        row.updated_by = performed_by if performed_by else row.updated_by
        row.save(update_fields=["assigned_team", "assigned_user_ids", "assigned_user", "updated_by", "updated_at"])
        restored_counts["deals"] += 1

    order_rows = {row.id: row for row in CrmSalesOrder.objects.filter(organization=org, is_deleted=False, id__in=[item.get("id") for item in (payload.get("sales_orders") or [])]).select_related("assigned_user")}
    for item in payload.get("sales_orders") or []:
        row = order_rows.get(item.get("id"))
        if not row:
            continue
        row.assigned_user_id = _coerce_positive_int(item.get("assigned_user_id")) or None
        row.updated_by = performed_by if performed_by else row.updated_by
        row.save(update_fields=["assigned_user", "updated_by", "updated_at"])
        restored_counts["sales_orders"] += 1

    meeting_rows = {row.id: row for row in CrmMeeting.objects.filter(organization=org, is_deleted=False, id__in=[item.get("id") for item in (payload.get("meetings") or [])]).select_related("updated_by")}
    for item in payload.get("meetings") or []:
        row = meeting_rows.get(item.get("id"))
        if not row:
            continue
        row.owner_user_ids = _crm_clean_user_id_list(item.get("owner_user_ids"))
        row.owner_names = str(item.get("owner_names") or "")
        row.updated_by = performed_by if performed_by else row.updated_by
        row.save(update_fields=["owner_user_ids", "owner_names", "updated_by", "updated_at"])
        restored_counts["meetings"] += 1

    snapshot.reverted_at = timezone.now()
    snapshot.save(update_fields=["reverted_at", "updated_at"])
    return {"restored": True, "summary": restored_counts}


def _crm_collect_contacts_linked_to_contact(row: CrmContact):
    company = str(getattr(row, "company", "") or "").strip().lower()
    name = str(getattr(row, "name", "") or "").strip().lower()
    email = str(getattr(row, "email", "") or "").strip().lower()
    phone = str(getattr(row, "phone", "") or "").strip()
    phone_code = str(getattr(row, "phone_country_code", "") or "+91").strip() or "+91"
    leads = []
    for lead in CrmLead.objects.filter(organization=row.organization, is_deleted=False).select_related("assigned_user", "created_by").order_by("-created_at", "-id"):
        lead_company = str(getattr(lead, "company", "") or "").strip().lower()
        lead_name = str(getattr(lead, "lead_name", "") or "").strip().lower()
        lead_phone = str(getattr(lead, "phone", "") or "").strip()
        if company and (lead_company == company or lead_name == company):
            leads.append(lead)
            continue
        if name and (lead_name == name or lead_company == name):
            leads.append(lead)
            continue
        if phone and lead_phone == phone:
            leads.append(lead)
            continue
    if email:
        # Email is stored on contacts, but leads do not currently persist it.
        # Keep the helper ready for future schema support without exposing stale assumptions.
        pass
    return leads


def _serialize_crm_deal(row: CrmDeal):
    crm_reference_id = str(row.crm_reference_id or "").strip()
    if not crm_reference_id and row.lead_id:
        crm_reference_id = str(getattr(row.lead, "crm_reference_id", "") or "").strip()
    assigned_user_name = ""
    if row.assigned_user_id and row.assigned_user:
        assigned_user_name = str(row.assigned_user.get_full_name() or row.assigned_user.email or "").strip()
    return {
        "id": row.id,
        "crm_reference_id": crm_reference_id,
        "lead_id": row.lead_id,
        "deal_name": row.deal_name,
        "company": row.company,
        "phone": row.phone,
        "deal_value": float(row.deal_value or 0),
        "won_amount_final": float(row.won_amount_final or 0),
        "stage": row.stage,
        "status": row.status,
        "assigned_user_id": row.assigned_user_id,
        "assigned_user_name": assigned_user_name,
        "assigned_user_ids": _crm_clean_user_id_list(row.assigned_user_ids),
        "assigned_team": row.assigned_team,
        "is_deleted": bool(row.is_deleted),
        "deleted_at": row.deleted_at.isoformat() if row.deleted_at else None,
        "created_by_id": row.created_by_id,
        "created_by_name": _get_org_user_display_name(row.created_by) if row.created_by_id else "",
        "updated_by_id": row.updated_by_id,
        "updated_by_name": _get_org_user_display_name(row.updated_by) if row.updated_by_id else "",
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "updated_at": row.updated_at.isoformat() if row.updated_at else "",
    }


def _serialize_crm_sales_order(row: CrmSalesOrder):
    crm_reference_id = str(row.crm_reference_id or "").strip()
    if not crm_reference_id and row.deal_id:
        deal_ref = str(getattr(row.deal, "crm_reference_id", "") or "").strip()
        lead_ref = str(getattr(getattr(row.deal, "lead", None), "crm_reference_id", "") or "").strip()
        crm_reference_id = deal_ref or lead_ref
    payload = _crm_sales_order_payload_dict(row)
    items = payload.get("items") if isinstance(payload.get("items"), list) else []
    totals = _crm_sales_order_totals(items)
    issue_date = str(payload.get("issueDate") or payload.get("issue_date") or "").strip()
    if not issue_date:
        issue_date = _crm_issue_date_from_order_id(row.order_id)
    if not issue_date and row.created_at:
        issue_date = row.created_at.date().isoformat()
    return {
        "id": row.id,
        "crm_reference_id": crm_reference_id,
        "deal_id": row.deal_id,
        "order_id": row.order_id,
        "customer_name": row.customer_name,
        "company": row.company,
        "phone": row.phone,
        "amount": float(row.amount or 0),
        "products": payload,
        "items": items,
        "quantity": int(row.quantity or 0),
        "price": float(row.price or 0),
        "tax": float(row.tax or 0),
        "total_amount": float(row.total_amount or 0),
        "status": row.status,
        "payment_status": row.payment_status,
        "paid_amount": float(row.paid_amount or 0),
        "payment_mode": row.payment_mode or "",
        "payment_date": row.payment_date.isoformat() if row.payment_date else "",
        "transaction_id": row.transaction_id or "",
        "balance_amount": float(max(Decimal("0"), (row.total_amount or Decimal("0")) - Decimal(str(row.paid_amount or 0)))),
        "issue_date": issue_date,
        "due_date": str(payload.get("dueDate") or payload.get("due_date") or ""),
        "gst_template_id": str(payload.get("gstTemplateId") or payload.get("gst_template_id") or ""),
        "billing_template_id": str(payload.get("billingTemplateId") or payload.get("billing_template_id") or ""),
        "salesperson": str(payload.get("salesperson") or ""),
        "customer_gstin": str(payload.get("customerGstin") or payload.get("customer_gstin") or ""),
        "billing_address": str(payload.get("billingAddress") or payload.get("billing_address") or ""),
        "notes": str(payload.get("notes") or ""),
        "terms_text": str(payload.get("termsText") or payload.get("terms_text") or ""),
        "payment_status_notes": str(payload.get("paymentStatusNotes") or payload.get("payment_status_notes") or ""),
        "source_deal_id": str(payload.get("sourceDealId") or payload.get("source_deal_id") or row.deal_id or ""),
        "converted_to_invoice": bool(payload.get("convertedToInvoice") or payload.get("converted_to_invoice")),
        "converted_invoice_id": str(payload.get("convertedInvoiceId") or payload.get("converted_invoice_id") or ""),
        "subtotal": float(totals["subtotal"]),
        "tax_total": float(totals["tax_total"]),
        "grand_total": float(totals["grand_total"]),
        "assigned_user_id": row.assigned_user_id,
        "assigned_user_name": _get_org_user_display_name(row.assigned_user) if row.assigned_user_id else "",
        "is_deleted": bool(row.is_deleted),
        "deleted_at": row.deleted_at.isoformat() if row.deleted_at else None,
        "created_by_id": row.created_by_id,
        "created_by_name": _get_org_user_display_name(row.created_by) if row.created_by_id else "",
        "updated_by_id": row.updated_by_id,
        "updated_by_name": _get_org_user_display_name(row.updated_by) if row.updated_by_id else "",
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "updated_at": row.updated_at.isoformat() if row.updated_at else "",
    }


def _serialize_crm_meeting(row: CrmMeeting):
    return {
        "id": row.id,
        "crm_reference_id": str(row.crm_reference_id or "").strip(),
        "title": row.title,
        "company_or_client_name": row.company_or_client_name,
        "related_to": row.related_to,
        "meeting_date": row.meeting_date.isoformat() if row.meeting_date else "",
        "meeting_time": row.meeting_time.strftime("%H:%M") if row.meeting_time else "",
        "owner": row.owner_names,
        "owner_user_ids": _crm_clean_user_id_list(row.owner_user_ids),
        "meeting_mode": row.meeting_mode,
        "reminder_channel": row.reminder_channels if isinstance(row.reminder_channels, list) else [],
        "reminder_days": _crm_clean_int_list(row.reminder_days),
        "reminder_minutes": _crm_clean_int_list(row.reminder_minutes),
        "reminder_summary": row.reminder_summary,
        "status": row.status,
        "is_deleted": bool(row.is_deleted),
        "deleted_at": row.deleted_at.isoformat() if row.deleted_at else None,
        "created_by_id": row.created_by_id,
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "updated_at": row.updated_at.isoformat() if row.updated_at else "",
    }


def _crm_meeting_recipient_emails(org: Organization, row: CrmMeeting):
    emails = []
    owner_ids = _crm_clean_user_id_list(row.owner_user_ids)
    if owner_ids:
        users = (
        OrganizationUser.objects
        .filter(organization=org, user_id__in=owner_ids, is_active=True, is_deleted=False)
        .select_related("user")
        )
        for membership in users:
            email = str(getattr(membership.user, "email", "") or "").strip().lower()
            if email:
                emails.append(email)
    if not emails and row.created_by_id:
        fallback_email = str(getattr(row.created_by, "email", "") or "").strip().lower()
        if fallback_email:
            emails.append(fallback_email)
    return list(dict.fromkeys(emails))


def _dispatch_due_crm_meeting_reminders(org: Organization = None, now=None, limit_per_run: int = 500):
    now_value = now or timezone.now()
    queryset = CrmMeeting.objects.filter(is_deleted=False).exclude(status__in=["Completed", "Cancelled", "Missed"])
    if org:
        queryset = queryset.filter(organization=org)
    rows = list(queryset.select_related("organization", "created_by").order_by("meeting_date", "meeting_time", "id")[:limit_per_run])
    if not rows:
        return {"checked": 0, "sent": 0}

    sent = 0
    for row in rows:
        if not row.meeting_date or not row.meeting_time:
            continue
        meeting_dt = datetime.combine(row.meeting_date, row.meeting_time)
        if timezone.is_naive(meeting_dt):
            meeting_dt = timezone.make_aware(meeting_dt, timezone.get_current_timezone())
        channels = {str(item or "").strip().lower() for item in (row.reminder_channels if isinstance(row.reminder_channels, list) else [])}
        if "email" not in channels:
            continue
        day_offsets = _crm_clean_int_list(row.reminder_days)
        minute_offsets = _crm_clean_int_list(row.reminder_minutes)
        schedule_points = []
        schedule_points.extend([("day", days, meeting_dt - timedelta(days=days)) for days in day_offsets])
        schedule_points.extend([("minute", minutes, meeting_dt - timedelta(minutes=minutes)) for minutes in minute_offsets])
        if not schedule_points:
            continue
        sent_map = row.reminder_email_sent_map if isinstance(row.reminder_email_sent_map, dict) else {}
        updated_map = dict(sent_map)
        recipients = _crm_meeting_recipient_emails(row.organization, row)
        if not recipients:
            continue
        for offset_type, offset_value, trigger_dt in schedule_points:
            key = f"{offset_type}:{offset_value}"
            if updated_map.get(key):
                continue
            if now_value < trigger_dt:
                continue
            if now_value > (meeting_dt + timedelta(minutes=15)):
                continue
            subject = f"Meeting Reminder: {row.title or 'Meeting'}"
            context = {
                "meeting_title": row.title or "Meeting",
                "company_or_client_name": row.company_or_client_name or "-",
                "related_to": row.related_to or "-",
                "meeting_date": row.meeting_date.isoformat() if row.meeting_date else "",
                "meeting_time": row.meeting_time.strftime("%I:%M %p") if row.meeting_time else "",
                "meeting_mode": row.meeting_mode or "-",
                "offset_type": "days" if offset_type == "day" else "minutes",
                "offset_value": offset_value,
                "owner": row.owner_names or "-",
            }
            mail_sent = send_templated_email(recipients, subject, "emails/crm_meeting_reminder.txt", context)
            if mail_sent:
                updated_map[key] = now_value.isoformat()
                sent += 1
        if updated_map != sent_map:
            row.reminder_email_sent_map = updated_map
            row.save(update_fields=["reminder_email_sent_map", "updated_at"])
    return {"checked": len(rows), "sent": sent}


def _crm_report_parse_date(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    parsed = _parse_iso_date(raw)
    if parsed:
        return parsed
    iso_candidate = raw.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(iso_candidate).date()
    except ValueError:
        pass
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _crm_report_issue_date_from_doc_no(doc_no):
    raw_doc_no = str(doc_no or "").strip()
    if not raw_doc_no:
        return None
    dashed_match = re.match(r"^(?:SO|EST|INV)-(\d{2})-(\d{2})-(\d{4})-\d+$", raw_doc_no, re.IGNORECASE)
    if dashed_match:
        try:
            return date(int(dashed_match.group(3)), int(dashed_match.group(2)), int(dashed_match.group(1)))
        except ValueError:
            return None
    compact_match = re.match(r"^(?:SO|EST|INV)-(\d{2})(\d{2})(\d{4})-\d+$", raw_doc_no, re.IGNORECASE)
    if compact_match:
        try:
            return date(int(compact_match.group(3)), int(compact_match.group(2)), int(compact_match.group(1)))
        except ValueError:
            return None
    return None


def _crm_report_extract_doc_issue_date(row):
    if not isinstance(row, dict):
        return None
    for key in ("issueDate", "issue_date", "createdDate", "created_date", "invoiceDate", "invoice_date", "date"):
        parsed = _crm_report_parse_date(row.get(key))
        if parsed:
            return parsed
    return _crm_report_issue_date_from_doc_no(
        row.get("docNo")
        or row.get("doc_no")
        or row.get("invoiceNo")
        or row.get("invoice_no")
        or row.get("estimateNo")
        or row.get("estimate_no")
    )


def _crm_report_extract_source_sales_order_refs(row):
    if not isinstance(row, dict):
        return "", ""
    source_id = str(
        row.get("sourceSalesOrderId")
        or row.get("source_sales_order_id")
        or ""
    ).strip()
    source_no = str(
        row.get("sourceSalesOrderNo")
        or row.get("source_sales_order_no")
        or ""
    ).strip()
    return source_id, source_no.lower()


def _crm_report_default_group_label(group_by):
    return "Unassigned User" if group_by == "user" else "Direct / Users"


def _crm_report_group_label_for_lead(row: CrmLead, group_by: str):
    if group_by == "team":
        assigned_team = str(getattr(row, "assigned_team", "") or "").strip()
        if assigned_team:
            return assigned_team
        return _crm_report_default_group_label(group_by)
    assigned_name = _get_org_user_display_name(getattr(row, "assigned_user", None))
    if assigned_name:
        return assigned_name
    created_name = _get_org_user_display_name(getattr(row, "created_by", None))
    if created_name:
        return created_name
    return _crm_report_default_group_label(group_by)


def _crm_report_group_label_for_sales_order(row: CrmSalesOrder, group_by: str):
    if group_by == "team":
        assigned_team = str(getattr(getattr(row, "deal", None), "assigned_team", "") or "").strip()
        if assigned_team:
            return assigned_team
        return _crm_report_default_group_label(group_by)
    assigned_name = _get_org_user_display_name(getattr(row, "assigned_user", None))
    if assigned_name:
        return assigned_name
    created_name = _get_org_user_display_name(getattr(row, "created_by", None))
    if created_name:
        return created_name
    return _crm_report_default_group_label(group_by)


def _crm_report_group_label_for_deal(row: CrmDeal, group_by: str):
    if group_by == "team":
        assigned_team = str(getattr(row, "assigned_team", "") or "").strip()
        if assigned_team:
            return assigned_team
        return _crm_report_default_group_label(group_by)
    assigned_name = _get_org_user_display_name(getattr(row, "assigned_user", None))
    if assigned_name:
        return assigned_name
    created_name = _get_org_user_display_name(getattr(row, "created_by", None))
    if created_name:
        return created_name
    return _crm_report_default_group_label(group_by)


def _crm_report_sales_order_issue_date(row: CrmSalesOrder):
    payload = _crm_sales_order_payload_dict(row)
    issue_date = _crm_report_parse_date(payload.get("issueDate") or payload.get("issue_date"))
    if issue_date:
        return issue_date
    issue_from_order_id = _crm_report_parse_date(_crm_issue_date_from_order_id(row.order_id))
    if issue_from_order_id:
        return issue_from_order_id
    if row.created_at:
        return row.created_at.date()
    return None


def _crm_report_in_range(value, start_date, end_date):
    if not value:
        return False
    return start_date <= value <= end_date


def _crm_report_build_payload(request_user: User, org: Organization, start_date: date, end_date: date, group_by: str):
    visible_leads = [
        row
        for row in (
            CrmLead.objects
            # Extra safety: exclude inconsistent soft-deleted rows.
            .filter(organization=org, is_deleted=False, deleted_at__isnull=True)
            .select_related("assigned_user", "created_by")
            .order_by("-created_at", "-id")
        )
        if _crm_can_view_row(request_user, org, row)
    ]
    visible_deals = [
        row
        for row in (
            CrmDeal.objects
            .filter(organization=org, is_deleted=False, deleted_at__isnull=True)
            .select_related("assigned_user", "created_by")
            .order_by("-created_at", "-id")
        )
        if _crm_can_view_row(request_user, org, row)
    ]
    visible_sales_orders = [
        row
        for row in (
            CrmSalesOrder.objects
            .filter(organization=org, is_deleted=False, deleted_at__isnull=True)
            .select_related("assigned_user", "created_by", "deal")
            .order_by("-created_at", "-id")
        )
        if _crm_can_view_row(request_user, org, row)
    ]

    sales_order_by_id = {}
    sales_order_by_no = {}
    for row in visible_sales_orders:
        row_id = str(row.id or "").strip()
        row_no = str(row.order_id or "").strip().lower()
        if row_id:
            sales_order_by_id[row_id] = row
        if row_no:
            sales_order_by_no[row_no] = row

    has_unrestricted_doc_access = bool(
        _crm_is_admin(request_user, org)
        or _crm_has_product_view_access(request_user)
        or _crm_has_unrestricted_row_access(request_user, org)
    )

    workspace = _get_accounts_workspace(org)
    workspace_data = _normalize_accounts_workspace(workspace.data)
    estimate_rows = workspace_data.get("estimates") if isinstance(workspace_data.get("estimates"), list) else []
    invoice_rows = workspace_data.get("invoices") if isinstance(workspace_data.get("invoices"), list) else []

    def _empty_group_row(label):
        return {
            "group_name": label,
            "total_leads": 0,
            "new_leads": 0,
            "pending_leads": 0,
            "completed_leads": 0,
            "onhold_leads": 0,
            "sales_orders": 0,
            "estimate_converted": 0,
            "invoice_converted": 0,
            "pipeline_value": Decimal("0"),
            "won_amount": Decimal("0"),
            "sales_order_value": Decimal("0"),
        }

    group_rows_map = {}

    def _group_bucket(label):
        normalized_label = str(label or "").strip() or _crm_report_default_group_label(group_by)
        if normalized_label not in group_rows_map:
            group_rows_map[normalized_label] = _empty_group_row(normalized_label)
        return group_rows_map[normalized_label]

    lead_details = []
    range_leads = []
    pipeline_leads = []
    for row in visible_leads:
        created_date = row.created_at.date() if row.created_at else None
        if not _crm_report_in_range(created_date, start_date, end_date):
            continue
        range_leads.append(row)
        normalized_status = str(row.status or "").strip().lower()
        normalized_stage = str(row.stage or "").strip().lower()
        is_new = normalized_stage == "new" or normalized_status == "open"
        is_pending = normalized_status in {"open", "onhold", "pending"}
        is_completed = normalized_status in {"closed", "completed", "won", "converted"}
        is_onhold = normalized_status == "onhold"
        group_name = _crm_report_group_label_for_lead(row, group_by)
        bucket = _group_bucket(group_name)
        bucket["total_leads"] += 1
        if is_new:
            bucket["new_leads"] += 1
        if is_pending:
            bucket["pending_leads"] += 1
        if is_completed:
            bucket["completed_leads"] += 1
        if is_onhold:
            bucket["onhold_leads"] += 1
        # Pipeline value should reflect only active opportunities (open/pending/onhold),
        # not already closed/completed leads.
        if is_pending and not is_completed:
            bucket["pipeline_value"] += _crm_to_decimal(row.lead_amount)
            pipeline_leads.append(row)
        lead_details.append(
            {
                "date": created_date.isoformat() if created_date else "",
                "crm_reference_id": str(row.crm_reference_id or "").strip(),
                "lead_name": str(row.lead_name or "").strip(),
                "company": str(row.company or "").strip(),
                "phone": str(row.phone or "").strip(),
                "lead_amount": float(_crm_to_decimal(row.lead_amount)),
                "status": str(row.status or "").strip().title() or "-",
                "priority": str(row.priority or "").strip().title() or "-",
                "lead_source": str(row.lead_source or "").strip(),
                "assigned_to": _get_org_user_display_name(row.assigned_user) if row.assigned_user_id else "",
                "assigned_team": str(row.assigned_team or "").strip(),
                "created_by": _get_org_user_display_name(row.created_by) if row.created_by_id else "",
                "group_name": group_name,
            }
        )

    range_sales_orders = []
    for row in visible_sales_orders:
        issue_date = _crm_report_sales_order_issue_date(row)
        if not _crm_report_in_range(issue_date, start_date, end_date):
            continue
        range_sales_orders.append(row)
        group_name = _crm_report_group_label_for_sales_order(row, group_by)
        bucket = _group_bucket(group_name)
        bucket["sales_orders"] += 1
        bucket["sales_order_value"] += _crm_to_decimal(row.total_amount)

    range_won_deals = []
    for row in visible_deals:
        created_date = row.created_at.date() if row.created_at else None
        if not _crm_report_in_range(created_date, start_date, end_date):
            continue
        if str(row.status or "").strip().lower() != "won":
            continue
        range_won_deals.append(row)
        group_name = _crm_report_group_label_for_deal(row, group_by)
        bucket = _group_bucket(group_name)
        bucket["won_amount"] += _crm_to_decimal(row.won_amount_final or row.deal_value)

    converted_estimate_count = 0
    converted_invoice_count = 0

    def _process_conversion_docs(rows, target_key):
        nonlocal converted_estimate_count, converted_invoice_count
        for row in rows:
            if not isinstance(row, dict):
                continue
            issue_date = _crm_report_extract_doc_issue_date(row)
            if not _crm_report_in_range(issue_date, start_date, end_date):
                continue
            source_id, source_no = _crm_report_extract_source_sales_order_refs(row)
            converted_flag = _to_bool(row.get("convertedFromSalesOrder") or row.get("converted_from_sales_order"))
            has_conversion_ref = bool(source_id or source_no)
            if not converted_flag and not has_conversion_ref:
                continue
            source_sales_order = sales_order_by_id.get(source_id) if source_id else None
            if not source_sales_order and source_no:
                source_sales_order = sales_order_by_no.get(source_no)
            if not has_unrestricted_doc_access and not source_sales_order:
                continue
            group_name = (
                _crm_report_group_label_for_sales_order(source_sales_order, group_by)
                if source_sales_order
                else _crm_report_default_group_label(group_by)
            )
            bucket = _group_bucket(group_name)
            bucket[target_key] += 1
            if target_key == "estimate_converted":
                converted_estimate_count += 1
            elif target_key == "invoice_converted":
                converted_invoice_count += 1

    _process_conversion_docs(estimate_rows, "estimate_converted")
    _process_conversion_docs(invoice_rows, "invoice_converted")

    summary_new_leads = sum(
        1
        for row in range_leads
        if str(row.stage or "").strip().lower() == "new" or str(row.status or "").strip().lower() == "open"
    )
    summary_pending_leads = sum(
        1
        for row in range_leads
        if str(row.status or "").strip().lower() in {"open", "onhold", "pending"}
    )
    summary_completed_leads = sum(
        1
        for row in range_leads
        if str(row.status or "").strip().lower() in {"closed", "completed", "won", "converted"}
    )
    summary_onhold_leads = sum(
        1
        for row in range_leads
        if str(row.status or "").strip().lower() == "onhold"
    )
    pipeline_value = sum((_crm_to_decimal(row.lead_amount) for row in pipeline_leads), Decimal("0"))
    won_amount_total = sum((_crm_to_decimal(row.won_amount_final or row.deal_value) for row in range_won_deals), Decimal("0"))
    sales_order_total = sum((_crm_to_decimal(row.total_amount) for row in range_sales_orders), Decimal("0"))

    group_rows = []
    for label, row in group_rows_map.items():
        group_rows.append(
            {
                "group_name": label,
                "total_leads": int(row["total_leads"]),
                "new_leads": int(row["new_leads"]),
                "pending_leads": int(row["pending_leads"]),
                "completed_leads": int(row["completed_leads"]),
                "onhold_leads": int(row["onhold_leads"]),
                "sales_orders": int(row["sales_orders"]),
                "estimate_converted": int(row["estimate_converted"]),
                "invoice_converted": int(row["invoice_converted"]),
                "pipeline_value": float(row["pipeline_value"]),
                "won_amount": float(row["won_amount"]),
                "sales_order_value": float(row["sales_order_value"]),
            }
        )
    group_rows.sort(key=lambda item: (item["total_leads"], item["sales_orders"], item["pipeline_value"]), reverse=True)
    lead_details.sort(key=lambda item: (item.get("date") or "", item.get("lead_name") or ""), reverse=True)

    summary = {
        "total_leads": len(range_leads),
        "new_leads": summary_new_leads,
        "pending_leads": summary_pending_leads,
        "completed_leads": summary_completed_leads,
        "onhold_leads": summary_onhold_leads,
        "sales_orders": len(range_sales_orders),
        "converted_estimates": converted_estimate_count,
        "converted_invoices": converted_invoice_count,
        "won_deals": len(range_won_deals),
        "pipeline_value": float(pipeline_value),
        "sales_order_value": float(sales_order_total),
        "won_amount": float(won_amount_total),
    }

    chart_items = [
        {"key": "new_leads", "label": "New Leads", "value": int(summary["new_leads"])},
        {"key": "pending_leads", "label": "Pending Leads", "value": int(summary["pending_leads"])},
        {"key": "completed_leads", "label": "Completed Leads", "value": int(summary["completed_leads"])},
        {"key": "sales_orders", "label": "Sales Orders", "value": int(summary["sales_orders"])},
        {"key": "converted_estimates", "label": "Estimate Converted", "value": int(summary["converted_estimates"])},
        {"key": "converted_invoices", "label": "Invoice Converted", "value": int(summary["converted_invoices"])},
    ]

    return {
        "group_by": group_by,
        "period": {
            "from_date": start_date.isoformat(),
            "to_date": end_date.isoformat(),
            "days": (end_date - start_date).days + 1,
            "generated_at": timezone.now().isoformat(),
        },
        "summary": summary,
        "chart_items": chart_items,
        "group_rows": group_rows,
        "lead_details": lead_details,
    }


def _crm_report_available_range(org: Organization):
    today = timezone.localdate()
    if not org:
        return {"from_date": today.isoformat(), "to_date": today.isoformat()}
    lead_agg = (
        CrmLead.objects
        .filter(organization=org, is_deleted=False, deleted_at__isnull=True)
        .aggregate(min_created=Min("created_at"), max_created=Max("created_at"))
    )
    deal_agg = (
        CrmDeal.objects
        .filter(organization=org, is_deleted=False, deleted_at__isnull=True)
        .aggregate(min_created=Min("created_at"), max_created=Max("created_at"))
    )
    so_agg = (
        CrmSalesOrder.objects
        .filter(organization=org, is_deleted=False, deleted_at__isnull=True)
        .aggregate(min_created=Min("created_at"), max_created=Max("created_at"))
    )

    min_candidates = [lead_agg.get("min_created"), deal_agg.get("min_created"), so_agg.get("min_created")]
    max_candidates = [lead_agg.get("max_created"), deal_agg.get("max_created"), so_agg.get("max_created")]
    min_dt = min([dt for dt in min_candidates if dt], default=None)
    max_dt = max([dt for dt in max_candidates if dt], default=None)

    min_date = min_dt.date() if min_dt else today
    max_date = max_dt.date() if max_dt else today
    if min_date > max_date:
        min_date, max_date = max_date, min_date
    return {"from_date": min_date.isoformat(), "to_date": max_date.isoformat()}


def _crm_report_available_range_dates(org: Organization):
    today = timezone.localdate()
    payload = _crm_report_available_range(org)
    min_raw = str(payload.get("from_date") or "").strip()
    max_raw = str(payload.get("to_date") or "").strip()
    min_date = _crm_report_parse_date(min_raw) or today
    max_date = _crm_report_parse_date(max_raw) or today
    if min_date > max_date:
        min_date, max_date = max_date, min_date
    return min_date, max_date


def _ba_format_ddmmyyyy(value):
    if not value:
        return ""
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    parsed = _crm_report_parse_date(str(value))
    if parsed:
        return parsed.strftime("%d-%m-%Y")
    raw = str(value).strip()
    return raw


def _crm_report_xlsx_response(org: Organization, request_user: User, payload):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return JsonResponse({"detail": "excel_dependency_missing", "package": "openpyxl"}, status=500)

    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    period = payload.get("period") if isinstance(payload.get("period"), dict) else {}
    group_rows = payload.get("group_rows") if isinstance(payload.get("group_rows"), list) else []
    lead_rows = payload.get("lead_details") if isinstance(payload.get("lead_details"), list) else []
    group_by = str(payload.get("group_by") or "user").strip().lower()

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    theme_primary_hex = _ba_effective_theme_primary_hex(org)
    theme_primary_rgb = theme_primary_hex.lstrip("#").upper()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=theme_primary_rgb or "1F4E79")
    label_font = Font(bold=True, color="1F2937")

    def write_kv_sheet(ws, title, rows):
        ws.append([title])
        ws["A1"].font = Font(bold=True, size=14, color="111827")
        ws.append([])
        ws.append(["Field", "Value"])
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for label, value in rows:
            ws.append([label, value])
            ws.cell(row=ws.max_row, column=1).font = label_font
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 42

    write_kv_sheet(
        ws_summary,
        f"CRM Report ({'User' if group_by == 'user' else 'Team'})",
        [
            ("Organization", str(getattr(org, "name", "") or "Organization").strip()),
            ("From Date", _ba_format_ddmmyyyy(period.get("from_date"))),
            ("To Date", _ba_format_ddmmyyyy(period.get("to_date"))),
            ("Days", int(period.get("days") or 0)),
            ("Generated At", str(period.get("generated_at") or "")),
            ("Total Leads", int(summary.get("total_leads") or 0)),
            ("New Leads", int(summary.get("new_leads") or 0)),
            ("Pending Leads", int(summary.get("pending_leads") or 0)),
            ("Completed Leads", int(summary.get("completed_leads") or 0)),
            ("Onhold Leads", int(summary.get("onhold_leads") or 0)),
            ("Sales Orders", int(summary.get("sales_orders") or 0)),
            ("Estimate Converted", int(summary.get("converted_estimates") or 0)),
            ("Invoice Converted", int(summary.get("converted_invoices") or 0)),
            ("Pipeline Value", float(summary.get("pipeline_value") or 0)),
            ("Won Amount", float(summary.get("won_amount") or 0)),
            ("Sales Order Value", float(summary.get("sales_order_value") or 0)),
        ],
    )

    ws_perf = wb.create_sheet("Performance")
    perf_headers = ["User / Team", "Total", "New", "Pending", "Completed", "Onhold", "SO", "Estimate", "Invoice", "Pipeline", "Won", "SO Value"]
    ws_perf.append(perf_headers)
    for cell in ws_perf[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for row in group_rows:
        ws_perf.append(
            [
                str(row.get("group_name") or ""),
                int(row.get("total_leads") or 0),
                int(row.get("new_leads") or 0),
                int(row.get("pending_leads") or 0),
                int(row.get("completed_leads") or 0),
                int(row.get("onhold_leads") or 0),
                int(row.get("sales_orders") or 0),
                int(row.get("estimate_converted") or 0),
                int(row.get("invoice_converted") or 0),
                float(row.get("pipeline_value") or 0),
                float(row.get("won_amount") or 0),
                float(row.get("sales_order_value") or 0),
            ]
        )
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws_perf.column_dimensions[col].width = 16 if col != "A" else 34

    ws_leads = wb.create_sheet("Lead Details")
    lead_headers = ["Date", "CRM ID", "Lead Name", "Company", "Phone", "Amount", "Status", "Priority", "Lead Source", "Assigned To", "Assigned Team", "Created By", "Group"]
    ws_leads.append(lead_headers)
    for cell in ws_leads[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for row in lead_rows:
        ws_leads.append(
            [
                str(row.get("date") or ""),
                str(row.get("crm_reference_id") or ""),
                str(row.get("lead_name") or ""),
                str(row.get("company") or ""),
                str(row.get("phone") or ""),
                float(row.get("lead_amount") or 0),
                str(row.get("status") or ""),
                str(row.get("priority") or ""),
                str(row.get("lead_source") or ""),
                str(row.get("assigned_to") or ""),
                str(row.get("assigned_team") or ""),
                str(row.get("created_by") or ""),
                str(row.get("group_name") or ""),
            ]
        )
    ws_leads.freeze_panes = "A2"
    ws_leads.column_dimensions["A"].width = 12
    ws_leads.column_dimensions["B"].width = 18
    ws_leads.column_dimensions["C"].width = 22
    ws_leads.column_dimensions["D"].width = 22
    ws_leads.column_dimensions["E"].width = 14
    ws_leads.column_dimensions["F"].width = 12
    ws_leads.column_dimensions["G"].width = 12
    ws_leads.column_dimensions["H"].width = 12
    ws_leads.column_dimensions["I"].width = 18
    ws_leads.column_dimensions["J"].width = 18
    ws_leads.column_dimensions["K"].width = 16
    ws_leads.column_dimensions["L"].width = 16
    ws_leads.column_dimensions["M"].width = 24

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"crm_report_{period.get('from_date')}_to_{period.get('to_date')}.xlsx".replace(":", "-")
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _crm_report_pdf_response(org: Organization, request_user: User, payload):
    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    period = payload.get("period") if isinstance(payload.get("period"), dict) else {}
    group_rows = payload.get("group_rows") if isinstance(payload.get("group_rows"), list) else []
    lead_rows = payload.get("lead_details") if isinstance(payload.get("lead_details"), list) else []
    chart_items = payload.get("chart_items") if isinstance(payload.get("chart_items"), list) else []
    group_by = str(payload.get("group_by") or "user").strip().lower()

    company_name = str(org.name or "Organization").strip() or "Organization"
    generated_at = timezone.localtime(timezone.now())
    from_date = _ba_format_ddmmyyyy(period.get("from_date")) or "-"
    to_date = _ba_format_ddmmyyyy(period.get("to_date")) or "-"

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    left = 16 * mm
    right = width - 16 * mm
    page_bottom = 14 * mm
    row_gap = 4 * mm
    money_font = _ba_get_unicode_pdf_font() or "Helvetica"
    logo_image = _ba_resolve_org_logo_image(org, request_user)
    theme_primary_hex = _ba_effective_theme_primary_hex(org)
    theme_primary_rgb = _ba_hex_to_rgb01(theme_primary_hex)
    theme_primary_card_rgb = _ba_blend_rgb(theme_primary_rgb, (1.0, 1.0, 1.0), 0.90)
    metric_cards = [
        ("Total Leads", str(summary.get("total_leads", 0))),
        ("New Leads", str(summary.get("new_leads", 0))),
        ("Pending Leads", str(summary.get("pending_leads", 0))),
        ("Sales Orders", str(summary.get("sales_orders", 0))),
        ("Estimate Conv.", str(summary.get("converted_estimates", 0))),
        ("Invoice Conv.", str(summary.get("converted_invoices", 0))),
        ("Pipeline Value", _format_pdf_inr(summary.get("pipeline_value", 0))),
        ("Won Amount", _format_pdf_inr(summary.get("won_amount", 0))),
    ]

    y = height - 18 * mm

    def draw_page_header(include_title=True):
        nonlocal y
        y = height - 18 * mm
        logo_height = _ba_draw_pdf_logo(pdf, logo_image, left, y + 1 * mm, 22 * mm, 12 * mm)
        if include_title:
            pdf.setFillColorRGB(0.07, 0.12, 0.2)
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(left, y, "CRM Performance Report")
            pdf.drawRightString(right, y, f"Group By: {'User' if group_by == 'user' else 'Team'}")
            pdf.setStrokeColorRGB(0.82, 0.85, 0.91)
            pdf.setLineWidth(0.7)
            pdf.line(left, y - 4.5 * mm, right, y - 4.5 * mm)
            info_x = left + (24 * mm if logo_height else 0)
            pdf.setFont("Helvetica-Bold", 11)
            # Extra spacing below the title rule.
            pdf.drawString(info_x, y - 11 * mm, company_name)
            pdf.setFont("Helvetica", 10)
            pdf.drawString(info_x, y - 15 * mm, f"Period: {from_date} to {to_date}")
            pdf.drawString(info_x, y - 19 * mm, f"Generated: {generated_at.strftime('%d-%m-%Y %I:%M %p')}")
            y = y - (max(logo_height, 18 * mm) + 10 * mm)
        else:
            y = y - (max(logo_height, 6 * mm) + 4 * mm)

    def ensure_space(required_height):
        nonlocal y
        if y - required_height < page_bottom:
            pdf.showPage()
            draw_page_header(include_title=False)

    draw_page_header(include_title=True)

    card_columns = 4
    card_gap = 3 * mm
    card_width = (right - left - (card_gap * (card_columns - 1))) / card_columns
    card_height = 16 * mm
    for index, (label, value) in enumerate(metric_cards):
        row_index = index // card_columns
        col_index = index % card_columns
        card_x = left + col_index * (card_width + card_gap)
        card_y = y - row_index * (card_height + card_gap)
        ensure_space(card_height + 2 * mm)
        pdf.setFillColorRGB(*theme_primary_card_rgb)
        pdf.roundRect(card_x, card_y - card_height, card_width, card_height, 2 * mm, stroke=0, fill=1)
        pdf.setFillColorRGB(0.18, 0.27, 0.41)
        pdf.setFont("Helvetica", 8)
        pdf.drawString(card_x + 2 * mm, card_y - 5 * mm, label)
        if label in {"Pipeline Value", "Won Amount"}:
            pdf.setFont(money_font, 10)
        else:
            pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(card_x + 2 * mm, card_y - 11 * mm, str(value or "-"))
    y = y - (2 * (card_height + card_gap) + 3 * mm)

    ensure_space(38 * mm)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.setFillColorRGB(0.07, 0.12, 0.2)
    pdf.drawString(left, y, "Performance Snapshot")
    y -= 5 * mm
    max_chart_value = max([int(item.get("value") or 0) for item in chart_items] + [1])
    for item in chart_items[:6]:
        ensure_space(7 * mm)
        label = str(item.get("label") or "-")
        value = int(item.get("value") or 0)
        bar_x = left + 45 * mm
        bar_width = 84 * mm
        bar_height = 4 * mm
        fill_width = bar_width * (value / max_chart_value if max_chart_value else 0)
        pdf.setFont("Helvetica", 9)
        pdf.setFillColorRGB(0.12, 0.16, 0.25)
        pdf.drawString(left, y, label)
        pdf.setFillColorRGB(0.90, 0.92, 0.97)
        pdf.roundRect(bar_x, y - 3 * mm, bar_width, bar_height, 1 * mm, stroke=0, fill=1)
        pdf.setFillColorRGB(*theme_primary_rgb)
        if fill_width > 0:
            pdf.roundRect(bar_x, y - 3 * mm, fill_width, bar_height, 1 * mm, stroke=0, fill=1)
        pdf.setFillColorRGB(0.07, 0.12, 0.2)
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawRightString(right, y, str(value))
        y -= 6 * mm

    y -= 2 * mm

    # Lead Details (keep this before performance table so user performance is last section).
    ensure_space(15 * mm)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.setFillColorRGB(0.07, 0.12, 0.2)
    pdf.drawString(left, y, "Lead Details")
    y -= 5 * mm
    lead_headers = ["Date", "CRM ID", "Lead Name", "Company", "Status", "Assigned"]
    lead_col_widths = [20 * mm, 26 * mm, 44 * mm, 40 * mm, 18 * mm, 34 * mm]

    def draw_lead_header():
        nonlocal y
        x = left
        for header, col_width in zip(lead_headers, lead_col_widths):
            pdf.setFillColorRGB(*theme_primary_rgb)
            pdf.rect(x, y - 5 * mm, col_width, 5 * mm, stroke=0, fill=1)
            pdf.setFillColorRGB(1, 1, 1)
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(x + 1.0 * mm, y - 3.6 * mm, header)
            x += col_width
        y -= 5.4 * mm

    draw_lead_header()
    for row in lead_rows[:50]:
        ensure_space(5.5 * mm)
        if y - 5.5 * mm < page_bottom:
            pdf.showPage()
            draw_page_header(include_title=False)
            pdf.setFont("Helvetica-Bold", 11)
            pdf.drawString(left, y, "Lead Details")
            y -= 5 * mm
            draw_lead_header()
        values = [
            str(row.get("date") or "-"),
            str(row.get("crm_reference_id") or "-"),
            str(row.get("lead_name") or "-"),
            str(row.get("company") or "-"),
            str(row.get("status") or "-"),
            str(row.get("assigned_to") or row.get("group_name") or "-"),
        ]
        x = left
        for value, col_width in zip(values, lead_col_widths):
            pdf.setStrokeColorRGB(0.87, 0.89, 0.93)
            pdf.rect(x, y - 5 * mm, col_width, 5 * mm, stroke=1, fill=0)
            pdf.setFillColorRGB(0.09, 0.14, 0.22)
            pdf.setFont("Helvetica", 7.5)
            pdf.drawString(x + 1.0 * mm, y - 3.5 * mm, str(value)[:42])
            x += col_width
        y -= 5 * mm

    # Extra gap between lead details and the performance table title.
    y -= 6 * mm
    table_title = "User Performance Table" if group_by == "user" else "Team Performance Table"
    ensure_space(28 * mm)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left, y, table_title)
    y -= 6 * mm
    headers = ["User / Team", "Total", "Pending", "Completed", "SO", "Estimate", "Invoice"]
    col_widths = [56 * mm, 14 * mm, 18 * mm, 18 * mm, 12 * mm, 18 * mm, 18 * mm]

    def draw_table_header():
        nonlocal y
        x = left
        for header, col_width in zip(headers, col_widths):
            pdf.setFillColorRGB(*theme_primary_rgb)
            pdf.rect(x, y - 5 * mm, col_width, 5 * mm, stroke=0, fill=1)
            pdf.setFillColorRGB(1, 1, 1)
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(x + 1.2 * mm, y - 3.6 * mm, header)
            x += col_width
        y -= 5.4 * mm

    draw_table_header()
    for row in group_rows[:24]:
        ensure_space(5.5 * mm)
        if y - 5.5 * mm < page_bottom:
            pdf.showPage()
            draw_page_header(include_title=False)
            pdf.setFont("Helvetica-Bold", 11)
            pdf.drawString(left, y, table_title)
            y -= 5 * mm
            draw_table_header()
        values = [
            str(row.get("group_name") or "-"),
            str(row.get("total_leads", 0)),
            str(row.get("pending_leads", 0)),
            str(row.get("completed_leads", 0)),
            str(row.get("sales_orders", 0)),
            str(row.get("estimate_converted", 0)),
            str(row.get("invoice_converted", 0)),
        ]
        x = left
        for value, col_width in zip(values, col_widths):
            pdf.setStrokeColorRGB(0.87, 0.89, 0.93)
            pdf.rect(x, y - 5 * mm, col_width, 5 * mm, stroke=1, fill=0)
            pdf.setFillColorRGB(0.09, 0.14, 0.22)
            pdf.setFont("Helvetica", 8)
            pdf.drawString(x + 1.2 * mm, y - 3.6 * mm, str(value)[:48])
            x += col_width
        y -= 5 * mm

    pdf.setFont("Helvetica", 7)
    pdf.setFillColorRGB(0.35, 0.4, 0.48)
    pdf.drawRightString(right, 8 * mm, f"Generated by Work Zilla CRM on {generated_at.strftime('%d-%m-%Y %I:%M %p')}")
    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    ts = generated_at.strftime("%Y%m%d_%H%M%S")
    filename = f"crm_report_{from_date}_to_{to_date}_{group_by}_{ts}.pdf".replace(":", "-")
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_http_methods(["GET"])
def crm_reports(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    if not (_crm_is_admin(request.user, org) or _crm_has_product_view_access(request.user) or _crm_has_view_access(request.user, org)):
        return JsonResponse({"detail": "forbidden"}, status=403)

    today = timezone.localdate()
    preset = str(request.GET.get("preset") or "last_30_days").strip().lower()
    from_raw = str(request.GET.get("from_date") or request.GET.get("from") or "").strip()
    to_raw = str(request.GET.get("to_date") or request.GET.get("to") or "").strip()
    group_by = str(request.GET.get("group_by") or "user").strip().lower()
    if group_by not in {"user", "team"}:
        group_by = "user"

    preset_days_map = {
        "last_7_days": 7,
        "last_15_days": 15,
        "last_30_days": 30,
        "last_60_days": 60,
        "last_90_days": 90,
    }
    default_span = preset_days_map.get(preset, 30)
    default_start = today - timedelta(days=max(default_span - 1, 0))
    start_date = _crm_report_parse_date(from_raw) or default_start
    end_date = _crm_report_parse_date(to_raw) or today

    # If the selected preset range starts before CRM has any data,
    # clamp to the first/last available activity dates so the UI period label
    # doesn't show "empty" early days.
    is_custom = preset == "custom"
    available_min, available_max = _crm_report_available_range_dates(org)
    if not is_custom and not from_raw:
        start_date = max(start_date, available_min)
    if not is_custom and not to_raw:
        end_date = min(end_date, available_max)

    if start_date > end_date:
        start_date, end_date = end_date, start_date
    max_span_days = 365
    if (end_date - start_date).days > max_span_days:
        start_date = end_date - timedelta(days=max_span_days)

    report_payload = _crm_report_build_payload(request.user, org, start_date, end_date, group_by)
    requested_format = str(request.GET.get("format") or "json").strip().lower()
    if requested_format == "pdf":
        return _crm_report_pdf_response(org, request.user, report_payload)
    if requested_format in {"xlsx", "excel"}:
        return _crm_report_xlsx_response(org, request.user, report_payload)

    return JsonResponse(
        {
            "report": report_payload,
            "available_range": _crm_report_available_range(org),
        }
    )


@require_http_methods(["GET", "POST", "PATCH", "DELETE"])
def crm_leads(request, lead_id: int = None):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("__crm_action") or payload.get("action") or "").strip().upper()
        if body_action in {"PATCH", "DELETE"}:
            resolved_method = body_action
        elif override_method in {"PATCH", "DELETE"}:
            resolved_method = override_method

    if resolved_method in {"PATCH", "DELETE"} and not lead_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        lead_id = _crm_body_row_id(payload, "lead_id", "id")

    if resolved_method == "POST":
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        lead_name = str(payload.get("lead_name") or payload.get("name") or "").strip()
        if not lead_name:
            return JsonResponse({"detail": "lead_name_required"}, status=400)
        raw_assigned_user_ids = payload.get("assigned_user_ids")
        if raw_assigned_user_ids is None:
            raw_assigned_user_ids = payload.get("assignedUserIds")
        assigned_user_ids = _crm_clean_user_id_list(raw_assigned_user_ids)
        assigned_user_names_payload = _crm_extract_assigned_user_names_from_payload(payload)
        assigned_user_ids_from_names = _crm_resolve_user_ids_from_names(org, assigned_user_names_payload)
        for user_id in assigned_user_ids_from_names:
            if user_id not in assigned_user_ids:
                assigned_user_ids.append(user_id)
        assigned_user_id = _coerce_positive_int(payload.get("assigned_user_id") if "assigned_user_id" in payload else payload.get("assignedUserId"))
        assigned_user = None
        if assigned_user_id:
            assigned_user = User.objects.filter(id=assigned_user_id).first()
        elif assigned_user_ids:
            assigned_user = User.objects.filter(id=assigned_user_ids[0]).first()
        if getattr(assigned_user, "id", None) and assigned_user.id not in assigned_user_ids:
            assigned_user_ids.insert(0, int(assigned_user.id))
        row = CrmLead.objects.create(
            organization=org,
            lead_name=lead_name[:180],
            company=str(payload.get("company") or "").strip()[:180],
            phone=str(payload.get("phone") or "").strip()[:40],
            lead_amount=_crm_to_decimal(payload.get("lead_amount")),
            lead_source=str(payload.get("lead_source") or "").strip()[:120],
            assign_type="Team" if str(payload.get("assign_type") or "").strip().lower() == "team" else "Users",
            assigned_user=assigned_user,
            assigned_user_ids=assigned_user_ids,
            assigned_team=str(payload.get("assigned_team") or "").strip()[:180],
            stage=str(payload.get("stage") or "New").strip()[:30] or "New",
            priority=str(payload.get("priority") or "Medium").strip()[:30] or "Medium",
            status=str(payload.get("status") or "Open").strip()[:30] or "Open",
            created_by=request.user,
            updated_by=request.user,
        )
        snapshot_after = _crm_lead_snapshot(row)
        _crm_log_lead_modification(
            org,
            row,
            changed_by=request.user,
            action="create",
            changes=_crm_diff_snapshots({}, snapshot_after),
        )
        return JsonResponse({"lead": _serialize_crm_lead(row)}, status=201)

    if resolved_method == "GET" and not lead_id:
        rows = [
            row
            for row in CrmLead.objects.filter(organization=org).select_related("assigned_user", "created_by").order_by("-created_at")
            if _crm_can_view_row(request.user, org, row)
        ]
        pipeline_value = sum((_crm_to_decimal(row.lead_amount) for row in rows if not row.is_deleted), Decimal("0"))
        return JsonResponse(
            {
                "leads": [_serialize_crm_lead(row) for row in rows],
                "pipeline_value": float(pipeline_value),
            }
        )

    row = CrmLead.objects.filter(organization=org, id=lead_id).select_related("assigned_user", "created_by").first() if lead_id else None
    if not row:
        return JsonResponse({"detail": "lead_not_found"}, status=404)
    if not _crm_can_access_row(request.user, org, row):
        return JsonResponse({"detail": "forbidden"}, status=403)

    if resolved_method == "GET":
        return JsonResponse({"lead": _serialize_crm_lead(row)})

    if resolved_method == "PATCH":
        before_snapshot = _crm_lead_snapshot(row)
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        update_fields = ["updated_by", "updated_at"]
        sync_related_deal = False
        if "lead_name" in payload or "name" in payload:
            lead_name = str(payload.get("lead_name") or payload.get("name") or "").strip()
            if not lead_name:
                return JsonResponse({"detail": "lead_name_required"}, status=400)
            row.lead_name = lead_name[:180]
            update_fields.append("lead_name")
            sync_related_deal = True
        if "company" in payload:
            row.company = str(payload.get("company") or "").strip()[:180]
            update_fields.append("company")
            sync_related_deal = True
        if "phone" in payload:
            row.phone = str(payload.get("phone") or "").strip()[:40]
            update_fields.append("phone")
            sync_related_deal = True
        if "lead_amount" in payload or "leadAmount" in payload:
            row.lead_amount = _crm_to_decimal(payload.get("lead_amount") if "lead_amount" in payload else payload.get("leadAmount"))
            update_fields.append("lead_amount")
            sync_related_deal = True
        if "lead_source" in payload or "leadSource" in payload:
            row.lead_source = str(payload.get("lead_source") or payload.get("leadSource") or "").strip()[:120]
            update_fields.append("lead_source")
        if "assign_type" in payload or "assignType" in payload:
            assign_type = str(payload.get("assign_type") or payload.get("assignType") or "").strip().lower()
            row.assign_type = "Team" if assign_type == "team" else "Users"
            update_fields.append("assign_type")
        has_assigned_user_id_update = False
        if "assigned_user_id" in payload or "assignedUserId" in payload:
            assigned_user_id = _coerce_positive_int(payload.get("assigned_user_id") if "assigned_user_id" in payload else payload.get("assignedUserId"))
            row.assigned_user = User.objects.filter(id=assigned_user_id).first() if assigned_user_id else None
            update_fields.append("assigned_user")
            sync_related_deal = True
            has_assigned_user_id_update = True
        should_update_assigned_user_ids = (
            "assigned_user_ids" in payload
            or "assignedUserIds" in payload
            or "assigned_user_names" in payload
            or "assignedUserNames" in payload
            or "assignedUser" in payload
            or "assigned_user" in payload
            or "assignedTo" in payload
            or has_assigned_user_id_update
        )
        if should_update_assigned_user_ids:
            if "assigned_user_ids" in payload or "assignedUserIds" in payload:
                next_assigned_user_ids = _crm_clean_user_id_list(
                    payload.get("assigned_user_ids") if "assigned_user_ids" in payload else payload.get("assignedUserIds")
                )
            else:
                next_assigned_user_ids = []
            assigned_user_names_payload = _crm_extract_assigned_user_names_from_payload(payload)
            for user_id in _crm_resolve_user_ids_from_names(org, assigned_user_names_payload):
                if user_id not in next_assigned_user_ids:
                    next_assigned_user_ids.append(user_id)
            if getattr(row.assigned_user, "id", None) and row.assigned_user.id not in next_assigned_user_ids:
                next_assigned_user_ids.insert(0, int(row.assigned_user.id))
            row.assigned_user_ids = next_assigned_user_ids
            update_fields.append("assigned_user_ids")
            sync_related_deal = True
            if not row.assigned_user_id and next_assigned_user_ids:
                row.assigned_user = User.objects.filter(id=next_assigned_user_ids[0]).first()
                update_fields.append("assigned_user")
        if "assigned_team" in payload or "assignedTeam" in payload:
            row.assigned_team = str(payload.get("assigned_team") or payload.get("assignedTeam") or "").strip()[:180]
            update_fields.append("assigned_team")
            sync_related_deal = True
        if "stage" in payload:
            row.stage = str(payload.get("stage") or "New").strip()[:30] or "New"
            update_fields.append("stage")
        if "priority" in payload:
            row.priority = str(payload.get("priority") or "Medium").strip()[:30] or "Medium"
            update_fields.append("priority")
        if "status" in payload:
            row.status = str(payload.get("status") or "Open").strip()[:30] or "Open"
            update_fields.append("status")
        if "final_proposal_amount" in payload or "finalProposalAmount" in payload:
            row.final_proposal_amount = _crm_to_decimal(
                payload.get("final_proposal_amount") if "final_proposal_amount" in payload else payload.get("finalProposalAmount")
            )
            update_fields.append("final_proposal_amount")
        if "proposal_finalized" in payload or "proposalFinalized" in payload:
            should_finalize = bool(payload.get("proposal_finalized") if "proposal_finalized" in payload else payload.get("proposalFinalized"))
            # Reopening (clearing finalized state) is restricted to CRM Full Access.
            # This avoids allowing view/edit roles to reopen an already completed proposal process.
            if not should_finalize and getattr(row, "proposal_finalized_at", None) and not _crm_has_full_access(request.user, org):
                return JsonResponse({"detail": "forbidden"}, status=403)
            if should_finalize:
                if _crm_to_decimal(getattr(row, "final_proposal_amount", 0) or 0) <= 0:
                    return JsonResponse({"detail": "final_proposal_amount_required"}, status=400)
                row.proposal_finalized_at = timezone.now()
                row.proposal_finalized_by = request.user
                # Optional: who completed the proposal conversion (employees and/or teams).
                completed_by_user_ids = payload.get("completed_by_user_ids") if "completed_by_user_ids" in payload else payload.get("completedByUserIds")
                completed_by_teams = payload.get("completed_by_teams") if "completed_by_teams" in payload else payload.get("completedByTeams")
                if not isinstance(completed_by_user_ids, list):
                    completed_by_user_ids = []
                if not isinstance(completed_by_teams, list):
                    completed_by_teams = []
                completed_by_user_id = _coerce_positive_int(payload.get("completed_by_user_id") if "completed_by_user_id" in payload else payload.get("completedByUserId"))
                completed_by_team = str(payload.get("completed_by_team") if "completed_by_team" in payload else payload.get("completedByTeam") or "").strip()
                if completed_by_user_id:
                    completed_by_user_ids = [completed_by_user_id]
                if completed_by_team:
                    completed_by_teams = [completed_by_team]
                # Validate user ids
                cleaned_user_ids = []
                for raw_id in completed_by_user_ids:
                    user_id = _coerce_positive_int(raw_id)
                    if not user_id:
                        continue
                    if user_id in cleaned_user_ids:
                        continue
                    completed_profile = UserProfile.objects.filter(organization=org, user_id=user_id).select_related("user").first()
                    if not completed_profile:
                        return JsonResponse({"detail": "completed_by_user_invalid"}, status=400)
                    cleaned_user_ids.append(int(user_id))
                cleaned_team_names = []
                for name in completed_by_teams:
                    value = str(name or "").strip()[:180]
                    if not value or value in cleaned_team_names:
                        continue
                    cleaned_team_names.append(value)
                row.completed_by_user_ids = cleaned_user_ids
                row.completed_by_team_names = cleaned_team_names
                # Keep legacy fields populated with the first values for backward compatibility
                if cleaned_user_ids:
                    row.completed_by_type = "Users"
                    row.completed_by_user = User.objects.filter(id=cleaned_user_ids[0]).first()
                elif cleaned_team_names:
                    row.completed_by_type = "Team"
                    row.completed_by_user = None
                else:
                    row.completed_by_type = ""
                    row.completed_by_user = None
                row.completed_by_team = cleaned_team_names[0] if cleaned_team_names else ""
                update_fields.extend(["completed_by_type", "completed_by_user", "completed_by_team", "completed_by_user_ids", "completed_by_team_names"])
            else:
                row.proposal_finalized_at = None
                row.proposal_finalized_by = None
                row.completed_by_type = ""
                row.completed_by_team = ""
                row.completed_by_user = None
                row.completed_by_user_ids = []
                row.completed_by_team_names = []
                update_fields.extend(["completed_by_type", "completed_by_user", "completed_by_team", "completed_by_user_ids", "completed_by_team_names"])
            update_fields.extend(["proposal_finalized_at", "proposal_finalized_by"])
        # Allow updating "completed by" fields without toggling proposal_finalized,
        # but only when the proposal is already finalized.
        if (
            (
                "completed_by_user_id" in payload
                or "completedByUserId" in payload
                or "completed_by_team" in payload
                or "completedByTeam" in payload
                or "completed_by_user_ids" in payload
                or "completedByUserIds" in payload
                or "completed_by_teams" in payload
                or "completedByTeams" in payload
            )
            and not ("proposal_finalized" in payload or "proposalFinalized" in payload)
            and getattr(row, "proposal_finalized_at", None)
        ):
            completed_by_user_id = _coerce_positive_int(payload.get("completed_by_user_id") if "completed_by_user_id" in payload else payload.get("completedByUserId"))
            completed_by_team = str(payload.get("completed_by_team") if "completed_by_team" in payload else payload.get("completedByTeam") or "").strip()[:180]
            completed_by_user_ids = payload.get("completed_by_user_ids") if "completed_by_user_ids" in payload else payload.get("completedByUserIds")
            completed_by_teams = payload.get("completed_by_teams") if "completed_by_teams" in payload else payload.get("completedByTeams")
            if not isinstance(completed_by_user_ids, list):
                completed_by_user_ids = []
            if not isinstance(completed_by_teams, list):
                completed_by_teams = []
            if completed_by_user_id:
                completed_by_user_ids = [completed_by_user_id]
            if completed_by_team:
                completed_by_teams = [completed_by_team]
            cleaned_user_ids = []
            for raw_id in completed_by_user_ids:
                user_id = _coerce_positive_int(raw_id)
                if not user_id:
                    continue
                if user_id in cleaned_user_ids:
                    continue
                completed_profile = UserProfile.objects.filter(organization=org, user_id=user_id).select_related("user").first()
                if not completed_profile:
                    return JsonResponse({"detail": "completed_by_user_invalid"}, status=400)
                cleaned_user_ids.append(int(user_id))
            cleaned_team_names = []
            for name in completed_by_teams:
                value = str(name or "").strip()[:180]
                if not value or value in cleaned_team_names:
                    continue
                cleaned_team_names.append(value)
            row.completed_by_user_ids = cleaned_user_ids
            row.completed_by_team_names = cleaned_team_names
            if cleaned_user_ids:
                row.completed_by_type = "Users"
                row.completed_by_user = User.objects.filter(id=cleaned_user_ids[0]).first()
            elif cleaned_team_names:
                row.completed_by_type = "Team"
                row.completed_by_user = None
            else:
                row.completed_by_type = ""
                row.completed_by_user = None
            row.completed_by_team = cleaned_team_names[0] if cleaned_team_names else ""
            update_fields.extend(["completed_by_type", "completed_by_user", "completed_by_team", "completed_by_user_ids", "completed_by_team_names"])
        if "is_deleted" in payload:
            is_deleted = bool(payload.get("is_deleted"))
            row.is_deleted = is_deleted
            row.deleted_at = timezone.now() if is_deleted else None
            row.deleted_by = request.user if is_deleted else None
            update_fields.extend(["is_deleted", "deleted_at", "deleted_by"])
        row.updated_by = request.user
        row.save(update_fields=list(dict.fromkeys(update_fields)))
        if sync_related_deal:
            linked_deals = CrmDeal.objects.filter(organization=org, lead=row, is_deleted=False)
            for linked_deal in linked_deals:
                linked_deal.crm_reference_id = str(row.crm_reference_id or "").strip()
                linked_deal.deal_name = str(row.lead_name or "").strip()[:180]
                linked_deal.company = row.company
                linked_deal.phone = row.phone
                linked_deal.deal_value = _crm_to_decimal(row.lead_amount)
                linked_deal.assigned_user = row.assigned_user
                linked_deal.assigned_user_ids = _crm_clean_user_id_list(row.assigned_user_ids)
                linked_deal.assigned_team = row.assigned_team
                linked_deal.updated_by = request.user
                linked_deal.save(
                    update_fields=[
                        "crm_reference_id",
                        "deal_name",
                        "company",
                        "phone",
                        "deal_value",
                        "assigned_user",
                        "assigned_user_ids",
                        "assigned_team",
                        "updated_by",
                        "updated_at",
                    ]
                )
        after_snapshot = _crm_lead_snapshot(row)
        changes = _crm_diff_snapshots(before_snapshot, after_snapshot)
        if changes:
            action = "update"
            if bool(before_snapshot.get("is_deleted")) is False and bool(after_snapshot.get("is_deleted")) is True:
                action = "delete"
            elif bool(before_snapshot.get("is_deleted")) is True and bool(after_snapshot.get("is_deleted")) is False:
                action = "restore"
            _crm_log_lead_modification(org, row, changed_by=request.user, action=action, changes=changes)
        return JsonResponse({"lead": _serialize_crm_lead(row)})

    if resolved_method == "DELETE":
        if not _crm_can_edit_row(request.user, org, row):
            return JsonResponse({"detail": "forbidden"}, status=403)
        before_snapshot = _crm_lead_snapshot(row)
        permanent = (
            str(request.GET.get("permanent") or "").strip().lower() in {"1", "true", "yes"}
            or bool((payload or {}).get("__crm_permanent"))
        )
        if permanent:
            _crm_log_lead_modification(
                org,
                row,
                changed_by=request.user,
                action="delete_permanent",
                changes=_crm_diff_snapshots(before_snapshot, {}),
            )
            row.delete()
            return JsonResponse({"deleted": True, "permanent": True})
        row.is_deleted = True
        row.deleted_at = timezone.now()
        row.deleted_by = request.user
        row.updated_by = request.user
        row.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "updated_by", "updated_at"])
        after_snapshot = _crm_lead_snapshot(row)
        _crm_log_lead_modification(
            org,
            row,
            changed_by=request.user,
            action="delete",
            changes=_crm_diff_snapshots(before_snapshot, after_snapshot),
        )
        return JsonResponse({"deleted": True})

    return JsonResponse({"detail": "invalid_method"}, status=405)


@require_http_methods(["GET"])
def crm_lead_history(request, lead_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    lead = CrmLead.objects.filter(organization=org, id=lead_id).select_related("assigned_user", "created_by").first()
    if not lead:
        return JsonResponse({"detail": "lead_not_found"}, status=404)
    if not _crm_can_view_row(request.user, org, lead):
        return JsonResponse({"detail": "forbidden"}, status=403)

    reference_id = str(getattr(lead, "crm_reference_id", "") or "").strip()
    rows = (
        CrmLeadModification.objects
        .filter(organization=org)
        .filter(Q(lead_id=lead.id) | Q(lead_reference_id=reference_id))
        .select_related("changed_by")
        .order_by("-created_at")[:200]
    )
    return JsonResponse({"history": [_serialize_crm_lead_modification(row) for row in rows]})


def _crm_sanitize_proposal_base_name(value):
    raw_value = str(value or "").strip()
    raw_value = re.sub(r"\s+", " ", raw_value)
    safe_value = re.sub(r"[^\w\s\-.()]+", "", raw_value).strip()
    return safe_value[:180]


def _crm_proposal_display_name(base_name: str, existing_count: int):
    normalized = str(base_name or "").strip()
    if not normalized:
        return ""
    if existing_count <= 0:
        return normalized[:220]
    letter_index = existing_count - 1
    if 0 <= letter_index < 26:
        suffix = chr(ord("A") + letter_index)
    else:
        suffix = str(existing_count)
    return f"{normalized}-{suffix}"[:220]


def _serialize_crm_lead_proposal_document(doc: CrmLeadProposalDocument):
    if not doc:
        return {}
    uploaded_by = "-"
    if doc.uploaded_by_id:
        first = str(getattr(doc.uploaded_by, "first_name", "") or "").strip()
        last = str(getattr(doc.uploaded_by, "last_name", "") or "").strip()
        full = " ".join([part for part in [first, last] if part])
        uploaded_by = full or str(getattr(doc.uploaded_by, "username", "") or "").strip() or uploaded_by
    return {
        "id": doc.id,
        "lead_id": doc.lead_id,
        "base_name": doc.base_name,
        "version_index": doc.version_index,
        "display_name": doc.display_name,
        "original_filename": doc.original_filename,
        "file_type": doc.file_type,
        "file_size": doc.file_size,
        "uploaded_by": uploaded_by,
        "created_at": timezone.localtime(doc.created_at).isoformat() if doc.created_at else "",
    }


@require_http_methods(["GET", "POST"])
def crm_lead_proposals(request, lead_id: int, proposal_id: int = None):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    lead = CrmLead.objects.filter(organization=org, id=lead_id).first()
    if not lead:
        return JsonResponse({"detail": "lead_not_found"}, status=404)

    if request.method == "GET":
        rows = list(
            CrmLeadProposalDocument.objects.filter(organization=org, lead=lead)
            .select_related("uploaded_by")
            .order_by("-created_at", "-id")
        )
        return JsonResponse({"proposals": [_serialize_crm_lead_proposal_document(row) for row in rows]})

    override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
    if override_method == "DELETE":
        if not proposal_id:
            return JsonResponse({"detail": "proposal_id_required"}, status=400)
        doc = CrmLeadProposalDocument.objects.filter(organization=org, lead=lead, id=proposal_id).first()
        if not doc:
            return JsonResponse({"detail": "proposal_not_found"}, status=404)
        try:
            if doc.file:
                doc.file.delete(save=False)
        except Exception:
            pass
        doc.delete()
        return JsonResponse({"deleted": True})

    uploaded_file = request.FILES.get("file") or request.FILES.get("document")
    if not uploaded_file:
        return JsonResponse({"detail": "file_required"}, status=400)
    file_size = int(getattr(uploaded_file, "size", 0) or 0)
    if file_size <= 0:
        return JsonResponse({"detail": "file_empty"}, status=400)
    max_bytes = 2 * 1024 * 1024
    if file_size > max_bytes:
        return JsonResponse({"detail": "file_too_large", "max_bytes": max_bytes}, status=400)

    filename = str(getattr(uploaded_file, "name", "") or "").strip()
    ext = ""
    if filename and "." in filename:
        ext = "." + filename.split(".")[-1].lower()
    allowed_exts = {".pdf", ".doc", ".docx"}
    if ext not in allowed_exts:
        return JsonResponse({"detail": "invalid_file_type", "allowed": sorted(list(allowed_exts))}, status=400)

    base_name = _crm_sanitize_proposal_base_name(request.POST.get("proposal_name") or request.POST.get("base_name"))
    if not base_name:
        return JsonResponse({"detail": "proposal_name_required"}, status=400)

    existing_count = CrmLeadProposalDocument.objects.filter(organization=org, lead=lead, base_name=base_name).count()
    display_name = _crm_proposal_display_name(base_name, existing_count)
    doc = CrmLeadProposalDocument.objects.create(
        organization=org,
        lead=lead,
        base_name=base_name,
        version_index=existing_count,
        display_name=display_name,
        original_filename=filename[:255],
        file=uploaded_file,
        file_type=str(getattr(uploaded_file, "content_type", "") or "")[:80],
        file_size=file_size,
        uploaded_by=request.user,
    )
    doc = CrmLeadProposalDocument.objects.filter(id=doc.id).select_related("uploaded_by").first()
    return JsonResponse({"proposal": _serialize_crm_lead_proposal_document(doc)})


@require_http_methods(["GET"])
def crm_lead_proposal_download(request, lead_id: int, proposal_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    lead = CrmLead.objects.filter(organization=org, id=lead_id).first()
    if not lead:
        return JsonResponse({"detail": "lead_not_found"}, status=404)
    doc = (
        CrmLeadProposalDocument.objects
        .filter(organization=org, lead=lead, id=proposal_id)
        .first()
    )
    if not doc or not doc.file:
        return JsonResponse({"detail": "proposal_not_found"}, status=404)

    # Keep filename safe & readable; avoid regex ranges by placing '-' at the end.
    safe_display = re.sub(r"[^a-zA-Z0-9 _().-]+", "", str(doc.display_name or "proposal")).strip() or "proposal"
    original = str(doc.original_filename or "").strip()
    ext = ""
    if original and "." in original:
        ext = "." + original.split(".")[-1].lower()
    if ext not in {".pdf", ".doc", ".docx"}:
        ext = ""
    filename = f"{safe_display}{ext}"
    response = FileResponse(doc.file.open("rb"), as_attachment=True, filename=filename)
    if doc.file_type:
        response["Content-Type"] = doc.file_type
    return response


@require_http_methods(["GET", "POST", "PATCH", "DELETE"])
def crm_contacts(request, contact_id: int = None):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("__crm_action") or payload.get("action") or "").strip().upper()
        if body_action in {"PATCH", "DELETE"}:
            resolved_method = body_action
        elif override_method in {"PATCH", "DELETE"}:
            resolved_method = override_method

    if resolved_method in {"PATCH", "DELETE"} and not contact_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        contact_id = _crm_body_row_id(payload, "contact_id", "id")

    if resolved_method == "GET" and not contact_id:
        rows = CrmContact.objects.filter(organization=org).select_related("created_by").order_by("-created_at")
        return JsonResponse({"contacts": [_serialize_crm_contact(row) for row in rows]})

    if resolved_method == "POST" and not contact_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        name = str(payload.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name_required"}, status=400)
        company = str(payload.get("company") or "").strip()[:180]
        email = str(payload.get("email") or "").strip()[:180]
        phone_country_code = str(payload.get("phone_country_code") or payload.get("phoneCountryCode") or "+91").strip()[:10] or "+91"
        phone = str(payload.get("phone") or "").strip()[:40]
        duplicate_row, duplicate_fields = _crm_find_duplicate_contact(
            org,
            company=company,
            email=email,
            phone_country_code=phone_country_code,
            phone=phone,
        )
        if duplicate_row:
            return JsonResponse(
                {
                    "detail": "duplicate_contact",
                    "duplicate_fields": duplicate_fields,
                    "existing_contact": _serialize_crm_contact(duplicate_row),
                },
                status=409,
            )
        tag = str(payload.get("tag") or "Client").strip().title()
        if tag not in {"Client", "Prospect", "Vendor"}:
            tag = "Client"
        row = CrmContact.objects.create(
            organization=org,
            name=name[:180],
            company=company,
            email=email,
            phone_country_code=phone_country_code,
            phone=phone,
            tag=tag,
            created_by=request.user,
            updated_by=request.user,
        )
        return JsonResponse({"contact": _serialize_crm_contact(row)}, status=201)

    row = CrmContact.objects.filter(organization=org, id=contact_id).select_related("created_by").first() if contact_id else None
    if not row:
        return JsonResponse({"detail": "contact_not_found"}, status=404)

    if resolved_method == "PATCH":
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        if not _crm_can_edit_row(request.user, org, row):
            return JsonResponse({"detail": "forbidden"}, status=403)
        next_name = str(row.name or "").strip()
        next_company = str(row.company or "").strip()
        next_email = str(row.email or "").strip()
        next_phone_country_code = str(row.phone_country_code or "+91").strip() or "+91"
        next_phone = str(row.phone or "").strip()
        update_fields = ["updated_by", "updated_at"]
        if "name" in payload:
            name = str(payload.get("name") or "").strip()
            if not name:
                return JsonResponse({"detail": "name_required"}, status=400)
            next_name = name[:180]
            update_fields.append("name")
        if "company" in payload:
            next_company = str(payload.get("company") or "").strip()[:180]
            update_fields.append("company")
        if "email" in payload:
            next_email = str(payload.get("email") or "").strip()[:180]
            update_fields.append("email")
        if "phone_country_code" in payload or "phoneCountryCode" in payload:
            next_phone_country_code = str(payload.get("phone_country_code") or payload.get("phoneCountryCode") or "+91").strip()[:10] or "+91"
            update_fields.append("phone_country_code")
        if "phone" in payload:
            next_phone = str(payload.get("phone") or "").strip()[:40]
            update_fields.append("phone")
        duplicate_row, duplicate_fields = _crm_find_duplicate_contact(
            org,
            company=next_company,
            email=next_email,
            phone_country_code=next_phone_country_code,
            phone=next_phone,
            exclude_contact_id=row.id,
        )
        if duplicate_row:
            return JsonResponse(
                {
                    "detail": "duplicate_contact",
                    "duplicate_fields": duplicate_fields,
                    "existing_contact": _serialize_crm_contact(duplicate_row),
                },
                status=409,
            )
        row.name = next_name[:180]
        row.company = next_company[:180]
        row.email = next_email[:180]
        row.phone_country_code = next_phone_country_code[:10] or "+91"
        row.phone = next_phone[:40]
        if "tag" in payload:
            tag = str(payload.get("tag") or "Client").strip().title()
            row.tag = tag if tag in {"Client", "Prospect", "Vendor"} else "Client"
            update_fields.append("tag")
        if "is_deleted" in payload:
            is_deleted = bool(payload.get("is_deleted"))
            row.is_deleted = is_deleted
            row.deleted_at = timezone.now() if is_deleted else None
            row.deleted_by = request.user if is_deleted else None
            update_fields.extend(["is_deleted", "deleted_at", "deleted_by"])
        row.updated_by = request.user
        row.save(update_fields=list(dict.fromkeys(update_fields)))
        return JsonResponse({"contact": _serialize_crm_contact(row)})

    if resolved_method == "DELETE":
        if not _crm_can_edit_row(request.user, org, row):
            return JsonResponse({"detail": "forbidden"}, status=403)
        permanent = (
            str(request.GET.get("permanent") or "").strip().lower() in {"1", "true", "yes"}
            or bool((payload or {}).get("__crm_permanent"))
        )
        if permanent:
            row.delete()
            return JsonResponse({"deleted": True, "permanent": True})
        row.is_deleted = True
        row.deleted_at = timezone.now()
        row.deleted_by = request.user
        row.updated_by = request.user
        row.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "updated_by", "updated_at"])
        return JsonResponse({"deleted": True})

    return JsonResponse({"detail": "invalid_method"}, status=405)


@require_http_methods(["POST"])
def crm_convert_to_deal(request, lead_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    lead = CrmLead.objects.filter(organization=org, id=lead_id, is_deleted=False).first()
    if not lead:
        return JsonResponse({"detail": "lead_not_found"}, status=404)
    if not _crm_can_access_row(request.user, org, lead):
        return JsonResponse({"detail": "forbidden"}, status=403)
    if _crm_to_decimal(lead.lead_amount) <= Decimal("0"):
        return JsonResponse({"detail": "lead_amount_required_for_conversion"}, status=400)
    existing_deal = CrmDeal.objects.filter(organization=org, lead=lead, is_deleted=False).first()
    if existing_deal:
        # Keep converted deal in sync with latest lead data when user retries convert.
        existing_deal.crm_reference_id = str(lead.crm_reference_id or "").strip()
        existing_deal.deal_name = str(lead.lead_name or "").strip()[:180]
        existing_deal.company = lead.company
        existing_deal.phone = lead.phone
        existing_deal.deal_value = _crm_to_decimal(lead.lead_amount)
        existing_deal.assigned_user = lead.assigned_user
        existing_deal.assigned_user_ids = _crm_clean_user_id_list(lead.assigned_user_ids)
        existing_deal.assigned_team = lead.assigned_team
        existing_deal.updated_by = request.user
        existing_deal.save(
            update_fields=[
                "crm_reference_id",
                "deal_name",
                "company",
                "phone",
                "deal_value",
                "assigned_user",
                "assigned_user_ids",
                "assigned_team",
                "updated_by",
                "updated_at",
            ]
        )
        if lead.status != "Converted" or lead.stage != "Qualified":
            lead.status = "Converted"
            lead.stage = "Qualified"
            lead.updated_by = request.user
            lead.save(update_fields=["status", "stage", "updated_by", "updated_at"])
        return JsonResponse(
            {
                "deal": _serialize_crm_deal(existing_deal),
                "lead": _serialize_crm_lead(lead),
                "already_converted": True,
                "synced": True,
            }
        )
    with transaction.atomic():
        deal = CrmDeal.objects.create(
            organization=org,
            lead=lead,
            crm_reference_id=str(lead.crm_reference_id or "").strip(),
            deal_name=str(lead.lead_name or "").strip()[:180],
            company=lead.company,
            phone=lead.phone,
            deal_value=_crm_to_decimal(lead.lead_amount),
            stage="Qualified",
            status="Open",
            assigned_user=lead.assigned_user,
            assigned_user_ids=_crm_clean_user_id_list(lead.assigned_user_ids),
            assigned_team=lead.assigned_team,
            created_by=request.user,
            updated_by=request.user,
        )
        lead.status = "Converted"
        lead.stage = "Qualified"
        lead.updated_by = request.user
        lead.save(update_fields=["status", "stage", "updated_by", "updated_at"])
    return JsonResponse({"deal": _serialize_crm_deal(deal), "lead": _serialize_crm_lead(lead)})


@require_http_methods(["GET", "POST", "PATCH", "DELETE"])
def crm_deals(request, deal_id: int = None):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("__crm_action") or payload.get("action") or "").strip().upper()
        if body_action in {"PATCH", "DELETE"}:
            resolved_method = body_action
        elif override_method in {"PATCH", "DELETE"}:
            resolved_method = override_method

    if resolved_method in {"PATCH", "DELETE"} and not deal_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        deal_id = _crm_body_row_id(payload, "deal_id", "id")

    if resolved_method == "GET" and not deal_id:
        rows = [
            row
            for row in (
                CrmDeal.objects.filter(organization=org)
                .select_related("assigned_user", "lead", "created_by", "updated_by")
                .order_by("-created_at")
            )
            if _crm_can_view_row(request.user, org, row)
        ]
        pipeline_value = sum((_crm_to_decimal(row.deal_value) for row in rows if not row.is_deleted), Decimal("0"))
        return JsonResponse({"deals": [_serialize_crm_deal(row) for row in rows], "pipeline_value": float(pipeline_value)})

    if resolved_method == "POST" and not deal_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        deal_name = str(payload.get("deal_name") or "").strip()
        if not deal_name:
            return JsonResponse({"detail": "deal_name_required"}, status=400)
        lead = None
        lead_id = _coerce_positive_int(payload.get("lead_id"))
        if lead_id:
            lead = CrmLead.objects.filter(organization=org, id=lead_id).first()
        deal_crm_reference_id = str(payload.get("crm_reference_id") or "").strip()[:32]
        if lead and not deal_crm_reference_id:
            deal_crm_reference_id = str(lead.crm_reference_id or "").strip()[:32]
        assigned_user_id = _coerce_positive_int(payload.get("assigned_user_id"))
        assigned_user = User.objects.filter(id=assigned_user_id).first() if assigned_user_id else None
        row = CrmDeal.objects.create(
            organization=org,
            lead=lead,
            crm_reference_id=deal_crm_reference_id,
            deal_name=deal_name[:180],
            company=str(payload.get("company") or "").strip()[:180],
            phone=str(payload.get("phone") or "").strip()[:40],
            deal_value=_crm_to_decimal(payload.get("deal_value")),
            won_amount_final=_crm_to_decimal(payload.get("won_amount_final")),
            stage=str(payload.get("stage") or "Qualified").strip()[:30] or "Qualified",
            status=str(payload.get("status") or "Open").strip()[:30] or "Open",
            assigned_user=assigned_user,
            assigned_user_ids=_crm_clean_user_id_list(payload.get("assigned_user_ids")),
            assigned_team=str(payload.get("assigned_team") or "").strip()[:180],
            created_by=request.user,
            updated_by=request.user,
        )
        return JsonResponse({"deal": _serialize_crm_deal(row)}, status=201)

    row = CrmDeal.objects.filter(organization=org, id=deal_id).select_related("assigned_user", "lead", "created_by", "updated_by").first() if deal_id else None
    if not row:
        return JsonResponse({"detail": "deal_not_found"}, status=404)
    if not _crm_can_access_row(request.user, org, row):
        return JsonResponse({"detail": "forbidden"}, status=403)

    if resolved_method == "GET":
        return JsonResponse({"deal": _serialize_crm_deal(row)})

    if resolved_method == "PATCH":
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        update_fields = []
        if "is_deleted" in payload:
            if not _crm_is_admin(request.user, org):
                return JsonResponse({"detail": "forbidden"}, status=403)
            is_deleted = bool(payload.get("is_deleted"))
            row.is_deleted = is_deleted
            row.deleted_at = timezone.now() if is_deleted else None
            row.deleted_by = request.user if is_deleted else None
            update_fields.extend(["is_deleted", "deleted_at", "deleted_by"])
        if "stage" in payload:
            stage = str(payload.get("stage") or "").strip()
            if stage in {"Qualified", "Proposal", "Won", "Lost"}:
                row.stage = stage
                update_fields.append("stage")
        if "status" in payload:
            status = str(payload.get("status") or "").strip()
            if status in {"Open", "Won", "Lost"}:
                row.status = status
                update_fields.append("status")
        if "deal_value" in payload:
            row.deal_value = _crm_to_decimal(payload.get("deal_value"))
            update_fields.append("deal_value")
        if "won_amount_final" in payload or "wonAmountFinal" in payload:
            row.won_amount_final = _crm_to_decimal(
                payload.get("won_amount_final") if "won_amount_final" in payload else payload.get("wonAmountFinal")
            )
            update_fields.append("won_amount_final")
        if not row.crm_reference_id and row.lead_id:
            lead_ref = str(getattr(row.lead, "crm_reference_id", "") or "").strip()
            if lead_ref:
                row.crm_reference_id = lead_ref
                update_fields.append("crm_reference_id")
        row.updated_by = request.user
        update_fields.extend(["updated_by", "updated_at"])
        row.save(update_fields=list(dict.fromkeys(update_fields)))
        return JsonResponse({"deal": _serialize_crm_deal(row)})

    if resolved_method == "DELETE":
        if not _crm_is_admin(request.user, org):
            return JsonResponse({"detail": "forbidden"}, status=403)
        permanent = (
            str(request.GET.get("permanent") or "").strip().lower() in {"1", "true", "yes"}
            or bool((payload or {}).get("__crm_permanent"))
        )
        if permanent:
            row.delete()
            return JsonResponse({"deleted": True, "permanent": True})
        row.is_deleted = True
        row.deleted_at = timezone.now()
        row.deleted_by = request.user
        row.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "updated_at"])
        return JsonResponse({"deleted": True})

    return JsonResponse({"detail": "invalid_method"}, status=405)


@require_http_methods(["GET", "POST", "PATCH", "DELETE"])
def crm_meetings(request, meeting_id: int = None):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("__crm_action") or payload.get("action") or "").strip().upper()
        if body_action in {"PATCH", "DELETE"}:
            resolved_method = body_action
        elif override_method in {"PATCH", "DELETE"}:
            resolved_method = override_method

    if resolved_method in {"PATCH", "DELETE"} and not meeting_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        meeting_id = _crm_body_row_id(payload, "meeting_id", "id")

    if resolved_method == "GET" and not meeting_id:
        _dispatch_due_crm_meeting_reminders(org=org)
        rows = [
            row
            for row in CrmMeeting.objects.filter(organization=org).order_by("-created_at")
            if _crm_can_view_row(request.user, org, row)
        ]
        return JsonResponse({"meetings": [_serialize_crm_meeting(row) for row in rows]})

    if resolved_method == "POST" and not meeting_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        try:
            meeting_id = _crm_body_row_id(payload, "meeting_id", "id")
        except Exception:
            meeting_id = None
        title = str(payload.get("title") or payload.get("meeting_title") or "").strip()
        if not title:
            return JsonResponse({"detail": "title_required"}, status=400)
        related_to_value = str(payload.get("related_to") or payload.get("relatedTo") or "").strip()[:180]
        meeting_crm_reference_id = str(payload.get("crm_reference_id") or payload.get("crmReferenceId") or "").strip()[:32]
        if not meeting_crm_reference_id:
            meeting_crm_reference_id = _crm_reference_id_from_related_to(org, related_to_value)
        meeting_date = parse_date(str(payload.get("meeting_date") or payload.get("meetingDate") or "").strip() or "")
        meeting_time = parse_time(str(payload.get("meeting_time") or payload.get("meetingTime") or "").strip() or "")
        row = CrmMeeting.objects.create(
            organization=org,
            crm_reference_id=meeting_crm_reference_id,
            title=title[:180],
            company_or_client_name=str(payload.get("company_or_client_name") or payload.get("companyOrClientName") or "").strip()[:180],
            related_to=related_to_value,
            meeting_date=meeting_date,
            meeting_time=meeting_time,
            owner_names=str(payload.get("owner") or "").strip()[:500],
            owner_user_ids=_crm_clean_user_id_list(payload.get("owner_user_ids") or payload.get("ownerUserIds")),
            meeting_mode=str(payload.get("meeting_mode") or payload.get("meetingMode") or "").strip()[:30],
            reminder_channels=payload.get("reminder_channel") if isinstance(payload.get("reminder_channel"), list)
            else payload.get("reminderChannel") if isinstance(payload.get("reminderChannel"), list)
            else [],
            reminder_days=_crm_clean_int_list(payload.get("reminder_days") if payload.get("reminder_days") is not None else payload.get("reminderDays")),
            reminder_minutes=_crm_clean_int_list(payload.get("reminder_minutes") if payload.get("reminder_minutes") is not None else payload.get("reminderMinutes")),
            reminder_summary=str(payload.get("reminder_summary") or payload.get("reminderSummary") or "").strip()[:255],
            status=str(payload.get("status") or "Scheduled").strip()[:30] or "Scheduled",
            created_by=request.user,
            updated_by=request.user,
        )
        _dispatch_due_crm_meeting_reminders(org=org)
        return JsonResponse({"meeting": _serialize_crm_meeting(row)}, status=201)

    row = CrmMeeting.objects.filter(organization=org, id=meeting_id).first() if meeting_id else None
    if not row:
        return JsonResponse({"detail": "meeting_not_found"}, status=404)
    if not _crm_can_access_row(request.user, org, row):
        return JsonResponse({"detail": "forbidden"}, status=403)

    if resolved_method == "PATCH":
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        related_to_changed = False
        if "title" in payload or "meeting_title" in payload:
            title = str(payload.get("title") or payload.get("meeting_title") or "").strip()
            if title:
                row.title = title[:180]
        if "company_or_client_name" in payload or "companyOrClientName" in payload:
            row.company_or_client_name = str(payload.get("company_or_client_name") or payload.get("companyOrClientName") or "").strip()[:180]
        if "related_to" in payload or "relatedTo" in payload:
            row.related_to = str(payload.get("related_to") or payload.get("relatedTo") or "").strip()[:180]
            related_to_changed = True
        if "crm_reference_id" in payload or "crmReferenceId" in payload:
            row.crm_reference_id = str(payload.get("crm_reference_id") or payload.get("crmReferenceId") or "").strip()[:32]
        elif related_to_changed and not str(row.crm_reference_id or "").strip():
            row.crm_reference_id = _crm_reference_id_from_related_to(org, row.related_to)
        if "meeting_date" in payload or "meetingDate" in payload:
            row.meeting_date = parse_date(str(payload.get("meeting_date") or payload.get("meetingDate") or "").strip() or "")
        if "meeting_time" in payload or "meetingTime" in payload:
            row.meeting_time = parse_time(str(payload.get("meeting_time") or payload.get("meetingTime") or "").strip() or "")
        if "owner" in payload:
            row.owner_names = str(payload.get("owner") or "").strip()[:500]
        if "owner_user_ids" in payload or "ownerUserIds" in payload:
            row.owner_user_ids = _crm_clean_user_id_list(payload.get("owner_user_ids") or payload.get("ownerUserIds"))
        if "meeting_mode" in payload or "meetingMode" in payload:
            row.meeting_mode = str(payload.get("meeting_mode") or payload.get("meetingMode") or "").strip()[:30]
        if "reminder_channel" in payload or "reminderChannel" in payload:
            channels = payload.get("reminder_channel") if isinstance(payload.get("reminder_channel"), list) else payload.get("reminderChannel")
            row.reminder_channels = channels if isinstance(channels, list) else []
        if "reminder_days" in payload or "reminderDays" in payload:
            row.reminder_days = _crm_clean_int_list(payload.get("reminder_days") if payload.get("reminder_days") is not None else payload.get("reminderDays"))
        if "reminder_minutes" in payload or "reminderMinutes" in payload:
            row.reminder_minutes = _crm_clean_int_list(payload.get("reminder_minutes") if payload.get("reminder_minutes") is not None else payload.get("reminderMinutes"))
        if "reminder_summary" in payload or "reminderSummary" in payload:
            row.reminder_summary = str(payload.get("reminder_summary") or payload.get("reminderSummary") or "").strip()[:255]
        if "status" in payload:
            row.status = str(payload.get("status") or row.status).strip()[:30] or row.status
        if "is_deleted" in payload:
            if not _crm_is_admin(request.user, org):
                return JsonResponse({"detail": "forbidden"}, status=403)
            is_deleted = bool(payload.get("is_deleted"))
            row.is_deleted = is_deleted
            row.deleted_at = timezone.now() if is_deleted else None
            row.deleted_by = request.user if is_deleted else None
        row.updated_by = request.user
        row.save()
        _dispatch_due_crm_meeting_reminders(org=org)
        return JsonResponse({"meeting": _serialize_crm_meeting(row)})

    if resolved_method == "DELETE":
        if not _crm_is_admin(request.user, org):
            return JsonResponse({"detail": "forbidden"}, status=403)
        permanent = (
            str(request.GET.get("permanent") or "").strip() in {"1", "true", "yes"}
            or bool((payload or {}).get("__crm_permanent"))
        )
        if permanent:
            row.delete()
            return JsonResponse({"deleted": True, "permanent": True})
        row.is_deleted = True
        row.deleted_at = timezone.now()
        row.deleted_by = request.user
        row.updated_by = request.user
        row.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "updated_by", "updated_at"])
        return JsonResponse({"deleted": True})

    return JsonResponse({"detail": "invalid_method"}, status=405)


@require_http_methods(["GET", "POST", "PATCH", "DELETE"])
def crm_sales_orders(request, order_id: int = None):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)

    resolved_method = request.method
    payload = None
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"detail": "invalid_json"}, status=400)
        override_method = str(request.META.get("HTTP_X_HTTP_METHOD_OVERRIDE") or "").strip().upper()
        body_action = str(payload.get("__crm_action") or payload.get("action") or "").strip().upper()
        if body_action in {"PATCH", "DELETE"}:
            resolved_method = body_action
        elif override_method in {"PATCH", "DELETE"}:
            resolved_method = override_method

    if resolved_method in {"PATCH", "DELETE"} and not order_id:
        if payload is None:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                return JsonResponse({"detail": "invalid_json"}, status=400)
        order_id = _crm_body_row_id(payload, "sales_order_id", "id")

    if resolved_method == "GET" and not order_id:
        rows = CrmSalesOrder.objects.filter(organization=org).select_related("assigned_user", "deal", "deal__lead", "created_by").order_by("-created_at")
        visible_rows = [row for row in rows if _crm_can_view_row(request.user, org, row)]
        return JsonResponse({"sales_orders": [_serialize_crm_sales_order(row) for row in visible_rows]})

    row = CrmSalesOrder.objects.filter(organization=org, id=order_id).select_related("deal", "assigned_user", "created_by").first() if order_id else None
    if order_id and not row:
        return JsonResponse({"detail": "sales_order_not_found"}, status=404)
    if row and not _crm_can_access_row(request.user, org, row):
        return JsonResponse({"detail": "forbidden"}, status=403)
    can_edit_payment_details = _crm_is_admin(request.user, org)

    if resolved_method == "DELETE":
        if not _crm_is_admin(request.user, org):
            return JsonResponse({"detail": "forbidden"}, status=403)
        permanent = (
            str(request.GET.get("permanent") or "").strip().lower() in {"1", "true", "yes"}
            or bool((payload or {}).get("__crm_permanent"))
        )
        if permanent:
            row.delete()
            return JsonResponse({"deleted": True, "permanent": True})
        row.is_deleted = True
        row.deleted_at = timezone.now()
        row.deleted_by = request.user
        row.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "updated_at"])
        return JsonResponse({"deleted": True})

    try:
        if payload is None:
            payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"detail": "invalid_json"}, status=400)

    if resolved_method == "PATCH" and "is_deleted" in payload:
        if not _crm_is_admin(request.user, org):
            return JsonResponse({"detail": "forbidden"}, status=403)
        is_deleted = bool(payload.get("is_deleted"))
        row.is_deleted = is_deleted
        row.deleted_at = timezone.now() if is_deleted else None
        row.deleted_by = request.user if is_deleted else None
        row.updated_by = request.user
        row.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "updated_by", "updated_at"])
        return JsonResponse({"sales_order": _serialize_crm_sales_order(row)})

    customer_name = str(payload.get("customer_name") or "").strip()
    if not customer_name:
        return JsonResponse({"detail": "customer_name_required"}, status=400)
    crm_reference_id = str(payload.get("crm_reference_id") or payload.get("crmReferenceId") or "").strip()[:32]
    source_deal_id = _coerce_positive_int(payload.get("source_deal_id") or payload.get("sourceDealId") or payload.get("deal_id"))
    source_deal = None
    if source_deal_id:
        source_deal = CrmDeal.objects.filter(organization=org, id=source_deal_id).select_related("lead").first()
    if source_deal and not crm_reference_id:
        crm_reference_id = str(source_deal.crm_reference_id or getattr(source_deal.lead, "crm_reference_id", "") or "").strip()[:32]
    gst_template_id = str(payload.get("gst_template_id") or payload.get("gstTemplateId") or "").strip()
    accounts_workspace = _get_accounts_workspace(org)
    accounts_data = _normalize_accounts_workspace(accounts_workspace.data)
    gst_templates_by_id = {
        str(item.get("id")): item
        for item in accounts_data.get("gstTemplates", [])
        if isinstance(item, dict)
    }
    default_tax_percent = _document_template_total_percent(
        _resolve_gst_template_by_id(gst_templates_by_id, gst_template_id)
    )
    normalized_items = _crm_normalize_sales_order_items(payload.get("items"))
    totals = _crm_sales_order_totals(normalized_items, default_tax_percent=default_tax_percent)
    subtotal_amount = _crm_to_decimal(payload.get("amount")) if not normalized_items else totals["subtotal"]
    total_amount = _crm_to_decimal(payload.get("total_amount")) if not normalized_items else totals["grand_total"]
    quantity = _coerce_positive_int(payload.get("quantity")) or int(totals["total_qty"] or Decimal("1")) or 1
    price = _crm_to_decimal(payload.get("price")) if not normalized_items else totals["first_rate"]
    tax = _crm_to_decimal(payload.get("tax")) if not normalized_items else totals["first_tax"]
    assigned_user_id = _coerce_positive_int(payload.get("assigned_user_id"))
    assigned_user = User.objects.filter(id=assigned_user_id).first() if assigned_user_id else request.user
    deal = None
    source_deal_id = _coerce_positive_int(payload.get("source_deal_id") or payload.get("sourceDealId") or payload.get("deal_id"))
    if source_deal_id:
        deal = CrmDeal.objects.filter(organization=org, id=source_deal_id).first()
    paid_amount = _crm_to_decimal(payload.get("paid_amount") or payload.get("paidAmount")) if can_edit_payment_details else _crm_to_decimal(row.paid_amount if row else 0)
    calculated_total_amount = total_amount or (subtotal_amount + (subtotal_amount * (tax / Decimal("100"))))
    if paid_amount >= calculated_total_amount and calculated_total_amount > 0:
        payment_status = "paid"
    elif paid_amount > 0:
        payment_status = "partial"
    else:
        payment_status = "pending"
    existing_products_payload = _crm_sales_order_payload_dict(row) if row else {}
    requested_order_id = str(payload.get("order_id") or payload.get("orderId") or (row.order_id if row else "") or "").strip()
    resolved_issue_date = str(payload.get("issue_date") or payload.get("issueDate") or "").strip()
    if not resolved_issue_date:
        resolved_issue_date = str(existing_products_payload.get("issueDate") or existing_products_payload.get("issue_date") or "").strip()
    if not resolved_issue_date:
        resolved_issue_date = _crm_issue_date_from_order_id(requested_order_id)
    if not resolved_issue_date:
        resolved_issue_date = timezone.localdate().isoformat()
    resolved_due_date = str(payload.get("due_date") or payload.get("dueDate") or "").strip()
    if not resolved_due_date:
        resolved_due_date = str(existing_products_payload.get("dueDate") or existing_products_payload.get("due_date") or "").strip()
    if not resolved_due_date:
        resolved_due_date = resolved_issue_date
    products_payload = {
        "items": normalized_items,
        "issueDate": resolved_issue_date,
        "dueDate": resolved_due_date,
        "gstTemplateId": gst_template_id,
        "billingTemplateId": str(payload.get("billing_template_id") or payload.get("billingTemplateId") or "").strip(),
        "salesperson": str(payload.get("salesperson") or "").strip(),
        "customerGstin": str(payload.get("customer_gstin") or payload.get("customerGstin") or "").strip(),
        "billingAddress": str(payload.get("billing_address") or payload.get("billingAddress") or "").strip(),
        "notes": str(payload.get("notes") or "").strip(),
        "termsText": str(payload.get("terms_text") or payload.get("termsText") or "").strip(),
        "paymentStatusNotes": str(payload.get("payment_status_notes") or payload.get("paymentStatusNotes") or "").strip(),
        "paymentEntries": [
            {
                "id": str(entry.get("id") or "").strip(),
                "paymentDate": str(entry.get("paymentDate") or entry.get("payment_date") or "").strip(),
                "paymentMode": str(entry.get("paymentMode") or entry.get("payment_mode") or "").strip(),
                "amount": float(_crm_to_decimal(entry.get("amount") or entry.get("paidAmount") or entry.get("paid_amount"))),
                "transactionId": str(entry.get("transactionId") or entry.get("transaction_id") or "").strip(),
                "notes": str(entry.get("notes") or entry.get("paymentNotes") or "").strip(),
            }
            for entry in (payload.get("payment_entries") or payload.get("paymentEntries") or [])
            if isinstance(entry, dict)
        ],
        "paymentStatus": payment_status,
        "paidAmount": float(paid_amount),
        "paymentMode": str(payload.get("payment_mode") or payload.get("paymentMode") or "").strip() if can_edit_payment_details else str(row.payment_mode if row else ""),
        "paymentDate": str(payload.get("payment_date") or payload.get("paymentDate") or "").strip() if can_edit_payment_details else (row.payment_date.isoformat() if row and row.payment_date else ""),
        "transactionId": str(payload.get("transaction_id") or payload.get("transactionId") or "").strip() if can_edit_payment_details else str(row.transaction_id if row else ""),
        "balanceAmount": float(max(Decimal("0"), calculated_total_amount - paid_amount)),
        "sourceDealId": str(source_deal_id or (row.deal_id if row else "") or "").strip(),
        "convertedToInvoice": bool(payload.get("converted_to_invoice") or payload.get("convertedToInvoice")),
        "convertedInvoiceId": str(payload.get("converted_invoice_id") or payload.get("convertedInvoiceId") or "").strip(),
    }
    if resolved_method == "PATCH":
        row.customer_name = customer_name[:180]
        row.company = str(payload.get("company") or "").strip()[:180]
        row.phone = str(payload.get("phone") or "").strip()[:40]
        row.amount = subtotal_amount
        row.products = products_payload
        row.quantity = quantity
        row.price = price
        row.tax = tax
        row.total_amount = calculated_total_amount
        row.payment_status = payment_status
        row.paid_amount = float(paid_amount)
        if can_edit_payment_details:
            row.payment_mode = str(payload.get("payment_mode") or payload.get("paymentMode") or "").strip()[:50]
            row.payment_date = parse_date(str(payload.get("payment_date") or payload.get("paymentDate") or "").strip() or "")
            row.transaction_id = str(payload.get("transaction_id") or payload.get("transactionId") or "").strip()[:100]
        row.status = str(payload.get("status") or row.status or "Pending").strip()[:20] or "Pending"
        row.assigned_user = assigned_user
        if deal:
            row.deal = deal
        row.updated_by = request.user
        row.save()
        return JsonResponse({"sales_order": _serialize_crm_sales_order(row)})

    order_id = _crm_resolve_unique_order_id(org, payload.get("order_id") or payload.get("orderId"))
    try:
        row = CrmSalesOrder.objects.create(
            organization=org,
            crm_reference_id=crm_reference_id,
            deal=deal,
            order_id=order_id,
            customer_name=customer_name[:180],
            company=str(payload.get("company") or "").strip()[:180],
            phone=str(payload.get("phone") or "").strip()[:40],
            amount=subtotal_amount,
            products=products_payload,
            quantity=quantity,
            price=price,
            tax=tax,
            total_amount=calculated_total_amount,
            payment_status=payment_status,
            paid_amount=float(paid_amount),
            payment_mode=str(payload.get("payment_mode") or payload.get("paymentMode") or "").strip()[:50] if can_edit_payment_details else "",
            payment_date=parse_date(str(payload.get("payment_date") or payload.get("paymentDate") or "").strip() or "") if can_edit_payment_details else None,
            transaction_id=str(payload.get("transaction_id") or payload.get("transactionId") or "").strip()[:100] if can_edit_payment_details else "",
            status=str(payload.get("status") or "Pending").strip()[:20] or "Pending",
            assigned_user=assigned_user,
            created_by=request.user,
            updated_by=request.user,
        )
    except IntegrityError:
        row = CrmSalesOrder.objects.create(
            organization=org,
            crm_reference_id=crm_reference_id,
            deal=deal,
            order_id=_crm_resolve_unique_order_id(org, ""),
            customer_name=customer_name[:180],
            company=str(payload.get("company") or "").strip()[:180],
            phone=str(payload.get("phone") or "").strip()[:40],
            amount=subtotal_amount,
            products=products_payload,
            quantity=quantity,
            price=price,
            tax=tax,
            total_amount=calculated_total_amount,
            payment_status=payment_status,
            paid_amount=float(paid_amount),
            payment_mode=str(payload.get("payment_mode") or payload.get("paymentMode") or "").strip()[:50] if can_edit_payment_details else "",
            payment_date=parse_date(str(payload.get("payment_date") or payload.get("paymentDate") or "").strip() or "") if can_edit_payment_details else None,
            transaction_id=str(payload.get("transaction_id") or payload.get("transactionId") or "").strip()[:100] if can_edit_payment_details else "",
            status=str(payload.get("status") or "Pending").strip()[:20] or "Pending",
            assigned_user=assigned_user,
            created_by=request.user,
            updated_by=request.user,
        )
    return JsonResponse({"sales_order": _serialize_crm_sales_order(row)}, status=201)


@require_http_methods(["POST"])
def crm_convert_to_sales_order(request, deal_id: int):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False}, status=401)
    org = _resolve_org(request.user, request)
    if not org:
        return JsonResponse({"detail": "organization_not_found"}, status=404)
    deal = CrmDeal.objects.filter(organization=org, id=deal_id, is_deleted=False).select_related("lead").first()
    if not deal:
        return JsonResponse({"detail": "deal_not_found"}, status=404)
    if not _crm_can_access_row(request.user, org, deal):
        return JsonResponse({"detail": "forbidden"}, status=403)
    if str(deal.stage or "").strip().lower() != "won" and str(deal.status or "").strip().lower() != "won":
        return JsonResponse({"detail": "deal_not_won"}, status=400)
    existing = CrmSalesOrder.objects.filter(organization=org, deal=deal, is_deleted=False).first()
    if existing:
        if not str(existing.crm_reference_id or "").strip():
            existing.crm_reference_id = str(deal.crm_reference_id or getattr(deal.lead, "crm_reference_id", "") or "").strip()[:32]
            existing.save(update_fields=["crm_reference_id", "updated_at"])
        return JsonResponse({"sales_order": _serialize_crm_sales_order(existing), "already_converted": True})
    with transaction.atomic():
        amount = _crm_to_decimal(deal.deal_value)
        customer_name = str(deal.lead.lead_name if deal.lead_id else deal.deal_name).strip()[:180]
        conversion_date = timezone.localdate().isoformat()
        next_order_id = _crm_order_id(org)
        row = CrmSalesOrder.objects.create(
            organization=org,
            deal=deal,
            crm_reference_id=str(deal.crm_reference_id or getattr(deal.lead, "crm_reference_id", "") or "").strip()[:32],
            order_id=next_order_id,
            customer_name=customer_name or "Customer",
            company=str(deal.company or "").strip()[:180],
            phone=str(deal.phone or "").strip()[:40],
            amount=amount,
            products={
                "items": [],
                "issueDate": conversion_date,
                "dueDate": conversion_date,
                "sourceDealId": str(deal.id),
            },
            quantity=1,
            price=amount,
            tax=Decimal("0"),
            total_amount=amount,
            payment_status="pending",
            paid_amount=0,
            payment_mode="",
            payment_date=None,
            transaction_id="",
            status="Pending",
            assigned_user=deal.assigned_user,
            created_by=request.user,
            updated_by=request.user,
        )
    return JsonResponse({"sales_order": _serialize_crm_sales_order(row)})
